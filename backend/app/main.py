import csv
import asyncio
import base64
import hashlib
import html
import hmac
import io
import json
import os
import re
import secrets
import sqlite3
import tempfile
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Annotated
from urllib.parse import parse_qsl, quote, urlencode, urlsplit, urlunsplit

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import and_, func, or_, text
from sqlalchemy.orm import Session, joinedload, object_session
from openpyxl import Workbook
import httpx
import qrcode
from qrcode.image.svg import SvgPathImage

from . import auth as auth_module
from .auth import (
    AuthContext,
    auth_public_config,
    require_access,
    require_admin,
)
from .branding import APP_BACKUP_NAME, APP_BRAND_NAME, APP_SHOW_BRAND_LOGO
from .team import clear_team_media, router as team_router
from .mobile import router as mobile_router
from .eda import purge_expired_assets, router as eda_router
from .features import FEATURE_EDA_ENABLED
from .purchases import router as purchases_router
from .risks import router as risks_router
from .team_projects import router as team_projects_router
from .codex_integration import prune_expired_operation_snapshots, router as codex_integration_router
from .database import Base, DATABASE_URL, SessionLocal, engine, get_db
from .models import (
    ActivityLog,
    AppMigration,
    AiKnowledgeCard,
    AiTask,
    Category,
    Component,
    ComponentIdentityRegistry,
    CustomLabelAsset,
    CustomLabelTemplate,
    CompetitionActivityLog,
    CompetitionAiResult,
    CompetitionComponentMarker,
    CompetitionInvite,
    CompetitionLibrary,
    CompetitionLibraryComponent,
    CompetitionLibraryMember,
    CompetitionPcb,
    EdaAsset,
    EdaComponentBinding,
    EdaFootprint,
    EdaSymbol,
    ImportRecord,
    InventoryLot,
    OrderImportBatch,
    OrderImportLine,
    Project,
    ProjectBoard,
    ProjectBomImportBatch,
    ProjectBomImportCandidate,
    ProjectBomImportRow,
    ProjectBomItem,
    ProjectBomSolderPoint,
    SupplierPart,
    User,
)
from .component_identity import (
    V060_COMPONENT_IDENTITIES,
    allocate_component_identity,
    archive_component_identity,
    identity_by_code,
    public_identity_out,
    run_component_identity_migration,
    seed_category_prefixes,
)
from .schemas import (
    AiComponentSearchRequest,
    AiComponentInfoRequest,
    AiClassifyRequest,
    AiExplainRequest,
    ImageImportPreviewRow,
    AiProjectPlanRequest,
    AiKnowledgeCardOut,
    ComponentAiAskOut,
    ComponentAiAskRequest,
    AiRefreshRequest,
    AiTaskOut,
    AiTaskSummary,
    ActivityLogOut,
    BomItemCreate,
    BomImportRowSelection,
    BomSolderPointBulkUpdate,
    BomSolderPointLossUpdate,
    BomSolderPointUpdate,
    BomItemStatusRequest,
    BomItemOut,
    BomItemUpdate,
    BomMatchCommitRequest,
    BomMatchCommitResult,
    CategoryOut,
    CategoryPrefixUpdate,
    ComponentCreate,
    ComponentAiOut,
    ComponentExportRequest,
    ComponentGroup,
    ComponentGroupPage,
    ComponentConsumeRequest,
    CustomLabelExportRequest,
    CustomLabelTemplateCreate,
    CustomLabelTemplateOut,
    CustomLabelTemplateUpdate,
    InventoryLotCreate,
    InventoryLotOut,
    ComponentList,
    LcscPreviewRequest,
    LcscPreviewResponse,
    ComponentOut,
    ComponentUsageRecordOut,
    ComponentUpdate,
    ImportCommitRequest,
    ImportCommitResult,
    ImportPreviewRow,
    OrderImportBatchOut,
    ProjectAiPlanRequest,
    ProjectAiConsultRequest,
    ProjectCreate,
    ProjectOut,
    ProjectUpdate,
    UsageEventRequest,
)
from .seed import seed_categories
from .services.bom_match import inspect_bom_fields, match_bom_rows, parse_bom_excel
from .services.excel_import import (
    component_values,
    create_import_record,
    find_duplicate,
    find_import_record,
    merge_component,
    parse_excel,
)
from .services.external_order_import import (
    parse_external_order,
    find_external_duplicate,
    find_external_import_record,
    external_import_record_payload,
)
from .services.mimo_ai import (
    MimoNotConfiguredError,
    MimoRequestError,
    assist_bom_matches,
    classify_component,
    analyze_bom,
    component_to_dict,
    component_info,
    component_question,
    image_import_preview,
    lcsc_search_url,
    lookup_lcsc_fallback,
    component_search,
    explain_component,
    organize_component,
    organize_lcsc_draft,
    project_plan,
    project_consult,
    search_component_candidates,
    search_empty_suggestions,
)
from .services.component_normalizer import (
    clean_component_name,
    clean_lcsc_keyword,
    normalize_component_values,
    normalize_tag_text,
)
from .services.lcsc_lookup import (
    LcscLookupError,
    exact_lcsc_source_present,
    fetch_lcsc_product,
    local_category_from_text,
    normalize_lcsc_number,
    official_product_draft,
    parse_lcsc_copy_text,
    parsed_copy_draft,
    product_url as lcsc_product_url,
)
from .services.category_governance import ai_category_allowed
from .services.inventory import (
    reserved_quantities as inventory_reserved_quantities,
    sort_components_by_value,
)
from .services.substitutions import substitution_suggestions_for_bom_items
from .services.stock_ledger import ensure_component_lot, migrate_legacy_inventory_lots, reconcile_component_lots, record_stock_delta
from .services.eda_storage import storage_root as eda_storage_root
from .labels import (
    custom_label_font_keys,
    category_package_summary_from_records,
    data_uri_from_bytes,
    is_standard_category_label_group,
    label_document,
    print_timestamp,
    render_component_label_pdf,
    render_component_label_sheet,
    render_basic_custom_label_pdf_items,
    render_custom_label_cards,
    render_standard_category_label_pdf_items,
    render_standard_category_label_cards,
    render_custom_label_sheet,
    sanitize_svg_markup,
)


ENABLE_API_DOCS = os.getenv("ENABLE_API_DOCS", "0") == "1"
APP_VERSION = "0.7.1"
PROCESS_STARTED_AT = time.monotonic()
PUBLIC_STATUS_RANK = {
    "operational": 0,
    "maintenance": 1,
    "unknown": 2,
    "degraded": 3,
    "outage": 4,
}


@asynccontextmanager
async def lifespan(_: FastAPI):
    startup_application()
    yield


app = FastAPI(
    title=APP_BRAND_NAME,
    version=APP_VERSION,
    docs_url="/docs" if ENABLE_API_DOCS else None,
    redoc_url="/redoc" if ENABLE_API_DOCS else None,
    openapi_url="/openapi.json" if ENABLE_API_DOCS else None,
    lifespan=lifespan,
)

allowed_origins = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://127.0.0.1:5173").split(",")
allowed_hosts = os.getenv("ALLOWED_HOSTS", "localhost,127.0.0.1").split(",")
app.add_middleware(
    TrustedHostMiddleware,
    allowed_hosts=[host.strip() for host in allowed_hosts if host.strip()],
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in allowed_origins if origin.strip()],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(team_router)
app.include_router(mobile_router)
app.include_router(eda_router)
app.include_router(purchases_router)
app.include_router(risks_router)
app.include_router(team_projects_router)
app.include_router(codex_integration_router)


@app.api_route(
    "/api/contest/{legacy_path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
)
def retired_contest_api(legacy_path: str):
    raise HTTPException(
        status_code=410,
        detail="旧团队版 API 已停用，请使用 /component-warehouse/api/team/",
    )


@app.api_route(
    "/api/contest",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
)
def retired_contest_api_root():
    raise HTTPException(
        status_code=410,
        detail="旧团队版 API 已停用，请使用 /component-warehouse/api/team/",
    )


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response: Response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
    response.headers.setdefault("Cache-Control", "no-store" if request.url.path.startswith("/api/") else "no-cache")
    return response


def client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",", 1)[0].strip()
    return request.client.host if request.client else "unknown"


AUTH_PROXY_LIMITS = {
    "captcha_ip": (60, 300),
    "sms_ip": (8, 600),
    "sms_phone": (3, 600),
    "login_ip": (30, 600),
    "login_phone": (8, 600),
    "reset_ip": (8, 600),
    "reset_phone": (3, 600),
}
_AUTH_PROXY_EVENTS: dict[str, list[float]] = {}
_AUTH_PROXY_LOCK = threading.Lock()


def enforce_auth_proxy_limit(bucket: str, key: str) -> None:
    limit, window = AUTH_PROXY_LIMITS[bucket]
    now = time.monotonic()
    event_key = f"{bucket}:{key}"
    cutoff = now - window
    with _AUTH_PROXY_LOCK:
        events = [stamp for stamp in _AUTH_PROXY_EVENTS.get(event_key, []) if stamp >= cutoff]
        if len(events) >= limit:
            _AUTH_PROXY_EVENTS[event_key] = events
            raise HTTPException(status_code=429, detail="登录请求过于频繁，请稍后再试")
        events.append(now)
        _AUTH_PROXY_EVENTS[event_key] = events
        if len(_AUTH_PROXY_EVENTS) > 5000:
            stale_cutoff = now - max(window for _, window in AUTH_PROXY_LIMITS.values())
            for stale_key in list(_AUTH_PROXY_EVENTS):
                _AUTH_PROXY_EVENTS[stale_key] = [stamp for stamp in _AUTH_PROXY_EVENTS[stale_key] if stamp >= stale_cutoff]
                if not _AUTH_PROXY_EVENTS[stale_key]:
                    _AUTH_PROXY_EVENTS.pop(stale_key, None)


def auth_proxy_phone(payload: dict) -> str:
    return re.sub(r"\D", "", str(payload.get("phone") or ""))[:20] or "unknown"


def auth_proxy_text(value: object, limit: int = 500) -> str:
    return str(value or "").strip()[:limit]


def legacy_login_retired() -> None:
    raise HTTPException(status_code=410, detail="旧登录方式已停用，请使用 WXY LAB 统一账号 SSO 登录")


SSO_STATE_COOKIE = "cw_sso_pending"
SSO_STATE_MAX_AGE_SECONDS = 600


def b64url(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def b64url_decode(value: str) -> bytes:
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode("ascii"))


def sso_cookie_secret() -> str:
    secret = auth_module.LOCAL_AUTH_SECRET or auth_module.ACCOUNT_CLIENT_SECRET
    if not secret:
        raise HTTPException(status_code=503, detail="SSO 状态密钥未配置，请设置 LOCAL_AUTH_SECRET")
    return secret


def request_public_origin(request: Request) -> str:
    redirect_parts = urlsplit(auth_module.ACCOUNT_SSO_REDIRECT_URI or "")
    if redirect_parts.scheme and redirect_parts.netloc:
        return f"{redirect_parts.scheme}://{redirect_parts.netloc}"
    proto = request.headers.get("x-forwarded-proto", request.url.scheme).split(",", 1)[0].strip() or "https"
    host = request.headers.get("x-forwarded-host", request.headers.get("host", request.url.netloc)).split(",", 1)[0].strip()
    return f"{proto}://{host}"


def safe_sso_return_to(value: object, request: Request) -> str:
    origin = request_public_origin(request)
    default = f"{origin}/component-warehouse/personal/"
    raw = auth_proxy_text(value, 1200)
    if not raw:
        return default
    parsed = urlsplit(raw)
    if parsed.scheme or parsed.netloc:
        allowed_origin = urlsplit(origin)
        if parsed.scheme != allowed_origin.scheme or parsed.netloc != allowed_origin.netloc:
            return default
        path = parsed.path
    else:
        path = raw if raw.startswith("/") else f"/{raw}"
        parsed = urlsplit(path)
        path = parsed.path
    if not (
        path.startswith("/component-warehouse/personal/")
        or path.startswith("/component-warehouse/team/")
    ):
        return default
    return urlunsplit((urlsplit(origin).scheme, urlsplit(origin).netloc, path, parsed.query, ""))


def sign_sso_cookie(payload: dict) -> str:
    body = b64url(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    signature = hmac.new(sso_cookie_secret().encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest()
    return f"{body}.{b64url(signature)}"


def parse_sso_cookie(value: str | None) -> dict | None:
    if not value:
        return None
    try:
        body, signature = value.split(".", 1)
        expected = hmac.new(sso_cookie_secret().encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest()
        if not hmac.compare_digest(b64url_decode(signature), expected):
            return None
        payload = json.loads(b64url_decode(body))
        if int(payload.get("iat") or 0) + SSO_STATE_MAX_AGE_SECONDS < int(time.time()):
            return None
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def clear_sso_cookie(response: Response) -> None:
    response.delete_cookie(SSO_STATE_COOKIE, path="/component-warehouse")


def set_sso_cookie(response: Response, request: Request, payload: dict) -> None:
    response.set_cookie(
        SSO_STATE_COOKIE,
        sign_sso_cookie(payload),
        max_age=SSO_STATE_MAX_AGE_SECONDS,
        httponly=True,
        secure=request.headers.get("x-forwarded-proto", request.url.scheme).split(",", 1)[0].strip() == "https",
        samesite="lax",
        path="/component-warehouse",
    )


def sso_code_challenge(verifier: str) -> str:
    return b64url(hashlib.sha256(verifier.encode("ascii")).digest())


def build_sso_authorize_url(redirect_uri: str, state: str, code_challenge: str) -> str:
    parts = urlsplit(auth_module.ACCOUNT_SSO_AUTHORIZE_URL)
    params = dict(parse_qsl(parts.query, keep_blank_values=True))
    params.update({
        "client_id": auth_module.ACCOUNT_WEB_CLIENT_ID,
        "redirect_uri": redirect_uri,
        "state": state,
        "code_challenge": code_challenge,
        "code_challenge_method": "S256",
    })
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(params), parts.fragment))


def account_sso_start(payload: dict, request: Request, response: Response) -> dict:
    if auth_module.AUTH_MODE == "local-password":
        raise HTTPException(status_code=404, detail="当前账号模式不需要统一账号 SSO")
    if not auth_module.ACCOUNT_SSO_AUTHORIZE_URL or not auth_module.ACCOUNT_SSO_TOKEN_URL:
        raise HTTPException(status_code=503, detail="统一账号 SSO 未配置")
    redirect_uri = (
        auth_proxy_text(payload.get("redirectUri"), 800)
        or auth_module.ACCOUNT_SSO_REDIRECT_URI
        or f"{request_public_origin(request)}/component-warehouse/personal/auth/callback"
    )
    state = secrets.token_urlsafe(24)
    code_verifier = secrets.token_urlsafe(48)
    return_to = safe_sso_return_to(payload.get("returnTo"), request)
    set_sso_cookie(response, request, {
        "state": state,
        "code_verifier": code_verifier,
        "redirect_uri": redirect_uri,
        "return_to": return_to,
        "iat": int(time.time()),
    })
    return {
        "authorizeUrl": build_sso_authorize_url(redirect_uri, state, sso_code_challenge(code_verifier)),
        "state": state,
        "redirectUri": redirect_uri,
        "returnTo": return_to,
    }


async def account_v1_proxy_request(method: str, path: str, payload: dict | None = None, token: str | None = None) -> dict:
    if auth_module.AUTH_MODE == "local-password":
        raise HTTPException(status_code=404, detail="当前账号模式不需要统一账号代理")
    if not auth_module.ACCOUNT_BASE_URL:
        raise HTTPException(status_code=503, detail="统一账号地址未配置")
    headers = {
        "X-Account-Client-Id": auth_module.ACCOUNT_WEB_CLIENT_ID,
        auth_module.LEGACY_CLIENT_ID_HEADER: auth_module.ACCOUNT_WEB_CLIENT_ID,
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        async with httpx.AsyncClient(timeout=auth_module.AUTH_HTTP_TIMEOUT_SECONDS) as client:
            response = await client.request(method, f"{auth_module.ACCOUNT_BASE_URL}{path}", headers=headers, json=payload)
    except httpx.RequestError as error:
        raise HTTPException(status_code=503, detail="统一账号暂时不可用，请稍后重试") from error
    try:
        data = response.json()
    except ValueError as error:
        raise HTTPException(status_code=502, detail="统一账号返回格式异常") from error
    if response.status_code >= 400:
        message = data.get("error", {}).get("message") if isinstance(data, dict) else ""
        raise HTTPException(status_code=response.status_code, detail=message or "统一账号请求失败")
    return data


async def account_sso_token_request(payload: dict, request: Request | None = None) -> dict:
    if auth_module.AUTH_MODE == "local-password":
        raise HTTPException(status_code=404, detail="当前账号模式不需要统一账号 SSO")
    if not auth_module.ACCOUNT_SSO_TOKEN_URL:
        raise HTTPException(status_code=503, detail="统一账号 SSO 未配置")
    cookie_payload = parse_sso_cookie(request.cookies.get(SSO_STATE_COOKIE) if request else None)
    state = auth_proxy_text(payload.get("state"), 200)
    if cookie_payload:
        if not state or state != auth_proxy_text(cookie_payload.get("state"), 200):
            raise HTTPException(status_code=400, detail="统一账号登录状态校验失败，请重新登录")
    body = {
        "client_id": auth_module.ACCOUNT_WEB_CLIENT_ID,
        "redirect_uri": (
            auth_proxy_text(payload.get("redirectUri"), 800)
            or auth_proxy_text(cookie_payload.get("redirect_uri") if cookie_payload else "", 800)
            or auth_module.ACCOUNT_SSO_REDIRECT_URI
        ),
        "code": auth_proxy_text(payload.get("code"), 500),
        "code_verifier": (
            auth_proxy_text(payload.get("codeVerifier"), 200)
            or auth_proxy_text(cookie_payload.get("code_verifier") if cookie_payload else "", 200)
        ),
    }
    if not body["redirect_uri"] or not body["code"] or not body["code_verifier"]:
        raise HTTPException(status_code=400, detail="SSO 登录参数不完整")
    try:
        async with httpx.AsyncClient(timeout=auth_module.AUTH_HTTP_TIMEOUT_SECONDS) as client:
            response = await client.post(auth_module.ACCOUNT_SSO_TOKEN_URL, json=body)
    except httpx.RequestError as error:
        raise HTTPException(status_code=503, detail="统一账号 SSO 暂时不可用，请稍后重试") from error
    try:
        data = response.json()
    except ValueError as error:
        raise HTTPException(status_code=502, detail="统一账号 SSO 返回格式异常") from error
    if response.status_code >= 400:
        message = data.get("error", {}).get("message") if isinstance(data, dict) else ""
        raise HTTPException(status_code=response.status_code, detail=message or "统一账号 SSO 登录失败")
    if cookie_payload and cookie_payload.get("return_to"):
        data["returnTo"] = cookie_payload["return_to"]
    return data


def ensure_sqlite_columns(connection, table: str, columns: dict[str, str]) -> None:
    existing = [row[1] for row in connection.execute(text(f"PRAGMA table_info({table})"))]
    for name, definition in columns.items():
        if name not in existing:
            connection.execute(text(f"ALTER TABLE {table} ADD COLUMN {name} {definition}"))


def ensure_database_schema(connection) -> None:
    ensure_sqlite_columns(
        connection,
        "categories",
        {
            "color": "VARCHAR(40) DEFAULT '#eef6ff'",
            "code_prefix": "VARCHAR(3)",
            "code_prefix_locked": "BOOLEAN DEFAULT 0",
        },
    )
    ensure_sqlite_columns(
        connection,
        "users",
        {
            "account_id": "VARCHAR(36)",
            "nickname": "VARCHAR(80)",
            "avatar_url": "VARCHAR(500)",
            "password_hash": "VARCHAR(300)",
            "is_admin": "BOOLEAN DEFAULT 0",
            "last_login_at": "DATETIME",
            "updated_at": "DATETIME",
        },
    )
    ensure_sqlite_columns(
        connection,
        "components",
        {
            "owner_user_id": "INTEGER",
            "warehouse_code": "VARCHAR(80)",
            "manufacturer": "VARCHAR(200)",
            "description": "TEXT",
            "ai_summary": "TEXT",
            "ai_usage": "TEXT",
            "ai_risk_notes": "TEXT",
            "ai_pcb_notes": "TEXT",
            "ai_substitutes": "TEXT",
            "ai_tags": "VARCHAR(500)",
            "source_title": "TEXT",
            "ai_confidence": "VARCHAR(40)",
            "ai_cache_key": "VARCHAR(80)",
            "ai_status": "VARCHAR(40) DEFAULT 'pending'",
            "ai_error": "TEXT",
            "ai_updated_at": "DATETIME",
            "first_stocked_at": "DATETIME",
            "last_stocked_at": "DATETIME",
            "last_outbound_at": "DATETIME",
            "part_family": "VARCHAR(40) DEFAULT 'component'",
            "count_mode": "VARCHAR(40) DEFAULT 'exact'",
            "normalized_spec": "VARCHAR(160)",
            "is_hand_solder_friendly": "BOOLEAN DEFAULT 0",
            "is_power_component": "BOOLEAN DEFAULT 0",
            "is_signal_component": "BOOLEAN DEFAULT 0",
            "is_high_current": "BOOLEAN DEFAULT 0",
            "is_high_voltage": "BOOLEAN DEFAULT 0",
            "is_common": "BOOLEAN DEFAULT 0",
            "revoked_at": "DATETIME",
            "competition_name": "VARCHAR(120)",
            "competition_category": "VARCHAR(80)",
            "priority": "VARCHAR(20)",
            "target_quantity": "INTEGER DEFAULT 0",
            "safety_quantity": "INTEGER DEFAULT 0",
            "low_stock_exempt": "BOOLEAN DEFAULT 0",
            "manual_stock_status": "VARCHAR(40)",
            "usability_status": "VARCHAR(40)",
            "verify_status": "VARCHAR(40)",
            "location_code": "VARCHAR(120)",
            "buy_url": "VARCHAR(500)",
        },
    )
    ensure_sqlite_columns(connection, "project_bom_items", {"status": "VARCHAR(40) DEFAULT 'reserved'"})
    ensure_sqlite_columns(
        connection,
        "projects",
        {
            "scope_type": "VARCHAR(20) DEFAULT 'personal'",
            "owner_user_id": "INTEGER",
            "team_library_id": "VARCHAR(36)",
        },
    )
    ensure_sqlite_columns(connection, "order_import_batches", {"owner_user_id": "INTEGER"})
    ensure_sqlite_columns(connection, "order_import_lines", {"owner_user_id": "INTEGER"})
    ensure_sqlite_columns(connection, "import_records", {"owner_user_id": "INTEGER"})
    ensure_sqlite_columns(connection, "activity_logs", {"owner_user_id": "INTEGER"})
    ensure_sqlite_columns(
        connection,
        "competition_library_members",
        {
            "joined_invite_id": "VARCHAR(36)",
            "blocked_invite_id": "VARCHAR(36)",
        },
    )
    ensure_sqlite_columns(
        connection,
        "competition_library_components",
        {
            "source_user_id": "INTEGER",
            "sync_status": "VARCHAR(24) DEFAULT 'live'",
            "warehouse_code_snapshot": "VARCHAR(80)",
            "frozen_snapshot_json": "TEXT",
        },
    )
    connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_users_account_id_unique ON users(account_id) WHERE account_id IS NOT NULL AND account_id != ''"))
    connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_categories_code_prefix_unique ON categories(code_prefix) WHERE code_prefix IS NOT NULL AND code_prefix != ''"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_team_markers_component ON competition_component_markers(library_id, component_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_team_markers_filter ON competition_component_markers(library_id, category, color, flagged)"))
    ensure_sqlite_columns(
        connection,
        "project_bom_solder_points",
        {
            "board_id": "INTEGER",
            "stock_applied": "BOOLEAN DEFAULT 0",
            "lost": "BOOLEAN DEFAULT 0",
            "lost_at": "DATETIME",
            "loss_stock_applied": "BOOLEAN DEFAULT 0",
            "loss_note": "TEXT",
        },
    )
    ensure_sqlite_columns(connection, "ai_tasks", {"next_attempt_at": "DATETIME"})
    ensure_sqlite_columns(
        connection,
        "projects",
        {
            "project_code": "VARCHAR(80)",
            "ai_bom_analysis": "TEXT",
            "ai_bom_cache_key": "VARCHAR(80)",
            "ai_bom_updated_at": "DATETIME",
            "bom_match_total": "INTEGER DEFAULT 0",
            "bom_match_matched": "INTEGER DEFAULT 0",
            "bom_match_review": "INTEGER DEFAULT 0",
            "bom_match_missing": "INTEGER DEFAULT 0",
            "bom_match_missing_items": "TEXT",
            "bom_match_rows": "TEXT",
            "bom_match_updated_at": "DATETIME",
        },
    )
    ensure_sqlite_columns(
        connection,
        "project_bom_import_batches",
        {
            "source_sha256": "VARCHAR(64)",
            "field_mapping_json": "TEXT",
            "analysis_json": "TEXT",
        },
    )
    ensure_sqlite_columns(connection, "import_records", {"batch_id": "INTEGER"})
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_project_bom_items_status ON project_bom_items(status)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_import_records_batch_id ON import_records(batch_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_ai_status ON components(ai_status)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_ai_cache_key ON components(ai_cache_key)"))
    connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_components_warehouse_code_unique ON components(warehouse_code) WHERE warehouse_code IS NOT NULL AND warehouse_code != ''"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_owner_user_id ON components(owner_user_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_revoked_at ON components(revoked_at)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_projects_owner_user_id ON projects(owner_user_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_projects_team_library_id ON projects(team_library_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_projects_scope_type ON projects(scope_type)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_manufacturer ON components(manufacturer)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_bom_import_batches_sha256 ON project_bom_import_batches(source_sha256)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_competition_name ON components(competition_name)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_competition_category ON components(competition_category)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_priority ON components(priority)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_components_location_code ON components(location_code)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_ai_tasks_next_attempt_at ON ai_tasks(next_attempt_at)"))
    connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_projects_project_code_unique ON projects(project_code) WHERE project_code IS NOT NULL AND project_code != ''"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_project_boards_project_id ON project_boards(project_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_project_bom_solder_points_board_id ON project_bom_solder_points(board_id)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_contest_members_user_status ON competition_library_members(user_id, status)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_contest_components_library_name ON competition_library_components(library_id, name)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_contest_components_source_user ON competition_library_components(source_user_id, sync_status)"))
    connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_contest_components_library_cw_unique ON competition_library_components(library_id, cw_component_id) WHERE cw_component_id IS NOT NULL"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_contest_logs_library_created ON competition_activity_logs(library_id, created_at)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_custom_label_templates_personal ON custom_label_templates(scope_type, owner_user_id, status, updated_at)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_custom_label_templates_team ON custom_label_templates(scope_type, team_library_id, status, updated_at)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_custom_label_assets_template ON custom_label_assets(template_id, created_at)"))


V04_ACCOUNT_MIGRATION = "v0.4.0-account-owner"
V04_LEGACY_COMPETITION_CLEANUP = "v0.4.0-stop-legacy-competition-fields"
V041_CONTEST_LIVE_INVENTORY = "v0.4.1-contest-live-inventory-and-bom-release"
V070_EDA_ENGINEERING = "v0.7.0-eda-engineering"
V072_LCSC_SOURCE_NORMALIZATION = "v0.7.2-lcsc-source-normalization"
V073_ADMIN_DEFAULTS = "v0.7.3-admin-defaults"


def sqlite_database_path() -> Path | None:
    if not DATABASE_URL.startswith("sqlite:///"):
        return None
    path = Path(DATABASE_URL.replace("sqlite:///", "", 1))
    return path if path.is_absolute() else Path.cwd() / path


def ensure_v04_migration_backup() -> Path | None:
    database_path = sqlite_database_path()
    if not database_path or not database_path.exists():
        return None
    source = sqlite3.connect(str(database_path))
    try:
        has_migrations = source.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='app_migrations'"
        ).fetchone()
        if has_migrations:
            applied = source.execute(
                "SELECT 1 FROM app_migrations WHERE key = ?",
                (V04_ACCOUNT_MIGRATION,),
            ).fetchone()
            if applied:
                return None
        backup_dir = database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        target_path = backup_dir / f"pre-v0.4-account-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.db"
        target = sqlite3.connect(str(target_path))
        try:
            source.backup(target)
        finally:
            target.close()
        return target_path
    finally:
        source.close()


def ensure_v041_migration_backup() -> Path | None:
    database_path = sqlite_database_path()
    if not database_path or not database_path.exists():
        return None
    source = sqlite3.connect(str(database_path))
    try:
        has_migrations = source.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='app_migrations'"
        ).fetchone()
        if has_migrations and source.execute(
            "SELECT 1 FROM app_migrations WHERE key = ?",
            (V041_CONTEST_LIVE_INVENTORY,),
        ).fetchone():
            return None
        backup_dir = database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        target_path = backup_dir / f"pre-v0.4.1-live-inventory-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.db"
        target = sqlite3.connect(str(target_path))
        try:
            source.backup(target)
        finally:
            target.close()
        return target_path
    finally:
        source.close()


def ensure_v060_migration_backup() -> Path | None:
    database_path = sqlite_database_path()
    if not database_path or not database_path.exists():
        return None
    source = sqlite3.connect(str(database_path))
    try:
        has_migrations = source.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='app_migrations'"
        ).fetchone()
        if has_migrations and source.execute(
            "SELECT 1 FROM app_migrations WHERE key = ?",
            (V060_COMPONENT_IDENTITIES,),
        ).fetchone():
            return None
        backup_dir = database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        target_path = backup_dir / f"pre-v0.6.0-identities-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.db"
        target = sqlite3.connect(str(target_path))
        try:
            source.backup(target)
        finally:
            target.close()
        return target_path
    finally:
        source.close()


def ensure_v070_migration_backup() -> Path | None:
    database_path = sqlite_database_path()
    if not database_path or not database_path.exists():
        return None
    source = sqlite3.connect(str(database_path))
    try:
        has_migrations = source.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='app_migrations'"
        ).fetchone()
        if has_migrations and source.execute(
            "SELECT 1 FROM app_migrations WHERE key = ?",
            (V070_EDA_ENGINEERING,),
        ).fetchone():
            return None
        backup_dir = database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        target_path = backup_dir / f"pre-v0.7.0-eda-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.db"
        target = sqlite3.connect(str(target_path))
        try:
            source.backup(target)
        finally:
            target.close()
        return target_path
    finally:
        source.close()


def remove_legacy_component_lcsc_unique() -> bool:
    database_path = sqlite_database_path()
    if not database_path or not database_path.exists():
        return False
    connection = sqlite3.connect(str(database_path))
    try:
        table_row = connection.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='components'"
        ).fetchone()
        table_sql = str(table_row[0] if table_row else "")
        if not re.search(
            r"CONSTRAINT\s+uq_components_lcsc_number\s+UNIQUE\s*\(\s*lcsc_number\s*\)",
            table_sql,
            flags=re.IGNORECASE,
        ):
            return False
        rebuilt_sql = re.sub(
            r",\s*CONSTRAINT\s+uq_components_lcsc_number\s+UNIQUE\s*\(\s*lcsc_number\s*\)",
            "",
            table_sql,
            count=1,
            flags=re.IGNORECASE,
        )
        rebuilt_sql = re.sub(
            r"^\s*CREATE\s+TABLE\s+components\b",
            "CREATE TABLE components_v041_rebuild",
            rebuilt_sql,
            count=1,
            flags=re.IGNORECASE,
        )
        columns = [
            row[1]
            for row in connection.execute("PRAGMA table_info(components)").fetchall()
        ]
        quoted_columns = ", ".join(f'"{name}"' for name in columns)
        index_sql = [
            row[0]
            for row in connection.execute(
                "SELECT sql FROM sqlite_master "
                "WHERE type='index' AND tbl_name='components' AND sql IS NOT NULL"
            ).fetchall()
            if row[0]
        ]
        connection.execute("PRAGMA foreign_keys=ON")
        violations_before = {
            tuple(row) for row in connection.execute("PRAGMA foreign_key_check").fetchall()
        }
        connection.execute("PRAGMA foreign_keys=OFF")
        connection.execute("BEGIN IMMEDIATE")
        connection.execute("DROP TABLE IF EXISTS components_v041_rebuild")
        connection.execute(rebuilt_sql)
        connection.execute(
            f"INSERT INTO components_v041_rebuild ({quoted_columns}) "
            f"SELECT {quoted_columns} FROM components"
        )
        connection.execute("DROP TABLE components")
        connection.execute(
            "ALTER TABLE components_v041_rebuild RENAME TO components"
        )
        for statement in index_sql:
            connection.execute(statement)
        connection.commit()
        connection.execute("PRAGMA foreign_keys=ON")
        violations_after = {
            tuple(row) for row in connection.execute("PRAGMA foreign_key_check").fetchall()
        }
        new_violations = violations_after - violations_before
        if new_violations:
            raise RuntimeError(
                f"移除旧立创唯一约束后产生了新外键问题：{list(new_violations)[:5]}"
            )
        return True
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def ensure_migration_owner_user(db: Session) -> User:
    user = db.get(User, 1)
    if not user:
        user = User(
            id=1,
            phone="local-user-1",
            nickname="默认用户 1",
            is_admin=False,
        )
        db.add(user)
        db.flush()
    user.password_hash = None
    user.is_admin = False
    return user


def migrate_all_owner_data(db: Session, user_id: int) -> None:
    owner_models = [
        Component,
        Project,
        OrderImportBatch,
        OrderImportLine,
        ImportRecord,
        ActivityLog,
    ]
    for model in owner_models:
        db.query(model).update({model.owner_user_id: user_id}, synchronize_session=False)


def run_v04_account_migration(db: Session) -> None:
    changed = False
    if not db.get(AppMigration, V04_ACCOUNT_MIGRATION):
        owner = ensure_migration_owner_user(db)
        migrate_all_owner_data(db, owner.id)
        for table_name in [
            "user_tokens",
            "password_reset_codes",
            "component_substitutes",
            "purchase_records",
            "module_boards",
            "packing_items",
            "test_records",
            "stock_movements",
        ]:
            db.execute(text(f"DROP TABLE IF EXISTS {table_name}"))
        db.add(
            AppMigration(
                key=V04_ACCOUNT_MIGRATION,
                detail="现有数据归属到默认账号 ID 1，并移除旧本地认证与旧竞赛空表。",
            )
        )
        changed = True
    if not db.get(AppMigration, V04_LEGACY_COMPETITION_CLEANUP):
        db.query(Component).update(
            {
                Component.competition_name: None,
                Component.competition_category: None,
                Component.priority: None,
                Component.target_quantity: 0,
                Component.safety_quantity: 0,
                Component.manual_stock_status: None,
                Component.usability_status: None,
                Component.verify_status: None,
                Component.location_code: None,
                Component.buy_url: None,
            },
            synchronize_session=False,
        )
        db.add(
            AppMigration(
                key=V04_LEGACY_COMPETITION_CLEANUP,
                detail="保留 components 旧比赛列但清空并停用，竞赛数据改用独立团队表。",
            )
        )
        changed = True
    if changed:
        db.commit()


def run_v041_inventory_migration(db: Session) -> None:
    if db.get(AppMigration, V041_CONTEST_LIVE_INVENTORY):
        return
    for item in db.query(CompetitionLibraryComponent).all():
        component = db.get(Component, item.cw_component_id) if item.cw_component_id else None
        source_user_id = component.owner_user_id if component else item.created_by_user_id
        item.source_user_id = source_user_id
        if component:
            item.quantity = int(component.quantity or 0)
            item.warehouse_code_snapshot = component.warehouse_code
            item.name = component.name
            item.model = component.model
            item.lcsc_number = component.lcsc_number
            reserved = inventory_reserved_quantities(db, [component.id]).get(component.id, 0)
            item.frozen_snapshot_json = json.dumps(
                component_out(component, reserved),
                ensure_ascii=False,
                default=str,
            )
            active_member = (
                db.query(CompetitionLibraryMember)
                .filter(
                    CompetitionLibraryMember.library_id == item.library_id,
                    CompetitionLibraryMember.user_id == source_user_id,
                    CompetitionLibraryMember.status == "active",
                )
                .first()
            )
            item.sync_status = "live" if active_member else "frozen"
            if not active_member:
                item.cw_component_id = None
        else:
            item.sync_status = "frozen"
            item.warehouse_code_snapshot = item.warehouse_code_snapshot or f"EX-{item.id[:8].upper()}"
    db.query(ProjectBomItem).filter(ProjectBomItem.status == "released").update(
        {ProjectBomItem.status: "reserved"},
        synchronize_session=False,
    )
    db.add(
        AppMigration(
            key=V041_CONTEST_LIVE_INVENTORY,
            detail="移除旧立创 ID 全局唯一约束；团队器件改为来源账号个人版库存实时镜像；离队器件冻结快照；旧 released BOM 恢复预占。",
        )
    )
    db.commit()


def run_v070_eda_migration(db: Session) -> None:
    if db.get(AppMigration, V070_EDA_ENGINEERING):
        return
    db.query(Project).filter(
        or_(Project.scope_type.is_(None), Project.scope_type == "")
    ).update({Project.scope_type: "personal"}, synchronize_session=False)
    db.query(CompetitionLibraryMember).filter(
        CompetitionLibraryMember.role == "member"
    ).update({CompetitionLibraryMember.role: "editor"}, synchronize_session=False)
    db.execute(
        text(
            "UPDATE activity_logs SET project_id = NULL "
            "WHERE project_id IS NOT NULL "
            "AND NOT EXISTS (SELECT 1 FROM projects WHERE projects.id = activity_logs.project_id)"
        )
    )
    migrated_lots = migrate_legacy_inventory_lots(db)
    db.add(
        AppMigration(
            key=V070_EDA_ENGINEERING,
            detail=(
                "新增 EDA 文件、库版本、Symbol、Footprint、供应商料号、采购、风险、库存批次与同步令牌；"
                f"为 {migrated_lots} 个旧元器件建立兼容库存批次；团队 member 迁移为 editor；修复失效项目日志外键。"
            ),
        )
    )
    db.commit()


def run_v072_lcsc_source_normalization(db: Session) -> None:
    if db.get(AppMigration, V072_LCSC_SOURCE_NORMALIZATION):
        return
    db.execute(
        text(
            "UPDATE components SET source = '立创' "
            "WHERE source IS NOT NULL "
            "AND (source LIKE '%立创%' OR source LIKE '%LCSC%' OR source LIKE '%lcsc%')"
        )
    )
    db.add(
        AppMigration(
            key=V072_LCSC_SOURCE_NORMALIZATION,
            detail="规范历史立创来源显示为短标签“立创”，保留 LCSC 编号、导入批次和库存批次追溯。",
        )
    )
    db.commit()


def run_v073_admin_defaults(db: Session) -> None:
    if db.get(AppMigration, V073_ADMIN_DEFAULTS):
        return
    admin_phone = os.getenv(
        "ADMIN_PHONE_NUMBERS",
        os.getenv("LEGACY_COMPONENT_ADMIN_PHONE", ""),
    ).split(",", 1)[0].strip()
    changed = 0
    if admin_phone:
        changed = (
            db.query(User)
            .filter(User.phone == admin_phone)
            .update({User.is_admin: True}, synchronize_session=False)
        )
    db.add(
        AppMigration(
            key=V073_ADMIN_DEFAULTS,
            detail=f"将 {admin_phone or '指定账号'} 设置为元器件系统管理员；当前匹配用户 {changed} 个。",
        )
    )
    db.commit()


def startup_application():
    ensure_v04_migration_backup()
    ensure_v041_migration_backup()
    ensure_v060_migration_backup()
    ensure_v070_migration_backup()
    remove_legacy_component_lcsc_unique()
    Base.metadata.create_all(bind=engine)
    with engine.begin() as connection:
        ensure_database_schema(connection)
    db = next(get_db())
    try:
        run_v04_account_migration(db)
        run_v041_inventory_migration(db)
        run_v070_eda_migration(db)
        run_v072_lcsc_source_normalization(db)
        run_v073_admin_defaults(db)
        prune_expired_operation_snapshots(db)
        if FEATURE_EDA_ENABLED:
            purge_expired_assets(db)
        seed_categories(db)
        ensure_category(db, "连接件", "#e8fff8")
        ensure_category(db, "时钟源", "#eef2ff")
        seed_category_prefixes(db)
        run_component_identity_migration(db)
        ensure_project_codes(db)
        reconcile_pending_purchase_from_import_records(db)
        ensure_project_boards(db)
        ensure_bom_solder_points(db)
        resolve_superseded_ai_failures(db)
        enqueue_organize_component_tasks(db, force=False, limit=80)
        enqueue_missing_component_ai_tasks(db, include_failed=True)
        db.commit()
        ensure_auto_backup()
    finally:
        db.close()
    ensure_ai_worker()


Protected = Annotated[AuthContext, Depends(require_access)]
AdminProtected = Annotated[AuthContext, Depends(require_admin)]


def owner_id(auth: AuthContext | None) -> int | None:
    return auth.user_id if auth else None


def set_owner(record, auth: AuthContext | None) -> None:
    if hasattr(record, "owner_user_id") and owner_id(auth):
        record.owner_user_id = owner_id(auth)


def filter_owner(query, model, auth: AuthContext | None):
    if not auth:
        return query
    if not hasattr(model, "owner_user_id"):
        return query
    query = query.filter(model.owner_user_id == auth.user_id)
    if model is Component:
        query = query.filter(Component.revoked_at.is_(None))
    return query


def assert_owned(record, auth: AuthContext | None, detail: str = "Not found") -> None:
    if not record:
        raise HTTPException(status_code=404, detail=detail)
    if isinstance(record, Component) and record.revoked_at is not None:
        raise HTTPException(status_code=404, detail=detail)
    if auth and hasattr(record, "owner_user_id") and record.owner_user_id != auth.user_id:
        raise HTTPException(status_code=404, detail=detail)


def require_project_access(db: Session, project_id: int, auth: AuthContext | None, detail: str = "Project not found") -> Project:
    project = db.get(Project, project_id)
    assert_owned(project, auth, detail)
    return project
AI_ANALYSIS_VERSION = "component-ai-v2-design-insights"
AI_TASK_MAX_RETRIES = int(os.getenv("AI_TASK_MAX_RETRIES", "8"))
AI_TASK_RETRY_BASE_SECONDS = int(os.getenv("AI_TASK_RETRY_BASE_SECONDS", "45"))
AI_TASK_RETRY_MAX_SECONDS = int(os.getenv("AI_TASK_RETRY_MAX_SECONDS", "1800"))
AI_AUTO_REFRESH_ENABLED = os.getenv("AI_AUTO_REFRESH_ENABLED", "1") == "1"
AI_AUTO_REFRESH_INTERVAL_HOURS = int(os.getenv("AI_AUTO_REFRESH_INTERVAL_HOURS", "12"))
AI_AUTO_REFRESH_MAX_PER_RUN = int(os.getenv("AI_AUTO_REFRESH_MAX_PER_RUN", "5"))
AI_AUTO_REFRESH_AFTER_DAYS = int(os.getenv("AI_AUTO_REFRESH_AFTER_DAYS", "30"))
AI_LAST_AUTO_REFRESH_AT: datetime | None = None
WAREHOUSE_CODE_PREFIX = (os.getenv("WAREHOUSE_CODE_PREFIX", "CW").strip().upper() or "CW")[:2].ljust(2, "X")
WAREHOUSE_CODE_WIDTH = max(6, min(18, int(os.getenv("WAREHOUSE_CODE_WIDTH", "8"))))
PROJECT_CODE_PREFIX = (os.getenv("PROJECT_CODE_PREFIX", "PJ").strip().upper() or "PJ")[:3]
PROJECT_CODE_WIDTH = max(4, min(18, int(os.getenv("PROJECT_CODE_WIDTH", "8"))))
PUBLIC_PERSONAL_BASE_URL = os.getenv(
    "PUBLIC_PERSONAL_BASE_URL",
    "http://localhost:8080/component-warehouse/personal",
).strip().rstrip("/")
CUSTOM_LABEL_STORAGE_ROOT = Path(os.getenv("CUSTOM_LABEL_STORAGE_ROOT", "./data/custom-labels"))
CUSTOM_LABEL_ALLOWED_MIME = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/svg+xml": ".svg",
}
CUSTOM_LABEL_MAX_IMAGE_BYTES = int(os.getenv("CUSTOM_LABEL_MAX_IMAGE_BYTES", str(5 * 1024 * 1024)))
CUSTOM_LABEL_MAX_SVG_BYTES = int(os.getenv("CUSTOM_LABEL_MAX_SVG_BYTES", str(256 * 1024)))
FOREGROUND_AI_WORKERS = max(1, min(4, int(os.getenv("FOREGROUND_AI_WORKERS", "2"))))
FOREGROUND_AI_EXECUTOR = ThreadPoolExecutor(max_workers=FOREGROUND_AI_WORKERS, thread_name_prefix="cw-foreground-ai")
BACKUP_KEEP_COUNT = max(1, min(365, int(os.getenv("BACKUP_KEEP_COUNT", "30"))))
BACKUP_AUTO_INTERVAL_HOURS = max(1, min(168, int(os.getenv("BACKUP_AUTO_INTERVAL_HOURS", "24"))))
BACKUP_MAX_UPLOAD_BYTES = max(10, min(1024, int(os.getenv("BACKUP_MAX_UPLOAD_MB", "512")))) * 1024 * 1024
BACKUP_LOCK = threading.Lock()


async def run_foreground_ai(func, *args):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(FOREGROUND_AI_EXECUTOR, lambda: func(*args))


def infer_small_part_fields(values: dict) -> dict:
    values["part_family"] = values.get("part_family") or "component"
    values["count_mode"] = values.get("count_mode") or "exact"
    return values


def warehouse_prefix_for_component(component: Component | None = None, values: dict | None = None) -> str:
    source = str((values or {}).get("source") or getattr(component, "source", "") or "")
    status = str((values or {}).get("status") or getattr(component, "status", "") or "")
    location = str((values or {}).get("location") or getattr(component, "location", "") or "")
    if "BOM 待采购库" in source or status == "pending_purchase" or location == "待采购":
        return "BP"
    if "立创" in source:
        return "LC"
    if "外部订单" in source:
        return "EX"
    if "图片识别" in source:
        return "IM"
    if "手动" in source:
        return "MN"
    return WAREHOUSE_CODE_PREFIX


def make_warehouse_code(component_id: int, prefix: str | None = None) -> str:
    code_prefix = ((prefix or WAREHOUSE_CODE_PREFIX).strip().upper() or WAREHOUSE_CODE_PREFIX)[:2].ljust(2, "X")
    return f"{code_prefix}-{int(component_id):0{WAREHOUSE_CODE_WIDTH}d}"


def make_project_code(project_id: int) -> str:
    return f"{PROJECT_CODE_PREFIX}-{int(project_id):0{PROJECT_CODE_WIDTH}d}"


def normalize_warehouse_code(value: str | None) -> str | None:
    text_value = str(value or "").strip()
    return text_value[:80] or None


def normalize_project_code(value: str | None) -> str | None:
    text_value = str(value or "").strip()
    return text_value[:80] or None


def assert_unique_warehouse_code(db: Session, code: str | None, component_id: int | None = None) -> None:
    normalized = normalize_warehouse_code(code)
    if not normalized:
        return
    if component_id:
        component = db.get(Component, component_id)
        if component and component.warehouse_code == normalized:
            return
    raise HTTPException(status_code=409, detail="器件 ID 由系统分配，创建后不可修改")


def assert_unique_project_code(db: Session, code: str | None, project_id: int | None = None) -> None:
    normalized = normalize_project_code(code)
    if not normalized:
        return
    query = db.query(Project).filter(Project.project_code == normalized)
    if project_id:
        query = query.filter(Project.id != project_id)
    if query.first():
        raise HTTPException(status_code=409, detail="Project code already exists")


def assign_component_warehouse_code(db: Session, component: Component) -> None:
    allocate_component_identity(db, component)


def assign_project_code(db: Session, project: Project) -> None:
    existing = normalize_project_code(project.project_code)
    if existing and not existing.startswith(f"{PROJECT_CODE_PREFIX}-"):
        project.project_code = existing
        return
    base = make_project_code(project.id)
    code = base
    suffix = 1
    while db.query(Project).filter(Project.project_code == code, Project.id != project.id).first():
        code = f"{base}-{suffix}"
        suffix += 1
    project.project_code = code[:80]


def ensure_component_warehouse_codes(db: Session) -> int:
    components = (
        db.query(Component)
        .filter(Component.revoked_at.is_(None))
        .order_by(Component.id.asc())
        .all()
    )
    changed = 0
    for component in components:
        before = component.warehouse_code
        assign_component_warehouse_code(db, component)
        changed += int(before != component.warehouse_code)
    if changed:
        log_activity(
            db,
            "component.warehouse_code.backfill",
            "component",
            f"补齐器件 ID {changed} 项",
            detail={"count": changed},
        )
    return changed


def ensure_project_codes(db: Session) -> int:
    projects = (
        db.query(Project)
        .filter(or_(Project.project_code.is_(None), Project.project_code == "", Project.project_code.like(f"{PROJECT_CODE_PREFIX}-%")))
        .order_by(Project.id.asc())
        .all()
    )
    for project in projects:
        assign_project_code(db, project)
    if projects:
        log_activity(
            db,
            "project.code.backfill",
            "project",
            f"补齐项目唯一 ID {len(projects)} 个",
            detail={"count": len(projects)},
        )
    return len(projects)


def category_id_by_name(db: Session, name: str | None) -> int | None:
    if not name:
        return None
    category = db.query(Category).filter(Category.name == name).first()
    return category.id if category else None


def ensure_category(db: Session, name: str, color: str = "#e8fff8") -> Category:
    category = db.query(Category).filter(Category.name == name).first()
    if category:
        if not category.color:
            category.color = color
        return category
    category = Category(name=name, color=color)
    db.add(category)
    db.flush()
    seed_category_prefixes(db)
    return category


def reconcile_pending_purchase_from_import_records(db: Session) -> int:
    pending_components = (
        db.query(Component)
        .filter(
            or_(
                Component.status == "pending_purchase",
                Component.source == "BOM 待采购库",
            ),
            Component.lcsc_number.isnot(None),
            Component.lcsc_number != "",
        )
        .all()
    )
    fixed = 0
    for component in pending_components:
        imported_quantity = (
            db.query(func.coalesce(func.sum(ImportRecord.quantity), 0))
            .filter(ImportRecord.lcsc_number == component.lcsc_number)
            .scalar()
            or 0
        )
        imported_quantity = int(imported_quantity or 0)
        current_quantity = int(component.quantity or 0)
        if imported_quantity <= 0 and current_quantity <= 0:
            continue
        if current_quantity < imported_quantity:
            mark_stock_change(component, imported_quantity - current_quantity)
            component.quantity = imported_quantity
        component.status = "in_stock"
        if component.source == "BOM 待采购库":
            component.source = "立创"
        if component.location == "待采购":
            component.location = None
        component.tags = clean_pending_purchase_tags(component.tags)
        note = f"历史订单导入记录已匹配立创 ID {component.lcsc_number}，自动从待采购转为在库。"
        if note not in (component.remark or ""):
            component.remark = f"{component.remark or ''}\n{note}".strip()
        fixed += 1
    if fixed:
        log_activity(
            db,
            "import.excel.reconcile_pending",
            "component",
            f"自动修复待采购库状态 {fixed} 项",
            detail={"fixed": fixed},
        )
    return fixed


def clean_pending_purchase_tags(tags: str | None) -> str | None:
    tokens = [part.strip() for part in str(tags or "").replace("，", ",").replace("；", ",").replace(";", ",").split(",")]
    kept = []
    seen = set()
    blocked = {"待采购", "bom待采购", "bom待采购库", "待入库", "pending_purchase"}
    for token in tokens:
        key = token.replace(" ", "").lower()
        if not token or key in blocked or key in seen:
            continue
        seen.add(key)
        kept.append(token)
    return ",".join(kept) or None


def normalize_for_inventory(db: Session, values: dict, *, clean_name: bool = True) -> dict:
    normalized = normalize_component_values(values) if clean_name else dict(values)
    normalized = infer_small_part_fields(normalized)
    if normalized.get("lcsc_number"):
        normalized["lcsc_number"] = normalize_lcsc_number(normalized.get("lcsc_number")) or str(normalized.get("lcsc_number")).strip()[:120]
    if "warehouse_code" in normalized:
        normalized["warehouse_code"] = normalize_warehouse_code(normalized.get("warehouse_code"))
    if not normalized.get("datasheet_url") and normalized.get("lcsc_number") and not normalized.get("buy_url"):
        normalized["datasheet_url"] = f"https://www.lcsc.com/product-detail/{normalized['lcsc_number']}.html"
    return normalized


def organize_cache_key(component: Component) -> str:
    payload = {
        "version": "component-organize-v1-ai-name",
        "name": component.name,
        "model": component.model,
        "manufacturer": component.manufacturer,
        "description": component.description,
        "parameters": component.parameters,
        "package": component.package,
        "tags": component.tags,
        "category_id": component.category_id,
        "part_family": component.part_family,
        "normalized_spec": component.normalized_spec,
        "source_title": getattr(component, "source_title", None),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def looks_like_needs_organize(component: Component) -> bool:
    category_name = component.category.name if component.category else ""
    tag_count = len(str(component.tags or "").split(",")) if component.tags else 0
    normalized_tag_count = len(normalize_tag_text(component.tags))
    return any(
        [
            not component.category_id,
            category_name in {"未分类", "其他", ""},
            len(component.name or "") > 36,
            tag_count > normalized_tag_count,
        ]
    )


def log_activity(
    db: Session,
    action: str,
    entity_type: str,
    summary: str,
    *,
    owner_user_id: int | None = None,
    entity_id: int | None = None,
    component_id: int | None = None,
    project_id: int | None = None,
    quantity_delta: int | None = None,
    detail: dict | str | None = None,
) -> None:
    if isinstance(detail, dict):
        detail_text = json.dumps(detail, ensure_ascii=False)
    else:
        detail_text = detail
    db.add(
        ActivityLog(
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            owner_user_id=owner_user_id,
            component_id=component_id,
            project_id=project_id,
            quantity_delta=quantity_delta,
            summary=summary,
            detail=detail_text,
        )
    )


def _json_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def component_ai_payload(component: Component) -> dict:
    return {
        "ai_analysis_version": AI_ANALYSIS_VERSION,
        "name": component.name,
        "model": component.model,
        "parameters": component.parameters,
        "package": component.package,
        "category_id": component.category_id,
        "category": component.category.name if component.category else None,
        "lcsc_number": component.lcsc_number,
        "datasheet_url": component.datasheet_url,
        "remark": component.remark,
        "tags": component.tags,
        "source_title": getattr(component, "source_title", None),
        "part_family": component.part_family,
        "normalized_spec": component.normalized_spec,
    }


def component_organize_payload(component: Component) -> dict:
    return {
        "id": component.id,
        "warehouse_code": component.warehouse_code,
        "name": component.name,
        "model": component.model,
        "category": component.category.name if component.category else None,
        "parameters": component.parameters,
        "package": component.package,
        "quantity": component.quantity,
        "source": component.source,
        "lcsc_number": component.lcsc_number,
        "tags": component.tags,
        "source_title": getattr(component, "source_title", None),
        "part_family": component.part_family,
        "count_mode": component.count_mode,
        "normalized_spec": component.normalized_spec,
        "status": component.status,
        "remark": component.remark,
    }


def component_ai_cache_key(component: Component) -> str:
    raw = json.dumps(component_ai_payload(component), ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def project_bom_cache_key(project: Project) -> str:
    rows = [
        {
            "component_id": item.component_id,
            "required_quantity": item.required_quantity,
            "status": item.status,
            "remark": item.remark,
            "name": item.component.name if item.component else None,
            "model": item.component.model if item.component else None,
            "parameters": item.component.parameters if item.component else None,
            "package": item.component.package if item.component else None,
        }
        for item in project.bom_items
    ]
    raw = json.dumps({"project": project.name, "description": project.description, "bom": rows}, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def apply_component_ai_result(db: Session, component: Component, result: dict, cache_key: str) -> None:
    component.ai_summary = result.get("summary") or result.get("normalized_name") or component.ai_summary
    usage_payload = {
        "usage": result.get("usage"),
        "key_specs": result.get("key_specs"),
        "typical_applications": result.get("typical_applications") or result.get("applications"),
        "design_insights": result.get("design_insights"),
        "do_not_use_for": result.get("do_not_use_for"),
        "recommended_pairings": result.get("recommended_pairings"),
        "datasheet_notes": result.get("datasheet_notes"),
        "source_notes": result.get("source_notes"),
    }
    fallback = result.get("applications") or result.get("typical_applications")
    if isinstance(fallback, list):
        fallback = {"usage": "", "key_specs": [], "typical_applications": fallback}
    component.ai_usage = _json_text({key: value for key, value in usage_payload.items() if value}) or _json_text(fallback)
    component.ai_risk_notes = _json_text(result.get("risk_notes") or result.get("risks"))
    component.ai_pcb_notes = _json_text(result.get("pcb_notes"))
    component.ai_substitutes = _json_text(
        {
            "substitutes": result.get("substitutes"),
            "substitution_notes": result.get("substitution_notes"),
        }
    )
    component.ai_tags = ",".join(result.get("ai_tags") or result.get("recommended_tags") or result.get("completion_suggestions") or [])[:500]
    component.ai_confidence = result.get("confidence") or "medium"
    component.ai_cache_key = cache_key
    component.ai_status = "completed"
    component.ai_error = None
    component.ai_updated_at = datetime.now()
    if result.get("datasheet_url") and not component.datasheet_url:
        component.datasheet_url = str(result["datasheet_url"])[:500]
    if not component.datasheet_url and component.lcsc_number:
        component.datasheet_url = f"https://www.lcsc.com/product-detail/{component.lcsc_number}.html"
    component.is_hand_solder_friendly = bool(result.get("is_hand_solder_friendly", component.is_hand_solder_friendly))
    component.is_power_component = bool(result.get("is_power_component", component.is_power_component))
    component.is_signal_component = bool(result.get("is_signal_component", component.is_signal_component))
    component.is_high_current = bool(result.get("is_high_current", component.is_high_current))
    component.is_high_voltage = bool(result.get("is_high_voltage", component.is_high_voltage))
    component.is_common = bool(result.get("is_common", component.is_common))
    card = AiKnowledgeCard(
        component_id=component.id,
        title=f"{component.model or component.name} AI 知识卡片",
        content=json.dumps(result, ensure_ascii=False, indent=2),
        tags=component.ai_tags,
        source_type="ai",
        confidence=component.ai_confidence,
    )
    db.add(card)


AI_UNDO_FIELDS = {
    "name",
    "category_id",
    "parameters",
    "package",
    "tags",
    "source_title",
    "part_family",
    "count_mode",
    "normalized_spec",
    "ai_summary",
    "ai_usage",
    "ai_risk_notes",
    "ai_pcb_notes",
    "ai_substitutes",
    "ai_tags",
    "ai_confidence",
    "datasheet_url",
    "is_hand_solder_friendly",
    "is_power_component",
    "is_signal_component",
    "is_high_current",
    "is_high_voltage",
    "is_common",
}


def component_ai_undo_snapshot(component: Component) -> dict:
    return {field: getattr(component, field) for field in AI_UNDO_FIELDS}


def mark_component_ai_stale(component: Component) -> None:
    cache_key = component_ai_cache_key(component)
    if component.ai_cache_key and component.ai_cache_key != cache_key:
        component.ai_status = "stale"


UNKNOWN_PACKAGE_WORDS = {"未知", "不详", "未确认", "待确认", "unknown", "n/a", "na", "-"}


def ai_package_can_fill(component: Component, result: dict) -> bool:
    if (component.package or "").strip():
        return False
    package = str(result.get("package") or "").strip()
    if not package or package.lower() in UNKNOWN_PACKAGE_WORDS:
        return False
    confidence = str(result.get("package_confidence") or result.get("confidence") or "").lower()
    return confidence == "high"


def apply_component_organize_result(db: Session, component: Component, result: dict, *, source: str) -> dict:
    before = {
        "name": component.name,
        "category_id": component.category_id,
        "parameters": component.parameters,
        "package": component.package,
        "tags": component.tags,
        "part_family": component.part_family,
        "count_mode": component.count_mode,
        "normalized_spec": component.normalized_spec,
        "source_title": component.source_title,
    }
    category_name = result.get("category") or result.get("category_suggestion")
    suggested_category_id = category_id_by_name(db, category_name)
    category_id = (
        suggested_category_id
        if suggested_category_id and ai_category_allowed(component, result)
        else component.category_id
    )
    normalized_name = clean_component_name(result.get("normalized_name") or component.name, component.model, component.lcsc_number)
    if normalized_name and normalized_name != component.name:
        component.source_title = component.source_title or component.name
        component.name = normalized_name
    if category_id:
        component.category_id = category_id
    if result.get("parameters"):
        component.parameters = str(result.get("parameters"))[:1000]
    if result.get("package") and ai_package_can_fill(component, result):
        component.package = str(result.get("package"))[:120]
    tags = result.get("tags") or result.get("ai_tags")
    if isinstance(tags, str):
        tags = normalize_tag_text(tags, component.package)
    if isinstance(tags, list):
        component.tags = ",".join(normalize_tag_text(",".join(str(item) for item in tags), component.package))[:300]
    for field in ["part_family", "count_mode", "normalized_spec"]:
        if result.get(field):
            setattr(component, field, str(result[field])[:160])
    component.ai_status = component.ai_status or "pending"
    after = {
        "name": component.name,
        "category_id": component.category_id,
        "parameters": component.parameters,
        "package": component.package,
        "tags": component.tags,
        "part_family": component.part_family,
        "count_mode": component.count_mode,
        "normalized_spec": component.normalized_spec,
        "source_title": component.source_title,
    }
    log_activity(
        db,
        "ai.component.organize",
        "component",
        f"AI 规范化元器件 {component.name}",
        entity_id=component.id,
        component_id=component.id,
        detail={"source": source, "before": before, "after": after, "confidence": result.get("confidence"), "reason": result.get("reason")},
    )
    return {"before": before, "after": after, "result": result}


def enqueue_ai_task(db: Session, task_type: str, target_type: str, target_id: int, input_hash: str | None = None) -> AiTask:
    existing = (
        db.query(AiTask)
        .filter(
            AiTask.task_type == task_type,
            AiTask.target_type == target_type,
            AiTask.target_id == target_id,
            AiTask.status.in_(["pending", "processing", "stale", "failed"]),
        )
        .order_by(AiTask.id.desc())
        .first()
    )
    if existing:
        if input_hash:
            existing.input_hash = input_hash
        if existing.status == "failed":
            existing.status = "pending"
            existing.next_attempt_at = datetime.now()
            existing.error_message = None
        elif existing.status != "processing":
            existing.status = "pending"
        return existing
    task = AiTask(task_type=task_type, target_type=target_type, target_id=target_id, input_hash=input_hash, status="pending")
    db.add(task)
    return task


def retry_delay_for(task: AiTask) -> int:
    retry_count = max(1, task.retry_count or 1)
    delay = AI_TASK_RETRY_BASE_SECONDS * (2 ** max(0, retry_count - 1))
    return min(delay, AI_TASK_RETRY_MAX_SECONDS)


def resolve_superseded_ai_failures(db: Session) -> int:
    resolved = 0
    failed_tasks = db.query(AiTask).filter(AiTask.status == "failed").all()
    for task in failed_tasks:
        newer_success = (
            db.query(AiTask.id)
            .filter(
                AiTask.task_type == task.task_type,
                AiTask.target_type == task.target_type,
                AiTask.target_id == task.target_id,
                AiTask.status == "completed",
                AiTask.id > task.id,
            )
            .first()
        )
        if not newer_success:
            continue
        task.status = "superseded"
        task.error_message = f"已被后续成功任务覆盖：{task.error_message or ''}"[:1000]
        task.next_attempt_at = None
        resolved += 1
    completed_component_ids = [
        component_id
        for (component_id,) in db.query(AiTask.target_id)
        .filter(
            AiTask.task_type == "component_analyze",
            AiTask.target_type == "component",
            AiTask.status == "completed",
        )
        .distinct()
        .all()
    ]
    if completed_component_ids:
        db.query(Component).filter(
            Component.id.in_(completed_component_ids),
            Component.ai_status == "failed",
            Component.ai_summary.isnot(None),
        ).update({Component.ai_status: "completed", Component.ai_error: None}, synchronize_session=False)
    return resolved


def reserved_quantities(db: Session, component_ids: list[int] | None = None) -> dict[int, int]:
    return inventory_reserved_quantities(db, component_ids)


def component_out(component: Component, reserved: int = 0) -> dict:
    quantity = component.quantity or 0
    available = max(0, quantity - reserved)
    safety_quantity = max(0, int(component.safety_quantity or 0))
    warning_threshold = safety_quantity if safety_quantity > 0 else 5
    low_stock_warning = bool(
        component.is_common
        and not component.low_stock_exempt
        and available <= warning_threshold
    )
    return {
        "id": component.id,
        "warehouse_code": component.warehouse_code,
        "name": component.name,
        "model": component.model,
        "manufacturer": component.manufacturer,
        "description": component.description,
        "category_id": component.category_id,
        "parameters": component.parameters,
        "package": component.package,
        "quantity": quantity,
        "source": component.source,
        "lcsc_number": component.lcsc_number,
        "tags": component.tags,
        "source_title": getattr(component, "source_title", None),
        "part_family": component.part_family or "component",
        "count_mode": component.count_mode or "exact",
        "normalized_spec": component.normalized_spec,
        "status": component.status,
        "location": component.location,
        "remark": component.remark,
        "datasheet_url": component.datasheet_url,
        "buy_url": component.buy_url,
        "is_hand_solder_friendly": bool(component.is_hand_solder_friendly),
        "is_power_component": bool(component.is_power_component),
        "is_signal_component": bool(component.is_signal_component),
        "is_high_current": bool(component.is_high_current),
        "is_high_voltage": bool(component.is_high_voltage),
        "is_common": bool(component.is_common),
        "safety_quantity": safety_quantity,
        "low_stock_exempt": bool(component.low_stock_exempt),
        "low_stock_warning": low_stock_warning,
        "category": component.category,
        "reserved_quantity": reserved,
        "available_quantity": available,
        "ai_summary": component.ai_summary,
        "ai_usage": component.ai_usage,
        "ai_risk_notes": component.ai_risk_notes,
        "ai_pcb_notes": component.ai_pcb_notes,
        "ai_substitutes": component.ai_substitutes,
        "ai_tags": component.ai_tags,
        "ai_confidence": component.ai_confidence,
        "ai_cache_key": component.ai_cache_key,
        "ai_status": component.ai_status or "pending",
        "ai_error": component.ai_error,
        "ai_updated_at": component.ai_updated_at,
        "first_stocked_at": component.first_stocked_at,
        "last_stocked_at": component.last_stocked_at,
        "last_outbound_at": component.last_outbound_at,
        "created_at": component.created_at,
        "updated_at": component.updated_at,
    }


def inventory_lot_out(lot: InventoryLot) -> dict:
    return {
        "id": lot.id,
        "component_id": lot.component_id,
        "source_type": lot.source_type,
        "source_reference": lot.source_reference,
        "location": lot.location,
        "initial_quantity": int(lot.initial_quantity or 0),
        "remaining_quantity": int(lot.remaining_quantity or 0),
        "unit_cost": lot.unit_cost,
        "status": lot.status,
        "received_at": lot.received_at,
        "created_at": lot.created_at,
    }


def custom_label_asset_url(asset: CustomLabelAsset, *, team_library_id: str | None = None) -> str:
    if team_library_id:
        return f"/api/team/libraries/{quote(team_library_id)}/custom-labels/assets/{quote(asset.id)}"
    return f"/api/custom-labels/assets/{quote(asset.id)}"


def custom_label_asset_out(asset: CustomLabelAsset, *, team_library_id: str | None = None) -> dict:
    return {
        "id": asset.id,
        "template_id": asset.template_id,
        "file_name": asset.file_name,
        "mime_type": asset.mime_type,
        "sha256": asset.sha256,
        "size_bytes": int(asset.size_bytes or 0),
        "url": custom_label_asset_url(asset, team_library_id=team_library_id),
        "created_at": asset.created_at,
    }


def parse_custom_label_content(value: str | dict | None) -> dict:
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value or "{}")
    except Exception:
        parsed = {}
    return parsed if isinstance(parsed, dict) else {}


def custom_label_template_out(template: CustomLabelTemplate, db: Session) -> dict:
    assets = (
        db.query(CustomLabelAsset)
        .filter(CustomLabelAsset.template_id == template.id)
        .order_by(CustomLabelAsset.created_at.asc())
        .all()
    )
    return {
        "id": template.id,
        "scope_type": template.scope_type,
        "owner_user_id": template.owner_user_id,
        "team_library_id": template.team_library_id,
        "name": template.name,
        "content": parse_custom_label_content(template.content_json),
        "status": template.status,
        "assets": [
            custom_label_asset_out(asset, team_library_id=template.team_library_id if template.scope_type == "team" else None)
            for asset in assets
        ],
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


def custom_label_query_for_auth(db: Session, auth: AuthContext, *, include_archived: bool = False):
    query = db.query(CustomLabelTemplate).filter(
        CustomLabelTemplate.scope_type == "personal",
        CustomLabelTemplate.owner_user_id == auth.user_id,
    )
    if not include_archived:
        query = query.filter(CustomLabelTemplate.status == "active")
    return query


def require_custom_label_template(db: Session, template_id: str, auth: AuthContext) -> CustomLabelTemplate:
    template = (
        custom_label_query_for_auth(db, auth, include_archived=False)
        .filter(CustomLabelTemplate.id == template_id)
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="自定义标签不存在")
    return template


def clean_custom_label_content(content: dict | None) -> dict:
    raw = content if isinstance(content, dict) else {}
    elements = raw.get("elements")
    cleaned_elements: list[dict] = []
    if isinstance(elements, list):
        for index, item in enumerate(elements[:40]):
            if not isinstance(item, dict):
                continue
            element_type = str(item.get("type") or "text").strip()
            if element_type not in {"text", "image", "svg", "field", "qr", "shape", "category_badge"}:
                element_type = "text"
            if element_type in {"field", "category_badge"} and str(item.get("field") or "").strip() == "print_date":
                continue
            clean = {
                "id": str(item.get("id") or f"el-{index + 1}")[:80],
                "type": element_type,
                "x": item.get("x", 22),
                "y": item.get("y", 34),
                "width": item.get("width", 56),
                "height": item.get("height", 30),
                "x_mm": item.get("x_mm"),
                "y_mm": item.get("y_mm"),
                "width_mm": item.get("width_mm"),
                "height_mm": item.get("height_mm"),
                "rotate": item.get("rotate", 0),
                "font_size": item.get("font_size", 13),
                "font_weight": item.get("font_weight", 400),
                "font_family": str(item.get("font_family") or "system")[:40],
                "color": str(item.get("color") or "#111827")[:40],
                "align": str(item.get("align") or "center")[:16],
            }
            if item.get("role"):
                clean["role"] = str(item.get("role") or "")[:80]
            if element_type == "text":
                clean["text"] = str(item.get("text") or "自定义标签")[:1000]
            elif element_type in {"image", "svg"}:
                clean["asset_id"] = str(item.get("asset_id") or "")[:80]
                if element_type == "svg" and item.get("svg"):
                    clean["svg"] = sanitize_svg_markup(str(item.get("svg") or ""))
            elif element_type in {"field", "qr", "category_badge"}:
                clean["field"] = str(item.get("field") or ("scan_url" if element_type == "qr" else "name"))[:80]
                clean["prefix"] = str(item.get("prefix") or "")[:80]
            elif element_type == "shape":
                clean["fill"] = str(item.get("fill") or "#eff6ff")[:40]
                clean["stroke"] = str(item.get("stroke") or "#93c5fd")[:40]
                clean["radius"] = item.get("radius", 1)
            for mm_key in ("x_mm", "y_mm", "width_mm", "height_mm"):
                if clean.get(mm_key) is None:
                    clean.pop(mm_key, None)
            cleaned_elements.append(clean)
    if not cleaned_elements:
        cleaned_elements = [{"id": "text-1", "type": "text", "text": str(raw.get("text") or "自定义标签")[:1000], "x": 18, "y": 33, "width": 64, "height": 30, "font_size": 16, "font_family": "system", "color": "#111827", "align": "center"}]
    result = {"elements": cleaned_elements, "show_logo": raw.get("show_logo") is not False}
    if raw.get("kind"):
        result["kind"] = str(raw.get("kind") or "")[:80]
    styles = raw.get("styles")
    cleaned_styles: list[dict] = []
    if isinstance(styles, list):
        for index, style in enumerate(styles[:20]):
            if not isinstance(style, dict):
                continue
            style_content = clean_custom_label_content({"elements": style.get("elements")})
            cleaned_styles.append({
                "id": str(style.get("id") or f"style-{index + 1}")[:80],
                "name": str(style.get("name") or f"样式 {index + 1}")[:80],
                "category_name": str(style.get("category_name") or "")[:80],
                "elements": style_content["elements"],
            })
    if cleaned_styles:
        result["styles"] = cleaned_styles
        active_style_id = str(raw.get("active_style_id") or cleaned_styles[0]["id"])[:80]
        result["active_style_id"] = active_style_id
    return result


def custom_label_asset_root(template: CustomLabelTemplate) -> Path:
    scope = "team" if template.scope_type == "team" else "personal"
    owner = template.team_library_id if scope == "team" else str(template.owner_user_id or "unknown")
    return CUSTOM_LABEL_STORAGE_ROOT / scope / owner / template.id


def assert_custom_label_mime(file: UploadFile, payload: bytes) -> tuple[str, bytes, str]:
    mime_type = (file.content_type or "").split(";", 1)[0].strip().lower()
    suffix = CUSTOM_LABEL_ALLOWED_MIME.get(mime_type)
    if not suffix:
        raise HTTPException(status_code=400, detail="仅支持 PNG、JPG、WebP 或 SVG")
    max_bytes = CUSTOM_LABEL_MAX_SVG_BYTES if mime_type == "image/svg+xml" else CUSTOM_LABEL_MAX_IMAGE_BYTES
    if len(payload) > max_bytes:
        raise HTTPException(status_code=413, detail="自定义标签素材超过大小限制")
    data = payload
    if mime_type == "image/svg+xml":
        cleaned = sanitize_svg_markup(payload.decode("utf-8", errors="ignore"))
        if not cleaned:
            raise HTTPException(status_code=400, detail="SVG 内容不安全或无法解析")
        data = cleaned.encode("utf-8")
    else:
        signatures = {
            "image/png": b"\x89PNG\r\n\x1a\n",
            "image/jpeg": b"\xff\xd8\xff",
            "image/webp": b"RIFF",
        }
        signature = signatures[mime_type]
        if not payload.startswith(signature):
            raise HTTPException(status_code=400, detail="图片文件头与类型不匹配")
        if mime_type == "image/webp" and payload[8:12] != b"WEBP":
            raise HTTPException(status_code=400, detail="WebP 文件头无效")
    return mime_type, data, suffix


async def save_custom_label_asset(db: Session, template: CustomLabelTemplate, file: UploadFile) -> CustomLabelAsset:
    payload = await file.read()
    mime_type, data, suffix = assert_custom_label_mime(file, payload)
    sha256 = hashlib.sha256(data).hexdigest()
    asset_id = secrets.token_hex(16)
    root = custom_label_asset_root(template)
    root.mkdir(parents=True, exist_ok=True)
    file_name = Path(file.filename or f"asset{suffix}").name[:200] or f"asset{suffix}"
    storage_path = root / f"{asset_id}{suffix}"
    tmp_path = storage_path.with_suffix(storage_path.suffix + ".tmp")
    tmp_path.write_bytes(data)
    tmp_path.replace(storage_path)
    asset = CustomLabelAsset(
        id=asset_id,
        template_id=template.id,
        file_name=file_name,
        storage_path=str(storage_path),
        mime_type=mime_type,
        sha256=sha256,
        size_bytes=len(data),
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def require_custom_label_asset(db: Session, asset_id: str, auth: AuthContext) -> CustomLabelAsset:
    asset = db.get(CustomLabelAsset, asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="标签素材不存在")
    template = require_custom_label_template(db, asset.template_id, auth)
    if template.status != "active":
        raise HTTPException(status_code=404, detail="标签素材不存在")
    return asset


def custom_label_asset_resolver(db: Session, template: CustomLabelTemplate):
    allowed_assets = {
        asset.id: asset
        for asset in db.query(CustomLabelAsset).filter(CustomLabelAsset.template_id == template.id).all()
    }

    def resolve(asset_id: str) -> dict:
        asset = allowed_assets.get(asset_id)
        if not asset:
            return {}
        path = Path(asset.storage_path)
        if not path.exists() or not path.is_file():
            return {}
        data = path.read_bytes()
        if asset.mime_type == "image/svg+xml":
            return {"svg": sanitize_svg_markup(data.decode("utf-8", errors="ignore"))}
        return {"data_uri": data_uri_from_bytes(asset.mime_type, data)}

    return resolve


def component_export_custom_label_cards(db: Session, auth: AuthContext, items: list, printed_at: str, records: list[dict] | None = None) -> list[str]:
    cards: list[str] = []
    package_summary = category_package_summary_from_records(records)
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_custom_label_template(db, template_id, auth)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        if is_standard_category_label_group(content):
            cards.extend(
                render_standard_category_label_cards(
                    content,
                    package_summary,
                    copies=int(getattr(item, "copies", 1) or 1),
                    printed_at=printed_at,
                )
            )
            continue
        cards.extend(
            render_custom_label_cards(
                content,
                asset_resolver=custom_label_asset_resolver(db, template),
                copies=int(getattr(item, "copies", 1) or 1),
                printed_at=printed_at,
                include_all_styles=True,
            )
        )
    return cards


def component_export_custom_label_font_keys(db: Session, auth: AuthContext, items: list) -> set[str]:
    keys: set[str] = set()
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_custom_label_template(db, template_id, auth)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        keys.update(custom_label_font_keys(content))
    return keys


def component_export_custom_label_pdf_items(db: Session, auth: AuthContext, items: list, records: list[dict] | None = None) -> list[dict]:
    pdf_items: list[dict] = []
    package_summary = category_package_summary_from_records(records)
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_custom_label_template(db, template_id, auth)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        copies = int(getattr(item, "copies", 1) or 1)
        if is_standard_category_label_group(content):
            pdf_items.extend(render_standard_category_label_pdf_items(content, package_summary, copies=copies))
        else:
            pdf_items.extend(render_basic_custom_label_pdf_items(content, copies=copies))
    return pdf_items


def component_ai_question_context(db: Session, component: Component, auth: AuthContext) -> dict:
    supplier_rows = (
        filter_owner(db.query(SupplierPart), SupplierPart, auth)
        .filter(SupplierPart.component_id == component.id)
        .order_by(SupplierPart.is_preferred.desc(), SupplierPart.created_at.desc())
        .limit(8)
        .all()
    )
    binding_rows = (
        db.query(EdaComponentBinding)
        .filter(EdaComponentBinding.component_id == component.id)
        .order_by(EdaComponentBinding.updated_at.desc())
        .limit(8)
        .all()
    )
    bindings = []
    for binding in binding_rows:
        symbol = db.get(EdaSymbol, binding.symbol_id) if binding.symbol_id else None
        footprint = db.get(EdaFootprint, binding.footprint_id) if binding.footprint_id else None
        datasheet = db.get(EdaAsset, binding.datasheet_asset_id) if binding.datasheet_asset_id else None
        model = db.get(EdaAsset, binding.model_asset_id) if binding.model_asset_id else None
        bindings.append(
            {
                "symbol": symbol.name if symbol else None,
                "footprint": footprint.name if footprint else None,
                "verification_status": binding.verification_status,
                "datasheet": datasheet.original_name if datasheet else None,
                "model_3d": model.original_name if model else None,
            }
        )
    lots = (
        db.query(InventoryLot)
        .filter(InventoryLot.component_id == component.id, InventoryLot.status == "active")
        .order_by(InventoryLot.remaining_quantity.desc(), InventoryLot.received_at.desc())
        .limit(8)
        .all()
    )
    return {
        "supplier_parts": [
            {
                "supplier": row.supplier,
                "supplier_part_number": row.supplier_part_number,
                "purchase_url": row.purchase_url,
                "is_preferred": bool(row.is_preferred),
            }
            for row in supplier_rows
        ],
        "eda_bindings": bindings,
        "inventory_lots": [inventory_lot_out(row) for row in lots],
        "recent_usage": component_usage_records(component.id, auth, db, limit=10),
    }


def public_component_out(component: Component) -> dict:
    identity = (
        db_identity
        if (db_identity := getattr(component, "_identity_registry", None))
        else None
    )
    if identity:
        return public_identity_out(identity, component)
    return {
        "warehouse_code": component.warehouse_code,
        "name": component.name,
        "model": component.model,
        "normalized_spec": component.normalized_spec,
        "package": component.package,
        "category": component.category.name if component.category else None,
        "category_color": component.category.color if component.category else None,
        "lcsc_number": component.lcsc_number,
        "datasheet_url": component.datasheet_url,
        "archived": False,
        "updated_at": component.updated_at,
    }


def public_project_out(project: Project, request: Request | None = None) -> dict:
    items = []
    boards = sorted(getattr(project, "boards", []) or [], key=lambda board: (board.board_index or 0, board.id or 0))
    total_designators = soldered_designators = lost_designators = pending_designators = 0
    for item in sorted(project.bom_items, key=lambda row: ((row.component.category.name if row.component and row.component.category else ""), row.id)):
        component = item.component
        solder_points = sorted(getattr(item, "solder_points", []) or [], key=lambda point: point.id)
        total_designators += len(solder_points)
        soldered_designators += sum(1 for point in solder_points if point.soldered)
        lost_designators += sum(1 for point in solder_points if getattr(point, "lost", False))
        pending_designators += sum(1 for point in solder_points if not point.soldered)
        items.append(
            {
                "id": item.id,
                "required_quantity": int(item.required_quantity or 0),
                "status": item.status or "reserved",
                "remark": item.remark,
                "component": {
                    "warehouse_code": component.warehouse_code if component else None,
                    "name": component.name if component else None,
                    "model": component.model if component else None,
                    "category": component.category.name if component and component.category else None,
                    "category_color": component.category.color if component and component.category else None,
                    "parameters": component.parameters if component else None,
                    "package": component.package if component else None,
                    "lcsc_number": component.lcsc_number if component else None,
                    "normalized_spec": component.normalized_spec if component else None,
                    "location": component.location if component else None,
                },
                "solder_points": [solder_point_out(point) for point in solder_points],
                "soldered_count": sum(1 for point in solder_points if point.soldered),
                "lost_count": sum(1 for point in solder_points if getattr(point, "lost", False)),
                "pending_count": sum(1 for point in solder_points if not point.soldered),
                "solder_total": len(solder_points),
            }
        )
    points_by_board: dict[int, list[ProjectBomSolderPoint]] = {}
    for item in project.bom_items:
        for point in getattr(item, "solder_points", []) or []:
            if point.board_id:
                points_by_board.setdefault(point.board_id, []).append(point)
    public_url = frontend_project_url(request, project.project_code) if request and project.project_code else None
    return {
        "project_code": project.project_code,
        "name": project.name,
        "description": project.description,
        "status": project.status,
        "public_url": public_url,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
        "boards": [board_out(board, points_by_board.get(board.id, [])) for board in boards],
        "active_board_id": boards[0].id if boards else None,
        "board_count": len(boards),
        "total_items": len(items),
        "total_designators": total_designators,
        "soldered_designators": soldered_designators,
        "lost_designators": lost_designators,
        "pending_designators": pending_designators,
        "solder_progress": round((soldered_designators / total_designators) * 100) if total_designators else 0,
        "bom_items": items,
    }


def mark_stock_change(component: Component, quantity_delta: int, at: datetime | None = None) -> None:
    changed_at = at or datetime.utcnow()
    if quantity_delta > 0:
        if not component.first_stocked_at:
            component.first_stocked_at = changed_at
        component.last_stocked_at = changed_at
    elif quantity_delta < 0:
        component.last_outbound_at = changed_at


def keyword_unit_variants(keyword: str | None) -> list[str]:
    text_value = str(keyword or "").strip()
    if not text_value:
        return []
    variants = {text_value, text_value.replace("µ", "u").replace("μ", "u")}
    for value in list(variants):
        if re.search("u", value, flags=re.IGNORECASE):
            variants.add(re.sub("u", "µ", value, flags=re.IGNORECASE))
            variants.add(re.sub("u", "μ", value, flags=re.IGNORECASE))
    return [value for value in variants if value]


def component_keyword_filters(keyword: str | None) -> list:
    filters = []
    for variant in keyword_unit_variants(keyword):
        like = f"%{variant}%"
        filters.extend(
            [
                Component.name.ilike(like),
                Component.warehouse_code.ilike(like),
                Component.model.ilike(like),
                Component.parameters.ilike(like),
                Component.package.ilike(like),
                Component.lcsc_number.ilike(like),
                Component.location.ilike(like),
                Component.remark.ilike(like),
                Component.ai_summary.ilike(like),
                Component.ai_tags.ilike(like),
                Component.tags.ilike(like),
                Component.normalized_spec.ilike(like),
                Component.source_title.ilike(like),
            ]
        )
    return filters


def apply_component_filters(
    query,
    *,
    category_id: int | None = None,
    status: str | None = None,
    tag: str | None = None,
    keyword: str | None = None,
    package: str | None = None,
    ai_status: str | None = None,
    stock: str | None = None,
    is_hand_solder_friendly: bool | None = None,
    is_power_component: bool | None = None,
    is_signal_component: bool | None = None,
    is_high_current: bool | None = None,
    is_high_voltage: bool | None = None,
):
    if category_id:
        query = query.filter(Component.category_id == category_id)
    if status:
        query = query.filter(Component.status == status)
    if tag:
        query = query.filter(or_(Component.tags.ilike(f"%{tag}%"), Component.ai_tags.ilike(f"%{tag}%")))
    if package:
        query = query.filter(Component.package.ilike(f"%{package}%"))
    if ai_status:
        query = query.filter(Component.ai_status == ai_status)
    if stock == "empty":
        query = query.filter(Component.quantity <= 0)
    elif stock == "low":
        query = query.filter(Component.is_common == True, Component.low_stock_exempt == False)
    elif stock == "available":
        query = query.filter(Component.quantity > 0)
    bool_filters = {
        Component.is_hand_solder_friendly: is_hand_solder_friendly,
        Component.is_power_component: is_power_component,
        Component.is_signal_component: is_signal_component,
        Component.is_high_current: is_high_current,
        Component.is_high_voltage: is_high_voltage,
    }
    for column, value in bool_filters.items():
        if value is not None:
            query = query.filter(column == value)
    if keyword:
        filters = component_keyword_filters(keyword)
        if filters:
            query = query.filter(or_(*filters))
    return query


def filter_low_stock_components(
    db: Session,
    components: list[Component],
    stock: str | None,
) -> list[Component]:
    if stock != "low" or not components:
        return components
    reserved = reserved_quantities(db, [component.id for component in components])
    return [
        component
        for component in components
        if max(0, int(component.quantity or 0) - reserved.get(component.id, 0))
        <= (max(0, int(component.safety_quantity or 0)) or 5)
    ]


PASSIVE_COVERAGE = {
    "电阻": {"unit": "Ω", "dimension": "resistance"},
    "电容": {"unit": "F", "dimension": "capacitance"},
    "电感": {"unit": "H", "dimension": "inductance"},
}


def is_ferrite_bead(component: Component) -> bool:
    text_value = " ".join(
        str(part or "")
        for part in [
            component.name,
            component.model,
            component.normalized_spec,
            component.parameters,
            component.tags,
            component.ai_tags,
            component.part_family,
        ]
    ).lower()
    return "磁珠" in text_value or "ferrite" in text_value or "Ω@" in text_value or "ohm@" in text_value


def parse_json_value(value):
    if not value:
        return None
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return None


def list_value(value) -> list:
    parsed = parse_json_value(value)
    if parsed is None:
        return []
    if isinstance(parsed, list):
        return [str(item) for item in parsed if item is not None and str(item).strip()]
    if isinstance(parsed, dict):
        return [str(item) for item in parsed.values() if item is not None and str(item).strip()]
    text_value = str(parsed).strip()
    return [text_value] if text_value else []


def parse_unit_value(value: str, dimension: str) -> float | None:
    text_value = str(value or "").strip().replace("μ", "u").replace("µ", "u")
    if dimension == "resistance":
        match = re.search(r"(\d+(?:\.\d+)?)\s*(m|k|K|M)?\s*(?:Ω|ohm|R)", text_value, flags=re.IGNORECASE)
        if not match:
            return None
        multipliers = {"m": 0.001, "k": 1000, "K": 1000, "M": 1000000}
        return float(match.group(1)) * multipliers.get(match.group(2) or "", 1)
    if dimension == "capacitance":
        match = re.search(r"(\d+(?:\.\d+)?)\s*(p|n|u|m)?\s*F", text_value, flags=re.IGNORECASE)
        if not match:
            return None
        multipliers = {"p": 1e-12, "n": 1e-9, "u": 1e-6, "m": 1e-3}
        return float(match.group(1)) * multipliers.get((match.group(2) or "").lower(), 1)
    if dimension == "inductance":
        match = re.search(r"(\d+(?:\.\d+)?)\s*(n|u|m)?\s*H", text_value, flags=re.IGNORECASE)
        if not match:
            return None
        multipliers = {"n": 1e-9, "u": 1e-6, "m": 1e-3}
        return float(match.group(1)) * multipliers.get((match.group(2) or "").lower(), 1)
    return None


def key_specs_from_component(component: Component) -> list[dict]:
    parsed = parse_json_value(component.ai_usage)
    if not isinstance(parsed, dict):
        return []
    specs = parsed.get("key_specs")
    return specs if isinstance(specs, list) else []


def split_bom_designators(value: str | None) -> list[str]:
    text_value = str(value or "").strip()
    if not text_value or text_value == "-":
        return []
    parts = re.split(r"[,，、\s]+", text_value)
    result = []
    seen = set()
    for part in parts:
        designator = part.strip().strip(";；")
        if not designator:
            continue
        key = designator.upper()
        if key in seen:
            continue
        seen.add(key)
        result.append(designator)
    return result


def parse_bom_field(text_value: str, field: str) -> str:
    match = re.search(rf"{re.escape(field)}\s*[:：]\s*([^；;\n]+)", text_value)
    return match.group(1).strip() if match else ""


def solder_point_specs_from_remark(remark: str | None) -> list[dict]:
    rows = [line.strip() for line in str(remark or "").splitlines() if line.strip()]
    if not rows and remark:
        rows = [str(remark)]
    specs = []
    seen = set()
    for row in rows:
        designators = split_bom_designators(parse_bom_field(row, "BOM 位号"))
        if not designators:
            continue
        model = parse_bom_field(row, "BOM 型号")
        footprint = parse_bom_field(row, "BOM 封装")
        value = parse_bom_field(row, "BOM 参数") or parse_bom_field(row, "BOM 值")
        for designator in designators:
            key = designator.upper()
            if key in seen:
                continue
            seen.add(key)
            specs.append(
                {
                    "designator": designator,
                    "bom_value": value or None,
                    "bom_model": model or None,
                    "bom_footprint": footprint or None,
                }
            )
    return specs


def solder_point_specs_for_item(item: ProjectBomItem) -> list[dict]:
    specs = solder_point_specs_from_remark(item.remark)
    if specs:
        return specs
    count = max(1, int(item.required_quantity or 1))
    component = getattr(item, "component", None)
    return [
        {
            "designator": f"用量{index}",
            "bom_value": getattr(component, "normalized_spec", None) or getattr(component, "parameters", None),
            "bom_model": getattr(component, "model", None),
            "bom_footprint": getattr(component, "package", None),
        }
        for index in range(1, count + 1)
    ]


def default_project_board(db: Session, project_id: int) -> ProjectBoard:
    board = (
        db.query(ProjectBoard)
        .filter(ProjectBoard.project_id == project_id)
        .order_by(ProjectBoard.board_index.asc(), ProjectBoard.id.asc())
        .first()
    )
    if board:
        return board
    board = ProjectBoard(project_id=project_id, board_index=1, name="第 1 板", status="active")
    db.add(board)
    db.flush()
    return board


def project_boards_for_sync(db: Session, project_id: int, board_id: int | None = None) -> list[ProjectBoard]:
    if board_id:
        board = db.query(ProjectBoard).filter(ProjectBoard.project_id == project_id, ProjectBoard.id == board_id).first()
        if not board:
            raise HTTPException(status_code=404, detail="Project board not found")
        return [board]
    boards = db.query(ProjectBoard).filter(ProjectBoard.project_id == project_id).order_by(ProjectBoard.board_index.asc(), ProjectBoard.id.asc()).all()
    if boards:
        return boards
    return [default_project_board(db, project_id)]


def ensure_project_boards(db: Session) -> int:
    created = 0
    for project_id, in db.query(Project.id).all():
        before_id = db.query(ProjectBoard.id).filter(ProjectBoard.project_id == project_id).first()
        board = default_project_board(db, project_id)
        if not before_id:
            created += 1
        item_ids = [item_id for (item_id,) in db.query(ProjectBomItem.id).filter(ProjectBomItem.project_id == project_id).all()]
        if item_ids:
            db.query(ProjectBomSolderPoint).filter(
                ProjectBomSolderPoint.bom_item_id.in_(item_ids),
                ProjectBomSolderPoint.board_id.is_(None),
            ).update({ProjectBomSolderPoint.board_id: board.id}, synchronize_session=False)
    db.flush()
    if created:
        log_activity(
            db,
            "project.board.backfill",
            "project_board",
            f"补齐项目默认板 {created} 块",
            detail={"created": created},
        )
    return created


def sync_bom_solder_points(db: Session, item: ProjectBomItem, board_id: int | None = None) -> None:
    specs = solder_point_specs_for_item(item)
    boards = project_boards_for_sync(db, item.project_id, board_id)
    for board in boards:
        existing = {
            (point.designator or "").upper(): point
            for point in db.query(ProjectBomSolderPoint)
            .filter(ProjectBomSolderPoint.bom_item_id == item.id, ProjectBomSolderPoint.board_id == board.id)
            .all()
        }
        for spec in specs:
            key = spec["designator"].upper()
            point = existing.get(key)
            if point:
                point.bom_value = spec.get("bom_value") or point.bom_value
                point.bom_model = spec.get("bom_model") or point.bom_model
                point.bom_footprint = spec.get("bom_footprint") or point.bom_footprint
            else:
                db.add(ProjectBomSolderPoint(bom_item_id=item.id, board_id=board.id, **spec))


def ensure_bom_solder_points(db: Session) -> int:
    items = db.query(ProjectBomItem).options(joinedload(ProjectBomItem.component)).all()
    before = db.query(func.count(ProjectBomSolderPoint.id)).scalar() or 0
    for item in items:
        sync_bom_solder_points(db, item)
    db.flush()
    after = db.query(func.count(ProjectBomSolderPoint.id)).scalar() or 0
    created = int(after - before)
    if created:
        log_activity(
            db,
            "bom.solder.backfill",
            "project_bom_solder_point",
            f"补齐 BOM 焊接位号 {created} 项",
            detail={"created": created},
        )
    return created


def solder_point_out(point: ProjectBomSolderPoint) -> dict:
    return {
        "id": point.id,
        "board_id": point.board_id,
        "designator": point.designator,
        "bom_value": point.bom_value,
        "bom_model": point.bom_model,
        "bom_footprint": point.bom_footprint,
        "soldered": bool(point.soldered),
        "soldered_at": point.soldered_at,
        "stock_applied": bool(getattr(point, "stock_applied", False)),
        "lost": bool(getattr(point, "lost", False)),
        "lost_at": getattr(point, "lost_at", None),
        "loss_stock_applied": bool(getattr(point, "loss_stock_applied", False)),
        "loss_note": getattr(point, "loss_note", None),
        "note": point.note,
    }


def board_out(board: ProjectBoard, points: list[ProjectBomSolderPoint] | None = None) -> dict:
    board_points = points if points is not None else list(getattr(board, "solder_points", []) or [])
    total = len(board_points)
    soldered = sum(1 for point in board_points if point.soldered)
    lost = sum(1 for point in board_points if getattr(point, "lost", False))
    pending = sum(1 for point in board_points if not point.soldered)
    return {
        "id": board.id,
        "project_id": board.project_id,
        "board_index": int(board.board_index or 1),
        "name": board.name,
        "status": board.status or "active",
        "note": board.note,
        "solder_total": total,
        "soldered_count": soldered,
        "lost_count": lost,
        "pending_count": pending,
        "solder_progress": round((soldered / total) * 100) if total else 0,
        "completed_at": board.completed_at,
        "created_at": board.created_at,
        "updated_at": board.updated_at,
    }


def bom_item_out(
    item: ProjectBomItem,
    reserved_by_component: dict[int, int] | None = None,
    substitution_suggestions: list[dict] | None = None,
) -> dict:
    reserved_by_component = reserved_by_component or {}
    solder_points = sorted(getattr(item, "solder_points", []) or [], key=lambda point: point.id)
    solder_total = len(solder_points)
    soldered_count = sum(1 for point in solder_points if point.soldered)
    lost_count = sum(1 for point in solder_points if getattr(point, "lost", False))
    pending_count = sum(1 for point in solder_points if not point.soldered)
    status = item.status or "reserved"
    component_quantity = item.component.quantity if item.component else 0
    total_reserved = reserved_by_component.get(item.component_id, 0)
    own_reserved = pending_count if solder_points and status == "reserved" else (int(item.required_quantity or 0) if status == "reserved" else 0)
    reserved_by_others = max(0, total_reserved - own_reserved)
    available_for_item = max(0, component_quantity - reserved_by_others)
    free_quantity = max(0, component_quantity - total_reserved)
    shortage = max(0, own_reserved - available_for_item) if status == "reserved" else 0
    return {
        "id": item.id,
        "component_id": item.component_id,
        "required_quantity": item.required_quantity,
        "status": status,
        "remark": item.remark,
        "component": component_out(item.component, total_reserved) if item.component else None,
        "available_quantity": available_for_item,
        "reserved_quantity": total_reserved,
        "free_quantity": free_quantity,
        "shortage_quantity": shortage,
        "enough": shortage == 0,
        "solder_points": [solder_point_out(point) for point in solder_points],
        "soldered_count": soldered_count,
        "solder_total": solder_total,
        "lost_count": lost_count,
        "pending_count": pending_count,
        "solder_progress": round((soldered_count / solder_total) * 100) if solder_total else 0,
        "substitution_suggestions": substitution_suggestions or [],
    }


def project_out(project: Project, reserved_by_component: dict[int, int] | None = None) -> dict:
    reserved_by_component = reserved_by_component or {}
    db = object_session(project)
    substitutions = (
        substitution_suggestions_for_bom_items(db, project.bom_items, reserved_by_component)
        if db is not None
        else {}
    )
    boards = sorted(getattr(project, "boards", []) or [], key=lambda board: (board.board_index or 0, board.id or 0))
    all_points = [point for item in getattr(project, "bom_items", []) or [] for point in getattr(item, "solder_points", []) or []]
    points_by_board: dict[int, list[ProjectBomSolderPoint]] = {}
    for point in all_points:
        if point.board_id:
            points_by_board.setdefault(point.board_id, []).append(point)
    solder_total = len(all_points)
    soldered_count = sum(1 for point in all_points if point.soldered)
    lost_count = sum(1 for point in all_points if getattr(point, "lost", False))
    pending_count = sum(1 for point in all_points if not point.soldered)
    return {
        "id": project.id,
        "project_code": project.project_code,
        "name": project.name,
        "description": project.description,
        "status": project.status,
        "ai_bom_analysis": project.ai_bom_analysis,
        "ai_bom_cache_key": project.ai_bom_cache_key,
        "ai_bom_updated_at": project.ai_bom_updated_at,
        "bom_match_total": project.bom_match_total or 0,
        "bom_match_matched": project.bom_match_matched or 0,
        "bom_match_review": project.bom_match_review or 0,
        "bom_match_missing": project.bom_match_missing or 0,
        "bom_match_missing_items": project.bom_match_missing_items,
        "bom_match_rows": project.bom_match_rows,
        "bom_match_updated_at": project.bom_match_updated_at,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
        "boards": [board_out(board, points_by_board.get(board.id, [])) for board in boards],
        "active_board_id": boards[0].id if boards else None,
        "board_count": len(boards),
        "soldered_count": soldered_count,
        "solder_total": solder_total,
        "lost_count": lost_count,
        "pending_count": pending_count,
        "solder_progress": round((soldered_count / solder_total) * 100) if solder_total else 0,
        "bom_items": [bom_item_out(item, reserved_by_component, substitutions.get(item.id)) for item in project.bom_items],
    }


@app.get("/health")
def health():
    return {"status": "ok"}


def public_status_checked_at(value: datetime | None = None) -> str:
    current = value or datetime.utcnow()
    return (current + timedelta(hours=8)).replace(microsecond=0).isoformat() + "+08:00"


def public_status_datetime(value: datetime | None) -> str | None:
    if not value:
        return None
    if value.tzinfo:
        return value.replace(microsecond=0).isoformat()
    local_now = datetime.utcnow() + timedelta(hours=8)
    utc_shifted = value + timedelta(hours=8)
    if utc_shifted <= local_now + timedelta(minutes=5):
        value = utc_shifted
    return value.replace(microsecond=0).isoformat() + "+08:00"


def public_status_component(name: str, status: str, label: str, **extra) -> dict:
    safe_status = status if status in PUBLIC_STATUS_RANK else "unknown"
    payload = {
        "name": name,
        "status": safe_status,
        "label": str(label or "").strip()[:80] or name,
    }
    for key in ("latencyMs", "lastSuccessAt"):
        value = extra.get(key)
        if value is not None:
            payload[key] = value
    return payload


def public_status_overall(components: list[dict]) -> str:
    if not components:
        return "unknown"
    return max(
        (item.get("status", "unknown") for item in components),
        key=lambda item: PUBLIC_STATUS_RANK.get(item, PUBLIC_STATUS_RANK["unknown"]),
    )


def account_center_public_health_component() -> dict:
    if auth_module.AUTH_MODE != "account-v1":
        return public_status_component("auth", "operational", "本地账号模式")
    health_url = auth_module.account_public_url("/health")
    if not health_url:
        return public_status_component("auth", "degraded", "统一账号未配置")
    start = time.monotonic()
    try:
        response = httpx.get(health_url, timeout=min(2.0, auth_module.AUTH_HTTP_TIMEOUT_SECONDS))
        latency_ms = int((time.monotonic() - start) * 1000)
        if response.status_code >= 500:
            return public_status_component("auth", "outage", "统一账号不可用", latencyMs=latency_ms)
        if response.status_code >= 400:
            return public_status_component("auth", "degraded", "统一账号状态异常", latencyMs=latency_ms)
        try:
            payload = response.json()
        except ValueError:
            payload = {}
        ok = payload.get("ok") if isinstance(payload, dict) else None
        status = payload.get("status") if isinstance(payload, dict) else ""
        if ok is False or status in {"outage", "down"}:
            return public_status_component("auth", "outage", "统一账号不可用", latencyMs=latency_ms)
        if status in {"degraded", "maintenance"}:
            return public_status_component("auth", "degraded", "统一账号状态降级", latencyMs=latency_ms)
        return public_status_component("auth", "operational", "统一账号可响应", latencyMs=latency_ms)
    except httpx.RequestError:
        return public_status_component("auth", "degraded", "统一账号检查超时")


@app.get("/health/status")
def health_status(db: Session = Depends(get_db)):
    components = [public_status_component("web", "operational", "Web/API 可响应")]
    metrics: dict[str, int | str | None] = {
        "uptimeSeconds": max(0, int(time.monotonic() - PROCESS_STARTED_AT)),
        "p95LatencyMs": None,
        "queuedJobs": None,
        "lastJobSuccessAt": None,
    }
    try:
        db.execute(text("SELECT 1")).scalar()
        components.append(public_status_component("database", "operational", "数据库可查询"))

        component_count = db.query(func.count(Component.id)).scalar() or 0
        project_count = db.query(func.count(Project.id)).scalar() or 0
        import_count = db.query(func.count(OrderImportBatch.id)).scalar() or 0
        components.append(
            public_status_component(
                "warehouse",
                "operational",
                f"{component_count} 个器件，{project_count} 个项目，{import_count} 个导入批次",
            )
        )

        queued_statuses = ("pending", "processing", "stale")
        queued_jobs = db.query(func.count(AiTask.id)).filter(AiTask.status.in_(queued_statuses)).scalar() or 0
        failed_jobs = db.query(func.count(AiTask.id)).filter(AiTask.status == "failed").scalar() or 0
        last_success_at = (
            db.query(AiTask.finished_at)
            .filter(AiTask.status == "completed", AiTask.finished_at.isnot(None))
            .order_by(AiTask.finished_at.desc())
            .limit(1)
            .scalar()
        )
        metrics["queuedJobs"] = int(queued_jobs)
        metrics["lastJobSuccessAt"] = public_status_datetime(last_success_at)
        ai_status = "degraded" if failed_jobs else "operational"
        ai_label = f"AI 队列 {queued_jobs} 个，失败 {failed_jobs} 个" if queued_jobs or failed_jobs else "AI 队列空闲"
        components.append(
            public_status_component(
                "ai",
                ai_status,
                ai_label,
                lastSuccessAt=metrics["lastJobSuccessAt"],
            )
        )

        last_activity_at = db.query(func.max(ActivityLog.created_at)).scalar()
        components.append(
            public_status_component(
                "activity",
                "operational",
                "操作记录可查询",
                lastSuccessAt=public_status_datetime(last_activity_at),
            )
        )
    except Exception:
        components.append(public_status_component("database", "outage", "数据库不可用"))
        metrics["queuedJobs"] = 0

    components.append(account_center_public_health_component())
    status = public_status_overall(components)
    return {
        "ok": status not in {"outage", "unknown"},
        "status": status,
        "service": "component-warehouse",
        "version": APP_VERSION,
        "checkedAt": public_status_checked_at(),
        "components": components[:12],
        "metrics": metrics,
    }


@app.get("/api/auth/config")
def auth_config():
    return auth_public_config()


@app.get("/api/auth/account/health")
async def auth_account_health(request: Request):
    enforce_auth_proxy_limit("captcha_ip", client_ip(request))
    return await account_v1_proxy_request("GET", "/health")


@app.get("/api/auth/account/captcha")
async def auth_account_captcha(request: Request):
    enforce_auth_proxy_limit("captcha_ip", client_ip(request))
    return await account_v1_proxy_request("GET", "/captcha")


@app.post("/api/auth/account/sms/send")
async def auth_account_send_sms(payload: dict, request: Request, _: Protected):
    ip = client_ip(request)
    phone = auth_proxy_phone(payload)
    purpose = str(payload.get("purpose") or "")[:40]
    if purpose not in {"change_password", "change_phone"}:
        legacy_login_retired()
    enforce_auth_proxy_limit("sms_ip", ip)
    enforce_auth_proxy_limit("sms_phone", phone)
    clean_payload = {
        "phone": phone,
        "purpose": purpose,
        "captchaId": str(payload.get("captchaId") or "")[:120],
        "captchaAnswer": str(payload.get("captchaAnswer") or "")[:20],
    }
    return await account_v1_proxy_request("POST", "/sms/send", clean_payload)


@app.post("/api/auth/account/login/sms")
async def auth_account_login_sms():
    legacy_login_retired()


@app.post("/api/auth/account/login/password")
async def auth_account_login_password():
    legacy_login_retired()


@app.post("/api/auth/account/sso/start")
async def auth_account_sso_start(payload: dict, request: Request, response: Response):
    enforce_auth_proxy_limit("login_ip", client_ip(request))
    return account_sso_start(payload, request, response)


@app.post("/api/auth/account/sso/token")
async def auth_account_sso_token(payload: dict, request: Request, response: Response):
    enforce_auth_proxy_limit("login_ip", client_ip(request))
    try:
        return await account_sso_token_request(payload, request)
    finally:
        if request.cookies.get(SSO_STATE_COOKIE):
            clear_sso_cookie(response)


@app.post("/api/auth/account/register")
async def auth_account_register():
    legacy_login_retired()


@app.post("/api/auth/account/password/reset")
async def auth_account_password_reset():
    legacy_login_retired()


@app.post("/api/auth/account/token/refresh")
async def auth_account_token_refresh(payload: dict, request: Request):
    enforce_auth_proxy_limit("login_ip", client_ip(request))
    return await account_v1_proxy_request("POST", "/token/refresh", {
        "refreshToken": str(payload.get("refreshToken") or ""),
    })


@app.get("/api/auth/account/me")
async def auth_account_me(authorization: str | None = Header(default=None)):
    token = auth_module.extract_bearer_token(authorization)
    if not token:
        raise HTTPException(status_code=401, detail="请先登录")
    return await account_v1_proxy_request("GET", "/me", token=token)


@app.post("/api/auth/account/logout")
async def auth_account_logout(request: Request, authorization: str | None = Header(default=None)):
    enforce_auth_proxy_limit("login_ip", client_ip(request))
    token = authorization.split(" ", 1)[1].strip() if authorization and authorization.lower().startswith("bearer ") else ""
    return await account_v1_proxy_request("POST", "/logout", token=token)


@app.post("/api/auth/local/register")
def auth_local_register():
    legacy_login_retired()


@app.post("/api/auth/local/login")
def auth_local_login():
    legacy_login_retired()


@app.post("/api/auth/local/token/refresh")
def auth_local_refresh():
    legacy_login_retired()


@app.post("/api/auth/local/logout")
def auth_local_logout():
    legacy_login_retired()


@app.patch("/api/auth/local/me")
def auth_local_update_profile():
    legacy_login_retired()


@app.post("/api/auth/local/password/change")
def auth_local_change_password():
    legacy_login_retired()


@app.get("/api/auth/me")
def auth_me(auth: Protected):
    return {
        "ok": True,
        "user": {
            "id": auth.user_id,
            "accountId": auth.account_id,
            "phone": auth.phone,
            "displayName": auth.nickname,
            "nickname": auth.nickname,
            "avatarUrl": auth.avatar_url,
            "isAdmin": auth.is_admin,
        },
        "auth_degraded": auth.auth_degraded,
    }


@app.get("/api/categories", response_model=list[CategoryOut])
def list_categories(_: Protected, db: Session = Depends(get_db)):
    return db.query(Category).order_by(Category.id).all()


@app.get("/api/public/components/{warehouse_code}")
def public_component_detail(warehouse_code: str, db: Session = Depends(get_db)):
    identity = identity_by_code(db, warehouse_code)
    if not identity:
        raise HTTPException(status_code=404, detail="Component not found")
    component = None
    if identity.component_id and identity.status == "active":
        component = (
            db.query(Component)
            .options(joinedload(Component.category))
            .filter(Component.id == identity.component_id)
            .first()
        )
    return public_identity_out(identity, component)


@app.get("/api/components/access-context/{warehouse_code}")
def component_access_context(
    warehouse_code: str,
    auth: Protected,
    db: Session = Depends(get_db),
):
    identity = identity_by_code(db, warehouse_code)
    if not identity:
        raise HTTPException(status_code=404, detail="器件不存在")
    component = db.get(Component, identity.component_id) if identity.component_id else None
    owner = bool(component and component.owner_user_id == auth.user_id)
    teams = []
    if component:
        rows = (
            db.query(
                CompetitionLibraryComponent,
                CompetitionLibrary,
                CompetitionLibraryMember,
            )
            .join(
                CompetitionLibrary,
                CompetitionLibrary.id == CompetitionLibraryComponent.library_id,
            )
            .join(
                CompetitionLibraryMember,
                and_(
                    CompetitionLibraryMember.library_id == CompetitionLibrary.id,
                    CompetitionLibraryMember.user_id == auth.user_id,
                    CompetitionLibraryMember.status == "active",
                ),
            )
            .filter(CompetitionLibraryComponent.cw_component_id == component.id)
            .order_by(CompetitionLibrary.name.asc())
            .all()
        )
        teams = [
            {
                "library_id": library.id,
                "library_name": library.name,
                "item_id": item.id,
                "role": member.role,
                "can_edit_quantity": bool(
                    item.source_user_id == auth.user_id or member.role == "captain"
                ),
            }
            for item, library, member in rows
        ]
    return {
        "warehouse_code": identity.code,
        "archived": identity.status == "archived",
        "owner": owner,
        "personal_component_id": component.id if owner and component else None,
        "teams": teams,
    }


@app.get("/api/public/projects/{project_code}")
def public_project_detail(project_code: str, request: Request, db: Session = Depends(get_db)):
    code = normalize_project_code(project_code)
    if not code:
        raise HTTPException(status_code=404, detail="Project not found")
    project = (
        db.query(Project)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .filter(func.lower(Project.project_code) == code.lower())
        .first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return public_project_out(project, request)


@app.get("/api/public/projects/{project_code}/qr.svg")
def public_project_qr(project_code: str, request: Request, db: Session = Depends(get_db)):
    code = normalize_project_code(project_code)
    if not code:
        raise HTTPException(status_code=404, detail="Project not found")
    exists = db.query(Project.id).filter(func.lower(Project.project_code) == code.lower()).first()
    if not exists:
        raise HTTPException(status_code=404, detail="Project not found")
    return Response(
        content=qr_svg_markup(frontend_project_url(request, code)),
        media_type="image/svg+xml",
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/categories", response_model=CategoryOut)
def create_category(name: str, _: Protected, db: Session = Depends(get_db)):
    existing = db.query(Category).filter(Category.name == name).first()
    if existing:
        return existing
    raise HTTPException(status_code=403, detail="类别由系统统一维护，请使用已有类别，避免重复和误分类")


@app.patch("/api/categories/{category_id}/code-prefix", response_model=CategoryOut)
def update_category_code_prefix(
    category_id: int,
    payload: CategoryPrefixUpdate,
    _: Protected,
    db: Session = Depends(get_db),
):
    raise HTTPException(status_code=403, detail="类别编号由系统自动生成，不能手工修改")


@app.get("/api/components", response_model=ComponentList)
def list_components(
    auth: Protected,
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    category_id: int | None = None,
    status: str | None = None,
    tag: str | None = None,
    keyword: str | None = None,
    package: str | None = None,
    ai_status: str | None = None,
    stock: str | None = None,
    is_hand_solder_friendly: bool | None = None,
    is_power_component: bool | None = None,
    is_signal_component: bool | None = None,
    is_high_current: bool | None = None,
    is_high_voltage: bool | None = None,
):
    query = db.query(Component).options(joinedload(Component.category))
    query = filter_owner(query, Component, auth)
    query = apply_component_filters(
        query,
        category_id=category_id,
        status=status,
        tag=tag,
        keyword=keyword,
        package=package,
        ai_status=ai_status,
        stock=stock,
        is_hand_solder_friendly=is_hand_solder_friendly,
        is_power_component=is_power_component,
        is_signal_component=is_signal_component,
        is_high_current=is_high_current,
        is_high_voltage=is_high_voltage,
    )
    all_items = sort_components_by_value(filter_low_stock_components(db, query.all(), stock))
    total = len(all_items)
    items = all_items[(page - 1) * page_size : page * page_size]
    reserved = reserved_quantities(db, [item.id for item in items])
    return {"items": [component_out(item, reserved.get(item.id, 0)) for item in items], "total": total}


def group_component_page(db: Session, components: list[Component], total: int, page: int, page_size: int) -> dict:
    reserved = reserved_quantities(db, [item.id for item in components])
    items = [component_out(item, reserved.get(item.id, 0)) for item in components]
    by_category: dict[int | None, dict] = {}
    for item in items:
        category = item["category"]
        key = category.id if category else None
        if key not in by_category:
            by_category[key] = {"category": category, "items": [], "total": 0}
        by_category[key]["items"].append(item)
        by_category[key]["total"] += 1
    return {"groups": list(by_category.values()), "total": total, "page": page, "page_size": page_size}


def component_label_title(component: Component) -> str:
    return str(component.model or component.normalized_spec or component.name or component.warehouse_code or component.id)[:48]


def component_label_subtitle(component: Component) -> str:
    parts = [component.normalized_spec, component.package, component.lcsc_number]
    text_value = " / ".join(str(part) for part in parts if part)
    return text_value[:72]


def component_label_type(component: Component) -> str:
    return str(component.category.name if component.category else component.part_family or "元器件")[:12]


@lru_cache(maxsize=1)
def brand_logo_data_uri() -> str:
    if not APP_SHOW_BRAND_LOGO:
        return ""
    path = os.path.join(os.path.dirname(__file__), "assets", "brand-logo.png")
    try:
        with open(path, "rb") as logo_file:
            encoded = base64.b64encode(logo_file.read()).decode("ascii")
            return f"data:image/png;base64,{encoded}"
    except OSError:
        return ""


@app.api_route("/api/assets/{asset_name}", methods=["GET", "HEAD"])
def public_brand_asset(asset_name: str):
    allowed_assets = {
        "brand-logo.png": "brand-logo.png",
        "brand-logo-label.png": "brand-logo-label.png",
    }
    filename = allowed_assets.get(asset_name)
    if not filename:
        raise HTTPException(status_code=404, detail="Asset not found")
    if not APP_SHOW_BRAND_LOGO:
        raise HTTPException(status_code=404, detail="Brand logo is disabled")
    path = os.path.join(os.path.dirname(__file__), "assets", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Brand logo not found")
    return FileResponse(
        path,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=86400, stale-while-revalidate=604800"},
    )


@lru_cache(maxsize=1)
def nfc_icon_data_uri() -> str:
    path = os.path.join(os.path.dirname(__file__), "assets", "nfc.svg")
    try:
        with open(path, "rb") as icon_file:
            encoded = base64.b64encode(icon_file.read()).decode("ascii")
        return f"data:image/svg+xml;base64,{encoded}"
    except OSError:
        return ""


@lru_cache(maxsize=2048)
def qr_svg_markup(value: str) -> str:
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=5, border=1)
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(image_factory=SvgPathImage)
    buffer = io.BytesIO()
    image.save(buffer)
    markup = buffer.getvalue().decode("utf-8")
    markup = re.sub(r"^\s*<\?xml[^>]*>\s*", "", markup)
    if "preserveAspectRatio=" not in markup:
        markup = markup.replace("<svg ", '<svg preserveAspectRatio="xMidYMid meet" ', 1)
    if "shape-rendering=" not in markup:
        markup = markup.replace("<svg ", '<svg shape-rendering="crispEdges" ', 1)
    return markup


def frontend_component_url(request: Request, code: str) -> str:
    if PUBLIC_PERSONAL_BASE_URL:
        return f"{PUBLIC_PERSONAL_BASE_URL}/scan/{quote(code)}"
    forwarded_proto = request.headers.get("x-forwarded-proto")
    scheme = forwarded_proto or request.url.scheme
    host = request.headers.get("host") or request.url.netloc
    return f"{scheme}://{host}/component-warehouse/personal/scan/{quote(code)}"


def frontend_project_url(request: Request, code: str) -> str:
    if PUBLIC_PERSONAL_BASE_URL:
        return f"{PUBLIC_PERSONAL_BASE_URL}/public/projects/{quote(code)}"
    forwarded_proto = request.headers.get("x-forwarded-proto")
    scheme = forwarded_proto or request.url.scheme
    host = request.headers.get("host") or request.url.netloc
    return f"{scheme}://{host}/component-warehouse/personal/public/projects/{quote(code)}"


def export_components_from_ids(
    db: Session,
    ids: list[int],
    auth: AuthContext | None = None,
    export_all: bool = False,
    imported_from: str | None = None,
    imported_to: str | None = None,
    excluded_categories: list[str] | None = None,
) -> list[Component]:
    clean_ids = [int(item) for item in ids if int(item) > 0]
    if not clean_ids and not export_all:
        return []
    query = (
        filter_owner(db.query(Component), Component, auth)
        .options(joinedload(Component.category))
    )
    if not export_all:
        query = query.filter(Component.id.in_(clean_ids))
    query = apply_label_import_date_filter(query, Component, imported_from, imported_to)
    query = apply_label_category_exclusion(query, excluded_categories)
    rows = query.order_by(Component.id.asc()).all()
    changed = False
    for component in rows:
        if not component.warehouse_code:
            assign_component_warehouse_code(db, component)
            changed = True
    if changed:
        db.commit()
    return rows


def parse_label_export_datetime(value: str | None, *, end: bool = False) -> datetime | None:
    text_value = str(value or "").strip()
    if not text_value:
        return None
    is_date_only = bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", text_value))
    try:
        parsed = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="标签导出日期格式无效") from exc
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone().replace(tzinfo=None)
    if end and is_date_only:
        parsed = parsed + timedelta(days=1)
    return parsed


def apply_label_import_date_filter(query, model, imported_from: str | None, imported_to: str | None):
    start = parse_label_export_datetime(imported_from)
    end = parse_label_export_datetime(imported_to, end=True)
    if start and end and start >= end:
        raise HTTPException(status_code=400, detail="标签导出开始日期必须早于结束日期")
    if not start and not end:
        return query
    if hasattr(model, "first_stocked_at"):
        date_expr = func.coalesce(model.first_stocked_at, model.created_at)
    else:
        date_expr = model.created_at
    if start:
        query = query.filter(date_expr >= start)
    if end:
        query = query.filter(date_expr < end)
    return query


def normalize_label_excluded_categories(values: list[str] | None) -> set[str]:
    result: set[str] = set()
    for value in values or []:
        name = str(value or "").strip()
        if name:
            result.add(name)
    return result


def apply_label_category_exclusion(query, excluded_categories: list[str] | None):
    names = normalize_label_excluded_categories(excluded_categories)
    if not names:
        return query
    include_uncategorized = "未分类" in names
    concrete_names = {name for name in names if name != "未分类"}
    if concrete_names:
        query = query.filter(~Component.category.has(Category.name.in_(concrete_names)))
    if include_uncategorized:
        query = query.filter(Component.category_id.isnot(None))
    return query


@app.post("/api/components/export/id-table")
def export_component_id_table(payload: ComponentExportRequest, auth: Protected, db: Session = Depends(get_db)):
    components = export_components_from_ids(db, payload.ids, auth, payload.all)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["器件 ID", "型号", "名称", "分类", "封装", "核心规格", "立创 ID", "数量", "可用库存"])
    reserved = reserved_quantities(db, [item.id for item in components])
    for component in components:
        quantity = int(component.quantity or 0)
        writer.writerow(
            [
                component.warehouse_code or "",
                component.model or "",
                component.name or "",
                component.category.name if component.category else "",
                component.package or "",
                component.normalized_spec or "",
                component.lcsc_number or "",
                quantity,
                max(0, quantity - reserved.get(component.id, 0)),
            ]
        )
    data = "\ufeff" + output.getvalue()
    return Response(
        content=data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="component-ids.csv"'},
    )


@app.get("/api/components/export/inventory.xlsx")
def export_component_inventory(auth: Protected, db: Session = Depends(get_db)):
    components = export_components_from_ids(db, [], auth, True)
    reserved = reserved_quantities(db, [item.id for item in components])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "元器件库存"
    sheet.append(
        [
            "器件 ID", "名称", "分类", "参数值", "封装", "厂商", "MPN", "LCSC",
            "总库存", "预占", "可用库存", "最低库存", "常用", "位置", "描述", "标签", "备注",
        ]
    )
    for component in components:
        quantity = int(component.quantity or 0)
        reserved_quantity = int(reserved.get(component.id, 0))
        sheet.append(
            [
                component.warehouse_code or "",
                component.name or "",
                component.category.name if component.category else "",
                component.parameters or component.normalized_spec or "",
                component.package or "",
                component.manufacturer or "",
                component.model or "",
                component.lcsc_number or "",
                quantity,
                reserved_quantity,
                max(0, quantity - reserved_quantity),
                int(component.safety_quantity or 0),
                "是" if component.is_common else "否",
                component.location or "",
                component.description or "",
                component.tags or "",
                component.remark or "",
            ]
        )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="component-inventory.xlsx"'},
    )


@app.post("/api/components/export/label-sheet")
def export_component_label_sheet(payload: ComponentExportRequest, request: Request, auth: Protected, db: Session = Depends(get_db)):
    components = export_components_from_ids(
        db,
        payload.ids,
        auth,
        payload.all,
        imported_from=payload.imported_from,
        imported_to=payload.imported_to,
        excluded_categories=payload.excluded_categories,
    )
    records = [component_out(component, 0) for component in components]
    printed_at = print_timestamp()
    output_format = (payload.output_format or "html").lower()
    if output_format == "pdf":
        appended_pdf_items = [] if payload.calibration else component_export_custom_label_pdf_items(db, auth, payload.custom_labels, records)
        pdf_doc = render_component_label_pdf(
            records,
            PUBLIC_PERSONAL_BASE_URL,
            start_slot=payload.start_slot,
            copies=payload.copies,
            offset_x_mm=payload.offset_x_mm,
            offset_y_mm=payload.offset_y_mm,
            calibration=payload.calibration,
            printed_at=printed_at,
            safe_margin=payload.safe_margin,
            appended_items=appended_pdf_items,
        )
        return Response(
            content=pdf_doc,
            media_type="application/pdf",
            headers={"Content-Disposition": 'inline; filename="component-labels.pdf"'},
        )
    appended_cards = [] if payload.calibration else component_export_custom_label_cards(db, auth, payload.custom_labels, printed_at, records)
    appended_font_keys = set() if payload.calibration else component_export_custom_label_font_keys(db, auth, payload.custom_labels)
    html_doc = render_component_label_sheet(
        records,
        PUBLIC_PERSONAL_BASE_URL,
        start_slot=payload.start_slot,
        copies=payload.copies,
        offset_x_mm=payload.offset_x_mm,
        offset_y_mm=payload.offset_y_mm,
        calibration=payload.calibration,
        appended_cards=appended_cards,
        printed_at=printed_at,
        safe_margin=payload.safe_margin,
        font_keys=appended_font_keys,
    )
    return Response(
        content=html_doc,
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": 'inline; filename="component-labels.html"'},
    )


@app.get("/api/custom-labels", response_model=list[CustomLabelTemplateOut])
def list_custom_labels(auth: Protected, db: Session = Depends(get_db)):
    rows = (
        custom_label_query_for_auth(db, auth)
        .order_by(CustomLabelTemplate.updated_at.desc(), CustomLabelTemplate.created_at.desc())
        .all()
    )
    return [custom_label_template_out(row, db) for row in rows]


@app.get("/api/custom-labels/category-summary")
def custom_label_category_summary(auth: Protected, db: Session = Depends(get_db)):
    rows = (
        filter_owner(db.query(Component).options(joinedload(Component.category)), Component, auth)
        .order_by(Component.category_id.asc(), Component.id.asc())
        .all()
    )
    summary = category_package_summary_from_records([component_to_dict(row) for row in rows])
    return [{"category": category, "summary": value} for category, value in summary.items()]


@app.post("/api/custom-labels", response_model=CustomLabelTemplateOut)
def create_custom_label(payload: CustomLabelTemplateCreate, auth: Protected, db: Session = Depends(get_db)):
    template = CustomLabelTemplate(
        id=secrets.token_hex(16),
        scope_type="personal",
        owner_user_id=owner_id(auth),
        team_library_id=None,
        name=payload.name.strip(),
        content_json=json.dumps(clean_custom_label_content(payload.content), ensure_ascii=False),
        status="active",
        created_by_user_id=owner_id(auth) or 0,
    )
    db.add(template)
    log_activity(
        db,
        "custom_label.create",
        "custom_label",
        f"新增自定义标签模板 {template.name}",
        owner_user_id=owner_id(auth),
        entity_id=0,
        detail={"template_id": template.id},
    )
    db.commit()
    db.refresh(template)
    return custom_label_template_out(template, db)


@app.put("/api/custom-labels/{template_id}", response_model=CustomLabelTemplateOut)
def update_custom_label(template_id: str, payload: CustomLabelTemplateUpdate, auth: Protected, db: Session = Depends(get_db)):
    template = require_custom_label_template(db, template_id, auth)
    if payload.name is not None:
        template.name = payload.name.strip()
    if payload.content is not None:
        template.content_json = json.dumps(clean_custom_label_content(payload.content), ensure_ascii=False)
    log_activity(
        db,
        "custom_label.update",
        "custom_label",
        f"更新自定义标签模板 {template.name}",
        owner_user_id=owner_id(auth),
        entity_id=0,
        detail={"template_id": template.id},
    )
    db.commit()
    db.refresh(template)
    return custom_label_template_out(template, db)


@app.delete("/api/custom-labels/{template_id}")
def archive_custom_label(template_id: str, auth: Protected, db: Session = Depends(get_db)):
    template = require_custom_label_template(db, template_id, auth)
    template.status = "archived"
    template.archived_at = datetime.utcnow()
    log_activity(
        db,
        "custom_label.archive",
        "custom_label",
        f"归档自定义标签模板 {template.name}",
        owner_user_id=owner_id(auth),
        entity_id=0,
        detail={"template_id": template.id},
    )
    db.commit()
    return {"archived": True}


@app.post("/api/custom-labels/{template_id}/assets")
async def upload_custom_label_asset(template_id: str, auth: Protected, file: UploadFile = File(...), db: Session = Depends(get_db)):
    template = require_custom_label_template(db, template_id, auth)
    asset = await save_custom_label_asset(db, template, file)
    return custom_label_asset_out(asset)


@app.get("/api/custom-labels/assets/{asset_id}")
def get_custom_label_asset(asset_id: str, auth: Protected, db: Session = Depends(get_db)):
    asset = require_custom_label_asset(db, asset_id, auth)
    path = Path(asset.storage_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="标签素材文件不存在")
    return FileResponse(path, media_type=asset.mime_type, filename=asset.file_name)


@app.post("/api/custom-labels/export-sheet")
def export_custom_label_sheet(payload: CustomLabelExportRequest, auth: Protected, db: Session = Depends(get_db)):
    template = require_custom_label_template(db, payload.template_id, auth) if payload.template_id else None
    content = clean_custom_label_content(payload.content if payload.content is not None else parse_custom_label_content(template.content_json if template else None))
    html_doc = render_custom_label_sheet(
        content,
        asset_resolver=custom_label_asset_resolver(db, template) if template else None,
        start_slot=payload.start_slot,
        copies=payload.copies,
        offset_x_mm=payload.offset_x_mm,
        offset_y_mm=payload.offset_y_mm,
        calibration=payload.calibration,
        safe_margin=payload.safe_margin,
    )
    return Response(
        content=html_doc,
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": 'inline; filename="custom-labels.html"'},
    )


@app.get("/api/components/grouped-page", response_model=ComponentGroupPage)
def list_components_grouped_page(
    auth: Protected,
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(3, ge=1, le=10),
    keyword: str | None = None,
    category_id: int | None = None,
    status: str | None = None,
    ai_status: str | None = None,
    stock: str | None = None,
    tag: str | None = None,
    package: str | None = None,
    is_hand_solder_friendly: bool | None = None,
    is_power_component: bool | None = None,
    is_signal_component: bool | None = None,
    is_high_current: bool | None = None,
    is_high_voltage: bool | None = None,
):
    query = db.query(Component).outerjoin(Category).options(joinedload(Component.category))
    query = filter_owner(query, Component, auth)
    query = apply_component_filters(
        query,
        category_id=category_id,
        status=status,
        tag=tag,
        keyword=keyword,
        package=package,
        ai_status=ai_status,
        stock=stock,
        is_hand_solder_friendly=is_hand_solder_friendly,
        is_power_component=is_power_component,
        is_signal_component=is_signal_component,
        is_high_current=is_high_current,
        is_high_voltage=is_high_voltage,
    )
    all_components = sort_components_by_value(filter_low_stock_components(db, query.all(), stock))
    total = len(all_components)
    grouped_components: dict[int | None, list[Component]] = {}
    for component in all_components:
        grouped_components.setdefault(component.category_id, []).append(component)
    category_groups = list(grouped_components.values())
    category_total = len(category_groups)
    start = (page - 1) * page_size
    components = [
        component
        for group in category_groups[start : start + page_size]
        for component in group
    ]
    result = group_component_page(db, components, total, page, page_size)
    result["category_total"] = category_total
    result["has_more"] = start + page_size < category_total
    return result


@app.get("/api/components/grouped", response_model=list[ComponentGroup])
def list_components_grouped(
    auth: Protected,
    db: Session = Depends(get_db),
    keyword: str | None = None,
    category_id: int | None = None,
    status: str | None = None,
    ai_status: str | None = None,
    stock: str | None = None,
    tag: str | None = None,
    package: str | None = None,
    is_hand_solder_friendly: bool | None = None,
    is_power_component: bool | None = None,
    is_signal_component: bool | None = None,
    is_high_current: bool | None = None,
    is_high_voltage: bool | None = None,
):
    query = db.query(Component).outerjoin(Category).options(joinedload(Component.category))
    query = filter_owner(query, Component, auth)
    if category_id:
        query = query.filter(Component.category_id == category_id)
    if status:
        query = query.filter(Component.status == status)
    if tag:
        query = query.filter(or_(Component.tags.ilike(f"%{tag}%"), Component.ai_tags.ilike(f"%{tag}%")))
    if package:
        query = query.filter(Component.package.ilike(f"%{package}%"))
    if ai_status:
        query = query.filter(Component.ai_status == ai_status)
    if stock == "empty":
        query = query.filter(Component.quantity <= 0)
    elif stock == "low":
        query = query.filter(Component.quantity > 0, Component.quantity <= 5)
    elif stock == "available":
        query = query.filter(Component.quantity > 0)
    bool_filters = {
        Component.is_hand_solder_friendly: is_hand_solder_friendly,
        Component.is_power_component: is_power_component,
        Component.is_signal_component: is_signal_component,
        Component.is_high_current: is_high_current,
        Component.is_high_voltage: is_high_voltage,
    }
    for column, value in bool_filters.items():
        if value is not None:
            query = query.filter(column == value)
    if keyword:
        filters = component_keyword_filters(keyword)
        if filters:
            query = query.filter(or_(*filters))
    components = sort_components_by_value(query.limit(500).all())
    reserved = reserved_quantities(db, [item.id for item in components])
    items = [component_out(item, reserved.get(item.id, 0)) for item in components]
    by_category: dict[int | None, dict] = {}
    for item in items:
        category = item["category"]
        key = category.id if category else None
        if key not in by_category:
            by_category[key] = {"category": category, "items": [], "total": 0}
        by_category[key]["items"].append(item)
        by_category[key]["total"] += 1
    return list(by_category.values())


@app.get("/api/components/coverage")
def components_coverage(
    auth: Protected,
    db: Session = Depends(get_db),
    category: str | None = None,
    package: str | None = None,
    only_available: bool = False,
):
    category_names = [category] if category in PASSIVE_COVERAGE else list(PASSIVE_COVERAGE.keys())
    rows = (
        filter_owner(db.query(Component), Component, auth)
        .join(Category, Component.category_id == Category.id)
        .options(joinedload(Component.category))
        .filter(Category.name.in_(category_names))
        .order_by(Category.id.asc(), Component.id.asc())
        .all()
    )
    component_ids = [item.id for item in rows]
    reserved = reserved_quantities(db, component_ids)
    coverage: dict[str, dict] = {
        name: {
            "category": name,
            "unit": PASSIVE_COVERAGE[name]["unit"],
            "dimension": PASSIVE_COVERAGE[name]["dimension"],
            "items": [],
            "packages": set(),
            "unparsed_count": 0,
            "total": 0,
        }
        for name in category_names
    }
    for component in rows:
        category_name = component.category.name if component.category else ""
        bucket = coverage.get(category_name)
        if not bucket:
            continue
        if category_name == "电感" and is_ferrite_bead(component):
            continue
        available = max(0, int(component.quantity or 0) - reserved.get(component.id, 0))
        if only_available and available <= 0:
            continue
        if package and package not in str(component.package or component.normalized_spec or ""):
            continue
        bucket["total"] += 1
        if component.package:
            bucket["packages"].add(component.package)
        parsed_value = None
        parsed_spec = None
        for spec in key_specs_from_component(component):
            if not isinstance(spec, dict) or spec.get("confidence") != "high":
                continue
            parsed_value = parse_unit_value(str(spec.get("value") or ""), bucket["dimension"])
            if parsed_value is not None:
                parsed_spec = spec
                break
        if parsed_value is None:
            bucket["unparsed_count"] += 1
            continue
        bucket["items"].append(
            {
                "id": component.id,
                "name": component.name,
                "model": component.model,
                "value": parsed_value,
                "display_value": parsed_spec.get("value") if parsed_spec else "",
                "spec_name": parsed_spec.get("name") if parsed_spec else "",
                "package": component.package,
                "normalized_spec": component.normalized_spec,
                "quantity": int(component.quantity or 0),
                "available_quantity": available,
                "tags": component.tags,
            }
        )
    result = []
    for item in coverage.values():
        item["packages"] = sorted(item["packages"])
        item["items"].sort(key=lambda value: value["value"])
        result.append(item)
    return {"categories": result}


def _lcsc_ai_parameter_text(value) -> str | None:
    if isinstance(value, str):
        return value.strip()[:4000] or None
    if not isinstance(value, list):
        return None
    parts = []
    for item in value:
        if isinstance(item, dict):
            name = str(item.get("name") or "").strip()
            parameter_value = str(item.get("value") or "").strip()
            text_value = " ".join(part for part in (name, parameter_value) if part)
        else:
            text_value = str(item or "").strip()
        if text_value:
            parts.append(text_value)
    return "；".join(parts)[:4000] or None


def _safe_external_url(value: str | None, *, lcsc_only: bool = False) -> str | None:
    text_value = str(value or "").strip()
    if not text_value or len(text_value) > 1000:
        return None
    parsed = urlsplit(text_value)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname or parsed.username or parsed.password:
        return None
    if lcsc_only and parsed.hostname.lower() not in {"lcsc.com", "www.lcsc.com", "szlcsc.com", "item.szlcsc.com"}:
        return None
    return text_value


def _merge_lcsc_sources(*groups) -> list[dict]:
    result = []
    seen = set()
    for group in groups:
        for item in group or []:
            if not isinstance(item, dict):
                continue
            url = _safe_external_url(item.get("url"))
            if not url or url in seen:
                continue
            seen.add(url)
            result.append({
                "title": str(item.get("title") or item.get("site_name") or "资料来源")[:300],
                "url": url,
                "site_name": str(item.get("site_name") or "")[:120] or None,
                "summary": str(item.get("summary") or "")[:800] or None,
            })
    return result[:8]


def _apply_lcsc_local_fields(draft: dict, result: dict, category_names: set[str]) -> str | None:
    suggested_name = str(result.get("name") or result.get("normalized_name") or "").strip()
    if suggested_name:
        draft["name"] = clean_component_name(suggested_name, draft.get("model"), draft.get("lcsc_number"))
    category_name = str(result.get("category") or result.get("category_suggestion") or "").strip()
    if category_name not in category_names:
        category_name = ""
    tags = result.get("tags") or result.get("ai_tags")
    if tags:
        tag_text = ",".join(str(item) for item in tags) if isinstance(tags, list) else str(tags)
        draft["tags"] = ",".join(normalize_tag_text(tag_text, draft.get("package")))
    return category_name or None


@app.post("/api/components/lcsc/preview", response_model=LcscPreviewResponse)
def preview_lcsc_component(payload: LcscPreviewRequest, auth: Protected, db: Session = Depends(get_db)):
    parsed = parse_lcsc_copy_text(payload.raw_text)
    lcsc_number = parsed.get("lcsc_number")
    if not lcsc_number:
        raise HTTPException(status_code=422, detail="未识别到立创编号，请粘贴包含 Cxxxx 的完整器件信息")

    category_rows = db.query(Category).order_by(Category.id).all()
    categories = [item.name for item in category_rows]
    category_names = set(categories)
    status_value = "parsed_only"
    warnings: list[str] = []
    sources: list[dict] = []
    draft = parsed_copy_draft(parsed)
    category_name: str | None = None

    try:
        product = fetch_lcsc_product(lcsc_number)
        draft = official_product_draft(product, parsed)
        status_value = "official"
        sources = _merge_lcsc_sources([
            {
                "title": draft.get("source_title") or f"LCSC {lcsc_number}",
                "url": draft.get("buy_url") or lcsc_product_url(lcsc_number),
                "site_name": "LCSC",
                "summary": draft.get("description"),
            }
        ])
        try:
            organized = organize_lcsc_draft(draft, categories)
            category_name = _apply_lcsc_local_fields(draft, organized, category_names)
        except Exception:
            warnings.append("立创官方字段已取得，但 AI 命名与分类整理失败，已保留确定性生成的名称和分类。")
    except LcscLookupError as exc:
        warnings.append(str(exc))
        warnings.append("立创官方商品页未完成核验，正在使用 AI 联网结果；保存前请重点核对核心字段。")
        try:
            fallback = lookup_lcsc_fallback(parsed, categories)
            fallback_sources = fallback.get("sources") if isinstance(fallback.get("sources"), list) else []
            sources = _merge_lcsc_sources(fallback_sources)
            exact_match = bool(fallback.get("exact_lcsc_match"))
            exact_match = exact_match and normalize_lcsc_number(fallback.get("lcsc_number")) == lcsc_number
            exact_match = exact_match and exact_lcsc_source_present(sources, lcsc_number)
            category_name = _apply_lcsc_local_fields(draft, fallback, category_names)
            if exact_match:
                status_value = "ai_fallback"
                for key in ("model", "manufacturer", "package", "description"):
                    value = str(fallback.get(key) or "").strip()
                    if value and not draft.get(key):
                        draft[key] = value[:4000 if key == "description" else 200]
                parameter_text = _lcsc_ai_parameter_text(fallback.get("parameters"))
                if parameter_text:
                    draft["parameters"] = parameter_text
                datasheet_url = _safe_external_url(fallback.get("datasheet_url"))
                product_url = _safe_external_url(fallback.get("product_url"), lcsc_only=True)
                if datasheet_url:
                    draft["datasheet_url"] = datasheet_url
                if product_url:
                    draft["buy_url"] = product_url
                warnings.append("核心字段来自 AI 联网精确编号匹配，不等同于直接读取立创商品页。")
            else:
                warnings.append("AI 未提供带精确立创编号证据的结果，仅保留复制文本和低风险整理建议。")
        except Exception:
            warnings.append("AI 联网补全失败。")
            warnings.append("当前仅保留复制文本解析草稿。")

    if not category_name:
        category_name = local_category_from_text(
            " ".join(
                str(value or "")
                for value in (
                    draft.get("official_category"),
                    draft.get("description"),
                    draft.get("copied_name"),
                    draft.get("name"),
                )
            )
        )
    if category_name not in category_names:
        category_name = None
    draft["category_name"] = category_name
    draft["category_id"] = category_id_by_name(db, category_name)
    draft["lcsc_number"] = lcsc_number
    draft["quantity"] = 0
    draft["source"] = "立创"
    draft["buy_url"] = _safe_external_url(draft.get("buy_url"), lcsc_only=True) or lcsc_product_url(lcsc_number)
    draft["datasheet_url"] = _safe_external_url(draft.get("datasheet_url"))
    for internal_key in ("official_category", "official_properties", "copied_name"):
        draft.pop(internal_key, None)

    duplicate = (
        filter_owner(db.query(Component), Component, auth)
        .options(joinedload(Component.category))
        .filter(func.upper(Component.lcsc_number) == lcsc_number)
        .first()
    )
    existing_payload = None
    if duplicate:
        duplicate_reserved = reserved_quantities(db, [duplicate.id]).get(duplicate.id, 0)
        existing_payload = component_out(duplicate, duplicate_reserved)
        duplicate_category = existing_payload.get("category")
        if duplicate_category:
            existing_payload["category"] = {
                "id": duplicate_category.id,
                "name": duplicate_category.name,
                "color": duplicate_category.color,
                "code_prefix": duplicate_category.code_prefix,
                "code_prefix_locked": bool(duplicate_category.code_prefix_locked),
            }
    return {
        "status": status_value,
        "draft": draft,
        "existing_component": existing_payload,
        "warnings": warnings,
        "sources": sources,
    }


@app.post("/api/components", response_model=ComponentOut)
def create_component(payload: ComponentCreate, auth: Protected, db: Session = Depends(get_db)):
    values = payload.model_dump()
    normalized_lcsc = normalize_lcsc_number(payload.lcsc_number)
    if normalized_lcsc:
        values["lcsc_number"] = normalized_lcsc
        duplicate = filter_owner(db.query(Component), Component, auth).filter(func.upper(Component.lcsc_number) == normalized_lcsc).first()
        if duplicate:
            raise HTTPException(status_code=409, detail=f"立创编号 {normalized_lcsc} 已存在，请打开现有器件 {duplicate.warehouse_code or duplicate.id}")
    assert_unique_warehouse_code(db, payload.warehouse_code)
    values["source"] = values.get("source") or "手动新增"
    component = Component(**normalize_for_inventory(db, values, clean_name=True))
    set_owner(component, auth)
    component.ai_status = "pending"
    if int(component.quantity or 0) > 0:
        mark_stock_change(component, int(component.quantity or 0))
    db.add(component)
    db.flush()
    assign_component_warehouse_code(db, component)
    if int(component.quantity or 0) > 0:
        record_stock_delta(
            db,
            component,
            int(component.quantity or 0),
            movement_type="component_create",
            reason="新增元器件初始库存",
            actor_user_id=owner_id(auth),
        )
    enqueue_ai_task(db, "component_analyze", "component", component.id, component_ai_cache_key(component))
    enqueue_ai_task(db, "component_organize", "component", component.id, organize_cache_key(component))
    log_activity(
        db,
        "component.create",
        "component",
        f"新增元器件 {component.name}",
        owner_user_id=owner_id(auth),
        entity_id=component.id,
        component_id=component.id,
        quantity_delta=component.quantity,
    )
    db.commit()
    db.refresh(component)
    return component_out(component, 0)


@app.put("/api/components/{component_id}", response_model=ComponentOut)
def update_component(component_id: int, payload: ComponentUpdate, auth: Protected, db: Session = Depends(get_db)):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    old_quantity = component.quantity or 0
    old_cache_key = component_ai_cache_key(component)
    values = normalize_for_inventory(db, {**component_out(component), **payload.model_dump(exclude_unset=True)}, clean_name=True)
    values = {key: value for key, value in values.items() if key in payload.model_fields or key in {"part_family", "count_mode", "normalized_spec", "category_id"}}
    normalized_lcsc = normalize_lcsc_number(values.get("lcsc_number"))
    if normalized_lcsc:
        duplicate = (
            filter_owner(db.query(Component), Component, auth)
            .filter(func.upper(Component.lcsc_number) == normalized_lcsc, Component.id != component.id)
            .first()
        )
        if duplicate:
            raise HTTPException(status_code=409, detail=f"立创编号 {normalized_lcsc} 已存在，请打开现有器件 {duplicate.warehouse_code or duplicate.id}")
        values["lcsc_number"] = normalized_lcsc
    if "warehouse_code" in values:
        assert_unique_warehouse_code(db, values.get("warehouse_code"), component.id)
    for key, value in values.items():
        setattr(component, key, value)
    assign_component_warehouse_code(db, component)
    new_cache_key = component_ai_cache_key(component)
    if new_cache_key != old_cache_key:
        component.ai_status = "stale"
        enqueue_ai_task(db, "component_analyze", "component", component.id, new_cache_key)
        enqueue_ai_task(db, "component_organize", "component", component.id, organize_cache_key(component))
    new_quantity = component.quantity or 0
    if new_quantity != old_quantity:
        mark_stock_change(component, new_quantity - old_quantity)
        record_stock_delta(
            db,
            component,
            new_quantity - old_quantity,
            movement_type="manual_adjustment",
            reason="编辑元器件库存总量",
            actor_user_id=owner_id(auth),
        )
        log_activity(
            db,
            "component.quantity.update",
            "component",
            f"修改 {component.name} 库存数量：{old_quantity} -> {new_quantity}",
            owner_user_id=owner_id(auth),
            entity_id=component.id,
            component_id=component.id,
            quantity_delta=new_quantity - old_quantity,
        )
    db.commit()
    db.refresh(component)
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return component_out(component, reserved)


@app.post("/api/components/{component_id}/quantity/decrement", response_model=ComponentOut)
def decrement_component_quantity(
    component_id: int,
    auth: Protected,
    db: Session = Depends(get_db),
    payload: ComponentConsumeRequest | None = None,
):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    consume_quantity = payload.quantity if payload else 1
    if (component.quantity or 0) < consume_quantity:
        raise HTTPException(status_code=400, detail="Quantity is not enough")
    if payload and payload.lot_id:
        lot = db.get(InventoryLot, payload.lot_id)
        if not lot or lot.component_id != component.id or lot.status != "active":
            raise HTTPException(status_code=404, detail="库存批次不存在")
        if int(lot.remaining_quantity or 0) < consume_quantity:
            raise HTTPException(status_code=400, detail=f"指定库存批次库存不足：需要 {consume_quantity}，剩余 {int(lot.remaining_quantity or 0)}")
    component.quantity -= consume_quantity
    mark_stock_change(component, -consume_quantity)
    try:
        record_stock_delta(
            db,
            component,
            -consume_quantity,
            movement_type="manual_consume",
            reason=payload.remark if payload else None,
            project_id=payload.project_id if payload else None,
            actor_user_id=owner_id(auth),
            lot_id=payload.lot_id if payload else None,
        )
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    log_activity(
        db,
        "component.consume",
        "component",
        f"使用元器件 {component.name} x {consume_quantity}",
        owner_user_id=owner_id(auth),
        entity_id=component.id,
        component_id=component.id,
        project_id=payload.project_id if payload else None,
        quantity_delta=-consume_quantity,
        detail={"remark": payload.remark} if payload and payload.remark else None,
    )
    db.commit()
    db.refresh(component)
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return component_out(component, reserved)


@app.post("/api/components/{component_id}/quantity/increment", response_model=ComponentOut)
def increment_component_quantity(
    component_id: int,
    auth: Protected,
    db: Session = Depends(get_db),
    payload: ComponentConsumeRequest | None = None,
):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    add_quantity = payload.quantity if payload else 1
    component.quantity = (component.quantity or 0) + add_quantity
    mark_stock_change(component, add_quantity)
    component.first_stocked_at = component.first_stocked_at or datetime.utcnow()
    component.last_stocked_at = datetime.utcnow()
    record_stock_delta(
        db,
        component,
        add_quantity,
        movement_type="manual_restock",
        reason=payload.remark if payload else None,
        project_id=payload.project_id if payload else None,
        actor_user_id=owner_id(auth),
        location=payload.location if payload else None,
        unit_cost=payload.unit_cost if payload else None,
        source_type=payload.source_type if payload else None,
        source_reference=payload.source_reference if payload else None,
    )
    log_activity(
        db,
        "component.increment",
        "component",
        f"增加库存 {component.name} x {add_quantity}",
        owner_user_id=owner_id(auth),
        entity_id=component.id,
        component_id=component.id,
        quantity_delta=add_quantity,
        detail={"remark": payload.remark} if payload and payload.remark else None,
    )
    db.commit()
    db.refresh(component)
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return component_out(component, reserved)


@app.get("/api/components/{component_id}/lots", response_model=list[InventoryLotOut])
def list_component_lots(component_id: int, auth: Protected, db: Session = Depends(get_db)):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    existing_active = (
        db.query(func.count(InventoryLot.id))
        .filter(InventoryLot.component_id == component.id, InventoryLot.status == "active")
        .scalar()
        or 0
    )
    changed = 0
    if int(component.quantity or 0) > 0 and not existing_active:
        ensure_component_lot(db, component)
        changed += 1
    changed += abs(reconcile_component_lots(db, component))
    if changed:
        db.commit()
    rows = (
        db.query(InventoryLot)
        .filter(InventoryLot.component_id == component.id)
        .order_by(InventoryLot.status.asc(), InventoryLot.remaining_quantity.desc(), InventoryLot.received_at.desc(), InventoryLot.created_at.desc())
        .all()
    )
    return [inventory_lot_out(row) for row in rows]


@app.post("/api/components/{component_id}/lots", response_model=InventoryLotOut)
def create_component_lot(component_id: int, payload: InventoryLotCreate, auth: Protected, db: Session = Depends(get_db)):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    quantity = int(payload.quantity or 0)
    component.quantity = int(component.quantity or 0) + quantity
    component.first_stocked_at = component.first_stocked_at or datetime.utcnow()
    component.last_stocked_at = datetime.utcnow()
    mark_stock_change(component, quantity)
    movements = record_stock_delta(
        db,
        component,
        quantity,
        movement_type="manual_lot_create",
        reason=payload.note or "手动新增渠道库存批次",
        actor_user_id=owner_id(auth),
        location=payload.location,
        unit_cost=payload.unit_cost,
        source_type=(payload.source_type or "manual").strip() or "manual",
        source_reference=(payload.source_reference or "").strip() or None,
    )
    lot_id = movements[0].lot_id if movements else None
    log_activity(
        db,
        "component.lot.create",
        "component",
        f"新增库存批次 {component.name} x {quantity}",
        owner_user_id=owner_id(auth),
        entity_id=component.id,
        component_id=component.id,
        quantity_delta=quantity,
        detail={
            "lot_id": lot_id,
            "source_type": payload.source_type,
            "source_reference": payload.source_reference,
            "location": payload.location,
        },
    )
    db.commit()
    lot = db.get(InventoryLot, lot_id) if lot_id else None
    if not lot:
        raise HTTPException(status_code=500, detail="库存批次创建失败")
    return inventory_lot_out(lot)


@app.get("/api/components/{component_id}/usage-records", response_model=list[ComponentUsageRecordOut])
def component_usage_records(
    component_id: int,
    auth: Protected,
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=300),
):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    rows = (
        filter_owner(db.query(ActivityLog), ActivityLog, auth)
        .filter(
            ActivityLog.component_id == component_id,
            ActivityLog.action.in_(["component.consume", "component.loss", "component.restore"]),
        )
        .order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc())
        .limit(limit)
        .all()
    )
    project_ids = {row.project_id for row in rows if row.project_id}
    projects = {
        project.id: project
        for project in db.query(Project).filter(Project.id.in_(project_ids)).all()
    } if project_ids else {}
    labels = {
        "component.consume": "焊接/使用",
        "component.loss": "报损",
        "component.restore": "返还库存",
    }
    result = []
    for row in rows:
        detail = parse_json_value(row.detail)
        designators = []
        if isinstance(detail, dict):
            raw_designators = detail.get("designators") or detail.get("designator") or ""
            if isinstance(raw_designators, list):
                designators = [str(value) for value in raw_designators if value]
            else:
                designators = [
                    value.strip()
                    for value in re.split(r"[,，、]", str(raw_designators))
                    if value.strip()
                ]
        project = projects.get(row.project_id)
        result.append(
            {
                "id": row.id,
                "action": row.action,
                "action_label": labels.get(row.action, row.action),
                "project_id": row.project_id,
                "project_code": project.project_code if project else None,
                "project_name": project.name if project else None,
                "quantity_delta": row.quantity_delta,
                "designators": designators,
                "summary": row.summary,
                "created_at": row.created_at,
            }
        )
    bom_rows = (
        db.query(ProjectBomItem, Project)
        .join(Project, Project.id == ProjectBomItem.project_id)
        .filter(
            ProjectBomItem.component_id == component_id,
            Project.owner_user_id == auth.user_id,
        )
        .all()
    )
    for item, project in bom_rows:
        designators = [
            value.strip()
            for value in re.split(r"[,，、\s]+", str(item.remark or ""))
            if value.strip() and not value.startswith("BOM") and ":" not in value
        ]
        result.append(
            {
                "id": -item.id,
                "action": "project.bom",
                "action_label": "项目 BOM",
                "project_id": project.id,
                "project_code": project.project_code,
                "project_name": project.name,
                "quantity_delta": item.required_quantity,
                "designators": designators,
                "summary": f"项目 BOM 需求 {item.required_quantity}",
                "created_at": project.updated_at or project.created_at,
            }
        )
    result.sort(key=lambda item: item.get("created_at") or datetime.min, reverse=True)
    return result[:limit]


@app.get("/api/components/{component_id}/ai", response_model=ComponentAiOut)
def get_component_ai(component_id: int, auth: Protected, db: Session = Depends(get_db)):
    component = db.query(Component).options(joinedload(Component.category)).filter(Component.id == component_id).first()
    assert_owned(component, auth, "Component not found")
    mark_component_ai_stale(component)
    cards = (
        db.query(AiKnowledgeCard)
        .filter(AiKnowledgeCard.component_id == component_id)
        .order_by(AiKnowledgeCard.updated_at.desc(), AiKnowledgeCard.id.desc())
        .limit(20)
        .all()
    )
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return {"component": component_out(component, reserved), "knowledge_cards": cards}


@app.post("/api/components/{component_id}/ai/ask", response_model=ComponentAiAskOut)
def ask_component_ai(component_id: int, payload: ComponentAiAskRequest, auth: Protected, db: Session = Depends(get_db)):
    component = db.query(Component).options(joinedload(Component.category)).filter(Component.id == component_id).first()
    assert_owned(component, auth, "Component not found")
    question = payload.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="问题不能为空")
    context = component_ai_question_context(db, component, auth)
    try:
        result = component_question(
            component_to_dict(component),
            question,
            context,
            "auto" if payload.use_web_search else "off",
        )
    except Exception as error:
        handle_mimo_error(error)
    answer = str(result.get("answer") or "").strip()
    if not answer:
        answer = "当前资料不足，无法给出可靠回答。请补充数据手册、BOM 封装或厂商型号后再询问。"
    return {
        "answer": answer,
        "confidence": result.get("confidence") or "medium",
        "evidence": list_value(result.get("evidence")),
        "risks": list_value(result.get("risks")),
        "needs_datasheet_check": bool(result.get("needs_datasheet_check", True)),
        "sources": result.get("sources") or [],
    }


@app.post("/api/components/{component_id}/ai/refresh", response_model=ComponentAiOut)
def refresh_component_ai(component_id: int, payload: AiRefreshRequest, auth: Protected, db: Session = Depends(get_db)):
    component = db.query(Component).options(joinedload(Component.category)).filter(Component.id == component_id).first()
    assert_owned(component, auth, "Component not found")
    try:
        analyze_component_with_ai(db, component, payload.scope, payload.force)
        db.commit()
    except Exception as error:
        handle_mimo_error(error)
    cards = (
        db.query(AiKnowledgeCard)
        .filter(AiKnowledgeCard.component_id == component_id)
        .order_by(AiKnowledgeCard.updated_at.desc(), AiKnowledgeCard.id.desc())
        .limit(20)
        .all()
    )
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return {"component": component_out(component, reserved), "knowledge_cards": cards}


@app.post("/api/components/{component_id}/organize", response_model=ComponentOut)
def organize_single_component(component_id: int, auth: Protected, db: Session = Depends(get_db), force: bool = True):
    component = db.query(Component).options(joinedload(Component.category)).filter(Component.id == component_id).first()
    assert_owned(component, auth, "Component not found")
    try:
        organize_component_record(db, component, force=force)
        db.commit()
    except Exception as error:
        handle_mimo_error(error)
    db.refresh(component)
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return component_out(component, reserved)


@app.delete("/api/components/{component_id}")
def delete_component(component_id: int, auth: Protected, db: Session = Depends(get_db)):
    component = (
        db.query(Component)
        .options(joinedload(Component.category))
        .filter(Component.id == component_id)
        .first()
    )
    assert_owned(component, auth, "Component not found")
    removed_at = datetime.utcnow()
    snapshot = component_out(
        component,
        reserved_quantities(db, [component.id]).get(component.id, 0),
    )
    for item in (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.cw_component_id == component.id)
        .all()
    ):
        item.warehouse_code_snapshot = component.warehouse_code
        item.frozen_snapshot_json = json.dumps(snapshot, ensure_ascii=False, default=str)
        item.quantity = int(component.quantity or 0)
        item.name = component.name
        item.model = component.model
        item.lcsc_number = component.lcsc_number
        item.cw_component_id = None
        item.sync_status = "frozen"
    archive_component_identity(db, component)
    component.revoked_at = removed_at
    component.status = "removed"
    log_activity(
        db,
        "component.remove",
        "component",
        f"移除元器件记录 {component.name}（保留 ID {component.warehouse_code or component.id}）",
        owner_user_id=owner_id(auth),
        entity_id=component.id,
        component_id=component.id,
        quantity_delta=-(component.quantity or 0),
    )
    db.commit()
    return {"removed": True, "id_reserved": True, "warehouse_code": component.warehouse_code}


@app.get("/api/dashboard/summary")
def dashboard(auth: Protected, db: Session = Depends(get_db)):
    now = datetime.utcnow()
    component_base = filter_owner(db.query(Component), Component, auth)
    project_base = filter_owner(db.query(Project), Project, auth)
    total_kinds = component_base.with_entities(func.count(Component.id)).scalar() or 0
    total_quantity = component_base.with_entities(func.coalesce(func.sum(Component.quantity), 0)).scalar() or 0
    pending = component_base.with_entities(func.count(Component.id)).filter(Component.status == "pending").scalar() or 0
    recent_projects = (
        project_base
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .order_by(Project.updated_at.desc(), Project.id.desc())
        .limit(5)
        .all()
    )
    recent_component_ids = list({item.component_id for project in recent_projects for item in project.bom_items})
    recent_reserved = reserved_quantities(db, recent_component_ids)
    visible_component_ids = [row[0] for row in component_base.with_entities(Component.id).all()]
    all_reserved = reserved_quantities(db, visible_component_ids)
    reserved_total = sum(all_reserved.values())
    low_stock_candidates = (
        filter_owner(db.query(Component), Component, auth)
        .options(joinedload(Component.category))
        .filter(Component.is_common == True, Component.low_stock_exempt == False)
        .all()
    )
    low_stock_items = [
        item
        for item in low_stock_candidates
        if max(0, int(item.quantity or 0) - all_reserved.get(item.id, 0))
        <= (max(0, int(item.safety_quantity or 0)) or 5)
    ]
    low_stock_items.sort(
        key=lambda item: (
            max(0, int(item.quantity or 0) - all_reserved.get(item.id, 0)),
            -(item.updated_at.timestamp() if item.updated_at else 0),
        )
    )
    low_stock = len(low_stock_items)
    low_by_category: dict[str, int] = {}
    for item in low_stock_items:
        category_name = item.category.name if item.category else "未分类"
        low_by_category[category_name] = low_by_category.get(category_name, 0) + 1
    category_stats = [
        {"name": name or "未分类", "value": int(count or 0), "quantity": int(quantity or 0), "low_stock": low_by_category.get(name or "未分类", 0)}
        for name, count, quantity in (
            filter_owner(db.query(Category.name, func.count(Component.id), func.coalesce(func.sum(Component.quantity), 0)), Component, auth)
            .join(Component, Component.category_id == Category.id, isouter=True)
            .group_by(Category.id)
            .order_by(Category.id)
            .all()
        )
    ]
    status_stats = [{"name": status or "未知", "value": count} for status, count in component_base.with_entities(Component.status, func.count(Component.id)).group_by(Component.status).all()]
    ai_task_summary = {
        status: count for status, count in component_base.with_entities(Component.ai_status, func.count(Component.id)).group_by(Component.ai_status).all()
    }
    recent_project_rows = [project_out(project, recent_reserved) for project in recent_projects]
    project_snapshots = []
    for project in recent_project_rows:
        bom_items = project.get("bom_items") or []
        shortage_count = sum(1 for item in bom_items if not item.get("enough"))
        satisfied_count = sum(1 for item in bom_items if item.get("enough"))
        reserved_count = sum(int(item.get("required_quantity") or 0) for item in bom_items if item.get("status") == "reserved")
        solder_total = sum(int(item.get("solder_total") or 0) for item in bom_items)
        soldered_count = sum(int(item.get("soldered_count") or 0) for item in bom_items)
        project_snapshots.append(
            {
                "id": project["id"],
                "name": project["name"],
                "status": project["status"],
                "bom_total": len(bom_items),
                "satisfied": satisfied_count,
                "shortage": shortage_count,
                "reserved_quantity": reserved_count,
                "board_count": project.get("board_count") or 0,
                "solder_total": solder_total,
                "soldered_count": soldered_count,
                "lost_count": project.get("lost_count") or 0,
                "pending_count": project.get("pending_count") or 0,
                "solder_progress": round((soldered_count / solder_total) * 100) if solder_total else 0,
                "bom_match_total": project.get("bom_match_total") or 0,
                "bom_match_matched": project.get("bom_match_matched") or 0,
                "bom_match_review": project.get("bom_match_review") or 0,
                "bom_match_missing": project.get("bom_match_missing") or 0,
                "bom_match_rate": round(((project.get("bom_match_matched") or 0) / (project.get("bom_match_total") or 1)) * 100) if project.get("bom_match_total") else 0,
                "bom_match_updated_at": project.get("bom_match_updated_at"),
                "updated_at": project.get("updated_at"),
            }
        )
    shortage_projects = sum(1 for item in project_snapshots if item["shortage"] > 0)
    total_solder_points = db.query(func.count(ProjectBomSolderPoint.id)).scalar() or 0
    soldered_points = db.query(func.count(ProjectBomSolderPoint.id)).filter(ProjectBomSolderPoint.soldered == True).scalar() or 0
    lost_points = db.query(func.count(ProjectBomSolderPoint.id)).filter(ProjectBomSolderPoint.lost == True).scalar() or 0
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    soldered_today = (
        db.query(func.count(ProjectBomSolderPoint.id))
        .filter(ProjectBomSolderPoint.soldered == True, ProjectBomSolderPoint.soldered_at >= today_start)
        .scalar()
        or 0
    )
    soldered_this_week = (
        db.query(func.count(ProjectBomSolderPoint.id))
        .filter(ProjectBomSolderPoint.soldered == True, ProjectBomSolderPoint.soldered_at >= week_start)
        .scalar()
        or 0
    )
    milestones = [10, 25, 50, 100, 250, 500, 1000, 2500, 5000, 10000]
    next_milestone = next((value for value in milestones if value > soldered_points), ((soldered_points // 1000) + 1) * 1000)
    solder_project_rank = [
        {"project_id": project_id, "name": name, "soldered": int(count or 0)}
        for project_id, name, count in (
            db.query(Project.id, Project.name, func.count(ProjectBomSolderPoint.id))
            .join(ProjectBomItem, ProjectBomItem.project_id == Project.id)
            .join(ProjectBomSolderPoint, ProjectBomSolderPoint.bom_item_id == ProjectBomItem.id)
            .filter(ProjectBomSolderPoint.soldered == True)
            .group_by(Project.id, Project.name)
            .order_by(func.count(ProjectBomSolderPoint.id).desc(), Project.updated_at.desc())
            .limit(5)
            .all()
        )
    ]
    solder_stats = {
        "total_solder_points": int(total_solder_points or 0),
        "soldered_points": int(soldered_points or 0),
        "lost_points": int(lost_points or 0),
        "pending_points": max(0, int(total_solder_points or 0) - int(soldered_points or 0)),
        "solder_progress": round((soldered_points / total_solder_points) * 100) if total_solder_points else 0,
        "soldered_today": int(soldered_today or 0),
        "soldered_this_week": int(soldered_this_week or 0),
        "next_milestone": int(next_milestone),
        "milestone_progress": round((soldered_points / next_milestone) * 100) if next_milestone else 0,
        "project_rank": solder_project_rank,
    }
    usage_actions = ["component.consume", "component.loss"]
    daily_rows = (
        db.query(func.date(ActivityLog.created_at), func.coalesce(func.sum(-ActivityLog.quantity_delta), 0))
        .filter(ActivityLog.quantity_delta < 0, ActivityLog.action.in_(usage_actions), ActivityLog.created_at >= today_start - timedelta(days=6))
        .group_by(func.date(ActivityLog.created_at))
        .all()
    )
    daily_map = {str(day): int(total or 0) for day, total in daily_rows}
    last_7_days = []
    for offset in range(6, -1, -1):
        day = today_start - timedelta(days=offset)
        key = day.date().isoformat()
        last_7_days.append({"date": key, "quantity": daily_map.get(key, 0)})
    today_used = sum(item["quantity"] for item in last_7_days if item["date"] == today_start.date().isoformat())
    today_by_category = [
        {"name": name or "未分类", "quantity": int(quantity or 0)}
        for name, quantity in (
            db.query(Category.name, func.coalesce(func.sum(-ActivityLog.quantity_delta), 0))
            .join(Component, ActivityLog.component_id == Component.id)
            .outerjoin(Category, Component.category_id == Category.id)
            .filter(ActivityLog.quantity_delta < 0, ActivityLog.action.in_(usage_actions), ActivityLog.created_at >= today_start)
            .group_by(Category.name)
            .order_by(func.coalesce(func.sum(-ActivityLog.quantity_delta), 0).desc())
            .all()
        )
    ]
    loss_today = (
        db.query(func.coalesce(func.sum(-ActivityLog.quantity_delta), 0))
        .filter(ActivityLog.quantity_delta < 0, ActivityLog.action == "component.loss", ActivityLog.created_at >= today_start)
        .scalar()
        or 0
    )
    usage_stats = {
        "today_used": int(today_used or 0),
        "loss_today": int(loss_today or 0),
        "last_7_days": last_7_days,
        "today_by_category": today_by_category,
    }
    datasheet_missing = component_base.with_entities(func.count(Component.id)).filter(or_(Component.datasheet_url == None, Component.datasheet_url == "")).scalar() or 0
    ai_pending = ai_task_summary.get("pending", 0) + ai_task_summary.get("stale", 0)
    mechanical_stats = [
        {
            "family": family or "other",
            "spec": spec or "未归一",
            "quantity": int(quantity or 0),
            "count_mode": count_mode or "exact",
        }
        for family, spec, count_mode, quantity in (
            component_base.with_entities(Component.part_family, Component.normalized_spec, Component.count_mode, func.coalesce(func.sum(Component.quantity), 0))
            .filter(Component.part_family.in_(["screw", "nut", "standoff", "pin_header"]))
            .group_by(Component.part_family, Component.normalized_spec, Component.count_mode)
            .order_by(Component.part_family, Component.normalized_spec)
            .limit(20)
            .all()
        )
    ]
    ai_summary = (
        f"需要行动的库存提醒 {len(low_stock_items)} 项，最近项目中 {shortage_projects} 个存在缺料；"
        f"待 AI 整理 {ai_pending} 项，{datasheet_missing} 个器件缺数据手册链接。"
        f"机械件已归类 {len(mechanical_stats)} 个规格。"
    )
    action_items = []
    for project in recent_project_rows:
        shortage_count = sum(1 for item in project.get("bom_items", []) if not item.get("enough"))
        if shortage_count:
            action_items.append({"type": "project_shortage", "title": f"{project['name']} 有 {shortage_count} 项缺料", "hint": "进入 BOM 页确认替代料或采购", "severity": "danger"})
    for item in low_stock_items[:5]:
        action_items.append(
            {
                "type": "stock_watch",
                "title": item.name,
                "hint": f"{item.category.name if item.category else '未分类'} · 库存 {item.quantity}，建议确认是否需要补货",
                "severity": "warning" if (item.quantity or 0) > 0 else "danger",
                "component_id": item.id,
            }
        )
    if datasheet_missing:
        action_items.append({"type": "datasheet_missing", "title": f"{datasheet_missing} 个器件缺数据手册", "hint": "优先补齐电源、接口和保护器件资料", "severity": "info"})
    if ai_task_summary.get("failed", 0):
        action_items.append({"type": "ai_failed", "title": f"AI 整理失败 {ai_task_summary.get('failed', 0)} 项", "hint": "检查 API 或重试后台整理", "severity": "warning"})
    return {
        "total_kinds": total_kinds,
        "category_count": db.query(func.count(Category.id)).scalar() or 0,
        "total_quantity": total_quantity,
        "reserved_quantity": reserved_total,
        "available_quantity": max(0, total_quantity - reserved_total),
        "low_stock": low_stock,
        "pending": pending,
        "common_count": component_base.with_entities(func.count(Component.id)).filter(Component.is_common == True).scalar() or 0,
        "ai_pending": ai_task_summary.get("pending", 0) + ai_task_summary.get("stale", 0),
        "ai_failed": ai_task_summary.get("failed", 0),
        "shortage_projects": shortage_projects,
        "datasheet_missing": datasheet_missing,
        "mechanical_stats": mechanical_stats,
        "solder_stats": solder_stats,
        "usage_stats": usage_stats,
        "action_items": action_items[:10],
        "category_stats": category_stats,
        "status_stats": status_stats,
        "low_stock_items": [
            component_out(item, all_reserved.get(item.id, 0))
            for item in low_stock_items[:8]
        ],
        "ai_summary": ai_summary,
        "recent_projects": recent_project_rows,
        "project_snapshots": project_snapshots,
    }


@app.get("/api/projects", response_model=list[ProjectOut])
def list_projects(auth: Protected, db: Session = Depends(get_db)):
    projects = (
        filter_owner(db.query(Project), Project, auth)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .order_by(Project.updated_at.desc(), Project.id.desc())
        .all()
    )
    component_ids = list({item.component_id for project in projects for item in project.bom_items})
    reserved = reserved_quantities(db, component_ids)
    return [project_out(project, reserved) for project in projects]


@app.post("/api/projects", response_model=ProjectOut)
def create_project(payload: ProjectCreate, auth: Protected, db: Session = Depends(get_db)):
    values = payload.model_dump()
    values["project_code"] = normalize_project_code(values.get("project_code"))
    assert_unique_project_code(db, values.get("project_code"))
    project = Project(**values)
    set_owner(project, auth)
    db.add(project)
    db.flush()
    assign_project_code(db, project)
    default_project_board(db, project.id)
    log_activity(
        db,
        "project.create",
        "project",
        f"创建项目 {project.name}",
        owner_user_id=owner_id(auth),
        entity_id=project.id,
        project_id=project.id,
    )
    db.commit()
    db.refresh(project)
    return project_out(project, {})


@app.get("/api/projects/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    return project_out(project, reserved)


@app.put("/api/projects/{project_id}", response_model=ProjectOut)
def update_project(project_id: int, payload: ProjectUpdate, auth: Protected, db: Session = Depends(get_db)):
    project = filter_owner(db.query(Project), Project, auth).options(joinedload(Project.boards), joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points)).filter(Project.id == project_id).first()
    assert_owned(project, auth, "Project not found")
    values = payload.model_dump(exclude_unset=True)
    if "project_code" in values:
        values["project_code"] = normalize_project_code(values.get("project_code"))
        assert_unique_project_code(db, values.get("project_code"), project.id)
    for key, value in values.items():
        setattr(project, key, value)
    assign_project_code(db, project)
    db.commit()
    db.refresh(project)
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    return project_out(project, reserved)


@app.post("/api/projects/{project_id}/boards", response_model=ProjectOut)
def create_project_board(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    max_index = max([int(board.board_index or 0) for board in project.boards] or [0])
    board = ProjectBoard(project_id=project.id, board_index=max_index + 1, name=f"第 {max_index + 1} 板", status="active")
    db.add(board)
    db.flush()
    for item in project.bom_items:
        sync_bom_solder_points(db, item, board.id)
    log_activity(
        db,
        "project.board.create",
        "project_board",
        f"{project.name} 新增 {board.name}",
        owner_user_id=owner_id(auth),
        entity_id=board.id,
        project_id=project.id,
        detail={"board_index": board.board_index},
    )
    db.commit()
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .filter(Project.id == project_id)
        .first()
    )
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    result = project_out(project, reserved)
    result["active_board_id"] = board.id
    return result


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.boards), joinedload(Project.bom_items).joinedload(ProjectBomItem.component), joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    restored = 0
    for item in project.bom_items:
        if (item.status or "reserved") == "picked" or not item.component:
            continue
        solder_applied = sum(
            1 for point in item.solder_points if getattr(point, "stock_applied", False)
        )
        loss_applied = sum(
            1 for point in item.solder_points if getattr(point, "loss_stock_applied", False)
        )
        quantity = solder_applied + loss_applied
        if not quantity:
            continue
        item.component.quantity = int(item.component.quantity or 0) + quantity
        mark_stock_change(item.component, quantity)
        restored += quantity
        log_activity(
            db,
            "component.restore",
            "component",
            f"删除项目返还焊接扣减库存 {item.component.name} x {quantity}",
            owner_user_id=owner_id(auth),
            entity_id=item.component_id,
            component_id=item.component_id,
            project_id=project.id,
            quantity_delta=quantity,
            detail={
                "bom_item_id": item.id,
                "reason": "project.delete",
                "solder_applied": solder_applied,
                "loss_applied": loss_applied,
            },
        )
    log_activity(
        db,
        "project.delete",
        "project",
        f"删除项目 {project.name}",
        owner_user_id=owner_id(auth),
        entity_id=project.id,
        project_id=project.id,
        detail={"project_code": project.project_code, "auto_solder_stock_restored": restored},
    )
    batch_ids = [row_id for (row_id,) in db.query(ProjectBomImportBatch.id).filter(ProjectBomImportBatch.project_id == project.id).all()]
    import_row_ids = [row_id for (row_id,) in db.query(ProjectBomImportRow.id).filter(ProjectBomImportRow.project_id == project.id).all()]
    if import_row_ids:
        db.query(ProjectBomImportCandidate).filter(ProjectBomImportCandidate.import_row_id.in_(import_row_ids)).delete(synchronize_session=False)
        db.query(ProjectBomImportRow).filter(ProjectBomImportRow.id.in_(import_row_ids)).delete(synchronize_session=False)
    if batch_ids:
        db.query(ProjectBomImportBatch).filter(ProjectBomImportBatch.id.in_(batch_ids)).delete(synchronize_session=False)
    db.query(AiKnowledgeCard).filter(AiKnowledgeCard.project_id == project.id).delete(synchronize_session=False)
    db.query(ActivityLog).filter(ActivityLog.project_id == project.id).update({ActivityLog.project_id: None}, synchronize_session=False)
    db.delete(project)
    db.commit()
    return {"deleted": True, "restored_quantity": restored}


@app.post("/api/projects/{project_id}/bom", response_model=BomItemOut)
def add_bom_item(project_id: int, payload: BomItemCreate, auth: Protected, db: Session = Depends(get_db)):
    require_project_access(db, project_id, auth)
    component = db.get(Component, payload.component_id)
    assert_owned(component, auth, "Component not found")
    item = ProjectBomItem(project_id=project_id, **payload.model_dump())
    db.add(item)
    db.flush()
    sync_bom_solder_points(db, item)
    if item.status == "reserved":
        log_activity(
            db,
            "bom.reserve",
            "project_bom_item",
            f"项目 {project_id} 占用 {component.name} x {item.required_quantity}",
            owner_user_id=owner_id(auth),
            entity_id=item.id,
            component_id=component.id,
            project_id=project_id,
            quantity_delta=item.required_quantity,
            detail={"remark": item.remark},
        )
    db.commit()
    db.refresh(item)
    item.component = component
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.put("/api/projects/{project_id}/bom/{item_id}", response_model=BomItemOut)
def update_bom_item(project_id: int, item_id: int, payload: BomItemUpdate, auth: Protected, db: Session = Depends(get_db)):
    require_project_access(db, project_id, auth)
    item = db.query(ProjectBomItem).filter(ProjectBomItem.project_id == project_id, ProjectBomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    assert_owned(item.component, auth, "Component not found")
    old_quantity = item.required_quantity
    old_status = item.status or "reserved"
    old_component_id = item.component_id
    payload_data = payload.model_dump(exclude_unset=True)
    new_component_id = payload_data.get("component_id")
    if new_component_id is not None:
        new_component = db.get(Component, new_component_id)
        assert_owned(new_component, auth, "Component not found")
        if int(new_component_id) != int(old_component_id):
            old_component = db.get(Component, old_component_id)
            solder_applied = sum(1 for point in item.solder_points if getattr(point, "stock_applied", False))
            loss_applied = sum(1 for point in item.solder_points if getattr(point, "loss_stock_applied", False))
            transferred_quantity = solder_applied + loss_applied
            if transferred_quantity and new_component.quantity < transferred_quantity:
                raise HTTPException(status_code=400, detail=f"新元器件库存不足，无法承接已焊/报损扣库数量 {transferred_quantity}")
            if transferred_quantity:
                if old_component:
                    old_component.quantity += transferred_quantity
                new_component.quantity -= transferred_quantity
                log_activity(
                    db,
                    "bom.component.transfer",
                    "project_bom_item",
                    f"调整项目 {project_id} BOM 物料扣库归属：{old_component_id} -> {new_component_id}",
                    owner_user_id=owner_id(auth),
                    entity_id=item.id,
                    component_id=new_component.id,
                    project_id=project_id,
                    quantity_delta=-transferred_quantity,
                    detail={
                        "old_component_id": old_component_id,
                        "new_component_id": new_component_id,
                        "solder_applied": solder_applied,
                        "loss_applied": loss_applied,
                    },
                )
    for key, value in payload_data.items():
        setattr(item, key, value)
    sync_bom_solder_points(db, item)
    new_status = item.status or "reserved"
    if old_quantity != item.required_quantity or old_status != new_status:
        old_reserved = old_quantity if old_status == "reserved" else 0
        new_reserved = item.required_quantity if new_status == "reserved" else 0
        log_activity(
            db,
            "bom.reserve.update",
            "project_bom_item",
            f"调整项目 {project_id} 的 {item.component.name} BOM：{old_quantity}/{old_status} -> {item.required_quantity}/{new_status}",
            owner_user_id=owner_id(auth),
            entity_id=item.id,
            component_id=item.component_id,
            project_id=project_id,
            quantity_delta=new_reserved - old_reserved,
        )
    db.commit()
    db.refresh(item)
    item.component = db.get(Component, item.component_id)
    reserved = reserved_quantities(db, list({old_component_id, item.component_id}))
    return bom_item_out(item, reserved)


@app.delete("/api/projects/{project_id}/bom/{item_id}")
def delete_bom_item(project_id: int, item_id: int, auth: Protected, db: Session = Depends(get_db)):
    require_project_access(db, project_id, auth)
    item = db.query(ProjectBomItem).filter(ProjectBomItem.project_id == project_id, ProjectBomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    assert_owned(item.component, auth, "Component not found")
    applied = [
        point
        for point in item.solder_points
        if bool(getattr(point, "stock_applied", False))
        or bool(getattr(point, "loss_stock_applied", False))
    ]
    if applied:
        raise HTTPException(
            status_code=400,
            detail=f"该 BOM 行已有 {len(applied)} 个焊接或报损扣库记录，请先逐项撤销后再删除",
        )
    old_status = item.status or "reserved"
    log_activity(
        db,
        "bom.delete",
        "project_bom_item",
        f"删除项目 {project_id} 的 BOM 行 {item.component.name} x {item.required_quantity}",
        owner_user_id=owner_id(auth),
        entity_id=item.id,
        component_id=item.component_id,
        project_id=project_id,
        quantity_delta=-item.required_quantity if old_status == "reserved" else 0,
    )
    db.delete(item)
    db.commit()
    return {"deleted": True}


@app.post("/api/projects/{project_id}/bom/{item_id}/status", response_model=BomItemOut)
def mark_bom_item_status(
    project_id: int,
    item_id: int,
    payload: BomItemStatusRequest,
    auth: Protected,
    db: Session = Depends(get_db),
):
    require_project_access(db, project_id, auth)
    if payload.status not in {"reserved", "picked", "done"}:
        raise HTTPException(status_code=400, detail="Invalid BOM status")
    item = (
        db.query(ProjectBomItem)
        .options(joinedload(ProjectBomItem.component).joinedload(Component.category))
        .filter(ProjectBomItem.project_id == project_id, ProjectBomItem.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    assert_owned(item.component, auth, "Component not found")
    old_status = item.status or "reserved"
    old_reserved = item.required_quantity if old_status == "reserved" else 0
    new_reserved = item.required_quantity if payload.status == "reserved" else 0
    if payload.consume_stock:
        if (item.component.quantity or 0) < item.required_quantity:
            raise HTTPException(status_code=400, detail="Quantity is not enough")
        item.component.quantity -= item.required_quantity
        mark_stock_change(item.component, -item.required_quantity)
        record_stock_delta(
            db,
            item.component,
            -item.required_quantity,
            movement_type="project_pick",
            reason=payload.remark or "项目 BOM 取料",
            project_id=project_id,
            actor_user_id=owner_id(auth),
        )
        log_activity(
            db,
            "component.consume",
            "component",
            f"项目 {project_id} 取料 {item.component.name} x {item.required_quantity}",
            owner_user_id=owner_id(auth),
            entity_id=item.component_id,
            component_id=item.component_id,
            project_id=project_id,
            quantity_delta=-item.required_quantity,
            detail={"bom_item_id": item.id, "remark": payload.remark},
        )
    item.status = payload.status
    if payload.remark:
        item.remark = f"{item.remark or ''}\n{payload.remark}".strip()
    log_activity(
        db,
        "bom.status",
        "project_bom_item",
        f"项目 {project_id} 的 {item.component.name} BOM 状态：{old_status} -> {item.status}",
        owner_user_id=owner_id(auth),
        entity_id=item.id,
        component_id=item.component_id,
        project_id=project_id,
        quantity_delta=new_reserved - old_reserved,
        detail={"consume_stock": payload.consume_stock},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


def get_project_bom_item_for_solder(db: Session, project_id: int, item_id: int, auth: AuthContext | None = None) -> ProjectBomItem:
    require_project_access(db, project_id, auth)
    item = (
        db.query(ProjectBomItem)
        .options(joinedload(ProjectBomItem.component).joinedload(Component.category), joinedload(ProjectBomItem.solder_points))
        .filter(ProjectBomItem.project_id == project_id, ProjectBomItem.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    assert_owned(item.component, auth, "Component not found")
    sync_bom_solder_points(db, item)
    db.flush()
    db.expire(item, ["solder_points"])
    return item


def get_board_for_project(db: Session, project_id: int, board_id: int, auth: AuthContext | None = None) -> ProjectBoard:
    require_project_access(db, project_id, auth)
    board = db.query(ProjectBoard).filter(ProjectBoard.project_id == project_id, ProjectBoard.id == board_id).first()
    if not board:
        raise HTTPException(status_code=404, detail="Project board not found")
    return board


def apply_solder_inventory_change(db: Session, item: ProjectBomItem, changed_points: list[ProjectBomSolderPoint], soldered: bool, project_id: int, auth: AuthContext | None = None) -> None:
    if not changed_points:
        return
    if item.status == "picked":
        return
    if soldered:
        inventory_points = [point for point in changed_points if not getattr(point, "stock_applied", False)]
    else:
        inventory_points = [point for point in changed_points if getattr(point, "stock_applied", False)]
    if not inventory_points:
        return
    quantity = len(inventory_points)
    delta = -quantity if soldered else quantity
    if delta < 0 and int(item.component.quantity or 0) < quantity:
        current = int(item.component.quantity or 0)
        raise HTTPException(
            status_code=400,
            detail=f"库存不足：需要 {quantity}，现有 {current}，缺少 {quantity - current}，本次未扣减",
        )
    item.component.quantity = int(item.component.quantity or 0) + delta
    mark_stock_change(item.component, delta)
    record_stock_delta(
        db,
        item.component,
        delta,
        movement_type="solder_consume" if delta < 0 else "solder_restore",
        reason="项目焊接状态变更",
        project_id=project_id,
        actor_user_id=owner_id(auth),
    )
    for point in inventory_points:
        point.stock_applied = bool(soldered)
    designators = ", ".join(point.designator for point in inventory_points)
    log_activity(
        db,
        "component.consume" if delta < 0 else "component.restore",
        "component",
        f"{'焊接消耗' if delta < 0 else '取消焊接返还'} {item.component.name} x {quantity}",
        owner_user_id=owner_id(auth),
        entity_id=item.component_id,
        component_id=item.component_id,
        project_id=project_id,
        quantity_delta=delta,
        detail={"bom_item_id": item.id, "designators": designators},
    )


def apply_loss_inventory_change(db: Session, item: ProjectBomItem, changed_points: list[ProjectBomSolderPoint], lost: bool, project_id: int, auth: AuthContext | None = None) -> None:
    if not changed_points:
        return
    if item.status == "picked":
        return
    if lost:
        inventory_points = [point for point in changed_points if not getattr(point, "loss_stock_applied", False)]
    else:
        inventory_points = [point for point in changed_points if getattr(point, "loss_stock_applied", False)]
    if not inventory_points:
        return
    quantity = len(inventory_points)
    delta = -quantity if lost else quantity
    if delta < 0 and int(item.component.quantity or 0) < quantity:
        current = int(item.component.quantity or 0)
        raise HTTPException(
            status_code=400,
            detail=f"库存不足：报损需要 {quantity}，现有 {current}，缺少 {quantity - current}，本次未扣减",
        )
    item.component.quantity = int(item.component.quantity or 0) + delta
    mark_stock_change(item.component, delta)
    record_stock_delta(
        db,
        item.component,
        delta,
        movement_type="loss" if delta < 0 else "loss_restore",
        reason="项目报损状态变更",
        project_id=project_id,
        actor_user_id=owner_id(auth),
    )
    for point in inventory_points:
        point.loss_stock_applied = bool(lost)
    designators = ", ".join(point.designator for point in inventory_points)
    log_activity(
        db,
        "component.loss" if delta < 0 else "component.restore",
        "component",
        f"{'报损消耗' if delta < 0 else '取消报损返还'} {item.component.name} x {quantity}",
        owner_user_id=owner_id(auth),
        entity_id=item.component_id,
        component_id=item.component_id,
        project_id=project_id,
        quantity_delta=delta,
        detail={"bom_item_id": item.id, "designators": designators, "reason": "board.loss"},
    )


@app.post("/api/projects/{project_id}/bom/{item_id}/solder-points/{point_id}", response_model=BomItemOut)
def update_bom_solder_point(
    project_id: int,
    item_id: int,
    point_id: int,
    payload: BomSolderPointUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    item = get_project_bom_item_for_solder(db, project_id, item_id, auth)
    point = next((candidate for candidate in item.solder_points if candidate.id == point_id), None)
    if not point:
        raise HTTPException(status_code=404, detail="BOM solder point not found")
    changed_points = [point] if bool(point.soldered) != bool(payload.soldered) else []
    apply_solder_inventory_change(db, item, changed_points, bool(payload.soldered), project_id, auth)
    point.soldered = bool(payload.soldered)
    point.soldered_at = datetime.utcnow() if point.soldered else None
    if payload.note is not None:
        point.note = payload.note.strip() or None
    log_activity(
        db,
        "bom.solder.update",
        "project_bom_solder_point",
        f"{'标记已焊' if point.soldered else '取消已焊'} {point.designator}",
        owner_user_id=owner_id(auth),
        entity_id=point.id,
        component_id=item.component_id,
        project_id=project_id,
        detail={"bom_item_id": item.id, "designator": point.designator, "note": point.note},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.post("/api/projects/{project_id}/bom/{item_id}/solder-points/bulk", response_model=BomItemOut)
def update_bom_solder_points_bulk(
    project_id: int,
    item_id: int,
    payload: BomSolderPointBulkUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    item = get_project_bom_item_for_solder(db, project_id, item_id, auth)
    point_ids = set(payload.point_ids or [])
    points = [point for point in item.solder_points if not point_ids or point.id in point_ids]
    if not points:
        raise HTTPException(status_code=404, detail="BOM solder point not found")
    changed_at = datetime.utcnow()
    changed_points = [point for point in points if bool(point.soldered) != bool(payload.soldered)]
    apply_solder_inventory_change(db, item, changed_points, bool(payload.soldered), project_id, auth)
    for point in points:
        point.soldered = bool(payload.soldered)
        point.soldered_at = changed_at if point.soldered else None
        if payload.note is not None:
            point.note = payload.note.strip() or None
    log_activity(
        db,
        "bom.solder.bulk_update",
        "project_bom_item",
        f"{'批量标记已焊' if payload.soldered else '批量取消已焊'} {len(points)} 个位号",
        owner_user_id=owner_id(auth),
        entity_id=item.id,
        component_id=item.component_id,
        project_id=project_id,
        detail={"point_ids": [point.id for point in points], "soldered": payload.soldered},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.post("/api/projects/{project_id}/boards/{board_id}/bom/{item_id}/solder-points/{point_id}", response_model=BomItemOut)
def update_board_bom_solder_point(
    project_id: int,
    board_id: int,
    item_id: int,
    point_id: int,
    payload: BomSolderPointUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    get_board_for_project(db, project_id, board_id, auth)
    item = get_project_bom_item_for_solder(db, project_id, item_id, auth)
    point = next((candidate for candidate in item.solder_points if candidate.id == point_id and candidate.board_id == board_id), None)
    if not point:
        raise HTTPException(status_code=404, detail="BOM solder point not found")
    changed_points = [point] if bool(point.soldered) != bool(payload.soldered) else []
    apply_solder_inventory_change(db, item, changed_points, bool(payload.soldered), project_id, auth)
    point.soldered = bool(payload.soldered)
    point.soldered_at = datetime.utcnow() if point.soldered else None
    if payload.note is not None:
        point.note = payload.note.strip() or None
    log_activity(
        db,
        "bom.solder.update",
        "project_bom_solder_point",
        f"{'标记已焊' if point.soldered else '取消已焊'} {point.designator}",
        owner_user_id=owner_id(auth),
        entity_id=point.id,
        component_id=item.component_id,
        project_id=project_id,
        detail={"bom_item_id": item.id, "board_id": board_id, "designator": point.designator, "note": point.note},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.post("/api/projects/{project_id}/boards/{board_id}/bom/{item_id}/solder-points/bulk", response_model=BomItemOut)
def update_board_bom_solder_points_bulk(
    project_id: int,
    board_id: int,
    item_id: int,
    payload: BomSolderPointBulkUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    get_board_for_project(db, project_id, board_id, auth)
    item = get_project_bom_item_for_solder(db, project_id, item_id, auth)
    point_ids = set(payload.point_ids or [])
    points = [point for point in item.solder_points if point.board_id == board_id and (not point_ids or point.id in point_ids)]
    if not points:
        raise HTTPException(status_code=404, detail="BOM solder point not found")
    changed_at = datetime.utcnow()
    changed_points = [point for point in points if bool(point.soldered) != bool(payload.soldered)]
    apply_solder_inventory_change(db, item, changed_points, bool(payload.soldered), project_id, auth)
    for point in points:
        point.soldered = bool(payload.soldered)
        point.soldered_at = changed_at if point.soldered else None
        if payload.note is not None:
            point.note = payload.note.strip() or None
    log_activity(
        db,
        "bom.solder.bulk_update",
        "project_bom_item",
        f"{'批量标记已焊' if payload.soldered else '批量取消已焊'} {len(points)} 个位号",
        owner_user_id=owner_id(auth),
        entity_id=item.id,
        component_id=item.component_id,
        project_id=project_id,
        detail={"board_id": board_id, "point_ids": [point.id for point in points], "soldered": payload.soldered},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.post("/api/projects/{project_id}/boards/{board_id}/bom/{item_id}/solder-points/{point_id}/loss", response_model=BomItemOut)
def update_board_bom_solder_point_loss(
    project_id: int,
    board_id: int,
    item_id: int,
    point_id: int,
    payload: BomSolderPointLossUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    get_board_for_project(db, project_id, board_id, auth)
    item = get_project_bom_item_for_solder(db, project_id, item_id, auth)
    point = next((candidate for candidate in item.solder_points if candidate.id == point_id and candidate.board_id == board_id), None)
    if not point:
        raise HTTPException(status_code=404, detail="BOM solder point not found")
    changed_points = [point] if bool(getattr(point, "lost", False)) != bool(payload.lost) else []
    apply_loss_inventory_change(db, item, changed_points, bool(payload.lost), project_id, auth)
    point.lost = bool(payload.lost)
    point.lost_at = datetime.utcnow() if point.lost else None
    point.loss_note = payload.note.strip() if payload.note else None
    log_activity(
        db,
        "bom.solder.loss",
        "project_bom_solder_point",
        f"{'标记报损' if point.lost else '取消报损'} {point.designator}",
        owner_user_id=owner_id(auth),
        entity_id=point.id,
        component_id=item.component_id,
        project_id=project_id,
        detail={"bom_item_id": item.id, "board_id": board_id, "designator": point.designator, "note": point.loss_note},
    )
    db.commit()
    db.refresh(item)
    reserved = reserved_quantities(db, [item.component_id])
    return bom_item_out(item, reserved)


@app.post("/api/projects/{project_id}/bom/import-matches", response_model=BomMatchCommitResult)
def import_matched_bom_items(
    project_id: int,
    payload: BomMatchCommitRequest,
    auth: Protected,
    db: Session = Depends(get_db),
):
    require_project_access(db, project_id, auth)
    added = updated = skipped = 0
    for item in payload.items:
        component = db.get(Component, item.component_id)
        try:
            assert_owned(component, auth, "Component not found")
        except HTTPException:
            skipped += 1
            continue
        target_bom_item = None
        existing = (
            db.query(ProjectBomItem)
            .filter(ProjectBomItem.project_id == project_id, ProjectBomItem.component_id == item.component_id)
            .first()
        )
        if existing:
            old_quantity = existing.required_quantity
            old_status = existing.status or "reserved"
            existing.required_quantity += item.required_quantity
            if item.remark:
                existing.remark = f"{existing.remark or ''}\n{item.remark}".strip()
            existing.status = "reserved"
            sync_bom_solder_points(db, existing)
            target_bom_item = existing
            log_activity(
                db,
                "bom.reserve.update",
                "project_bom_item",
                f"导入 BOM 更新占用 {component.name}：{old_quantity}/{old_status} -> {existing.required_quantity}/reserved",
                owner_user_id=owner_id(auth),
                entity_id=existing.id,
                component_id=component.id,
                project_id=project_id,
                quantity_delta=item.required_quantity if old_status == "reserved" else existing.required_quantity,
            )
            updated += 1
        else:
            bom_item = ProjectBomItem(
                project_id=project_id,
                component_id=item.component_id,
                required_quantity=item.required_quantity,
                remark=item.remark,
                status="reserved",
            )
            db.add(bom_item)
            db.flush()
            sync_bom_solder_points(db, bom_item)
            target_bom_item = bom_item
            log_activity(
                db,
                "bom.reserve",
                "project_bom_item",
                f"导入 BOM 占用 {component.name} x {item.required_quantity}",
                owner_user_id=owner_id(auth),
                entity_id=bom_item.id,
                component_id=component.id,
                project_id=project_id,
                quantity_delta=item.required_quantity,
                detail={"remark": item.remark},
            )
            added += 1
        if item.import_row_id:
            row = (
                db.query(ProjectBomImportRow)
                .filter(ProjectBomImportRow.project_id == project_id, ProjectBomImportRow.id == item.import_row_id)
                .first()
            )
            if row:
                row.status = "imported"
                row.selected_component_id = item.component_id
                row.match_confidence = max(int(row.match_confidence or 0), 100)
                row.auto_imported = True
                row.auto_import_note = f"已导入正式 BOM：{target_bom_item.id if target_bom_item else '-'}"
                row.ai_reason = "已进入正式项目 BOM，不再计入待采购。"
                row.updated_at = datetime.utcnow()
                batch = db.get(ProjectBomImportBatch, row.batch_id)
                if batch:
                    recompute_bom_import_batch(db, batch)
    db.commit()
    return {"added": added, "updated": updated, "skipped": skipped}


@app.post("/api/projects/{project_id}/bom/import-rows/{row_id}/ignore")
def ignore_bom_import_row(project_id: int, row_id: int, auth: Protected, db: Session = Depends(get_db)):
    require_project_access(db, project_id, auth)
    row = (
        db.query(ProjectBomImportRow)
        .filter(ProjectBomImportRow.project_id == project_id, ProjectBomImportRow.id == row_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="BOM import row not found")
    if row.auto_imported:
        raise HTTPException(status_code=400, detail="该行已自动导入正式 BOM，请先删除对应 BOM 行后再处理")
    row.status = "ignored"
    row.selected_component_id = None
    row.match_confidence = 0
    row.ai_reason = "已手动忽略，不参与匹配进度和待采购统计。"
    row.updated_at = datetime.utcnow()
    batch = db.get(ProjectBomImportBatch, row.batch_id)
    if batch:
        recompute_bom_import_batch(db, batch)
    log_activity(
        db,
        "bom.import.ignore",
        "project_bom_import_row",
        f"忽略 BOM 导入行 {row.designator or row.source_row or row.id}",
        owner_user_id=owner_id(auth),
        entity_id=row.id,
        project_id=project_id,
        detail={
            "source_row": row.source_row,
            "designator": row.designator,
            "manufacturer_part": row.manufacturer_part,
            "value": row.value,
            "footprint": row.footprint,
        },
    )
    db.commit()
    return latest_bom_import_batch_out(db, project_id) or {"rows": []}


@app.post("/api/projects/{project_id}/bom/import-rows/{row_id}/selection")
def update_bom_import_row_selection(
    project_id: int,
    row_id: int,
    payload: BomImportRowSelection,
    auth: Protected,
    db: Session = Depends(get_db),
):
    require_project_access(db, project_id, auth)
    row = (
        db.query(ProjectBomImportRow)
        .filter(ProjectBomImportRow.project_id == project_id, ProjectBomImportRow.id == row_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="BOM import row not found")
    if row.auto_imported:
        return latest_bom_import_batch_out(db, project_id) or {"rows": []}
    component = db.get(Component, payload.component_id) if payload.component_id else None
    if payload.component_id:
        assert_owned(component, auth, "Component not found")
    if component:
        row.selected_component_id = component.id
        row.status = "manual_selected" if row.status not in {"exact_lcsc", "pending_purchase"} else row.status
        row.match_confidence = max(int(row.match_confidence or 0), 100)
        row.ai_reason = "已手动选择库内元器件，等待导入确认。"
        exists = (
            db.query(ProjectBomImportCandidate)
            .filter(ProjectBomImportCandidate.import_row_id == row.id, ProjectBomImportCandidate.component_id == component.id)
            .first()
        )
        if not exists:
            db.add(
                ProjectBomImportCandidate(
                    import_row_id=row.id,
                    component_id=component.id,
                    score=100,
                    match_type="manual",
                    reason="手动从库存选择",
                    flags="手动匹配",
                    available_quantity=int(component.quantity or 0),
                    shortage_quantity=max(0, int(row.required_quantity or 1) - int(component.quantity or 0)),
                    enough=int(component.quantity or 0) >= int(row.required_quantity or 1),
                    rank=-1,
                )
            )
    else:
        if not row.selected_component_id and row.status != "manual_selected":
            return latest_bom_import_batch_out(db, project_id) or {"rows": []}
        row.selected_component_id = None
        row.match_confidence = 0
        if row.status == "manual_selected":
            row.status = "supplier_missing" if row.supplier_part else "review"
        row.ai_reason = "已清空手动匹配，请重新选择库存或加入待采购库。"
    row.updated_at = datetime.utcnow()
    batch = db.get(ProjectBomImportBatch, row.batch_id)
    if batch:
        recompute_bom_import_batch(db, batch)
    log_activity(
        db,
        "bom.import.selection",
        "project_bom_import_row",
        f"{'选择' if component else '清空'} BOM 导入行匹配 {row.designator or row.source_row or row.id}",
        owner_user_id=owner_id(auth),
        entity_id=row.id,
        component_id=component.id if component else None,
        project_id=project_id,
        detail={"source_row": row.source_row, "designator": row.designator, "component_id": payload.component_id},
    )
    db.commit()
    return latest_bom_import_batch_out(db, project_id) or {"rows": []}


@app.post("/api/projects/{project_id}/bom/import-rows/{row_id}/pending-component")
def create_pending_component_from_bom_row(project_id: int, row_id: int, auth: Protected, db: Session = Depends(get_db)):
    require_project_access(db, project_id, auth)
    row = (
        db.query(ProjectBomImportRow)
        .filter(ProjectBomImportRow.project_id == project_id, ProjectBomImportRow.id == row_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="BOM import row not found")
    project = db.get(Project, project_id)
    assert_owned(project, auth, "Project not found")
    if not row.supplier_part and not row.manufacturer_part:
        raise HTTPException(status_code=400, detail="该行缺少立创 ID 或型号，无法加入待采购库")

    component = None
    if row.supplier_part:
        component = filter_owner(db.query(Component), Component, auth).filter(Component.lcsc_number == row.supplier_part).first()
    if not component and row.manufacturer_part:
        component = filter_owner(db.query(Component), Component, auth).filter(Component.model == row.manufacturer_part, Component.source == "BOM 待采购库").first()

    if not component:
        category = None
        if row.comment:
            category = db.query(Category).filter(Category.name == row.comment).first()
        display_name = row.value or row.manufacturer_part or row.supplier_part or f"BOM 行 {row.source_row}"
        component = Component(
            name=display_name,
            model=row.manufacturer_part or row.supplier_part,
            category_id=category.id if category else None,
            parameters=row.value,
            package=row.footprint,
            quantity=0,
            source="BOM 待采购库",
            lcsc_number=row.supplier_part,
            tags="待采购",
            source_title=" / ".join(part for part in [row.manufacturer_part, row.value, row.footprint, row.supplier_part] if part),
            normalized_spec=row.value or row.footprint,
            status="pending_purchase",
            location="待采购",
            remark=f"由项目 {project.name} 的 BOM 导入行创建；位号 {row.designator or '-'}；数量 {row.required_quantity}",
            ai_status="pending",
        )
        db.add(component)
        db.flush()
        set_owner(component, auth)
        assign_component_warehouse_code(db, component)
        enqueue_ai_task(db, "component_organize", "component", component.id, organize_cache_key(component))
        enqueue_ai_task(db, "component_analyze", "component", component.id, component_ai_cache_key(component))

    remark = "；".join(
        [
            "待采购库占位",
            f"BOM 位号: {row.designator or '-'}",
            f"BOM 型号: {row.manufacturer_part or '-'}",
            f"BOM 参数: {row.value or '-'}",
            f"BOM 立创 ID: {row.supplier_part or '-'}",
            f"BOM 封装: {row.footprint or '-'}",
        ]
    )
    existing = (
        db.query(ProjectBomItem)
        .filter(ProjectBomItem.project_id == project_id, ProjectBomItem.component_id == component.id)
        .first()
    )
    if existing:
        existing.required_quantity += int(row.required_quantity or 1)
        existing.status = "reserved"
        existing.remark = f"{existing.remark or ''}\n{remark}".strip()
        sync_bom_solder_points(db, existing)
    else:
        bom_item = ProjectBomItem(
            project_id=project_id,
            component_id=component.id,
            required_quantity=int(row.required_quantity or 1),
            remark=remark,
            status="reserved",
        )
        db.add(bom_item)
        db.flush()
        sync_bom_solder_points(db, bom_item)
    row.selected_component_id = component.id
    row.status = "pending_purchase"
    row.match_confidence = 100
    row.ai_reason = "已加入待采购库并匹配到项目 BOM，后续采购入库后补充库存数量。"
    row.auto_import_note = "待采购库占位，库存数量为 0"
    batch = db.get(ProjectBomImportBatch, row.batch_id)
    if batch:
        recompute_bom_import_batch(db, batch)
    log_activity(
        db,
        "bom.import.pending_purchase",
        "project_bom_import_row",
        f"加入待采购库并匹配 BOM：{component.name}",
        owner_user_id=owner_id(auth),
        entity_id=row.id,
        component_id=component.id,
        project_id=project_id,
        quantity_delta=int(row.required_quantity or 1),
        detail={"lcsc_number": row.supplier_part, "source_row": row.source_row, "designator": row.designator},
    )
    db.commit()
    return latest_bom_import_batch_out(db, project_id) or {"rows": []}


@app.get("/api/projects/{project_id}/shortage")
def project_shortage(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.bom_items).joinedload(ProjectBomItem.component))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    rows = [bom_item_out(item, reserved) for item in project.bom_items]
    return [row for row in rows if not row["enough"]]


@app.get("/api/projects/{project_id}/export")
def export_project_bom(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    writer.writerow(["项目", "名称", "型号", "分类", "参数", "封装", "需求数量", "总库存", "已占用", "可用库存", "缺料数量", "立创 ID", "备注"])
    for item in project.bom_items:
        component = item.component
        row = bom_item_out(item, reserved)
        writer.writerow(
            [
                project.name,
                component.name,
                component.model or "",
                component.category.name if component.category else "",
                component.parameters or "",
                component.package or "",
                item.required_quantity,
                component.quantity,
                row["reserved_quantity"],
                row["available_quantity"],
                row["shortage_quantity"],
                component.lcsc_number or "",
                item.remark or "",
            ]
        )
    buffer.seek(0)
    filename = f"project-{project_id}-bom.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def purchase_export_key(row: dict) -> tuple[str, str, str, str]:
    return (
        str(row.get("lcsc_number") or "").strip(),
        str(row.get("model") or "").strip(),
        str(row.get("package") or "").strip(),
        str(row.get("parameters") or "").strip(),
    )


def add_purchase_export_row(rows: dict, row: dict) -> None:
    key = purchase_export_key(row)
    if key in rows:
        rows[key]["quantity"] += int(row.get("quantity") or 1)
        designator = str(row.get("designator") or "").strip()
        if designator:
            existing = {item.strip() for item in str(rows[key].get("designator") or "").split(",") if item.strip()}
            existing.add(designator)
            rows[key]["designator"] = ",".join(sorted(existing))
        return
    rows[key] = {
        "model": row.get("model") or "",
        "package": row.get("package") or "",
        "parameters": row.get("parameters") or "",
        "quantity": int(row.get("quantity") or 1),
        "brand": row.get("brand") or "",
        "category": row.get("category") or "",
        "lcsc_number": row.get("lcsc_number") or "",
        "material_code": row.get("material_code") or "",
        "designator": row.get("designator") or "",
    }


@app.get("/api/projects/{project_id}/purchase-bom/export")
def export_project_purchase_bom(project_id: int, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")

    export_rows: dict[tuple[str, str, str, str], dict] = {}
    batch = (
        db.query(ProjectBomImportBatch)
        .filter(ProjectBomImportBatch.project_id == project_id)
        .order_by(ProjectBomImportBatch.created_at.desc(), ProjectBomImportBatch.id.desc())
        .first()
    )
    if batch:
        import_rows = (
            db.query(ProjectBomImportRow)
            .filter(ProjectBomImportRow.batch_id == batch.id, ProjectBomImportRow.status != "ignored")
            .order_by(ProjectBomImportRow.source_row.asc(), ProjectBomImportRow.id.asc())
            .all()
        )
        for row in import_rows:
            selected = db.get(Component, row.selected_component_id) if row.selected_component_id else None
            needs_purchase = row.status in {"supplier_missing", "missing", "review", "pending_purchase"}
            if selected and selected.status == "pending_purchase":
                needs_purchase = True
            if not needs_purchase:
                continue
            add_purchase_export_row(
                export_rows,
                {
                    "model": row.manufacturer_part or (selected.model if selected else "") or row.supplier_part,
                    "package": row.footprint or (selected.package if selected else ""),
                    "parameters": row.value or row.comment or (selected.parameters if selected else ""),
                    "quantity": row.required_quantity or 1,
                    "brand": "",
                    "category": row.comment or (selected.category.name if selected and selected.category else ""),
                    "lcsc_number": row.supplier_part or (selected.lcsc_number if selected else ""),
                    "material_code": "",
                    "designator": row.designator or "",
                },
            )

    if not export_rows:
        reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
        for item in project.bom_items:
            row = bom_item_out(item, reserved)
            if row["shortage_quantity"] <= 0 and item.component.status != "pending_purchase":
                continue
            component = item.component
            add_purchase_export_row(
                export_rows,
                {
                    "model": component.model or component.name,
                    "package": component.package,
                    "parameters": component.normalized_spec or component.parameters or component.name,
                    "quantity": row["shortage_quantity"] or item.required_quantity,
                    "brand": "",
                    "category": component.category.name if component.category else "",
                    "lcsc_number": component.lcsc_number,
                    "material_code": "",
                    "designator": item.remark or "",
                },
            )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购BOM"
    headers = ["序号", "型号", "封装", "参数/规格", "单套用量", "品牌", "分类", "商品编号", "物料编码", "位号"]
    sheet.append(headers)
    for index, row in enumerate(export_rows.values(), start=1):
        sheet.append(
            [
                index,
                row["model"],
                row["package"],
                row["parameters"],
                row["quantity"],
                row["brand"],
                row["category"],
                row["lcsc_number"],
                row["material_code"],
                row["designator"],
            ]
        )
    for column, width in zip("ABCDEFGHIJ", [8, 26, 18, 36, 12, 18, 16, 18, 18, 24]):
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"project-{project_id}-purchase-bom.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/projects/{project_id}/ai/analyze-bom")
def analyze_project_bom(project_id: int, auth: Protected, db: Session = Depends(get_db), force: bool = False):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    cache_key = project_bom_cache_key(project)
    if not force and project.ai_bom_analysis and project.ai_bom_cache_key == cache_key:
        return json.loads(project.ai_bom_analysis)
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    bom_items = [bom_item_out(item, reserved) for item in project.bom_items]
    try:
        result = analyze_bom(
            {"id": project.id, "name": project.name, "description": project.description, "status": project.status},
            bom_items,
        )
    except Exception as error:
        handle_mimo_error(error)
    result["cache_key"] = cache_key
    result["generated_at"] = datetime.now().isoformat()
    project.ai_bom_analysis = json.dumps(result, ensure_ascii=False)
    project.ai_bom_cache_key = cache_key
    project.ai_bom_updated_at = datetime.now()
    db.add(
        AiKnowledgeCard(
            project_id=project.id,
            title=f"{project.name} BOM AI 分析",
            content=json.dumps(result, ensure_ascii=False, indent=2),
            tags="BOM,缺料,风险",
            source_type="ai",
            confidence=result.get("confidence"),
        )
    )
    db.commit()
    return result


@app.post("/api/projects/{project_id}/ai/plan")
def project_ai_plan(project_id: int, payload: ProjectAiPlanRequest, auth: Protected, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    assert_owned(project, auth, "Project not found")
    candidates = filter_owner(db.query(Component), Component, auth).filter(Component.id.in_([item.id for item in search_component_candidates(db, payload.goal, 30)])).all()
    try:
        result = project_plan(payload.goal, candidates)
    except Exception as error:
        handle_mimo_error(error)
    result["generated_at"] = datetime.now().isoformat()
    db.add(
        AiKnowledgeCard(
            project_id=project.id,
            title=f"{project.name} 项目规划",
            content=json.dumps(result, ensure_ascii=False, indent=2),
            tags="项目规划,推荐BOM,风险",
            source_type="ai",
            confidence=result.get("confidence"),
        )
    )
    db.commit()
    return result


@app.post("/api/projects/{project_id}/ai/consult")
def project_ai_consult(project_id: int, payload: ProjectAiConsultRequest, auth: Protected, db: Session = Depends(get_db)):
    project = (
        filter_owner(db.query(Project), Project, auth)
        .options(joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category))
        .filter(Project.id == project_id)
        .first()
    )
    assert_owned(project, auth, "Project not found")
    reserved = reserved_quantities(db, list({item.component_id for item in project.bom_items}))
    bom_items = [bom_item_out(item, reserved) for item in project.bom_items]
    candidates = filter_owner(db.query(Component), Component, auth).filter(Component.id.in_([item.id for item in search_component_candidates(db, payload.question, 30)])).all()
    try:
        result = project_consult(
            payload.question,
            {"id": project.id, "name": project.name, "description": project.description, "status": project.status},
            bom_items,
            candidates,
        )
    except Exception as error:
        handle_mimo_error(error)
    result["generated_at"] = datetime.now().isoformat()
    db.add(
        AiKnowledgeCard(
            project_id=project.id,
            title=f"{project.name} AI 咨询",
            content=json.dumps(result, ensure_ascii=False, indent=2),
            tags="项目咨询,BOM,推荐",
            source_type="ai",
            confidence=result.get("confidence"),
        )
    )
    db.commit()
    return result


@app.post("/api/import/excel/preview", response_model=list[ImportPreviewRow])
async def preview_excel(auth: Protected, file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    parsed = parse_excel(content, db, file.filename)
    result: list[dict] = []
    for row in parsed:
        duplicate = find_duplicate(db, row.data)
        if duplicate and owner_id(auth) and duplicate.owner_user_id != owner_id(auth):
            duplicate = None
        import_record = find_import_record(db, row.data)
        if import_record and owner_id(auth) and import_record.owner_user_id != owner_id(auth):
            import_record = None
        already_imported = import_record is not None
        suggested_action = "skip" if already_imported else "merge" if duplicate else "create"
        result.append(
            {
                **row.data,
                "source_row": row.source_row,
                "duplicate": duplicate is not None,
                "duplicate_component_id": duplicate.id if duplicate else None,
                "already_imported": already_imported,
                "suggested_action": suggested_action,
                "action": suggested_action,
            }
        )
    return result


ORDER_IMPORT_SNAPSHOT_FIELDS = [
    "warehouse_code",
    "name",
    "model",
    "category_id",
    "parameters",
    "package",
    "quantity",
    "source",
    "lcsc_number",
    "tags",
    "source_title",
    "part_family",
    "count_mode",
    "normalized_spec",
    "status",
    "location",
    "remark",
    "datasheet_url",
    "ai_status",
    "first_stocked_at",
    "last_stocked_at",
    "last_outbound_at",
]


def component_snapshot(component: Component | None) -> dict | None:
    if not component:
        return None
    snapshot = {}
    for field in ORDER_IMPORT_SNAPSHOT_FIELDS:
        value = getattr(component, field, None)
        snapshot[field] = value.isoformat() if isinstance(value, datetime) else value
    return snapshot


def restore_component_snapshot(component: Component, previous: dict) -> None:
    date_fields = {"first_stocked_at", "last_stocked_at", "last_outbound_at"}
    for field in ORDER_IMPORT_SNAPSHOT_FIELDS:
        if field not in previous:
            continue
        value = previous[field]
        if field in date_fields and isinstance(value, str):
            try:
                value = datetime.fromisoformat(value)
            except ValueError:
                value = None
        setattr(component, field, value)


def add_order_import_line(
    db: Session,
    *,
    batch: OrderImportBatch,
    values: dict,
    operation: str,
    component: Component | None,
    quantity_delta: int = 0,
    previous_component: dict | None = None,
    import_record: ImportRecord | None = None,
    note: str | None = None,
) -> None:
    db.add(
        OrderImportLine(
            batch_id=batch.id,
            import_record_id=import_record.id if import_record else None,
            component_id=component.id if component else None,
            source_row=values.get("source_row"),
            order_number=values.get("order_number"),
            lcsc_number=values.get("lcsc_number"),
            operation=operation,
            quantity_delta=int(quantity_delta or 0),
            previous_component=json.dumps(previous_component, ensure_ascii=False, default=str) if previous_component else None,
            row_data=json.dumps(values, ensure_ascii=False, default=str),
            note=note,
            owner_user_id=batch.owner_user_id,
        )
    )


@app.post("/api/import/excel/commit", response_model=ImportCommitResult)
def commit_excel(payload: ImportCommitRequest, auth: Protected, db: Session = Depends(get_db)):
    created = merged = skipped = already_imported = resolved_pending_purchase = 0
    touched_component_ids: set[int] = set()
    first_row = payload.rows[0].model_dump() if payload.rows else {}
    batch = OrderImportBatch(
        source_file=first_row.get("source_file"),
        order_number=first_row.get("order_number"),
        status="active",
    )
    set_owner(batch, auth)
    db.add(batch)
    db.flush()
    for row in payload.rows:
        values = row.model_dump()
        if not values.get("name"):
            skipped += 1
            add_order_import_line(db, batch=batch, values=values, operation="skip", component=None, note="缺少名称，已跳过")
            continue
        existing_import_record = find_import_record(db, values)
        if existing_import_record and owner_id(auth) and existing_import_record.owner_user_id != owner_id(auth):
            existing_import_record = None
        if existing_import_record:
            already_imported += 1
            skipped += 1
            add_order_import_line(db, batch=batch, values=values, operation="skip", component=None, note="同订单同立创 ID 已导入，自动跳过")
            continue
        component_data = normalize_for_inventory(db, component_values(values), clean_name=True)
        duplicate = find_duplicate(db, values)
        if duplicate and owner_id(auth) and duplicate.owner_user_id != owner_id(auth):
            duplicate = None
        if duplicate:
            previous = component_snapshot(duplicate)
            was_pending_purchase = duplicate.status == "pending_purchase"
            quantity_delta = int(values.get("quantity") or 0)
            merge_component(duplicate, component_data)
            mark_stock_change(duplicate, quantity_delta)
            if quantity_delta > 0:
                record_stock_delta(
                    db,
                    duplicate,
                    quantity_delta,
                    movement_type="order_import_merge",
                    reason="立创订单导入合并库存",
                    actor_user_id=owner_id(auth),
                    source_type="lcsc",
                    source_reference=values.get("order_number") or batch.order_number or values.get("source_file"),
                    location=component_data.get("location"),
                )
            duplicate.ai_status = "stale"
            import_record = create_import_record(db, values, duplicate.id, batch.id)
            if import_record:
                set_owner(import_record, auth)
            add_order_import_line(
                db,
                batch=batch,
                values=values,
                operation="merge",
                component=duplicate,
                quantity_delta=quantity_delta,
                previous_component=previous,
                import_record=import_record,
                note="立创 ID 一致，自动合并库存" if values.get("lcsc_number") else "名称/型号/封装一致，自动合并库存",
            )
            touched_component_ids.add(duplicate.id)
            merged += 1
            if was_pending_purchase:
                resolved_pending_purchase += 1
        else:
            component = Component(**component_data)
            set_owner(component, auth)
            component.ai_status = "pending"
            quantity_delta = int(values.get("quantity") or 0)
            if quantity_delta > 0:
                mark_stock_change(component, quantity_delta)
            db.add(component)
            db.flush()
            assign_component_warehouse_code(db, component)
            if quantity_delta > 0:
                record_stock_delta(
                    db,
                    component,
                    quantity_delta,
                    movement_type="order_import_create",
                    reason="立创订单导入新增库存",
                    actor_user_id=owner_id(auth),
                    source_type="lcsc",
                    source_reference=values.get("order_number") or batch.order_number or values.get("source_file"),
                    location=component_data.get("location"),
                )
            import_record = create_import_record(db, values, component.id, batch.id)
            if import_record:
                set_owner(import_record, auth)
            add_order_import_line(
                db,
                batch=batch,
                values=values,
                operation="create",
                component=component,
                quantity_delta=quantity_delta,
                import_record=import_record,
                note="立创 ID 未入库，自动新增库存器件",
            )
            touched_component_ids.add(component.id)
            created += 1
        log_activity(
            db,
            "import.excel.row",
            "component",
            f"Excel 导入 {values.get('name')} x {values.get('quantity') or 0}",
            component_id=(duplicate.id if duplicate else component.id),
            owner_user_id=owner_id(auth),
            quantity_delta=int(values.get("quantity") or 0),
            detail={
                "order_number": values.get("order_number"),
                "lcsc_number": values.get("lcsc_number"),
                "action": "merge" if duplicate else "create",
                "source_file": values.get("source_file"),
                "resolved_pending_purchase": bool(duplicate and duplicate.status == "in_stock"),
                "batch_id": batch.id,
            },
        )
    batch.created_count = created
    batch.merged_count = merged
    batch.skipped_count = skipped
    batch.already_imported_count = already_imported
    batch.resolved_pending_count = resolved_pending_purchase
    for component_id in touched_component_ids:
        component = db.get(Component, component_id)
        if component:
            enqueue_ai_task(db, "component_organize", "component", component.id, organize_cache_key(component))
            enqueue_ai_task(db, "component_analyze", "component", component.id, component_ai_cache_key(component))
    db.commit()
    log_activity(
        db,
        "import.excel.commit",
        "import",
        f"Excel 导入完成：新增 {created}，合并 {merged}，跳过 {skipped}，已导入 {already_imported}",
        owner_user_id=owner_id(auth),
        entity_id=batch.id,
        detail={
            "batch_id": batch.id,
            "created": created,
            "merged": merged,
            "skipped": skipped,
            "already_imported": already_imported,
            "resolved_pending_purchase": resolved_pending_purchase,
        },
    )
    db.commit()
    return {
        "created": created,
        "merged": merged,
        "skipped": skipped,
        "already_imported": already_imported,
        "resolved_pending_purchase": resolved_pending_purchase,
        "batch_id": batch.id,
    }


def order_import_batch_out(batch: OrderImportBatch, lines: list[OrderImportLine] | None = None) -> dict:
    return {
        "id": batch.id,
        "source_file": batch.source_file,
        "order_number": batch.order_number,
        "status": batch.status,
        "created_count": batch.created_count,
        "merged_count": batch.merged_count,
        "skipped_count": batch.skipped_count,
        "already_imported_count": batch.already_imported_count,
        "resolved_pending_count": batch.resolved_pending_count,
        "rollback_summary": batch.rollback_summary,
        "created_at": batch.created_at,
        "rolled_back_at": batch.rolled_back_at,
        "lines": lines or [],
    }


@app.get("/api/import/excel/batches", response_model=list[OrderImportBatchOut])
def list_order_import_batches(auth: Protected, db: Session = Depends(get_db), limit: int = Query(20, ge=1, le=100)):
    batches = filter_owner(db.query(OrderImportBatch), OrderImportBatch, auth).order_by(OrderImportBatch.created_at.desc(), OrderImportBatch.id.desc()).limit(limit).all()
    return [order_import_batch_out(batch) for batch in batches]


@app.post("/api/import/excel/batches/{batch_id}/rollback", response_model=OrderImportBatchOut)
def rollback_order_import_batch(batch_id: int, auth: Protected, db: Session = Depends(get_db)):
    batch = db.get(OrderImportBatch, batch_id)
    assert_owned(batch, auth, "Import batch not found")
    if batch.status == "rolled_back" or batch.rolled_back_at:
        raise HTTPException(status_code=400, detail="该导入批次已经撤销，不能重复撤销")
    lines = (
        db.query(OrderImportLine)
        .filter(OrderImportLine.batch_id == batch.id)
        .order_by(OrderImportLine.id.desc())
        .all()
    )
    restored = marked_created = skipped = 0
    now = datetime.utcnow()
    for line in lines:
        if line.rolled_back_at:
            skipped += 1
            continue
        component = db.get(Component, line.component_id) if line.component_id else None
        if line.operation == "merge" and component and line.previous_component:
            previous = json.loads(line.previous_component)
            restore_component_snapshot(component, previous)
            restored += 1
        elif line.operation == "create" and component:
            component.quantity = 0
            component.status = "rolled_back"
            component.location = None
            note = f"订单导入批次 {batch.id} 已撤销，保留记录但不计库存。"
            component.remark = f"{component.remark or ''}\n{note}".strip()
            marked_created += 1
        else:
            skipped += 1
        if line.import_record_id:
            record = db.get(ImportRecord, line.import_record_id)
            if record:
                db.delete(record)
        line.rolled_back_at = now
    batch.status = "rolled_back"
    batch.rolled_back_at = now
    batch.rollback_summary = f"已还原合并 {restored} 行，撤销新增 {marked_created} 行，跳过 {skipped} 行。"
    log_activity(
        db,
            "import.excel.rollback",
            "import",
            f"撤销 Excel 导入批次 {batch.id}",
            owner_user_id=owner_id(auth),
            entity_id=batch.id,
        detail={"restored": restored, "marked_created": marked_created, "skipped": skipped},
    )
    db.commit()
    db.refresh(batch)
    lines = db.query(OrderImportLine).filter(OrderImportLine.batch_id == batch.id).order_by(OrderImportLine.id.asc()).all()
    return order_import_batch_out(batch, lines)


@app.post("/api/import/external-order/preview", response_model=list[ImportPreviewRow])
async def preview_external_order(auth: Protected, file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    categories = [name for (name,) in db.query(Category.name).order_by(Category.id).all()]
    category_map = {name: cid for name, cid in db.query(Category.name, Category.id).all()}
    try:
        parsed = await run_foreground_ai(parse_external_order, content, file.filename, categories)
    except MimoNotConfiguredError as exc:
        raise HTTPException(status_code=503, detail="外部订单导入需要配置 MIMO_API_KEY 才能进行 AI 表格解析") from exc
    except MimoRequestError as exc:
        raise HTTPException(status_code=502, detail=f"AI 外部订单解析失败：{exc}") from exc
    result: list[dict] = []
    for row in parsed:
        if row.data.get("category_name") in category_map:
            row.data["category_id"] = category_map[row.data["category_name"]]
        duplicate = find_external_duplicate(db, row.data)
        if duplicate and owner_id(auth) and duplicate.owner_user_id != owner_id(auth):
            duplicate = None
        import_record = find_external_import_record(db, row.data)
        if import_record and owner_id(auth) and import_record.owner_user_id != owner_id(auth):
            import_record = None
        already_imported = import_record is not None
        suggested_action = "skip" if already_imported else "merge" if duplicate else "create"
        result.append({
            **row.data,
            "source_row": row.source_row,
            "duplicate": duplicate is not None,
            "duplicate_component_id": duplicate.id if duplicate else None,
            "already_imported": already_imported,
            "suggested_action": suggested_action,
            "action": suggested_action,
        })
    return result


@app.post("/api/import/external-order/commit", response_model=ImportCommitResult)
def commit_external_order(payload: ImportCommitRequest, auth: Protected, db: Session = Depends(get_db)):
    category_map = {name: cid for name, cid in db.query(Category.name, Category.id).all()}
    created = merged = skipped = 0
    touched_component_ids: set[int] = set()
    first_row = payload.rows[0].model_dump() if payload.rows else {}
    batch = OrderImportBatch(
        source_file=first_row.get("source_file") or "外部订单导入",
        order_number=first_row.get("order_number"),
        status="active",
    )
    set_owner(batch, auth)
    db.add(batch)
    db.flush()
    batch_id = batch.id
    db.commit()
    batch = db.get(OrderImportBatch, batch_id)

    for row in payload.rows:
        values = row.model_dump()
        if not values.get("name") or values.get("action") == "skip":
            skipped += 1
            add_order_import_line(db, batch=batch, values=values, operation="skip", component=None, note="跳过")
            db.commit()
            continue
        cat_name = values.pop("category_name", None)
        if cat_name and cat_name in category_map:
            values["category_id"] = category_map[cat_name]
        if not isinstance(values.get("category_id"), int):
            values["category_id"] = None
        existing_record = find_external_import_record(db, values)
        if existing_record and owner_id(auth) and existing_record.owner_user_id != owner_id(auth):
            existing_record = None
        if existing_record:
            skipped += 1
            add_order_import_line(db, batch=batch, values=values, operation="skip", component=None, note="该外部订单行已导入，自动跳过")
            db.commit()
            continue
        try:
            existing = find_external_duplicate(db, values)
            if existing and owner_id(auth) and existing.owner_user_id != owner_id(auth):
                existing = None
            if existing:
                previous = component_snapshot(existing)
                component_data = normalize_for_inventory(db, component_values(values), clean_name=True)
                merge_component(existing, component_data)
                quantity_delta = int(values.get("quantity") or 0)
                mark_stock_change(existing, quantity_delta)
                if quantity_delta > 0:
                    record_stock_delta(
                        db,
                        existing,
                        quantity_delta,
                        movement_type="external_order_merge",
                        reason="外部订单导入合并库存",
                        actor_user_id=owner_id(auth),
                        source_type="external_order",
                        source_reference=values.get("order_number") or batch.order_number or values.get("source_file"),
                        location=component_data.get("location"),
                    )
                existing.ai_status = "stale"
                touched_component_ids.add(existing.id)
                merged += 1
                add_order_import_line(
                    db,
                    batch=batch,
                    values=values,
                    operation="merge",
                    component=existing,
                    quantity_delta=int(values.get("quantity") or 0),
                    previous_component=previous,
                    note="外部订单 AI 解析后精确合并",
                )
                component_for_record = existing
            else:
                component_data = normalize_for_inventory(db, component_values(values), clean_name=True)
                if not component_data.get("source"):
                    component_data["source"] = "外部订单 AI 导入"
                component = Component(**component_data)
                set_owner(component, auth)
                if int(component.quantity or 0) > 0:
                    mark_stock_change(component, int(component.quantity or 0))
                db.add(component)
                db.flush()
                assign_component_warehouse_code(db, component)
                if int(component.quantity or 0) > 0:
                    record_stock_delta(
                        db,
                        component,
                        int(component.quantity or 0),
                        movement_type="external_order_create",
                        reason="外部订单导入新增库存",
                        actor_user_id=owner_id(auth),
                        source_type="external_order",
                        source_reference=values.get("order_number") or batch.order_number or values.get("source_file"),
                        location=component_data.get("location"),
                    )
                touched_component_ids.add(component.id)
                created += 1
                add_order_import_line(
                    db,
                    batch=batch,
                    values=values,
                    operation="create",
                    component=component,
                    quantity_delta=int(values.get("quantity") or 0),
                    note="外部订单 AI 解析后新增",
                )
                component_for_record = component
            record_payload = external_import_record_payload(values)
            if not find_import_record(db, record_payload):
                import_record = create_import_record(db, record_payload, component_for_record.id, batch.id)
                if import_record:
                    set_owner(import_record, auth)
            log_activity(
                db,
                "import.external_order.row",
                "component",
                f"外部订单 AI 导入 {values.get('name')} x {values.get('quantity') or 0}",
                component_id=component_for_record.id,
                owner_user_id=owner_id(auth),
                quantity_delta=int(values.get("quantity") or 0),
                detail={
                    "source_file": values.get("source_file"),
                    "source_row": values.get("source_row"),
                    "order_number": values.get("order_number"),
                    "ai_confidence": values.get("ai_confidence"),
                    "ai_reason": values.get("ai_reason"),
                    "action": "merge" if existing else "create",
                    "batch_id": batch.id,
                },
            )
            db.commit()
        except Exception:
            db.rollback()
            skipped += 1
        batch = db.get(OrderImportBatch, batch_id)

    batch = db.get(OrderImportBatch, batch_id)
    for cid in touched_component_ids:
        component = db.get(Component, cid)
        if component:
            mark_component_ai_stale(component)
            enqueue_ai_task(db, "component_organize", "component", component.id, organize_cache_key(component))
            enqueue_ai_task(db, "component_analyze", "component", component.id, component_ai_cache_key(component))
    batch.created_count = created
    batch.merged_count = merged
    batch.skipped_count = skipped
    db.commit()
    return {"created": created, "merged": merged, "skipped": skipped, "batch_id": batch_id, "resolved_pending_purchase": 0}


@app.get("/api/import/external-order/template")
def get_external_order_template():
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "外部订单样表"
    headers = ["订单号", "订单提交时间", "订单状态", "店铺名称", "商品名称", "商品链接", "型号款式", "商品数量", "商品金额", "实付金额", "运费", "物流公司", "物流单号"]
    ws.append(headers)
    ws.append(["TB20240101001", "2024-01-01 10:30:00", "已完成", "某电子旗舰店", "0805 100nF 贴片电容 100只", "https://item.taobao.com/xxx", "0805 100nF 50V X7R", "1", "9.90", "8.50", "0", "圆通", "YT1234567890"])
    ws.append(["168820240102002", "2024-01-02 14:00:00", "已完成", "深圳某电子", "AO3400 N-MOS SOT-23", "https://detail.1688.com/xxx", "AO3400", "50", "25.00", "22.00", "3", "顺丰", "SF9876543210"])
    ws.append(["PDD20240103003", "2024-01-03 09:15:00", "待发货", "某元器件专营", "USB Type-C 16P 母座 沉板", "https://mobile.yangkeduo.com/xxx", "USB-C 16P 沉板", "20", "15.00", "13.00", "0", "", ""])
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=external_order_template.xlsx"})


def sqlite_database_path() -> Path | None:
    if not DATABASE_URL.startswith("sqlite:///"):
        return None
    raw_path = DATABASE_URL.replace("sqlite:///", "", 1)
    if raw_path == ":memory:":
        return None
    return Path(raw_path).expanduser().resolve()


def backup_root() -> Path | None:
    db_path = sqlite_database_path()
    if not db_path:
        return None
    root = db_path.parent / "backups"
    root.mkdir(parents=True, exist_ok=True)
    return root


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_backup_file(archive: zipfile.ZipFile, path: Path, arcname: str, files: list[dict]) -> None:
    archive.write(path, arcname)
    files.append({"path": arcname, "bytes": path.stat().st_size, "sha256": file_sha256(path)})


def backup_type_from_name(path: Path) -> str:
    name = path.name
    if name.startswith("pre-restore-"):
        return "pre_restore"
    if name.startswith("pre-clear-"):
        return "pre_clear"
    if name.startswith("auto-"):
        return "auto"
    return "manual"


def create_backup_archive(backup_type: str = "manual") -> tuple[bytes, dict]:
    created_at = datetime.utcnow()
    db_path = sqlite_database_path()
    buffer = io.BytesIO()
    manifest = {
        "app": APP_BACKUP_NAME,
        "created_at": created_at.isoformat() + "Z",
        "backup_type": backup_type,
        "database_url": "sqlite" if DATABASE_URL.startswith("sqlite") else "non-sqlite",
        "files": [],
        "warnings": [],
    }
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        if db_path and db_path.exists():
            snapshot = tempfile.NamedTemporaryFile(prefix="cw-backup-", suffix=".db", delete=False)
            snapshot_path = Path(snapshot.name)
            snapshot.close()
            try:
                source = sqlite3.connect(str(db_path))
                target = sqlite3.connect(str(snapshot_path))
                try:
                    source.backup(target)
                finally:
                    target.close()
                    source.close()
                add_backup_file(archive, snapshot_path, "database/component_warehouse.snapshot.db", manifest["files"])
            finally:
                try:
                    snapshot_path.unlink(missing_ok=True)
                except OSError:
                    pass

            data_root = db_path.parent
            if data_root.exists():
                backups_dir = data_root / "backups"
                eda_root = eda_storage_root()
                for path in sorted(data_root.rglob("*")):
                    if (
                        path.is_file()
                        and backups_dir not in path.parents
                        and eda_root != path
                        and eda_root not in path.parents
                    ):
                        add_backup_file(archive, path, f"data/{path.relative_to(data_root).as_posix()}", manifest["files"])
                manifest["warnings"].append(
                    "EDA 大文件未包含在常规数据库备份中；请使用 /api/admin/eda/archive 生成完整或增量归档。"
                )
        else:
            manifest["warnings"].append("当前 DATABASE_URL 不是可备份的 SQLite 文件路径，未生成数据库文件快照。")
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    return buffer.getvalue(), manifest


def write_named_backup(prefix: str) -> Path | None:
    root = backup_root()
    if not root:
        return None
    backup_bytes, _ = create_backup_archive(prefix.rstrip("-"))
    path = root / f"{prefix}-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"
    path.write_bytes(backup_bytes)
    prune_backups()
    return path


def list_local_backups() -> list[dict]:
    root = backup_root()
    if not root:
        return []
    rows = []
    for path in sorted(root.glob("*.zip"), key=lambda item: item.stat().st_mtime, reverse=True):
        stat = path.stat()
        rows.append(
            {
                "filename": path.name,
                "type": backup_type_from_name(path),
                "bytes": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime),
            }
        )
    return rows


def prune_backups() -> None:
    root = backup_root()
    if not root:
        return
    backups = sorted(root.glob("*.zip"), key=lambda item: item.stat().st_mtime, reverse=True)
    for path in backups[BACKUP_KEEP_COUNT:]:
        try:
            path.unlink()
        except OSError:
            pass


def ensure_auto_backup() -> None:
    root = backup_root()
    if not root:
        return
    auto_backups = sorted(root.glob("auto-*.zip"), key=lambda item: item.stat().st_mtime, reverse=True)
    if auto_backups:
        age = datetime.utcnow() - datetime.utcfromtimestamp(auto_backups[0].stat().st_mtime)
        if age < timedelta(hours=BACKUP_AUTO_INTERVAL_HOURS):
            prune_backups()
            return
    write_named_backup("auto")


def validate_backup_zip(content: bytes) -> dict:
    if not content:
        raise HTTPException(status_code=400, detail="备份文件为空")
    if len(content) > BACKUP_MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="备份文件过大")
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="不是有效的 ZIP 备份文件")
    with archive:
        names = archive.namelist()
        for name in names:
            parts = Path(name).parts
            if name.startswith("/") or ".." in parts:
                raise HTTPException(status_code=400, detail="备份文件包含不安全路径")
        if "manifest.json" not in names:
            raise HTTPException(status_code=400, detail="缺少 manifest.json，不是系统备份")
        if "database/component_warehouse.snapshot.db" not in names:
            raise HTTPException(status_code=400, detail="缺少数据库快照")
        try:
            manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            raise HTTPException(status_code=400, detail="manifest.json 无法读取")
        if manifest.get("app") not in {APP_BACKUP_NAME, "Component Warehouse"}:
            raise HTTPException(status_code=400, detail=f"不是 {APP_BRAND_NAME} 备份")
        snapshot_info = archive.getinfo("database/component_warehouse.snapshot.db")
        if snapshot_info.file_size <= 0:
            raise HTTPException(status_code=400, detail="数据库快照为空")
        snapshot_bytes = archive.read("database/component_warehouse.snapshot.db")
    snapshot = tempfile.NamedTemporaryFile(prefix="cw-inspect-", suffix=".db", delete=False)
    snapshot_path = Path(snapshot.name)
    snapshot.write(snapshot_bytes)
    snapshot.close()
    try:
        connection = sqlite3.connect(str(snapshot_path))
        try:
            ok = connection.execute("PRAGMA integrity_check").fetchone()
            if not ok or ok[0] != "ok":
                raise HTTPException(status_code=400, detail="数据库快照校验失败")
            table_count = connection.execute("SELECT count(*) FROM sqlite_master WHERE type='table'").fetchone()[0]
        finally:
            connection.close()
    finally:
        snapshot_path.unlink(missing_ok=True)
    return {
        "manifest": manifest,
        "snapshot_bytes": len(snapshot_bytes),
        "snapshot_sha256": hashlib.sha256(snapshot_bytes).hexdigest(),
        "file_count": len(names),
        "table_count": table_count,
        "warnings": manifest.get("warnings") or [],
    }


@app.get("/api/admin/backup")
def export_data_backup(_: AdminProtected):
    backup_bytes, _ = create_backup_archive("manual")
    created_at = datetime.utcnow()
    filename = f"component-warehouse-backup-{created_at.strftime('%Y%m%d-%H%M%S')}.zip"
    return Response(
        content=backup_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/backups")
def list_data_backups(_: AdminProtected):
    return {"items": list_local_backups(), "keep_count": BACKUP_KEEP_COUNT}


@app.post("/api/admin/backup/inspect")
async def inspect_data_backup(_: AdminProtected, file: UploadFile = File(...)):
    content = await file.read()
    result = validate_backup_zip(content)
    return {
        "filename": file.filename,
        "created_at": result["manifest"].get("created_at"),
        "backup_type": result["manifest"].get("backup_type"),
        "snapshot_bytes": result["snapshot_bytes"],
        "snapshot_sha256": result["snapshot_sha256"],
        "file_count": result["file_count"],
        "table_count": result["table_count"],
        "warnings": result["warnings"],
    }


@app.post("/api/admin/restore")
async def restore_data_backup(
    _: AdminProtected,
    file: UploadFile = File(...),
    confirm_text: str = Form(...),
):
    if confirm_text.strip() != "恢复数据库":
        raise HTTPException(status_code=400, detail="确认文本不正确")
    db_path = sqlite_database_path()
    if not db_path:
        raise HTTPException(status_code=400, detail="当前只支持 SQLite 文件数据库恢复")
    content = await file.read()
    inspection = validate_backup_zip(content)
    with BACKUP_LOCK:
        pre_restore_path = write_named_backup("pre-restore")
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            snapshot_bytes = archive.read("database/component_warehouse.snapshot.db")
        snapshot = tempfile.NamedTemporaryFile(prefix="cw-restore-", suffix=".db", delete=False)
        snapshot_path = Path(snapshot.name)
        snapshot.write(snapshot_bytes)
        snapshot.close()
        try:
            engine.dispose()
            source = sqlite3.connect(str(snapshot_path))
            target = sqlite3.connect(str(db_path))
            try:
                source.backup(target)
            finally:
                target.close()
                source.close()
            Base.metadata.create_all(bind=engine)
            with engine.begin() as connection:
                ensure_database_schema(connection)
            migration_db = SessionLocal()
            try:
                run_v04_account_migration(migration_db)
                run_v041_inventory_migration(migration_db)
                run_v070_eda_migration(migration_db)
                seed_categories(migration_db)
                seed_category_prefixes(migration_db)
                run_component_identity_migration(migration_db)
                ensure_project_codes(migration_db)
                ensure_project_boards(migration_db)
                ensure_bom_solder_points(migration_db)
                migration_db.commit()
            finally:
                migration_db.close()
        finally:
            snapshot_path.unlink(missing_ok=True)
        db = SessionLocal()
        try:
            log_activity(
                db,
                "admin.restore",
                "system",
                f"恢复数据库备份 {file.filename or ''}".strip(),
                detail={
                    "backup_created_at": inspection["manifest"].get("created_at"),
                    "snapshot_sha256": inspection["snapshot_sha256"],
                    "pre_restore_backup": pre_restore_path.name if pre_restore_path else None,
                },
            )
            db.commit()
        finally:
            db.close()
    return {
        "restored": True,
        "message": "数据库已恢复。建议刷新页面；如仍看到旧数据，请重启后端服务。",
        "pre_restore_backup": pre_restore_path.name if pre_restore_path else None,
    }


@app.post("/api/admin/clear-database")
def clear_database(payload: dict, _: AdminProtected, db: Session = Depends(get_db)):
    if payload.get("confirm_text") != "清空数据库":
        raise HTTPException(status_code=400, detail="确认文本不正确")
    write_named_backup("pre-clear")
    clear_team_media()
    for component in (
        db.query(Component)
        .options(joinedload(Component.category))
        .order_by(Component.id.asc())
        .all()
    ):
        archive_component_identity(db, component)
    db.flush()
    for model in [
        CompetitionAiResult,
        CompetitionActivityLog,
        CompetitionComponentMarker,
        CompetitionInvite,
        CompetitionLibraryComponent,
        CompetitionPcb,
        CompetitionLibraryMember,
        CompetitionLibrary,
        ProjectBomImportCandidate,
        ProjectBomImportRow,
        ProjectBomImportBatch,
        ProjectBomSolderPoint,
        ProjectBomItem,
        Project,
        OrderImportLine,
        ImportRecord,
        OrderImportBatch,
        AiKnowledgeCard,
        AiTask,
        ActivityLog,
        Component,
    ]:
        db.query(model).delete(synchronize_session=False)
    seed_categories(db)
    log_activity(
        db,
        "system.clear_database",
        "system",
        "已清空业务数据库并保留默认分类",
        detail={"confirm_text": payload.get("confirm_text")},
    )
    db.commit()
    return {"cleared": True}


def bom_match_bucket(row: dict) -> str:
    if row.get("status") == "ignored":
        return "ignored"
    if row.get("auto_imported") or row.get("status") in {"pending_purchase", "imported"}:
        return "matched"
    if row.get("selected_component_id"):
        return "matched"
    if row.get("status") == "supplier_missing":
        return "missing"
    if row.get("matches"):
        return "review"
    return "missing"


def recompute_bom_import_batch(db: Session, batch: ProjectBomImportBatch) -> None:
    rows = db.query(ProjectBomImportRow).filter(ProjectBomImportRow.batch_id == batch.id).all()
    candidate_counts = {
        row_id: count
        for row_id, count in (
            db.query(ProjectBomImportCandidate.import_row_id, func.count(ProjectBomImportCandidate.id))
            .filter(ProjectBomImportCandidate.import_row_id.in_([row.id for row in rows] or [0]))
            .group_by(ProjectBomImportCandidate.import_row_id)
            .all()
        )
    }
    buckets = {"matched": 0, "review": 0, "missing": 0}
    for row in rows:
        if row.status == "ignored":
            continue
        if row.auto_imported or row.status in {"pending_purchase", "imported"} or row.selected_component_id:
            buckets["matched"] += 1
        elif row.status == "supplier_missing":
            buckets["missing"] += 1
        elif candidate_counts.get(row.id, 0) > 0:
            buckets["review"] += 1
        else:
            buckets["missing"] += 1
    batch.total_count = sum(buckets.values())
    batch.matched_count = buckets["matched"]
    batch.review_count = buckets["review"]
    batch.missing_count = buckets["missing"]
    batch.auto_imported_count = sum(1 for row in rows if row.auto_imported)
    project = db.get(Project, batch.project_id)
    if project:
        project.bom_match_total = batch.total_count
        project.bom_match_matched = batch.matched_count
        project.bom_match_review = batch.review_count
        project.bom_match_missing = batch.missing_count
        project.bom_match_missing_items = None
        project.bom_match_rows = None
        project.bom_match_updated_at = datetime.utcnow()


def auto_import_exact_lcsc_rows(db: Session, project_id: int | None, rows: list[dict]) -> dict[str, int]:
    if not project_id:
        return {"added": 0, "updated": 0, "skipped": 0}
    if not db.get(Project, project_id):
        return {"added": 0, "updated": 0, "skipped": 0}
    added = updated = skipped = 0
    for row in rows:
        if row.get("status") != "exact_lcsc" or not row.get("selected_component_id"):
            continue
        component_id = int(row["selected_component_id"])
        component = db.get(Component, component_id)
        if not component:
            skipped += 1
            continue
        marker = f"BOM自动导入行:{row.get('source_row')}"
        remark = "；".join(
            [
                marker,
                "编号一致",
                f"BOM 位号: {row.get('designator') or '-'}",
                f"BOM 型号: {row.get('manufacturer_part') or '-'}",
                f"BOM 参数: {row.get('value') or '-'}",
                f"BOM 封装: {row.get('footprint') or '-'}",
            ]
        )
        existing = (
            db.query(ProjectBomItem)
            .filter(ProjectBomItem.project_id == project_id, ProjectBomItem.component_id == component_id)
            .first()
        )
        if existing and marker in (existing.remark or ""):
            sync_bom_solder_points(db, existing)
            row["auto_imported"] = True
            row["auto_import_note"] = "编号一致，已自动导入过"
            skipped += 1
            continue
        if existing:
            old_status = existing.status or "reserved"
            existing.required_quantity += int(row.get("required_quantity") or 1)
            existing.status = "reserved"
            existing.remark = f"{existing.remark or ''}\n{remark}".strip()
            sync_bom_solder_points(db, existing)
            row["auto_imported"] = True
            row["auto_import_note"] = "编号一致，已自动合并到项目 BOM"
            log_activity(
                db,
                "bom.reserve.auto_update",
                "project_bom_item",
                f"编号一致自动合并 {component.name} x {row.get('required_quantity') or 1}",
                entity_id=existing.id,
                component_id=component.id,
                project_id=project_id,
                quantity_delta=int(row.get("required_quantity") or 1) if old_status == "reserved" else existing.required_quantity,
                detail={"source_row": row.get("source_row"), "reason": "编号一致"},
            )
            updated += 1
        else:
            bom_item = ProjectBomItem(
                project_id=project_id,
                component_id=component_id,
                required_quantity=int(row.get("required_quantity") or 1),
                remark=remark,
                status="reserved",
            )
            db.add(bom_item)
            db.flush()
            sync_bom_solder_points(db, bom_item)
            row["auto_imported"] = True
            row["auto_import_note"] = "编号一致，已自动导入项目 BOM"
            log_activity(
                db,
                "bom.reserve.auto",
                "project_bom_item",
                f"编号一致自动导入 {component.name} x {row.get('required_quantity') or 1}",
                entity_id=bom_item.id,
                component_id=component.id,
                project_id=project_id,
                quantity_delta=int(row.get("required_quantity") or 1),
                detail={"source_row": row.get("source_row"), "reason": "编号一致"},
            )
            added += 1
    return {"added": added, "updated": updated, "skipped": skipped}


def save_bom_match_snapshot(
    db: Session,
    project_id: int | None,
    rows: list[dict],
    source_file: str | None = None,
    source_sha256: str | None = None,
    field_mapping: dict | None = None,
) -> None:
    if not project_id:
        return
    project = db.get(Project, project_id)
    if not project:
        return
    batch = ProjectBomImportBatch(
        project_id=project_id,
        source_file=source_file,
        source_sha256=source_sha256,
        field_mapping_json=json.dumps(
            field_mapping or {"mode": "auto_exact_headers"},
            ensure_ascii=False,
        ),
        analysis_json=json.dumps({"matching_policy": "exact_only"}, ensure_ascii=False),
        status="pending",
    )
    db.add(batch)
    db.flush()
    buckets = {"matched": 0, "review": 0, "missing": 0}
    for row in rows:
        bucket = bom_match_bucket(row)
        buckets[bucket] += 1
        suggestion = row.get("missing_suggestion") or {}
        alternatives = suggestion.get("alternatives") or []
        reason_parts = [suggestion.get("reason")]
        reason_parts.extend(item.get("description") for item in alternatives if isinstance(item, dict))
        import_row = ProjectBomImportRow(
            batch_id=batch.id,
            project_id=project_id,
            source_row=row.get("source_row"),
            designator=row.get("designator"),
            required_quantity=int(row.get("required_quantity") or 1),
            comment=row.get("comment"),
            footprint=row.get("footprint"),
            value=row.get("value"),
            manufacturer_part=row.get("manufacturer_part"),
            supplier_part=row.get("supplier_part"),
            status=row.get("status") or "missing",
            selected_component_id=row.get("selected_component_id"),
            match_confidence=int(row.get("match_confidence") or 0),
            role=row.get("role"),
            ai_reason=row.get("ai_reason"),
            ai_confidence=row.get("ai_confidence"),
            ai_error=row.get("ai_error"),
            missing_description=suggestion.get("description"),
            missing_reason="\n".join(part for part in reason_parts if part),
            lcsc_search_keyword=suggestion.get("lcsc_search_keyword"),
            lcsc_search_url=suggestion.get("lcsc_search_url"),
            auto_imported=bool(row.get("auto_imported")),
            auto_import_note=row.get("auto_import_note"),
        )
        db.add(import_row)
        db.flush()
        for rank, match in enumerate(row.get("matches") or []):
            component = match.get("component") or {}
            if not component.get("id"):
                continue
            db.add(
                ProjectBomImportCandidate(
                    import_row_id=import_row.id,
                    component_id=int(component["id"]),
                    score=int(match.get("score") or 0),
                    match_type=match.get("match_type"),
                    reason=match.get("reason"),
                    flags=",".join(match.get("flags") or []),
                    available_quantity=int(match.get("available_quantity") or 0),
                    shortage_quantity=int(match.get("shortage_quantity") or 0),
                    enough=bool(match.get("enough")),
                    rank=rank,
                )
            )
    batch.total_count = len(rows)
    batch.matched_count = buckets["matched"]
    batch.review_count = buckets["review"]
    batch.missing_count = buckets["missing"]
    batch.auto_imported_count = sum(1 for row in rows if row.get("auto_imported"))
    project.bom_match_total = len(rows)
    project.bom_match_matched = buckets["matched"]
    project.bom_match_review = buckets["review"]
    project.bom_match_missing = buckets["missing"]
    project.bom_match_missing_items = None
    project.bom_match_rows = None
    project.bom_match_updated_at = datetime.utcnow()
    db.commit()


def bom_import_row_out(db: Session, row: ProjectBomImportRow) -> dict:
    candidates = (
        db.query(ProjectBomImportCandidate)
        .filter(ProjectBomImportCandidate.import_row_id == row.id)
        .order_by(ProjectBomImportCandidate.rank.asc(), ProjectBomImportCandidate.score.desc())
        .all()
    )
    matches = []
    selected_seen = False
    for candidate in candidates:
        component = db.get(Component, candidate.component_id)
        if not component:
            continue
        if row.selected_component_id and component.id == row.selected_component_id:
            selected_seen = True
        matches.append(
            {
                "component": component_to_dict(component),
                "score": candidate.score,
                "match_type": candidate.match_type,
                "reason": candidate.reason,
                "flags": [flag for flag in (candidate.flags or "").split(",") if flag],
                "available_quantity": candidate.available_quantity,
                "shortage_quantity": candidate.shortage_quantity,
                "enough": candidate.enough,
            }
        )
    selected_component = db.get(Component, row.selected_component_id) if row.selected_component_id else None
    if selected_component and not selected_seen:
        matches.insert(
            0,
            {
                "component": component_to_dict(selected_component),
                "score": row.match_confidence or 100,
                "match_type": row.status or "selected",
                "reason": "已选中：待采购库/手动匹配",
                "flags": ["待采购库" if selected_component.status == "pending_purchase" or row.status == "pending_purchase" else "手动匹配"],
                "available_quantity": int(selected_component.quantity or 0),
                "shortage_quantity": max(0, int(row.required_quantity or 1) - int(selected_component.quantity or 0)),
                "enough": int(selected_component.quantity or 0) >= int(row.required_quantity or 1),
            },
        )
    alternatives = []
    reason_lines = [line.strip() for line in (row.missing_reason or "").splitlines() if line.strip()]
    if len(reason_lines) > 1:
        alternatives = [{"description": line, "reason": "组合替代建议"} for line in reason_lines[1:]]
    return {
        "id": row.id,
        "source_row": row.source_row,
        "designator": row.designator,
        "required_quantity": row.required_quantity,
        "comment": row.comment,
        "footprint": row.footprint,
        "value": row.value,
        "manufacturer_part": row.manufacturer_part,
        "supplier_part": row.supplier_part,
        "status": row.status,
        "selected_component_id": row.selected_component_id,
        "selected_component": component_to_dict(selected_component) if selected_component else None,
        "match_confidence": row.match_confidence,
        "matches": matches,
        "role": row.role,
        "ai_reason": row.ai_reason,
        "ai_confidence": row.ai_confidence,
        "ai_error": row.ai_error,
        "auto_imported": row.auto_imported,
        "auto_import_note": row.auto_import_note,
        "missing_suggestion": {
            "description": row.missing_description,
            "reason": reason_lines[0] if reason_lines else None,
            "lcsc_search_keyword": row.lcsc_search_keyword,
            "lcsc_search_url": row.lcsc_search_url,
            "alternatives": alternatives,
        }
        if row.missing_description or row.lcsc_search_keyword or alternatives
        else None,
        "lcsc_search_url": row.lcsc_search_url,
    }


def bom_item_designator_set(item: ProjectBomItem) -> set[str]:
    points = getattr(item, "solder_points", []) or []
    values = [point.designator for point in points]
    if not values:
        specs = solder_point_specs_from_remark(item.remark)
        values = [spec.get("designator") for spec in specs]
    return {str(value).strip().upper() for value in values if str(value or "").strip()}


def reconcile_bom_import_batch_with_project(db: Session, batch: ProjectBomImportBatch) -> bool:
    rows = db.query(ProjectBomImportRow).filter(ProjectBomImportRow.batch_id == batch.id).all()
    bom_items = (
        db.query(ProjectBomItem)
        .options(joinedload(ProjectBomItem.solder_points), joinedload(ProjectBomItem.component))
        .filter(ProjectBomItem.project_id == batch.project_id)
        .all()
    )
    indexed_items = [(item, bom_item_designator_set(item)) for item in bom_items]
    changed = False
    for row in rows:
        if row.status == "ignored" or row.auto_imported:
            continue
        row_designators = {value.upper() for value in split_bom_designators(row.designator)}
        if not row_designators:
            continue
        matched_item = next((item for item, designators in indexed_items if row_designators & designators), None)
        if not matched_item:
            continue
        row.status = "imported"
        row.selected_component_id = matched_item.component_id
        row.match_confidence = max(int(row.match_confidence or 0), 100)
        row.auto_imported = True
        row.auto_import_note = f"已在正式 BOM 中：{matched_item.id}"
        row.ai_reason = "该 BOM 位号已进入正式项目 BOM，不再计入待采购。"
        row.updated_at = datetime.utcnow()
        changed = True
    if changed:
        recompute_bom_import_batch(db, batch)
    return changed


def latest_bom_import_batch_out(db: Session, project_id: int) -> dict | None:
    batch = (
        db.query(ProjectBomImportBatch)
        .filter(ProjectBomImportBatch.project_id == project_id)
        .order_by(ProjectBomImportBatch.created_at.desc(), ProjectBomImportBatch.id.desc())
        .first()
    )
    if not batch:
        return None
    if reconcile_bom_import_batch_with_project(db, batch):
        db.commit()
        db.refresh(batch)
    rows = (
        db.query(ProjectBomImportRow)
        .filter(ProjectBomImportRow.batch_id == batch.id)
        .order_by(ProjectBomImportRow.source_row.asc(), ProjectBomImportRow.id.asc())
        .all()
    )
    return {
        "id": batch.id,
        "project_id": batch.project_id,
        "status": batch.status,
        "total_count": batch.total_count,
        "matched_count": batch.matched_count,
        "review_count": batch.review_count,
        "missing_count": batch.missing_count,
        "auto_imported_count": batch.auto_imported_count,
        "created_at": batch.created_at,
        "updated_at": batch.updated_at,
        "rows": [bom_import_row_out(db, row) for row in rows],
    }


@app.get("/api/projects/{project_id}/bom/import-batch/latest")
def latest_bom_import_batch(project_id: int, _: Protected, db: Session = Depends(get_db)):
    result = latest_bom_import_batch_out(db, project_id)
    if not result:
        return {"rows": [], "total_count": 0, "matched_count": 0, "review_count": 0, "missing_count": 0, "auto_imported_count": 0}
    return result


@app.post("/api/ai/bom-match/preview")
async def preview_bom_match(
    auth: Protected,
    file: UploadFile = File(...),
    project_id: int | None = Form(None),
    field_mapping_json: str | None = Form(None),
    db: Session = Depends(get_db),
):
    content = await file.read()
    try:
        field_mapping = json.loads(field_mapping_json) if field_mapping_json else None
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="BOM 字段映射格式无效") from exc
    rows = parse_bom_excel(content, file.filename, field_mapping)
    component_ids = [
        component_id
        for (component_id,) in filter_owner(
            db.query(Component.id),
            Component,
            auth,
        ).all()
    ]
    matched = match_bom_rows(
        db,
        rows,
        component_ids=component_ids,
        supplier_scope_type="personal",
        supplier_owner_user_id=auth.user_id,
    )
    for row in matched:
        row["ai_checked"] = False
        row["ai_error"] = None
    low_rows = [row for row in matched if row.get("status") in {"missing", "review", "supplier_missing"}][:10]
    if low_rows:
        inventory = db.query(Component).options(joinedload(Component.category)).order_by(Component.quantity.desc(), Component.updated_at.desc()).limit(120).all()
        try:
            ai_result = await run_foreground_ai(assist_bom_matches, low_rows, inventory)
            ai_rows = {int(row.get("source_row")): row for row in ai_result.get("rows", []) if row.get("source_row") is not None}
            for row in matched:
                ai_row = ai_rows.get(int(row.get("source_row") or 0))
                if not ai_row:
                    continue
                row["ai_checked"] = True
                row["role"] = ai_row.get("role") or row.get("role")
                row["ai_reason"] = ai_row.get("reason") or row.get("ai_reason")
                row["ai_confidence"] = ai_row.get("confidence")
                if ai_row.get("selected_component_id") and not row.get("selected_component_id"):
                    row["ai_reason"] = (
                        f"{row.get('ai_reason') or ''} AI 推荐了候选元件，但按精确匹配策略必须人工确认。"
                    ).strip()
                if ai_row.get("missing_suggestion"):
                    row["missing_suggestion"] = ai_row["missing_suggestion"]
        except Exception as exc:
            message = f"AI 辅助失败，已使用库存预匹配结果：{exc}"
            for row in low_rows:
                row["ai_checked"] = False
                row["ai_error"] = message
    auto_result = auto_import_exact_lcsc_rows(db, project_id, matched)
    if any(auto_result.values()):
        for row in matched:
            row["auto_import_result"] = auto_result
    save_bom_match_snapshot(
        db,
        project_id,
        matched,
        file.filename,
        hashlib.sha256(content).hexdigest(),
        field_mapping,
    )
    return matched


@app.post("/api/ai/bom-match/inspect")
async def inspect_bom_upload(
    auth: Protected,
    file: UploadFile = File(...),
):
    content = await file.read()
    if not content or len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="BOM 文件为空或超过 20MB")
    try:
        return inspect_bom_fields(content, file.filename)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"BOM 字段读取失败：{exc}") from exc


@app.get("/api/activity-logs", response_model=list[ActivityLogOut])
def list_activity_logs(
    auth: Protected,
    db: Session = Depends(get_db),
    limit: int = Query(80, ge=1, le=300),
    component_id: int | None = None,
    project_id: int | None = None,
    action: str | None = None,
):
    query = filter_owner(db.query(ActivityLog), ActivityLog, auth)
    if component_id:
        query = query.filter(ActivityLog.component_id == component_id)
    if project_id:
        query = query.filter(ActivityLog.project_id == project_id)
    if action:
        query = query.filter(ActivityLog.action == action)
    return query.order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc()).limit(limit).all()


@app.post("/api/usage-events")
def record_usage_event(payload: UsageEventRequest, request: Request, auth: Protected, db: Session = Depends(get_db)):
    event = re.sub(r"[^a-zA-Z0-9_.:-]+", ".", payload.event).strip(".")[:70]
    if not event.startswith("ui."):
        event = f"ui.{event}"[:80]
    detail = {
        "page": payload.page,
        "entry": payload.entry,
        "target_type": payload.target_type,
        "target_id": payload.target_id,
        "viewport_width": payload.viewport_width,
        "viewport_height": payload.viewport_height,
        "client_ip": client_ip(request),
        "detail": payload.detail or {},
    }
    entity_id = None
    if isinstance(payload.target_id, int):
        entity_id = payload.target_id
    elif isinstance(payload.target_id, str) and payload.target_id.isdigit():
        entity_id = int(payload.target_id)
    log_activity(
        db,
        event,
        payload.target_type or "ui",
        f"界面操作：{event}",
        owner_user_id=auth.user_id,
        entity_id=entity_id,
        detail=detail,
    )
    db.commit()
    return {"ok": True}


def ui_action_label(action: str | None) -> str:
    mapping = {
        "ui.page.view": "页面访问",
        "ui.nav.click": "导航点击",
        "ui.back_to_top.click": "返回顶部",
        "ui.components.label_export": "导出标签",
        "ui.components.auto_load": "自动加载",
        "ui.components.create_open": "新增元器件",
        "ui.components.detail_open": "打开元器件详情",
        "ui.components.ai_quick_create": "AI 补全元器件",
        "ui.components.lot_create": "新增库存批次",
        "ui.components.lot_consume": "批次扣减",
        "ui.components.ai_ask": "元器件 AI 问答",
        "ui.components.remove": "移除元器件记录",
        "ui.team_components.auto_load": "团队元器件自动加载",
        "ui.team_components.ai_quick_create": "团队 AI 补全元器件",
        "ui.team_components.detail_open": "团队元器件详情",
        "ui.team_components.lot_create": "团队新增库存批次",
        "ui.team_components.lot_consume": "团队批次扣减",
        "ui.team_components.ai_ask": "团队元器件 AI 问答",
    }
    return mapping.get(action or "", action or "未知操作")


def _activity_detail_value(detail: str | None, key: str) -> str | None:
    if not detail:
        return None
    try:
        data = json.loads(detail)
    except Exception:
        return None
    value = data.get(key)
    if value is None and isinstance(data.get("detail"), dict):
        value = data["detail"].get(key)
    return str(value) if value is not None else None


@app.get("/api/admin/usage-dashboard")
def admin_usage_dashboard(auth: AdminProtected, db: Session = Depends(get_db)):
    now = datetime.utcnow()
    today_start = datetime(now.year, now.month, now.day)
    month_start = now - timedelta(days=30)
    week_start = now - timedelta(days=7)
    day_rows = []
    for offset in range(29, -1, -1):
        day = today_start - timedelta(days=offset)
        day_rows.append({"date": day.date().isoformat(), "users": 0, "events": 0})
    day_index = {row["date"]: row for row in day_rows}

    personal_daily = (
        db.query(func.date(ActivityLog.created_at), func.count(ActivityLog.id), func.count(func.distinct(ActivityLog.owner_user_id)))
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= month_start)
        .group_by(func.date(ActivityLog.created_at))
        .all()
    )
    for day, event_count, user_count in personal_daily:
        key = str(day)
        if key in day_index:
            day_index[key]["events"] += int(event_count or 0)
            day_index[key]["users"] += int(user_count or 0)

    team_daily = (
        db.query(func.date(CompetitionActivityLog.created_at), func.count(CompetitionActivityLog.id), func.count(func.distinct(CompetitionActivityLog.actor_user_id)))
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= month_start)
        .group_by(func.date(CompetitionActivityLog.created_at))
        .all()
    )
    for day, event_count, user_count in team_daily:
        key = str(day)
        if key in day_index:
            day_index[key]["events"] += int(event_count or 0)
            day_index[key]["users"] += int(user_count or 0)

    personal_month_users = {
        row[0]
        for row in db.query(ActivityLog.owner_user_id)
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= month_start, ActivityLog.owner_user_id.isnot(None))
        .distinct()
        .all()
    }
    team_month_users = {
        row[0]
        for row in db.query(CompetitionActivityLog.actor_user_id)
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= month_start)
        .distinct()
        .all()
    }
    week_users = {
        row[0]
        for row in db.query(ActivityLog.owner_user_id)
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= week_start, ActivityLog.owner_user_id.isnot(None))
        .distinct()
        .all()
    } | {
        row[0]
        for row in db.query(CompetitionActivityLog.actor_user_id)
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= week_start)
        .distinct()
        .all()
    }
    today_users = {
        row[0]
        for row in db.query(ActivityLog.owner_user_id)
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= today_start, ActivityLog.owner_user_id.isnot(None))
        .distinct()
        .all()
    } | {
        row[0]
        for row in db.query(CompetitionActivityLog.actor_user_id)
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= today_start)
        .distinct()
        .all()
    }

    top_counts: dict[str, int] = {}
    for action, count in (
        db.query(ActivityLog.action, func.count(ActivityLog.id))
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= month_start)
        .group_by(ActivityLog.action)
        .all()
    ):
        top_counts[action] = top_counts.get(action, 0) + int(count or 0)
    for action, count in (
        db.query(CompetitionActivityLog.action, func.count(CompetitionActivityLog.id))
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= month_start)
        .group_by(CompetitionActivityLog.action)
        .all()
    ):
        top_counts[action] = top_counts.get(action, 0) + int(count or 0)
    top_features = [
        {"action": action, "label": ui_action_label(action), "count": count}
        for action, count in sorted(top_counts.items(), key=lambda item: item[1], reverse=True)[:12]
    ]

    last_personal = {
        row[0]: row[1]
        for row in db.query(ActivityLog.owner_user_id, func.max(ActivityLog.created_at))
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.owner_user_id.isnot(None))
        .group_by(ActivityLog.owner_user_id)
        .all()
    }
    last_team = {
        row[0]: row[1]
        for row in db.query(CompetitionActivityLog.actor_user_id, func.max(CompetitionActivityLog.created_at))
        .filter(CompetitionActivityLog.action.like("ui.%"))
        .group_by(CompetitionActivityLog.actor_user_id)
        .all()
    }
    event_counts = {
        row[0]: int(row[1] or 0)
        for row in db.query(ActivityLog.owner_user_id, func.count(ActivityLog.id))
        .filter(ActivityLog.action.like("ui.%"), ActivityLog.created_at >= month_start, ActivityLog.owner_user_id.isnot(None))
        .group_by(ActivityLog.owner_user_id)
        .all()
    }
    for user_id, count in (
        db.query(CompetitionActivityLog.actor_user_id, func.count(CompetitionActivityLog.id))
        .filter(CompetitionActivityLog.action.like("ui.%"), CompetitionActivityLog.created_at >= month_start)
        .group_by(CompetitionActivityLog.actor_user_id)
        .all()
    ):
        event_counts[user_id] = event_counts.get(user_id, 0) + int(count or 0)
    last_by_user = {}
    for user_id, ts in {**last_personal, **last_team}.items():
        last_by_user[user_id] = max([value for value in [last_personal.get(user_id), last_team.get(user_id)] if value])
    users = db.query(User).filter(User.id.in_(list(last_by_user.keys()) or [-1])).all()
    users_by_id = {user.id: user for user in users}
    recent_users = []
    for user_id, last_seen in sorted(last_by_user.items(), key=lambda item: item[1] or datetime.min, reverse=True)[:16]:
        user = users_by_id.get(user_id)
        recent_users.append(
            {
                "user_id": user_id,
                "nickname": user.nickname if user else f"用户 {user_id}",
                "phone": user.phone if user else "",
                "last_seen_at": last_seen,
                "last_login_at": user.last_login_at if user else None,
                "event_count_30d": event_counts.get(user_id, 0),
            }
        )

    recent_personal = (
        db.query(ActivityLog)
        .filter(ActivityLog.action.like("ui.%"))
        .order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc())
        .limit(40)
        .all()
    )
    recent_team = (
        db.query(CompetitionActivityLog)
        .filter(CompetitionActivityLog.action.like("ui.%"))
        .order_by(CompetitionActivityLog.created_at.desc(), CompetitionActivityLog.id.desc())
        .limit(40)
        .all()
    )
    recent_events = []
    user_cache: dict[int, User] = {}
    for row in recent_personal:
        user = None
        if row.owner_user_id:
            user = user_cache.get(row.owner_user_id)
            if user is None:
                user = db.get(User, row.owner_user_id)
                if user:
                    user_cache[row.owner_user_id] = user
        recent_events.append(
            {
                "id": f"p-{row.id}",
                "scope": "个人",
                "action": row.action,
                "label": ui_action_label(row.action),
                "page": _activity_detail_value(row.detail, "page"),
                "entry": _activity_detail_value(row.detail, "entry"),
                "user": user.nickname if user else f"用户 {row.owner_user_id or '-'}",
                "phone": user.phone if user else "",
                "created_at": row.created_at,
            }
        )
    for row in recent_team:
        user = user_cache.get(row.actor_user_id)
        if user is None:
            user = db.get(User, row.actor_user_id)
            if user:
                user_cache[row.actor_user_id] = user
        recent_events.append(
            {
                "id": f"t-{row.id}",
                "scope": "团队",
                "action": row.action,
                "label": ui_action_label(row.action),
                "page": _activity_detail_value(row.after_json, "page"),
                "entry": _activity_detail_value(row.after_json, "entry"),
                "user": user.nickname if user else row.actor_nickname or f"用户 {row.actor_user_id}",
                "phone": user.phone if user else "",
                "created_at": row.created_at,
            }
        )
    recent_events = sorted(recent_events, key=lambda item: item["created_at"] or datetime.min, reverse=True)[:50]

    return {
        "generated_at": now,
        "registered_users": int(db.query(func.count(User.id)).scalar() or 0),
        "monthly_active_users": len(personal_month_users | team_month_users),
        "weekly_active_users": len(week_users),
        "today_active_users": len(today_users),
        "ui_events_30d": int(sum(row["events"] for row in day_rows)),
        "daily": day_rows,
        "top_features": top_features,
        "recent_users": recent_users,
        "recent_events": recent_events,
    }


@app.get("/api/ai/tasks/summary", response_model=AiTaskSummary)
def ai_tasks_summary(_: AdminProtected, db: Session = Depends(get_db)):
    if resolve_superseded_ai_failures(db):
        db.commit()
    counts = {status: 0 for status in ["pending", "processing", "completed", "failed", "stale"]}
    for status, count in db.query(AiTask.status, func.count(AiTask.id)).group_by(AiTask.status).all():
        if status in counts:
            counts[status] = count
    missing = db.query(func.count(Component.id)).filter(Component.ai_status.in_(["pending", "failed", "stale"])).scalar() or 0
    current_task = db.query(AiTask).filter(AiTask.status == "processing").order_by(AiTask.started_at.desc()).first()
    current_component = None
    if current_task and current_task.target_type == "component":
        component = db.get(Component, current_task.target_id)
        current_component = component.name if component else None
    last_finished_at = db.query(func.max(AiTask.finished_at)).scalar()
    return {
        **counts,
        "missing_components": missing,
        "current_component": current_component,
        "last_finished_at": last_finished_at,
        "running": AI_WORKER_RUNNING,
        "paused": AI_WORKER_PAUSED,
    }


@app.post("/api/ai/tasks/enqueue-missing", response_model=AiTaskSummary)
def enqueue_missing_ai_tasks(_: AdminProtected, db: Session = Depends(get_db)):
    enqueue_missing_component_ai_tasks(db, include_failed=True)
    db.commit()
    return ai_tasks_summary(_, db)


@app.post("/api/ai/tasks/enqueue-organize", response_model=AiTaskSummary)
def enqueue_organize_ai_tasks(_: AdminProtected, db: Session = Depends(get_db), force: bool = False):
    enqueue_organize_component_tasks(db, force=force)
    db.commit()
    return ai_tasks_summary(_, db)


@app.post("/api/ai/reset-and-reorganize")
def reset_and_reorganize(_: AdminProtected, db: Session = Depends(get_db)):
    db.query(Component).update(
        {
            Component.ai_summary: None,
            Component.ai_usage: None,
            Component.ai_risk_notes: None,
            Component.ai_pcb_notes: None,
            Component.ai_substitutes: None,
            Component.ai_tags: None,
            Component.ai_confidence: None,
            Component.ai_cache_key: None,
            Component.ai_status: "pending",
            Component.ai_error: None,
            Component.ai_updated_at: None,
            Component.tags: None,
        },
        synchronize_session=False,
    )
    db.query(AiKnowledgeCard).filter(AiKnowledgeCard.source_type == "ai").delete(synchronize_session=False)
    db.query(AiTask).delete(synchronize_session=False)
    db.commit()
    organize_count = enqueue_organize_component_tasks(db, force=True)
    analyze_count = enqueue_missing_component_ai_tasks(db, include_failed=True)
    db.commit()
    return {"reset": True, "organize_queued": organize_count, "analyze_queued": analyze_count}


@app.post("/api/ai/tasks/start", response_model=AiTaskSummary)
def start_ai_tasks(_: AdminProtected, db: Session = Depends(get_db)):
    global AI_WORKER_PAUSED
    ensure_ai_worker()
    AI_WORKER_PAUSED = False
    return ai_tasks_summary(_, db)


@app.post("/api/ai/tasks/pause", response_model=AiTaskSummary)
def pause_ai_tasks(_: AdminProtected, db: Session = Depends(get_db)):
    global AI_WORKER_PAUSED
    AI_WORKER_PAUSED = True
    return ai_tasks_summary(_, db)


@app.get("/api/ai/tasks", response_model=list[AiTaskOut])
def list_ai_tasks(_: AdminProtected, db: Session = Depends(get_db), limit: int = Query(80, ge=1, le=300)):
    return db.query(AiTask).order_by(AiTask.created_at.desc(), AiTask.id.desc()).limit(limit).all()


@app.get("/api/integrations/components", response_model=ComponentList)
def integration_components(auth: Protected, db: Session = Depends(get_db), keyword: str | None = None, limit: int = Query(50, ge=1, le=200)):
    query = filter_owner(db.query(Component), Component, auth).options(joinedload(Component.category))
    if keyword:
        filters = component_keyword_filters(keyword)
        if filters:
            query = query.filter(or_(*filters))
    items = query.order_by(Component.updated_at.desc(), Component.id.desc()).limit(limit).all()
    reserved = reserved_quantities(db, [item.id for item in items])
    return {"items": [component_out(item, reserved.get(item.id, 0)) for item in items], "total": len(items)}


@app.get("/api/integrations/projects", response_model=list[ProjectOut])
def integration_projects(auth: Protected, db: Session = Depends(get_db), limit: int = Query(50, ge=1, le=200)):
    projects = (
        filter_owner(db.query(Project), Project, auth)
        .options(
            joinedload(Project.boards),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.component).joinedload(Component.category),
            joinedload(Project.bom_items).joinedload(ProjectBomItem.solder_points),
        )
        .order_by(Project.updated_at.desc(), Project.id.desc())
        .limit(limit)
        .all()
    )
    reserved = reserved_quantities(db, list({item.component_id for project in projects for item in project.bom_items}))
    return [project_out(project, reserved) for project in projects]


@app.post("/api/integrations/components/{component_id}/consume", response_model=ComponentOut)
def integration_consume_component(
    component_id: int,
    payload: ComponentConsumeRequest,
    auth: Protected,
    db: Session = Depends(get_db),
):
    return decrement_component_quantity(component_id, auth, db, payload)


@app.get("/api/integrations/activity-logs", response_model=list[ActivityLogOut])
def integration_activity_logs(auth: Protected, db: Session = Depends(get_db), limit: int = Query(80, ge=1, le=300)):
    return filter_owner(db.query(ActivityLog), ActivityLog, auth).order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc()).limit(limit).all()


def enqueue_missing_component_ai_tasks(db: Session, include_failed: bool = False) -> int:
    count = 0
    components = db.query(Component).options(joinedload(Component.category)).all()
    for component in components:
        mark_component_ai_stale(component)
        status = component.ai_status or "pending"
        cache_key = component_ai_cache_key(component)
        needs_initial_run = not component.ai_summary and status != "failed"
        needs_refresh = status in {"pending", "stale"}
        needs_manual_retry = include_failed and status == "failed"
        if needs_initial_run or needs_refresh or needs_manual_retry:
            enqueue_ai_task(db, "component_analyze", "component", component.id, cache_key)
            if status == "failed" and include_failed:
                component.ai_status = "pending"
            count += 1
    return count


def enqueue_auto_refresh_component_ai_tasks(db: Session) -> int:
    global AI_LAST_AUTO_REFRESH_AT
    if not AI_AUTO_REFRESH_ENABLED:
        return 0
    now = datetime.now()
    if AI_LAST_AUTO_REFRESH_AT and now - AI_LAST_AUTO_REFRESH_AT < timedelta(hours=AI_AUTO_REFRESH_INTERVAL_HOURS):
        return 0
    cutoff = now - timedelta(days=AI_AUTO_REFRESH_AFTER_DAYS)
    components = (
        db.query(Component)
        .filter(
            Component.ai_status == "completed",
            Component.ai_summary.isnot(None),
            or_(Component.ai_updated_at == None, Component.ai_updated_at < cutoff),
        )
        .order_by(Component.ai_updated_at.asc().nullsfirst(), Component.id.asc())
        .limit(AI_AUTO_REFRESH_MAX_PER_RUN)
        .all()
    )
    for component in components:
        component.ai_status = "stale"
        enqueue_ai_task(db, "component_analyze", "component", component.id, component_ai_cache_key(component))
    AI_LAST_AUTO_REFRESH_AT = now
    return len(components)


def enqueue_organize_component_tasks(db: Session, force: bool = False, limit: int | None = None) -> int:
    count = 0
    query = db.query(Component).options(joinedload(Component.category)).order_by(Component.id.asc())
    if limit:
        query = query.limit(limit)
    for component in query.all():
        cache_key = organize_cache_key(component)
        if not force:
            completed = (
                db.query(AiTask)
                .filter(
                    AiTask.task_type == "component_organize",
                    AiTask.target_type == "component",
                    AiTask.target_id == component.id,
                    AiTask.input_hash == cache_key,
                    AiTask.status == "completed",
                )
                .first()
            )
            if completed or not looks_like_needs_organize(component):
                continue
        enqueue_ai_task(db, "component_organize", "component", component.id, cache_key)
        count += 1
    return count


def analyze_component_with_ai(db: Session, component: Component, scope: str = "full", force: bool = False) -> dict:
    cache_key = component_ai_cache_key(component)
    if not force and scope == "full" and component.ai_status == "completed" and component.ai_cache_key == cache_key:
        return {"cached": True, "summary": component.ai_summary, "cache_key": cache_key}
    known_specs = json.dumps(component_ai_payload(component), ensure_ascii=False)
    result = component_info(component.model or component.name, known_specs, "auto")
    result["generated_at"] = datetime.now().isoformat()
    result["cache_key"] = cache_key
    result["scope"] = scope
    before = component_ai_undo_snapshot(component)
    apply_component_ai_result(db, component, result, cache_key)
    after = component_ai_undo_snapshot(component)
    log_activity(
        db,
        "ai.component.analyze",
        "component",
        f"AI 整理元器件 {component.name}",
        entity_id=component.id,
        component_id=component.id,
        detail={
            "scope": scope,
            "confidence": result.get("confidence"),
            "before": before,
            "after": after,
        },
    )
    return result


@app.post("/api/components/{component_id}/ai/undo-latest")
def undo_latest_component_ai_change(
    component_id: int,
    auth: Protected,
    db: Session = Depends(get_db),
):
    component = db.get(Component, component_id)
    assert_owned(component, auth, "Component not found")
    logs = (
        filter_owner(db.query(ActivityLog), ActivityLog, auth)
        .filter(
            ActivityLog.component_id == component_id,
            ActivityLog.action.in_(["ai.component.analyze", "ai.component.organize"]),
        )
        .order_by(ActivityLog.id.desc())
        .all()
    )
    target = None
    target_detail = None
    for row in logs:
        detail = parse_json_value(row.detail)
        if isinstance(detail, dict) and detail.get("before") and not detail.get("undone_at"):
            target = row
            target_detail = detail
            break
    if not target or not target_detail:
        raise HTTPException(status_code=404, detail="没有可撤销的 AI 修改")
    before = target_detail["before"]
    for field, value in before.items():
        if field in AI_UNDO_FIELDS and hasattr(component, field):
            setattr(component, field, value)
    target_detail["undone_at"] = datetime.utcnow().isoformat() + "Z"
    target_detail["undone_by_user_id"] = auth.user_id
    target.detail = json.dumps(target_detail, ensure_ascii=False, default=str)
    component.ai_status = "stale"
    log_activity(
        db,
        "ai.component.undo",
        "component",
        f"撤销 AI 修改 {component.name}",
        entity_id=component.id,
        component_id=component.id,
        detail={"source_log_id": target.id, "restored": before},
    )
    db.commit()
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return component_out(component, reserved)


def organize_component_record(db: Session, component: Component, force: bool = False) -> dict:
    cache_key = organize_cache_key(component)
    completed = (
        db.query(AiTask)
        .filter(
            AiTask.task_type == "component_organize",
            AiTask.target_type == "component",
            AiTask.target_id == component.id,
            AiTask.input_hash == cache_key,
            AiTask.status == "completed",
        )
        .first()
    )
    if completed and not force:
        return {"cached": True, "cache_key": cache_key}

    current_fields = {
        "current_name": component.name,
        "current_category": component.category.name if component.category else None,
        "current_part_family": component.part_family,
        "current_count_mode": component.count_mode,
        "current_normalized_spec": component.normalized_spec,
        "current_tags": component.tags,
    }
    categories = [name for (name,) in db.query(Category.name).order_by(Category.id).all()]
    result = organize_component(component_organize_payload(component), categories, current_fields)
    source = "ai"
    apply_result = apply_component_organize_result(db, component, result, source=source)
    return {"cache_key": cache_key, **apply_result}


def process_ai_task(task_id: int) -> None:
    db = SessionLocal()
    try:
        task = db.get(AiTask, task_id)
        if not task:
            return
        task.status = "processing"
        task.started_at = datetime.now()
        db.commit()
        if task.task_type == "component_analyze" and task.target_type == "component":
            component = db.get(Component, task.target_id)
            if not component:
                raise RuntimeError("Component not found")
            component.ai_status = "processing"
            component.ai_error = None
            db.commit()
            result = analyze_component_with_ai(db, component, "full", False)
        elif task.task_type == "component_organize" and task.target_type == "component":
            component = db.query(Component).options(joinedload(Component.category)).filter(Component.id == task.target_id).first()
            if not component:
                raise RuntimeError("Component not found")
            result = organize_component_record(db, component, False)
        else:
            raise RuntimeError(f"Unsupported AI task: {task.task_type}")
        task.result_json = json.dumps(result, ensure_ascii=False)
        task.status = "completed"
        task.finished_at = datetime.now()
        task.error_message = None
        task.next_attempt_at = None
        db.query(AiTask).filter(
            AiTask.task_type == task.task_type,
            AiTask.target_type == task.target_type,
            AiTask.target_id == task.target_id,
            AiTask.status == "failed",
            AiTask.id != task.id,
        ).update({AiTask.status: "superseded", AiTask.next_attempt_at: None}, synchronize_session=False)
        db.commit()
    except Exception as error:
        db.rollback()
        task = db.get(AiTask, task_id)
        if task:
            task.retry_count = (task.retry_count or 0) + 1
            task.error_message = str(error)
            if task.retry_count < AI_TASK_MAX_RETRIES:
                task.status = "pending"
                task.next_attempt_at = datetime.now() + timedelta(seconds=retry_delay_for(task))
            else:
                task.status = "failed"
                task.next_attempt_at = None
            task.finished_at = datetime.now()
            if task.target_type == "component":
                component = db.get(Component, task.target_id)
                if component:
                    component.ai_status = "pending" if task.status == "pending" else "failed"
                    component.ai_error = str(error)
            db.commit()
    finally:
        db.close()


AI_WORKER_RUNNING = False
AI_WORKER_PAUSED = False
AI_WORKER_THREAD: threading.Thread | None = None
AI_WORKER_CONCURRENCY = max(1, min(3, int(os.getenv("AI_WORKER_CONCURRENCY", "1"))))


def ai_worker_loop() -> None:
    global AI_WORKER_RUNNING
    while AI_WORKER_RUNNING:
        if AI_WORKER_PAUSED:
            time.sleep(1)
            continue
        db = SessionLocal()
        try:
            tasks = (
                db.query(AiTask.id)
                .filter(
                    AiTask.status.in_(["pending", "stale"]),
                    or_(AiTask.next_attempt_at.is_(None), AiTask.next_attempt_at <= datetime.now()),
                )
                .order_by(
                    (AiTask.task_type == "component_organize").desc(),
                    AiTask.next_attempt_at.asc().nullsfirst(),
                    AiTask.created_at.asc(),
                    AiTask.id.asc(),
                )
                .limit(AI_WORKER_CONCURRENCY * 2)
                .all()
            )
            task_ids = [t.id for t in tasks]
        finally:
            db.close()
        if task_ids:
            with ThreadPoolExecutor(max_workers=AI_WORKER_CONCURRENCY) as pool:
                futures = {pool.submit(process_ai_task, tid): tid for tid in task_ids}
                for future in as_completed(futures):
                    try:
                        future.result()
                    except Exception:
                        pass
        else:
            db = SessionLocal()
            try:
                enqueue_missing_component_ai_tasks(db, include_failed=True)
                enqueue_auto_refresh_component_ai_tasks(db)
                db.commit()
            finally:
                db.close()
            time.sleep(2)


def ensure_ai_worker() -> None:
    global AI_WORKER_RUNNING, AI_WORKER_THREAD
    if AI_WORKER_THREAD and AI_WORKER_THREAD.is_alive():
        return
    AI_WORKER_RUNNING = True
    AI_WORKER_THREAD = threading.Thread(target=ai_worker_loop, daemon=True)
    AI_WORKER_THREAD.start()


def handle_mimo_error(error: Exception):
    if isinstance(error, MimoNotConfiguredError):
        raise HTTPException(status_code=503, detail="MIMO_API_KEY is not configured") from error
    if isinstance(error, MimoRequestError):
        raise HTTPException(status_code=502, detail=str(error)) from error
    raise error


@app.post("/api/ai/classify")
def ai_classify(payload: AiClassifyRequest, _: Protected, db: Session = Depends(get_db)):
    categories = [name for (name,) in db.query(Category.name).order_by(Category.id).all()]
    try:
        result = classify_component(payload.model_dump(), categories)
    except Exception as error:
        handle_mimo_error(error)
    result["requires_confirmation"] = True
    return result


@app.post("/api/ai/explain")
def ai_explain(payload: AiExplainRequest, _: Protected):
    try:
        return explain_component(payload.model_dump())
    except Exception as error:
        handle_mimo_error(error)


@app.post("/api/ai/project-plan")
def ai_project_plan(payload: AiProjectPlanRequest, _: Protected, db: Session = Depends(get_db)):
    candidates = search_component_candidates(db, payload.goal, 20)
    try:
        return project_plan(payload.goal, candidates)
    except Exception as error:
        handle_mimo_error(error)


@app.post("/api/ai/component-search")
def ai_component_search(payload: AiComponentSearchRequest, _: Protected, db: Session = Depends(get_db)):
    candidates = search_component_candidates(db, payload.requirement, payload.limit)
    try:
        return component_search(payload.requirement, candidates)
    except Exception as error:
        handle_mimo_error(error)


@app.get("/api/ai/search-suggestions")
def ai_search_suggestions(
    _: Protected,
    db: Session = Depends(get_db),
    keyword: str = Query(..., min_length=1, max_length=120),
    category_id: int | None = None,
    status: str | None = None,
    ai_status: str | None = None,
    stock: str | None = None,
):
    categories = [name for (name,) in db.query(Category.name).order_by(Category.id).all()]
    filters = {
        "category": db.get(Category, category_id).name if category_id and db.get(Category, category_id) else "",
        "status": status or "",
        "ai_status": ai_status or "",
        "stock": stock or "",
    }
    try:
        return search_empty_suggestions(keyword, filters, categories)
    except Exception as error:
        handle_mimo_error(error)


@app.post("/api/ai/component-info")
def ai_component_info(payload: AiComponentInfoRequest, _: Protected):
    try:
        return component_info(payload.query, payload.known_specs, payload.web_search)
    except Exception as error:
        handle_mimo_error(error)


@app.post("/api/ai/image-import/preview", response_model=list[ImageImportPreviewRow])
async def ai_image_import_preview(_: Protected, files: list[UploadFile] = File(...), db: Session = Depends(get_db)):
    if not files:
        raise HTTPException(status_code=400, detail="No image files uploaded")
    if len(files) > 8:
        raise HTTPException(status_code=400, detail="At most 8 images are allowed")
    images: list[dict[str, str]] = []
    for file in files:
        content_type = file.content_type or "application/octet-stream"
        if content_type not in {"image/jpeg", "image/png", "image/webp"}:
            raise HTTPException(status_code=400, detail=f"Unsupported image type: {content_type}")
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"Image is too large: {file.filename}")
        images.append({"content_type": content_type, "base64": base64.b64encode(content).decode("ascii")})
    inventory = db.query(Component).options(joinedload(Component.category)).order_by(Component.updated_at.desc()).limit(80).all()
    categories = [name for (name,) in db.query(Category.name).order_by(Category.id).all()]
    try:
        result = await run_foreground_ai(image_import_preview, images, inventory, categories)
    except Exception as error:
        handle_mimo_error(error)
    rows = []
    for item in result.get("items", []):
        if not isinstance(item, dict):
            continue
        values = {
            "name": clean_component_name(item.get("normalized_name") or item.get("name") or item.get("model"), item.get("model")),
            "model": item.get("model"),
            "category_id": category_id_by_name(db, item.get("category_suggestion")),
            "parameters": item.get("parameters"),
            "package": item.get("package"),
            "quantity": int(item.get("quantity") or 1),
            "source": item.get("source") or "图片识别导入",
            "lcsc_number": item.get("lcsc_number"),
            "tags": ",".join(normalize_tag_text(item.get("tags"))),
            "source_title": item.get("source_title") or item.get("name") or item.get("evidence_text"),
            "status": "in_stock",
            "location": None,
            "remark": item.get("evidence_text"),
            "datasheet_url": None,
            "part_family": item.get("part_family") or "component",
            "count_mode": item.get("count_mode") or "exact",
            "normalized_spec": item.get("normalized_spec"),
        }
        values.update(
            {
                "confidence": item.get("confidence") or "low",
                "evidence_text": item.get("evidence_text"),
                "category_suggestion": item.get("category_suggestion"),
                "matched_component_id": item.get("matched_component_id"),
                "match_score": int(item.get("match_score") or 0),
                "lcsc_search_url": item.get("lcsc_search_url") or lcsc_search_url(item.get("model") or item.get("name")),
                "action": (
                    "skip"
                    if (item.get("confidence") == "low" or values.get("part_family") == "other")
                    else "merge"
                    if item.get("matched_component_id")
                    else "create"
                ),
            }
        )
        rows.append(values)
    return rows
