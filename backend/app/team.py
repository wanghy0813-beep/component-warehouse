import base64
import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import shutil
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

import qrcode
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse
from openpyxl import Workbook, load_workbook
from qrcode.image.svg import SvgPathImage
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from .auth import AuthContext, require_access
from .team_schemas import (
    ContestAiRequest,
    ContestComponentBulkAdd,
    ContestComponentCreate,
    ContestComponentLink,
    ContestComponentRebind,
    ContestComponentUpdate,
    ContestLibraryCreate,
    ContestLibraryUpdate,
    ContestPcbCreate,
    ContestPcbUpdate,
    TeamComponentQuantityUpdate,
    TeamMarkerCreate,
    TeamMarkerUpdate,
)
from .schemas import ComponentAiAskRequest, ComponentAiAskOut, ComponentConsumeRequest, ComponentExportRequest, CustomLabelExportRequest, CustomLabelTemplateCreate, CustomLabelTemplateOut, CustomLabelTemplateUpdate, InventoryLotCreate, InventoryLotOut, UsageEventRequest
from .database import get_db
from .branding import APP_BRAND_NAME
from .models import (
    CompetitionActivityLog,
    CompetitionAiResult,
    CompetitionComponentMarker,
    CompetitionInvite,
    CompetitionLibrary,
    CompetitionLibraryComponent,
    CompetitionLibraryMember,
    CompetitionPcb,
    CustomLabelAsset,
    CustomLabelTemplate,
    ActivityLog,
    Category,
    Component,
    InventoryLot,
    Project,
    ProjectBomItem,
    User,
)
from .services.inventory import category_sort_key, component_value_sort_key, reserved_quantities
from .services.component_search import find_unit_conversion_match, keyword_unit_variants
from .services.stock_ledger import ensure_component_lot, reconcile_component_lots, record_stock_delta
from .services.lcsc_lookup import normalize_lcsc_number
from .component_identity import allocate_component_identity
from .services.mimo_ai import (
    MimoNotConfiguredError,
    MimoRequestError,
    component_question,
    component_to_dict,
    contest_library_assist,
)
from .labels import (
    custom_label_font_keys,
    category_package_summary_from_records,
    data_uri_from_bytes,
    is_standard_category_label_group,
    label_document,
    print_timestamp,
    render_component_label_pdf,
    render_component_label_sheet,
    render_basic_custom_label_pdf_items,
    render_custom_label_cards,
    render_standard_category_label_pdf_items,
    render_standard_category_label_cards,
    render_custom_label_sheet,
    sanitize_svg_markup,
)


router = APIRouter(prefix="/api/team", tags=["team"])

INVITE_SECRET = os.getenv("TEAM_INVITE_SECRET", "").strip()
TEAM_SECRET_FILE = Path(
    os.getenv("TEAM_SECRET_FILE", "./data/.contest-invite-secret")
)
_INVITE_SECRET_LOCK = threading.Lock()
PUBLIC_TEAM_BASE_URL = os.getenv(
    "PUBLIC_TEAM_BASE_URL",
    "http://localhost:8080/component-warehouse/team",
).rstrip("/")
PUBLIC_PERSONAL_BASE_URL = os.getenv(
    "PUBLIC_PERSONAL_BASE_URL",
    "http://localhost:8080/component-warehouse/personal",
).rstrip("/")
TEAM_MEDIA_ROOT = Path(os.getenv("TEAM_MEDIA_ROOT", "./data/contest-media"))
MAX_PCB_IMAGE_BYTES = 2 * 1024 * 1024
PCB_STATUSES = {"可用", "待确认", "停用"}
CUSTOM_LABEL_STORAGE_ROOT = Path(os.getenv("CUSTOM_LABEL_STORAGE_ROOT", "./data/custom-labels"))
CUSTOM_LABEL_ALLOWED_MIME = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/svg+xml": ".svg",
}
CUSTOM_LABEL_MAX_IMAGE_BYTES = int(os.getenv("CUSTOM_LABEL_MAX_IMAGE_BYTES", str(5 * 1024 * 1024)))
CUSTOM_LABEL_MAX_SVG_BYTES = int(os.getenv("CUSTOM_LABEL_MAX_SVG_BYTES", str(256 * 1024)))


def new_uuid() -> str:
    return str(uuid.uuid4())


def parse_label_export_datetime(value: str | None, *, end: bool = False) -> datetime | None:
    text_value = str(value or "").strip()
    if not text_value:
        return None
    is_date_only = bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", text_value))
    try:
        parsed = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="标签导出日期格式无效") from exc
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone().replace(tzinfo=None)
    if end and is_date_only:
        parsed = parsed + timedelta(days=1)
    return parsed


def apply_label_import_date_filter(query, model, imported_from: str | None, imported_to: str | None):
    start = parse_label_export_datetime(imported_from)
    end = parse_label_export_datetime(imported_to, end=True)
    if start and end and start >= end:
        raise HTTPException(status_code=400, detail="标签导出开始日期必须早于结束日期")
    if not start and not end:
        return query
    date_expr = model.created_at
    if start:
        query = query.filter(date_expr >= start)
    if end:
        query = query.filter(date_expr < end)
    return query


def normalize_label_excluded_categories(values: list[str] | None) -> set[str]:
    result: set[str] = set()
    for value in values or []:
        name = str(value or "").strip()
        if name:
            result.add(name)
    return result


def apply_team_label_category_exclusion(query, excluded_categories: list[str] | None):
    names = normalize_label_excluded_categories(excluded_categories)
    if not names:
        return query
    include_uncategorized = "未分类" in names
    concrete_names = {name for name in names if name != "未分类"}
    if concrete_names:
        query = query.filter(
            or_(
                CompetitionLibraryComponent.category.is_(None),
                CompetitionLibraryComponent.category == "",
                ~CompetitionLibraryComponent.category.in_(concrete_names),
            )
        )
    if include_uncategorized:
        query = query.filter(CompetitionLibraryComponent.category.isnot(None), CompetitionLibraryComponent.category != "")
    return query


def label_record_category_name(record: dict) -> str:
    category = record.get("category") if isinstance(record, dict) else None
    if isinstance(category, dict):
        name = str(category.get("name") or "").strip()
    else:
        name = str(category or "").strip()
    return name or "未分类"


def filter_label_records_by_excluded_categories(records: list[dict], excluded_categories: list[str] | None) -> list[dict]:
    names = normalize_label_excluded_categories(excluded_categories)
    if not names:
        return records
    return [record for record in records if label_record_category_name(record) not in names]


def get_invite_secret() -> str:
    global INVITE_SECRET
    if INVITE_SECRET:
        return INVITE_SECRET
    with _INVITE_SECRET_LOCK:
        if INVITE_SECRET:
            return INVITE_SECRET
        if TEAM_SECRET_FILE.exists():
            INVITE_SECRET = TEAM_SECRET_FILE.read_text(encoding="utf-8").strip()
        if not INVITE_SECRET:
            TEAM_SECRET_FILE.parent.mkdir(parents=True, exist_ok=True)
            INVITE_SECRET = secrets.token_urlsafe(48)
            TEAM_SECRET_FILE.write_text(INVITE_SECRET, encoding="utf-8")
            try:
                TEAM_SECRET_FILE.chmod(0o600)
            except OSError:
                pass
        return INVITE_SECRET


def request_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    return (forwarded.split(",")[0].strip() or (request.client.host if request.client else ""))[:80]


def phone_last4(phone: str | None) -> str:
    value = str(phone or "")
    return value[-4:] if len(value) >= 4 else value


def json_value(value) -> str | None:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)


def record_dict(record) -> dict:
    return {
        column.name: getattr(record, column.name)
        for column in record.__table__.columns
    }


def append_text(first: str | None, second: str | None) -> str | None:
    values = []
    for value in [first, second]:
        text = str(value or "").strip()
        if text and text not in values:
            values.append(text)
    return "\n".join(values) or None


def log_action(
    db: Session,
    library_id: str,
    auth: AuthContext,
    request: Request,
    action: str,
    entity_type: str,
    summary: str,
    *,
    entity_id: str | None = None,
    before=None,
    after=None,
) -> None:
    db.add(
        CompetitionActivityLog(
            library_id=library_id,
            actor_user_id=auth.user_id,
            actor_nickname=auth.nickname[:80],
            actor_phone_last4=phone_last4(auth.phone),
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            summary=summary[:300],
            before_json=json_value(before),
            after_json=json_value(after),
            ip_address=request_ip(request),
        )
    )


def membership(
    db: Session,
    library_id: str,
    user_id: int,
) -> CompetitionLibraryMember | None:
    return (
        db.query(CompetitionLibraryMember)
        .filter(
            CompetitionLibraryMember.library_id == library_id,
            CompetitionLibraryMember.user_id == user_id,
        )
        .first()
    )


def require_library_member(
    db: Session,
    library_id: str,
    auth: AuthContext,
) -> tuple[CompetitionLibrary, CompetitionLibraryMember]:
    library = db.get(CompetitionLibrary, library_id)
    member = membership(db, library_id, auth.user_id)
    if (
        not library
        or library.status != "active"
        or not member
        or member.status != "active"
    ):
        raise HTTPException(status_code=404, detail="团队器件库不存在或你还没有加入")
    return library, member


def require_library_captain(
    db: Session,
    library_id: str,
    auth: AuthContext,
) -> tuple[CompetitionLibrary, CompetitionLibraryMember]:
    library, member = require_library_member(db, library_id, auth)
    if member.role != "captain":
        raise HTTPException(status_code=403, detail="只有队长可以操作")
    return library, member


def require_library_editor(
    db: Session,
    library_id: str,
    auth: AuthContext,
) -> tuple[CompetitionLibrary, CompetitionLibraryMember]:
    library, member = require_library_member(db, library_id, auth)
    if member.role not in {"captain", "editor", "member"}:
        raise HTTPException(status_code=403, detail="当前角色只有查看权限")
    return library, member


def user_out(user: User | None, role: str | None = None) -> dict:
    if not user:
        return {
            "id": None,
            "nickname": "未知成员",
            "phone_last4": "",
            "role": role,
        }
    return {
        "id": user.id,
        "nickname": user.nickname or f"用户{phone_last4(user.phone)}",
        "phone_last4": phone_last4(user.phone),
        "role": role,
    }


def library_out(
    db: Session,
    library: CompetitionLibrary,
    member: CompetitionLibraryMember | None = None,
) -> dict:
    component_count = (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.library_id == library.id)
        .count()
    )
    pcb_count = (
        db.query(CompetitionPcb)
        .filter(CompetitionPcb.library_id == library.id)
        .count()
    )
    member_count = (
        db.query(CompetitionLibraryMember)
        .filter(
            CompetitionLibraryMember.library_id == library.id,
            CompetitionLibraryMember.status == "active",
        )
        .count()
    )
    return {
        "id": library.id,
        "name": library.name,
        "competition_type": library.competition_type,
        "description": library.description,
        "status": library.status,
        "creator_user_id": library.creator_user_id,
        "role": member.role if member else None,
        "component_count": component_count,
        "pcb_count": pcb_count,
        "member_count": member_count,
        "created_at": library.created_at,
        "updated_at": library.updated_at,
    }


def cw_component_out(
    component: Component | None,
    reserved: int = 0,
    search_keyword: str | None = None,
) -> dict | None:
    if not component:
        return None
    available = max(0, int(component.quantity or 0) - int(reserved or 0))
    safety_quantity = max(0, int(component.safety_quantity or 0))
    return {
        "id": component.id,
        "owner_user_id": component.owner_user_id,
        "warehouse_code": component.warehouse_code,
        "name": component.name,
        "model": component.model,
        "manufacturer": component.manufacturer,
        "description": component.description,
        "lcsc_number": component.lcsc_number,
        "parameters": component.parameters,
        "package": component.package,
        "datasheet_url": component.datasheet_url,
        "buy_url": component.buy_url,
        "source_title": component.source_title,
        "tags": component.tags,
        "ai_summary": component.ai_summary,
        "ai_usage": component.ai_usage,
        "ai_risk_notes": component.ai_risk_notes,
        "ai_pcb_notes": component.ai_pcb_notes,
        "ai_substitutes": component.ai_substitutes,
        "ai_tags": component.ai_tags,
        "ai_confidence": component.ai_confidence,
        "category": {
            "id": component.category.id,
            "name": component.category.name,
            "color": component.category.color,
        } if component.category else None,
        "quantity": int(component.quantity or 0),
        "reserved_quantity": int(reserved or 0),
        "available_quantity": available,
        "safety_quantity": safety_quantity,
        "low_stock_exempt": bool(component.low_stock_exempt),
        "low_stock_warning": bool(
            component.is_common
            and not component.low_stock_exempt
            and available <= (safety_quantity if safety_quantity > 0 else 5)
        ),
        "is_common": bool(component.is_common),
        "status": component.status,
        "source": component.source,
        "normalized_spec": component.normalized_spec,
        "location": component.location,
        "remark": component.remark,
        "first_stocked_at": component.first_stocked_at,
        "last_stocked_at": component.last_stocked_at,
        "created_at": component.created_at,
        "search_unit_conversion": find_unit_conversion_match(
            search_keyword,
            (
                component.name,
                component.warehouse_code,
                component.model,
                component.parameters,
                component.normalized_spec,
                component.package,
                component.lcsc_number,
                component.location,
                component.remark,
                component.ai_summary,
                component.ai_tags,
                component.tags,
                component.source_title,
            ),
        ),
    }


def team_components_out(
    db: Session,
    items: list[CompetitionLibraryComponent],
    auth: AuthContext | None = None,
    member: CompetitionLibraryMember | None = None,
) -> list[dict]:
    if not items:
        return []
    component_ids = sorted({
        item.cw_component_id
        for item in items
        if item.sync_status == "live" and item.cw_component_id
    })
    components = {
        component.id: component
        for component in (
            db.query(Component)
            .options(joinedload(Component.category))
            .filter(Component.id.in_(component_ids), Component.revoked_at.is_(None))
            .all()
            if component_ids
            else []
        )
    }
    reserved_by_component = reserved_quantities(db, component_ids)
    item_ids = [item.id for item in items]
    marker_rows = (
        db.query(CompetitionComponentMarker)
        .filter(CompetitionComponentMarker.component_id.in_(item_ids))
        .order_by(
            CompetitionComponentMarker.component_id,
            CompetitionComponentMarker.flagged.desc(),
            CompetitionComponentMarker.created_at.desc(),
        )
        .all()
    )
    creator_ids = sorted({marker.created_by_user_id for marker in marker_rows})
    creators = {
        user.id: user
        for user in (
            db.query(User).filter(User.id.in_(creator_ids)).all()
            if creator_ids
            else []
        )
    }
    markers_by_item: dict[str, list[dict]] = {}
    for marker in marker_rows:
        creator = creators.get(marker.created_by_user_id)
        markers_by_item.setdefault(marker.component_id, []).append({
            **record_dict(marker),
            "creator_name": creator.nickname if creator else "团队成员",
        })
    if auth and member is None:
        member = membership(db, items[0].library_id, auth.user_id)

    output = []
    for item in items:
        component = components.get(item.cw_component_id)
        if component and component.owner_user_id != item.source_user_id:
            component = None
        snapshot = {}
        if item.frozen_snapshot_json:
            try:
                snapshot = json.loads(item.frozen_snapshot_json)
            except json.JSONDecodeError:
                snapshot = {}
        source = (
            cw_component_out(component, reserved_by_component.get(component.id, 0))
            if component
            else snapshot
        ) or {}
        category = source.get("category")
        category_name = (
            category.get("name")
            if isinstance(category, dict)
            else item.category or None
        )
        quantity = int(source.get("quantity") if source.get("quantity") is not None else item.quantity or 0)
        reserved_quantity = int(source.get("reserved_quantity") or 0)
        output.append({
            **record_dict(item),
            "name": source.get("name") or item.name,
            "model": source.get("model") or item.model,
            "manufacturer": source.get("manufacturer"),
            "description": source.get("description"),
            "lcsc_number": source.get("lcsc_number") or item.lcsc_number,
            "warehouse_code": source.get("warehouse_code") or item.warehouse_code_snapshot,
            "parameters": source.get("parameters"),
            "package": source.get("package"),
            "datasheet_url": source.get("datasheet_url"),
            "buy_url": source.get("buy_url"),
            "source_title": source.get("source_title"),
            "normalized_spec": source.get("normalized_spec"),
            "ai_summary": source.get("ai_summary"),
            "ai_usage": source.get("ai_usage"),
            "ai_risk_notes": source.get("ai_risk_notes"),
            "ai_pcb_notes": source.get("ai_pcb_notes"),
            "ai_substitutes": source.get("ai_substitutes"),
            "ai_tags": source.get("ai_tags"),
            "ai_confidence": source.get("ai_confidence"),
            "source": source.get("source"),
            "status": source.get("status") or "in_stock",
            "category": category if isinstance(category, dict) else (
                {"id": None, "name": category_name, "color": "#eef2f7"} if category_name else None
            ),
            "quantity": quantity,
            "reserved_quantity": reserved_quantity,
            "available_quantity": max(0, quantity - reserved_quantity),
            "safety_quantity": int(source.get("safety_quantity") or 0),
            "low_stock_exempt": bool(source.get("low_stock_exempt")),
            "low_stock_warning": bool(source.get("low_stock_warning")),
            "is_common": bool(source.get("is_common")),
            "can_view_usage": bool(component),
            "can_edit_quantity": bool(
                auth
                and component
                and (item.source_user_id == auth.user_id or (member and member.role == "captain"))
            ),
            "markers": markers_by_item.get(item.id, []),
            "cw_component": source,
        })
    return output


def team_component_out(
    db: Session,
    item: CompetitionLibraryComponent,
    auth: AuthContext | None = None,
) -> dict:
    return team_components_out(db, [item], auth)[0]


def inventory_lot_out(lot: InventoryLot) -> dict:
    return {
        "id": lot.id,
        "component_id": lot.component_id,
        "source_type": lot.source_type,
        "source_reference": lot.source_reference,
        "location": lot.location,
        "initial_quantity": int(lot.initial_quantity or 0),
        "remaining_quantity": int(lot.remaining_quantity or 0),
        "unit_cost": lot.unit_cost,
        "status": lot.status,
        "received_at": lot.received_at,
        "created_at": lot.created_at,
    }


def custom_label_asset_url(asset: CustomLabelAsset, library_id: str) -> str:
    return f"/api/team/libraries/{quote(library_id)}/custom-labels/assets/{quote(asset.id)}"


def custom_label_asset_out(asset: CustomLabelAsset, library_id: str) -> dict:
    return {
        "id": asset.id,
        "template_id": asset.template_id,
        "file_name": asset.file_name,
        "mime_type": asset.mime_type,
        "sha256": asset.sha256,
        "size_bytes": int(asset.size_bytes or 0),
        "url": custom_label_asset_url(asset, library_id),
        "created_at": asset.created_at,
    }


def parse_custom_label_content(value: str | dict | None) -> dict:
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value or "{}")
    except Exception:
        parsed = {}
    return parsed if isinstance(parsed, dict) else {}


def clean_custom_label_content(content: dict | None) -> dict:
    raw = content if isinstance(content, dict) else {}
    elements = raw.get("elements")
    cleaned: list[dict] = []
    if isinstance(elements, list):
        for index, item in enumerate(elements[:40]):
            if not isinstance(item, dict):
                continue
            element_type = str(item.get("type") or "text").strip()
            if element_type not in {"text", "image", "svg", "field", "qr", "shape", "category_badge"}:
                element_type = "text"
            if element_type in {"field", "category_badge"} and str(item.get("field") or "").strip() == "print_date":
                continue
            row = {
                "id": str(item.get("id") or f"el-{index + 1}")[:80],
                "type": element_type,
                "x": item.get("x", 22),
                "y": item.get("y", 34),
                "width": item.get("width", 56),
                "height": item.get("height", 30),
                "x_mm": item.get("x_mm"),
                "y_mm": item.get("y_mm"),
                "width_mm": item.get("width_mm"),
                "height_mm": item.get("height_mm"),
                "rotate": item.get("rotate", 0),
                "font_size": item.get("font_size", 13),
                "font_weight": item.get("font_weight", 400),
                "font_family": str(item.get("font_family") or "system")[:40],
                "color": str(item.get("color") or "#111827")[:40],
                "align": str(item.get("align") or "center")[:16],
            }
            if item.get("role"):
                row["role"] = str(item.get("role") or "")[:80]
            if element_type == "text":
                row["text"] = str(item.get("text") or "自定义标签")[:1000]
            elif element_type in {"image", "svg"}:
                row["asset_id"] = str(item.get("asset_id") or "")[:80]
                if element_type == "svg" and item.get("svg"):
                    row["svg"] = sanitize_svg_markup(str(item.get("svg") or ""))
            elif element_type in {"field", "qr", "category_badge"}:
                row["field"] = str(item.get("field") or ("scan_url" if element_type == "qr" else "name"))[:80]
                row["prefix"] = str(item.get("prefix") or "")[:80]
            elif element_type == "shape":
                row["fill"] = str(item.get("fill") or "#eff6ff")[:40]
                row["stroke"] = str(item.get("stroke") or "#93c5fd")[:40]
                row["radius"] = item.get("radius", 1)
            for mm_key in ("x_mm", "y_mm", "width_mm", "height_mm"):
                if row.get(mm_key) is None:
                    row.pop(mm_key, None)
            cleaned.append(row)
    if not cleaned:
        cleaned = [{"id": "text-1", "type": "text", "text": str(raw.get("text") or "自定义标签")[:1000], "x": 18, "y": 33, "width": 64, "height": 30, "font_size": 16, "font_family": "system", "color": "#111827", "align": "center"}]
    result = {"elements": cleaned, "show_logo": raw.get("show_logo") is not False}
    if raw.get("kind"):
        result["kind"] = str(raw.get("kind") or "")[:80]
    styles = raw.get("styles")
    cleaned_styles: list[dict] = []
    if isinstance(styles, list):
        for index, style in enumerate(styles[:20]):
            if not isinstance(style, dict):
                continue
            style_content = clean_custom_label_content({"elements": style.get("elements")})
            cleaned_styles.append({
                "id": str(style.get("id") or f"style-{index + 1}")[:80],
                "name": str(style.get("name") or f"样式 {index + 1}")[:80],
                "category_name": str(style.get("category_name") or "")[:80],
                "elements": style_content["elements"],
            })
    if cleaned_styles:
        result["styles"] = cleaned_styles
        active_style_id = str(raw.get("active_style_id") or cleaned_styles[0]["id"])[:80]
        result["active_style_id"] = active_style_id
    return result


def custom_label_template_out(template: CustomLabelTemplate, db: Session) -> dict:
    assets = (
        db.query(CustomLabelAsset)
        .filter(CustomLabelAsset.template_id == template.id)
        .order_by(CustomLabelAsset.created_at.asc())
        .all()
    )
    return {
        "id": template.id,
        "scope_type": template.scope_type,
        "owner_user_id": template.owner_user_id,
        "team_library_id": template.team_library_id,
        "name": template.name,
        "content": parse_custom_label_content(template.content_json),
        "status": template.status,
        "assets": [custom_label_asset_out(asset, template.team_library_id or "") for asset in assets],
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


def require_team_custom_label_template(db: Session, library_id: str, template_id: str) -> CustomLabelTemplate:
    template = (
        db.query(CustomLabelTemplate)
        .filter(
            CustomLabelTemplate.scope_type == "team",
            CustomLabelTemplate.team_library_id == library_id,
            CustomLabelTemplate.status == "active",
            CustomLabelTemplate.id == template_id,
        )
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="自定义标签不存在")
    return template


def custom_label_asset_root(template: CustomLabelTemplate) -> Path:
    owner = template.team_library_id or "unknown"
    return CUSTOM_LABEL_STORAGE_ROOT / "team" / owner / template.id


def assert_custom_label_mime(file: UploadFile, payload: bytes) -> tuple[str, bytes, str]:
    mime_type = (file.content_type or "").split(";", 1)[0].strip().lower()
    suffix = CUSTOM_LABEL_ALLOWED_MIME.get(mime_type)
    if not suffix:
        raise HTTPException(status_code=400, detail="仅支持 PNG、JPG、WebP 或 SVG")
    max_bytes = CUSTOM_LABEL_MAX_SVG_BYTES if mime_type == "image/svg+xml" else CUSTOM_LABEL_MAX_IMAGE_BYTES
    if len(payload) > max_bytes:
        raise HTTPException(status_code=413, detail="自定义标签素材超过大小限制")
    data = payload
    if mime_type == "image/svg+xml":
        cleaned = sanitize_svg_markup(payload.decode("utf-8", errors="ignore"))
        if not cleaned:
            raise HTTPException(status_code=400, detail="SVG 内容不安全或无法解析")
        data = cleaned.encode("utf-8")
    else:
        signatures = {"image/png": b"\x89PNG\r\n\x1a\n", "image/jpeg": b"\xff\xd8\xff", "image/webp": b"RIFF"}
        if not payload.startswith(signatures[mime_type]):
            raise HTTPException(status_code=400, detail="图片文件头与类型不匹配")
        if mime_type == "image/webp" and payload[8:12] != b"WEBP":
            raise HTTPException(status_code=400, detail="WebP 文件头无效")
    return mime_type, data, suffix


async def save_custom_label_asset(db: Session, template: CustomLabelTemplate, file: UploadFile) -> CustomLabelAsset:
    payload = await file.read()
    mime_type, data, suffix = assert_custom_label_mime(file, payload)
    asset_id = secrets.token_hex(16)
    root = custom_label_asset_root(template)
    root.mkdir(parents=True, exist_ok=True)
    file_name = Path(file.filename or f"asset{suffix}").name[:200] or f"asset{suffix}"
    storage_path = root / f"{asset_id}{suffix}"
    tmp_path = storage_path.with_suffix(storage_path.suffix + ".tmp")
    tmp_path.write_bytes(data)
    tmp_path.replace(storage_path)
    asset = CustomLabelAsset(
        id=asset_id,
        template_id=template.id,
        file_name=file_name,
        storage_path=str(storage_path),
        mime_type=mime_type,
        sha256=hashlib.sha256(data).hexdigest(),
        size_bytes=len(data),
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def custom_label_asset_resolver(db: Session, template: CustomLabelTemplate):
    allowed = {asset.id: asset for asset in db.query(CustomLabelAsset).filter(CustomLabelAsset.template_id == template.id).all()}

    def resolve(asset_id: str) -> dict:
        asset = allowed.get(asset_id)
        if not asset:
            return {}
        path = Path(asset.storage_path)
        if not path.exists() or not path.is_file():
            return {}
        data = path.read_bytes()
        if asset.mime_type == "image/svg+xml":
            return {"svg": sanitize_svg_markup(data.decode("utf-8", errors="ignore"))}
        return {"data_uri": data_uri_from_bytes(asset.mime_type, data)}

    return resolve


def team_component_export_custom_label_cards(db: Session, library_id: str, items: list, printed_at: str, records: list[dict] | None = None) -> list[str]:
    cards: list[str] = []
    package_summary = category_package_summary_from_records(records)
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_team_custom_label_template(db, library_id, template_id)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        if is_standard_category_label_group(content):
            cards.extend(
                render_standard_category_label_cards(
                    content,
                    package_summary,
                    copies=int(getattr(item, "copies", 1) or 1),
                    printed_at=printed_at,
                )
            )
            continue
        cards.extend(
            render_custom_label_cards(
                content,
                asset_resolver=custom_label_asset_resolver(db, template),
                copies=int(getattr(item, "copies", 1) or 1),
                printed_at=printed_at,
                include_all_styles=True,
            )
        )
    return cards


def team_component_export_custom_label_font_keys(db: Session, library_id: str, items: list) -> set[str]:
    keys: set[str] = set()
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_team_custom_label_template(db, library_id, template_id)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        keys.update(custom_label_font_keys(content))
    return keys


def team_component_export_custom_label_pdf_items(db: Session, library_id: str, items: list, records: list[dict] | None = None) -> list[dict]:
    pdf_items: list[dict] = []
    package_summary = category_package_summary_from_records(records)
    for item in items or []:
        template_id = str(getattr(item, "template_id", "") or "").strip()
        if not template_id:
            continue
        template = require_team_custom_label_template(db, library_id, template_id)
        content = clean_custom_label_content(parse_custom_label_content(template.content_json))
        copies = int(getattr(item, "copies", 1) or 1)
        if is_standard_category_label_group(content):
            pdf_items.extend(render_standard_category_label_pdf_items(content, package_summary, copies=copies))
        else:
            pdf_items.extend(render_basic_custom_label_pdf_items(content, copies=copies))
    return pdf_items


def require_team_source_component(
    db: Session,
    library_id: str,
    item_id: str,
    auth: AuthContext,
    *,
    mutate: bool = False,
) -> tuple[CompetitionLibraryComponent, Component, CompetitionLibraryMember]:
    _, member = (require_library_editor if mutate else require_library_member)(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="器件不存在")
    if item.sync_status != "live" or not item.cw_component_id:
        raise HTTPException(status_code=409, detail="冻结器件需要重新绑定后才能查看或修改来源库存")
    component = db.get(Component, item.cw_component_id)
    if not component or component.owner_user_id != item.source_user_id:
        raise HTTPException(status_code=409, detail="来源器件不可用，请重新绑定")
    if mutate and item.source_user_id != auth.user_id and member.role != "captain":
        raise HTTPException(status_code=403, detail="仅来源成员或队长可以修改来源库存")
    return item, component, member


def marker_out(db: Session, marker: CompetitionComponentMarker) -> dict:
    creator = db.get(User, marker.created_by_user_id)
    return {
        **record_dict(marker),
        "creator_name": creator.nickname if creator else "团队成员",
    }


def pcb_out(item: CompetitionPcb) -> dict:
    data = record_dict(item)
    base = f"/api/team/libraries/{item.library_id}/pcbs/{item.id}/images"
    data["front_image_url"] = f"{base}/front" if item.front_image_path else None
    data["back_image_url"] = f"{base}/back" if item.back_image_path else None
    return data


def invite_signature(invite_id: str) -> str:
    digest = hmac.new(
        get_invite_secret().encode("utf-8"),
        invite_id.encode("utf-8"),
        hashlib.sha256,
    ).digest()
    return base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")[:32]


def invite_token(invite_id: str) -> str:
    return f"{invite_id}.{invite_signature(invite_id)}"


def invite_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def active_invite(db: Session, library_id: str) -> CompetitionInvite:
    invite = (
        db.query(CompetitionInvite)
        .filter(
            CompetitionInvite.library_id == library_id,
            CompetitionInvite.active == True,
        )
        .order_by(CompetitionInvite.created_at.desc())
        .first()
    )
    if invite:
        return invite
    raise HTTPException(status_code=404, detail="邀请二维码尚未创建")


def resolve_invite(db: Session, token: str) -> tuple[CompetitionInvite, CompetitionLibrary]:
    parts = str(token or "").split(".", 1)
    if len(parts) != 2 or not hmac.compare_digest(parts[1], invite_signature(parts[0])):
        raise HTTPException(status_code=404, detail="邀请链接无效")
    invite = db.get(CompetitionInvite, parts[0])
    if (
        not invite
        or not invite.active
        or not hmac.compare_digest(invite.token_hash, invite_hash(token))
    ):
        raise HTTPException(status_code=404, detail="邀请链接已失效")
    library = db.get(CompetitionLibrary, invite.library_id)
    if not library or library.status != "active":
        raise HTTPException(status_code=404, detail="团队器件库已停用")
    return invite, library


def create_invite(
    db: Session,
    library_id: str,
    user_id: int,
) -> tuple[CompetitionInvite, str]:
    invite = CompetitionInvite(
        id=new_uuid(),
        library_id=library_id,
        token_hash="",
        active=True,
        created_by_user_id=user_id,
    )
    token = invite_token(invite.id)
    invite.token_hash = invite_hash(token)
    db.add(invite)
    return invite, token


def find_cw_component(
    db: Session,
    *,
    owner_user_id: int,
    cw_component_id: int | None = None,
    lcsc_number: str | None = None,
) -> Component | None:
    if cw_component_id:
        return (
            db.query(Component)
            .filter(
                Component.id == int(cw_component_id),
                Component.owner_user_id == owner_user_id,
                Component.revoked_at.is_(None),
            )
            .first()
        )
    lcsc = normalize_lcsc_number(lcsc_number) or str(lcsc_number or "").strip()
    if lcsc:
        return (
            db.query(Component)
            .filter(
                func.lower(Component.lcsc_number) == lcsc.lower(),
                Component.owner_user_id == owner_user_id,
                Component.revoked_at.is_(None),
            )
            .first()
        )
    return None


def component_snapshot(db: Session, component: Component) -> dict:
    reserved = reserved_quantities(db, [component.id]).get(component.id, 0)
    return cw_component_out(component, reserved) or {}


def freeze_member_components(
    db: Session,
    library_id: str,
    user_id: int,
) -> int:
    rows = (
        db.query(CompetitionLibraryComponent)
        .filter(
            CompetitionLibraryComponent.library_id == library_id,
            CompetitionLibraryComponent.source_user_id == user_id,
            CompetitionLibraryComponent.sync_status == "live",
        )
        .all()
    )
    for item in rows:
        component = db.get(Component, item.cw_component_id) if item.cw_component_id else None
        if component:
            snapshot = component_snapshot(db, component)
            item.quantity = int(component.quantity or 0)
            item.name = component.name
            item.model = component.model
            item.lcsc_number = component.lcsc_number
            item.warehouse_code_snapshot = component.warehouse_code
            item.frozen_snapshot_json = json.dumps(snapshot, ensure_ascii=False, default=str)
        item.cw_component_id = None
        item.sync_status = "frozen"
    return len(rows)


def create_personal_component(
    db: Session,
    auth: AuthContext,
    values: dict,
) -> tuple[Component, bool]:
    normalized_lcsc = normalize_lcsc_number(values.get("lcsc_number")) or str(values.get("lcsc_number") or "").strip()[:120] or None
    exact = find_cw_component(
        db,
        owner_user_id=auth.user_id,
        lcsc_number=normalized_lcsc,
    )
    if exact:
        return exact, False
    category_name = str(values.get("category") or "").strip()
    category = (
        db.query(Category).filter(Category.name == category_name).first()
        if category_name
        else None
    )
    component = Component(
        owner_user_id=auth.user_id,
        name=str(values.get("name") or values.get("model") or "未命名物料").strip()[:200],
        model=str(values.get("model") or "").strip()[:200] or None,
        manufacturer=str(values.get("manufacturer") or "").strip()[:200] or None,
        description=str(values.get("description") or "").strip()[:4000] or None,
        lcsc_number=normalized_lcsc,
        quantity=max(0, int(values.get("quantity") or 0)),
        category_id=category.id if category else None,
        package=str(values.get("package") or "").strip()[:120] or None,
        parameters=str(values.get("parameters") or "").strip()[:4000] or None,
        datasheet_url=str(values.get("datasheet_url") or "").strip()[:1000] or None,
        buy_url=str(values.get("buy_url") or "").strip()[:500] or None,
        tags=str(values.get("tags") or "").strip()[:300] or None,
        source_title=str(values.get("source_title") or "").strip()[:4000] or None,
        location=str(values.get("location") or "").strip()[:200] or None,
        remark=str(values.get("remark") or "").strip() or None,
        source=str(values.get("source") or "团队版新增").strip()[:120],
        status="in_stock",
    )
    db.add(component)
    db.flush()
    allocate_component_identity(db, component)
    if int(component.quantity or 0) > 0:
        component.first_stocked_at = component.first_stocked_at or datetime.utcnow()
        component.last_stocked_at = datetime.utcnow()
        record_stock_delta(
            db,
            component,
            int(component.quantity or 0),
            movement_type="team_component_create",
            reason="团队版手动新增初始库存",
            actor_user_id=auth.user_id,
            source_type="team_manual",
            source_reference="团队版新增",
            location=component.location,
        )
    return component, True


def find_existing_team_component(
    db: Session,
    library_id: str,
    *,
    cw_component_id: int | None = None,
    lcsc_number: str | None = None,
    exclude_id: str | None = None,
) -> CompetitionLibraryComponent | None:
    query = db.query(CompetitionLibraryComponent).filter(
        CompetitionLibraryComponent.library_id == library_id
    )
    if exclude_id:
        query = query.filter(CompetitionLibraryComponent.id != exclude_id)
    if cw_component_id:
        return query.filter(
            CompetitionLibraryComponent.cw_component_id == cw_component_id
        ).first()
    lcsc = str(lcsc_number or "").strip()
    if lcsc:
        return query.filter(CompetitionLibraryComponent.lcsc_number == lcsc).first()
    return None


def add_or_merge_component(
    db: Session,
    library_id: str,
    auth: AuthContext,
    request: Request,
    values: dict,
) -> tuple[CompetitionLibraryComponent, bool]:
    cw = find_cw_component(
        db,
        owner_user_id=auth.user_id,
        cw_component_id=values.get("cw_component_id"),
        lcsc_number=values.get("lcsc_number"),
    )
    if not cw:
        cw, _ = create_personal_component(db, auth, values)
    cw_id = cw.id if cw else None
    existing = find_existing_team_component(
        db,
        library_id,
        cw_component_id=cw_id,
        lcsc_number=values.get("lcsc_number") if not cw_id else None,
    )
    if existing:
        before = record_dict(existing)
        existing.location = existing.location or values.get("location")
        existing.remark = append_text(existing.remark, values.get("remark"))
        existing.tags = existing.tags or values.get("tags")
        existing.category = existing.category or values.get("category")
        existing.updated_by_user_id = auth.user_id
        log_action(
            db,
            library_id,
            auth,
            request,
            "component.merge",
            "component",
            f"器件 {existing.name} 已在团队器件库，保留实时库存绑定",
            entity_id=existing.id,
            before=before,
            after=record_dict(existing),
        )
        return existing, True

    item = CompetitionLibraryComponent(
        id=new_uuid(),
        library_id=library_id,
        cw_component_id=cw_id,
        source_user_id=auth.user_id,
        sync_status="live",
        warehouse_code_snapshot=cw.warehouse_code,
        frozen_snapshot_json=json.dumps(component_snapshot(db, cw), ensure_ascii=False, default=str),
        name=str(values.get("name") or (cw.name if cw else "")).strip()[:200],
        model=(values.get("model") or (cw.model if cw else None)),
        lcsc_number=(values.get("lcsc_number") or (cw.lcsc_number if cw else None)),
        quantity=int(cw.quantity or 0),
        location=values.get("location"),
        category=values.get("category"),
        tags=values.get("tags"),
        remark=values.get("remark"),
        created_by_user_id=auth.user_id,
        updated_by_user_id=auth.user_id,
    )
    if not item.name:
        raise HTTPException(status_code=400, detail="物料名称不能为空")
    db.add(item)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.create",
        "component",
        f"加入物料 {item.name}（{cw.warehouse_code or cw.id}）并启用实时库存",
        entity_id=item.id,
        after=record_dict(item),
    )
    return item, False


@router.get("/config")
def team_config():
    return {
        "version": "0.7.1",
        "public_base_url": PUBLIC_TEAM_BASE_URL,
        "max_pcb_image_bytes": MAX_PCB_IMAGE_BYTES,
    }


@router.get("/session")
def team_session(auth: AuthContext = Depends(require_access)):
    return {
        "user": {
            "id": auth.user_id,
            "account_id": auth.account_id,
            "nickname": auth.nickname,
            "avatar_url": auth.avatar_url,
            "phone": auth.phone,
            "phone_last4": phone_last4(auth.phone),
            "is_admin": auth.is_admin,
        },
        "auth_degraded": auth.auth_degraded,
    }


@router.get("/libraries")
def list_libraries(
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    memberships = (
        db.query(CompetitionLibraryMember)
        .filter(
            CompetitionLibraryMember.user_id == auth.user_id,
            CompetitionLibraryMember.status == "active",
        )
        .all()
    )
    libraries = {
        item.id: item
        for item in db.query(CompetitionLibrary)
        .filter(
            CompetitionLibrary.id.in_([member.library_id for member in memberships]),
            CompetitionLibrary.status == "active",
        )
        .all()
    } if memberships else {}
    return [
        library_out(db, libraries[member.library_id], member)
        for member in memberships
        if member.library_id in libraries
    ]


@router.post("/libraries")
def create_library(
    payload: ContestLibraryCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    library = CompetitionLibrary(
        id=new_uuid(),
        name=payload.name.strip(),
        competition_type=(payload.competition_type or "").strip() or None,
        description=(payload.description or "").strip() or None,
        creator_user_id=auth.user_id,
        status="active",
    )
    db.add(library)
    db.flush()
    member = CompetitionLibraryMember(
        library_id=library.id,
        user_id=auth.user_id,
        role="captain",
        status="active",
    )
    db.add(member)
    _, token = create_invite(db, library.id, auth.user_id)
    log_action(
        db,
        library.id,
        auth,
        request,
        "library.create",
        "library",
        f"创建团队器件库 {library.name}",
        entity_id=library.id,
        after=record_dict(library),
    )
    db.commit()
    db.refresh(library)
    return {
        **library_out(db, library, member),
        "invite_url": f"{PUBLIC_TEAM_BASE_URL}/join/{quote(token)}",
    }


@router.get("/libraries/{library_id}")
def get_library(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    library, member = require_library_member(db, library_id, auth)
    return library_out(db, library, member)


@router.put("/libraries/{library_id}")
def update_library(
    library_id: str,
    payload: ContestLibraryUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    library, member = require_library_captain(db, library_id, auth)
    before = record_dict(library)
    values = payload.model_dump(exclude_unset=True)
    if values.get("status") not in {None, "active", "archived"}:
        raise HTTPException(status_code=400, detail="团队器件库状态不正确")
    for key, value in values.items():
        if isinstance(value, str):
            value = value.strip() or None
        setattr(library, key, value)
    log_action(
        db,
        library_id,
        auth,
        request,
        "library.update",
        "library",
        f"更新团队器件库 {library.name}",
        entity_id=library.id,
        before=before,
        after=record_dict(library),
    )
    db.commit()
    db.refresh(library)
    return library_out(db, library, member)


@router.get("/invites/{token}")
def invite_preview(token: str, db: Session = Depends(get_db)):
    _, library = resolve_invite(db, token)
    captain = db.get(User, library.creator_user_id)
    return {
        "library": {
            "id": library.id,
            "name": library.name,
        },
        "captain": {
            "nickname": captain.nickname if captain else "队长",
        },
    }


@router.post("/invites/{token}/join")
def join_library(
    token: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    invite, library = resolve_invite(db, token)
    member = membership(db, library.id, auth.user_id)
    if member and member.blocked_at:
        raise HTTPException(status_code=403, detail="你已被移出该团队器件库，请联系队长")
    if member and member.blocked_invite_id == invite.id:
        raise HTTPException(status_code=403, detail="原邀请已不能用于重新加入，请联系队长获取新链接")
    if member and member.status == "active":
        return {"joined": True, "library": library_out(db, library, member)}
    if member:
        member.status = "active"
        member.removed_at = None
        member.joined_at = datetime.utcnow()
        member.joined_invite_id = invite.id
        member.blocked_invite_id = None
    else:
        member = CompetitionLibraryMember(
            library_id=library.id,
            user_id=auth.user_id,
            role="editor",
            status="active",
            joined_invite_id=invite.id,
        )
        db.add(member)
    invite.use_count = int(invite.use_count or 0) + 1
    log_action(
        db,
        library.id,
        auth,
        request,
        "member.join",
        "member",
        f"{auth.nickname} 加入团队器件库",
        entity_id=str(auth.user_id),
        after={"user_id": auth.user_id, "role": "editor"},
    )
    db.commit()
    return {"joined": True, "library": library_out(db, library, member)}


@router.get("/libraries/{library_id}/invite")
def get_library_invite(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_captain(db, library_id, auth)
    invite = active_invite(db, library_id)
    token = invite_token(invite.id)
    return {
        "token": token,
        "url": f"{PUBLIC_TEAM_BASE_URL}/join/{quote(token)}",
        "use_count": invite.use_count,
        "created_at": invite.created_at,
    }


@router.get("/libraries/{library_id}/invite/qr.svg")
def get_library_invite_qr(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_captain(db, library_id, auth)
    invite = active_invite(db, library_id)
    url = f"{PUBLIC_TEAM_BASE_URL}/join/{quote(invite_token(invite.id))}"
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=6,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    buffer = io.BytesIO()
    qr.make_image(image_factory=SvgPathImage).save(buffer)
    return Response(buffer.getvalue(), media_type="image/svg+xml")


@router.post("/libraries/{library_id}/invite/reset")
def reset_library_invite(
    library_id: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_captain(db, library_id, auth)
    for invite in (
        db.query(CompetitionInvite)
        .filter(
            CompetitionInvite.library_id == library_id,
            CompetitionInvite.active == True,
        )
        .all()
    ):
        invite.active = False
        invite.revoked_at = datetime.utcnow()
    invite, token = create_invite(db, library_id, auth.user_id)
    log_action(
        db,
        library_id,
        auth,
        request,
        "invite.reset",
        "invite",
        "重置团队二维码与普通邀请链接",
        entity_id=invite.id,
    )
    db.commit()
    return {
        "token": token,
        "url": f"{PUBLIC_TEAM_BASE_URL}/join/{quote(token)}",
        "use_count": 0,
    }


@router.get("/libraries/{library_id}/members")
def list_members(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    rows = (
        db.query(CompetitionLibraryMember)
        .filter(CompetitionLibraryMember.library_id == library_id)
        .order_by(CompetitionLibraryMember.role.asc(), CompetitionLibraryMember.joined_at.asc())
        .all()
    )
    users = {
        user.id: user
        for user in db.query(User)
        .filter(User.id.in_([row.user_id for row in rows]))
        .all()
    } if rows else {}
    return [
        {
            **user_out(users.get(row.user_id), row.role),
            "status": row.status,
            "joined_at": row.joined_at,
            "removed_at": row.removed_at,
            "blocked": bool(row.blocked_at),
        }
        for row in rows
    ]


@router.delete("/libraries/{library_id}/members/{user_id}")
def remove_member(
    library_id: str,
    user_id: int,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    library, _ = require_library_captain(db, library_id, auth)
    if user_id == library.creator_user_id:
        raise HTTPException(status_code=400, detail="不能移除队长")
    member = membership(db, library_id, user_id)
    if not member:
        raise HTTPException(status_code=404, detail="成员不存在")
    before = record_dict(member)
    member.status = "removed"
    member.removed_at = datetime.utcnow()
    member.blocked_at = datetime.utcnow()
    member.blocked_invite_id = member.joined_invite_id or active_invite(db, library_id).id
    frozen_count = freeze_member_components(db, library_id, user_id)
    user = db.get(User, user_id)
    log_action(
        db,
        library_id,
        auth,
        request,
        "member.remove",
        "member",
        f"移除成员 {(user.nickname if user else user_id)}，冻结其提供的 {frozen_count} 条库存镜像",
        entity_id=str(user_id),
        before=before,
        after={**record_dict(member), "frozen_component_count": frozen_count},
    )
    db.commit()
    return {"ok": True}


@router.put("/libraries/{library_id}/members/{user_id}/role")
def update_member_role(
    library_id: str,
    user_id: int,
    payload: dict,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    library, _ = require_library_captain(db, library_id, auth)
    if user_id == library.creator_user_id:
        raise HTTPException(status_code=400, detail="队长角色不能修改")
    role = str(payload.get("role") or "").strip()
    if role not in {"editor", "viewer"}:
        raise HTTPException(status_code=400, detail="角色必须为 editor 或 viewer")
    member = membership(db, library_id, user_id)
    if not member or member.status != "active":
        raise HTTPException(status_code=404, detail="成员不存在")
    before = record_dict(member)
    member.role = role
    log_action(
        db,
        library_id,
        auth,
        request,
        "member.role.update",
        "member",
        f"修改成员角色为 {role}",
        entity_id=str(user_id),
        before=before,
        after=record_dict(member),
    )
    db.commit()
    return {"ok": True, "user_id": user_id, "role": role}


@router.post("/libraries/{library_id}/members/{user_id}/unblock")
def unblock_member(
    library_id: str,
    user_id: int,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_captain(db, library_id, auth)
    member = membership(db, library_id, user_id)
    if not member:
        raise HTTPException(status_code=404, detail="成员不存在")
    before = record_dict(member)
    member.blocked_at = None
    member.status = "removed"
    log_action(
        db,
        library_id,
        auth,
        request,
        "member.unblock",
        "member",
        "解除禁入；该成员需使用重置后的新邀请重新加入",
        entity_id=str(user_id),
        before=before,
        after=record_dict(member),
    )
    db.commit()
    return {"ok": True}


@router.get("/libraries/{library_id}/components")
def list_components(
    library_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(3, ge=1, le=10),
    keyword: str | None = None,
    category: str | None = None,
    location: str | None = None,
    linked: bool | None = None,
    marker_category: str | None = None,
    marker_color: str | None = None,
    flagged: bool | None = None,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _library, member = require_library_member(db, library_id, auth)
    rows = db.query(CompetitionLibraryComponent).filter(
        CompetitionLibraryComponent.library_id == library_id
    ).order_by(
        CompetitionLibraryComponent.created_at.asc(),
        CompetitionLibraryComponent.id.asc(),
    ).all()
    all_items = team_components_out(db, rows, auth, member)
    category_options = sorted({
        (item.get("category") or {}).get("name") or "未分类"
        for item in all_items
    }, key=category_sort_key)
    marker_category_options = sorted({
        marker.get("category")
        for item in all_items
        for marker in item.get("markers") or []
        if marker.get("category")
    })
    marker_color_options = sorted({
        marker.get("color")
        for item in all_items
        for marker in item.get("markers") or []
        if marker.get("color")
    })
    items = list(all_items)
    search = str(keyword or "").strip().lower()
    if search:
        items = [
            item for item in items
            if any(
                search in str(item.get(field) or "").lower()
                for field in ["warehouse_code", "name", "model", "lcsc_number", "location", "tags", "remark", "normalized_spec"]
            )
        ]
    if category:
        items = [item for item in items if (item.get("category") or {}).get("name") == category]
    if location:
        items = [item for item in items if str(location).lower() in str(item.get("location") or "").lower()]
    if linked is True:
        items = [item for item in items if item.get("sync_status") == "live"]
    elif linked is False:
        items = [item for item in items if item.get("sync_status") != "live"]
    if marker_category or marker_color or flagged is not None:
        filtered_items = []
        for item in items:
            markers = item.get("markers") or []
            candidates = [
                marker for marker in markers
                if (not marker_category or marker.get("category") == marker_category)
                and (not marker_color or marker.get("color") == marker_color)
            ]
            if (marker_category or marker_color) and not candidates:
                continue
            if flagged is True and not any(bool(marker.get("flagged")) for marker in candidates):
                continue
            if flagged is False:
                if marker_category or marker_color:
                    if not any(not bool(marker.get("flagged")) for marker in candidates):
                        continue
                elif any(bool(marker.get("flagged")) for marker in markers):
                    continue
            filtered_items.append(item)
        items = filtered_items
    items.sort(
        key=lambda item: component_value_sort_key(
            type("ContestSortItem", (), item)(),
            (item.get("category") or {}).get("name"),
        )
    )
    grouped: dict[str, dict] = {}
    for item in items:
        category_data = item.get("category") or {}
        name = category_data.get("name") or "未分类"
        if name not in grouped:
            grouped[name] = {
                "name": name,
                "color": category_data.get("color") or "#eef2f7",
                "items": [],
            }
        grouped[name]["items"].append(item)
    groups = sorted(grouped.values(), key=lambda group: category_sort_key(group["name"]))
    category_total = len(groups)
    start = (page - 1) * page_size
    page_groups = groups[start : start + page_size]
    page_items = [
        item
        for group in page_groups
        for item in group["items"]
    ]
    return {
        "items": page_items,
        "groups": page_groups,
        "total": len(items),
        "total_quantity": sum(int(item.get("quantity") or 0) for item in items),
        "page": page,
        "page_size": page_size,
        "category_total": category_total,
        "has_more": start + page_size < category_total,
        "filter_options": {
            "categories": category_options,
            "marker_categories": marker_category_options,
            "marker_colors": marker_color_options,
        },
    }


@router.post("/libraries/{library_id}/components/export/label-sheet")
def export_library_component_labels(
    library_id: str,
    payload: ComponentExportRequest | None = None,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _library, member = require_library_captain(db, library_id, auth)
    options = payload or ComponentExportRequest()
    query = (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.library_id == library_id)
    )
    query = apply_label_import_date_filter(query, CompetitionLibraryComponent, options.imported_from, options.imported_to)
    rows = query.order_by(CompetitionLibraryComponent.created_at.asc()).all()
    records = filter_label_records_by_excluded_categories(
        team_components_out(db, rows, auth, member),
        options.excluded_categories,
    )
    printed_at = print_timestamp()
    output_format = (options.output_format or "html").lower()
    if output_format == "pdf":
        appended_pdf_items = [] if options.calibration else team_component_export_custom_label_pdf_items(db, library_id, options.custom_labels, records)
        pdf_doc = render_component_label_pdf(
            records,
            PUBLIC_PERSONAL_BASE_URL,
            start_slot=options.start_slot,
            copies=options.copies,
            offset_x_mm=options.offset_x_mm,
            offset_y_mm=options.offset_y_mm,
            calibration=options.calibration,
            printed_at=printed_at,
            safe_margin=options.safe_margin,
            appended_items=appended_pdf_items,
        )
        return Response(
            content=pdf_doc,
            media_type="application/pdf",
            headers={"Content-Disposition": 'inline; filename="team-component-labels.pdf"'},
        )
    appended_cards = [] if options.calibration else team_component_export_custom_label_cards(db, library_id, options.custom_labels, printed_at, records)
    appended_font_keys = set() if options.calibration else team_component_export_custom_label_font_keys(db, library_id, options.custom_labels)
    html_doc = render_component_label_sheet(
        records,
        PUBLIC_PERSONAL_BASE_URL,
        start_slot=options.start_slot,
        copies=options.copies,
        offset_x_mm=options.offset_x_mm,
        offset_y_mm=options.offset_y_mm,
        calibration=options.calibration,
        appended_cards=appended_cards,
        printed_at=printed_at,
        safe_margin=options.safe_margin,
        font_keys=appended_font_keys,
    )
    return Response(
        content=html_doc,
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": 'inline; filename="team-component-labels.html"'},
    )


@router.get("/libraries/{library_id}/custom-labels", response_model=list[CustomLabelTemplateOut])
def list_team_custom_labels(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    rows = (
        db.query(CustomLabelTemplate)
        .filter(
            CustomLabelTemplate.scope_type == "team",
            CustomLabelTemplate.team_library_id == library_id,
            CustomLabelTemplate.status == "active",
        )
        .order_by(CustomLabelTemplate.updated_at.desc(), CustomLabelTemplate.created_at.desc())
        .all()
    )
    return [custom_label_template_out(row, db) for row in rows]


@router.get("/libraries/{library_id}/custom-labels/category-summary")
def team_custom_label_category_summary(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _library, member = require_library_member(db, library_id, auth)
    items = (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.library_id == library_id)
        .order_by(CompetitionLibraryComponent.created_at.asc())
        .all()
    )
    records = team_components_out(db, items, auth, member)
    summary = category_package_summary_from_records(records)
    return [{"category": category, "summary": value} for category, value in summary.items()]


@router.post("/libraries/{library_id}/custom-labels", response_model=CustomLabelTemplateOut)
def create_team_custom_label(
    library_id: str,
    payload: CustomLabelTemplateCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    template = CustomLabelTemplate(
        id=secrets.token_hex(16),
        scope_type="team",
        owner_user_id=None,
        team_library_id=library_id,
        name=payload.name.strip(),
        content_json=json.dumps(clean_custom_label_content(payload.content), ensure_ascii=False),
        status="active",
        created_by_user_id=auth.user_id,
    )
    db.add(template)
    log_action(db, library_id, auth, request, "custom_label.create", "custom_label", f"新增自定义标签模板 {template.name}", entity_id=template.id)
    db.commit()
    db.refresh(template)
    return custom_label_template_out(template, db)


@router.put("/libraries/{library_id}/custom-labels/{template_id}", response_model=CustomLabelTemplateOut)
def update_team_custom_label(
    library_id: str,
    template_id: str,
    payload: CustomLabelTemplateUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    template = require_team_custom_label_template(db, library_id, template_id)
    if payload.name is not None:
        template.name = payload.name.strip()
    if payload.content is not None:
        template.content_json = json.dumps(clean_custom_label_content(payload.content), ensure_ascii=False)
    log_action(db, library_id, auth, request, "custom_label.update", "custom_label", f"更新自定义标签模板 {template.name}", entity_id=template.id)
    db.commit()
    db.refresh(template)
    return custom_label_template_out(template, db)


@router.delete("/libraries/{library_id}/custom-labels/{template_id}")
def archive_team_custom_label(
    library_id: str,
    template_id: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    template = require_team_custom_label_template(db, library_id, template_id)
    template.status = "archived"
    template.archived_at = datetime.utcnow()
    log_action(db, library_id, auth, request, "custom_label.archive", "custom_label", f"归档自定义标签模板 {template.name}", entity_id=template.id)
    db.commit()
    return {"archived": True}


@router.post("/libraries/{library_id}/custom-labels/{template_id}/assets")
async def upload_team_custom_label_asset(
    library_id: str,
    template_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    require_library_editor(db, library_id, auth)
    template = require_team_custom_label_template(db, library_id, template_id)
    asset = await save_custom_label_asset(db, template, file)
    return custom_label_asset_out(asset, library_id)


@router.get("/libraries/{library_id}/custom-labels/assets/{asset_id}")
def get_team_custom_label_asset(
    library_id: str,
    asset_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    asset = db.get(CustomLabelAsset, asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="标签素材不存在")
    template = require_team_custom_label_template(db, library_id, asset.template_id)
    if template.status != "active":
        raise HTTPException(status_code=404, detail="标签素材不存在")
    path = Path(asset.storage_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="标签素材文件不存在")
    return FileResponse(path, media_type=asset.mime_type, filename=asset.file_name)


@router.post("/libraries/{library_id}/custom-labels/export-sheet")
def export_team_custom_label_sheet(
    library_id: str,
    payload: CustomLabelExportRequest,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    template = require_team_custom_label_template(db, library_id, payload.template_id) if payload.template_id else None
    content = clean_custom_label_content(payload.content if payload.content is not None else parse_custom_label_content(template.content_json if template else None))
    return Response(
        content=render_custom_label_sheet(
            content,
            asset_resolver=custom_label_asset_resolver(db, template) if template else None,
            start_slot=payload.start_slot,
            copies=payload.copies,
            offset_x_mm=payload.offset_x_mm,
            offset_y_mm=payload.offset_y_mm,
            calibration=payload.calibration,
            safe_margin=payload.safe_margin,
        ),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": 'inline; filename="team-custom-labels.html"'},
    )


@router.get("/libraries/{library_id}/components/export/inventory.xlsx")
def export_library_component_inventory(
    library_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _library, member = require_library_member(db, library_id, auth)
    items = (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.library_id == library_id)
        .order_by(CompetitionLibraryComponent.created_at.asc())
        .all()
    )
    rows = team_components_out(db, items, auth, member)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "团队元器件库存"
    sheet.append(
        [
            "团队器件 ID", "器件 ID", "名称", "型号", "LCSC", "数量", "来源成员 ID",
            "同步状态", "位置", "标签", "备注",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row["id"],
                row.get("warehouse_code") or "",
                row.get("name") or "",
                row.get("model") or "",
                row.get("lcsc_number") or "",
                int(row.get("quantity") or 0),
                row.get("source_user_id") or "",
                row.get("sync_status") or "",
                row.get("location") or "",
                row.get("tags") or "",
                row.get("remark") or "",
            ]
        )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="team-component-inventory.xlsx"'},
    )


@router.post("/libraries/{library_id}/components")
def create_component(
    library_id: str,
    payload: ContestComponentCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item, merged = add_or_merge_component(
        db,
        library_id,
        auth,
        request,
        payload.model_dump(),
    )
    db.commit()
    db.refresh(item)
    return {"item": team_component_out(db, item, auth), "merged": merged}


@router.post("/libraries/{library_id}/components/bulk")
def bulk_add_components(
    library_id: str,
    payload: ContestComponentBulkAdd,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    created = merged = 0
    for raw in payload.items[:500]:
        cw = find_cw_component(
            db,
            owner_user_id=auth.user_id,
            cw_component_id=raw.get("cw_component_id"),
        )
        if not cw:
            continue
        _, was_merged = add_or_merge_component(
            db,
            library_id,
            auth,
            request,
            {
                "cw_component_id": cw.id,
                "name": cw.name,
                "model": cw.model,
                "lcsc_number": cw.lcsc_number,
                "quantity": int(cw.quantity or 0),
                "location": raw.get("location"),
                "category": raw.get("category"),
                "tags": raw.get("tags"),
                "remark": raw.get("remark"),
            },
        )
        merged += int(was_merged)
        created += int(not was_merged)
    db.commit()
    return {"created": created, "merged": merged}


@router.post("/libraries/{library_id}/components/import-all-mine")
def retired_import_all_my_components(library_id: str):
    raise HTTPException(
        status_code=410,
        detail="一键导入个人全部器件功能已移除，请按需选择器件加入团队器件库",
    )


@router.put("/libraries/{library_id}/components/{item_id}")
def update_component(
    library_id: str,
    item_id: str,
    payload: ContestComponentUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="物料不存在")
    before = record_dict(item)
    allowed = {"location", "category", "tags", "remark"}
    for key, value in payload.model_dump(exclude_unset=True).items():
        if key not in allowed:
            continue
        if isinstance(value, str):
            value = value.strip() or None
        setattr(item, key, value)
    item.updated_by_user_id = auth.user_id
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.update",
        "component",
        f"更新物料 {item.name}",
        entity_id=item.id,
        before=before,
        after=record_dict(item),
    )
    db.commit()
    db.refresh(item)
    return team_component_out(db, item, auth)


@router.get("/libraries/{library_id}/components/{item_id}/lots", response_model=list[InventoryLotOut])
def list_team_component_lots(
    library_id: str,
    item_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _, component, _ = require_team_source_component(db, library_id, item_id, auth, mutate=False)
    existing_active = (
        db.query(func.count(InventoryLot.id))
        .filter(InventoryLot.component_id == component.id, InventoryLot.status == "active")
        .scalar()
        or 0
    )
    changed = 0
    if int(component.quantity or 0) > 0 and not existing_active:
        ensure_component_lot(db, component)
        changed += 1
    changed += abs(reconcile_component_lots(db, component))
    if changed:
        db.commit()
    rows = (
        db.query(InventoryLot)
        .filter(InventoryLot.component_id == component.id)
        .order_by(InventoryLot.status.asc(), InventoryLot.remaining_quantity.desc(), InventoryLot.received_at.desc(), InventoryLot.created_at.desc())
        .all()
    )
    return [inventory_lot_out(row) for row in rows]


@router.post("/libraries/{library_id}/components/{item_id}/lots", response_model=InventoryLotOut)
def create_team_component_lot(
    library_id: str,
    item_id: str,
    payload: InventoryLotCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    item, component, _ = require_team_source_component(db, library_id, item_id, auth, mutate=True)
    quantity = int(payload.quantity or 0)
    old_quantity = int(component.quantity or 0)
    component.quantity = old_quantity + quantity
    item.quantity = component.quantity
    component.first_stocked_at = component.first_stocked_at or datetime.utcnow()
    component.last_stocked_at = datetime.utcnow()
    movements = record_stock_delta(
        db,
        component,
        quantity,
        movement_type="team_lot_create",
        reason=payload.note or f"团队库 {library_id} 新增渠道库存批次",
        actor_user_id=auth.user_id,
        location=payload.location,
        unit_cost=payload.unit_cost,
        source_type=(payload.source_type or "manual").strip() or "manual",
        source_reference=(payload.source_reference or "").strip() or None,
    )
    lot_id = movements[0].lot_id if movements else None
    item.frozen_snapshot_json = json.dumps(component_snapshot(db, component), ensure_ascii=False, default=str)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.lot.create",
        "component",
        f"新增库存批次 {component.name} x {quantity}",
        entity_id=item.id,
        before={"quantity": old_quantity},
        after={"quantity": component.quantity, "lot_id": lot_id},
    )
    db.commit()
    lot = db.get(InventoryLot, lot_id) if lot_id else None
    if not lot:
        raise HTTPException(status_code=500, detail="库存批次创建失败")
    return inventory_lot_out(lot)


@router.post("/libraries/{library_id}/components/{item_id}/quantity/decrement")
def decrement_team_component_quantity(
    library_id: str,
    item_id: str,
    payload: ComponentConsumeRequest,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    item, component, _ = require_team_source_component(db, library_id, item_id, auth, mutate=True)
    quantity = int(payload.quantity or 1)
    if int(component.quantity or 0) < quantity:
        raise HTTPException(status_code=400, detail="库存不足")
    if payload.lot_id:
        lot = db.get(InventoryLot, payload.lot_id)
        if not lot or lot.component_id != component.id or lot.status != "active":
            raise HTTPException(status_code=404, detail="库存批次不存在")
        if int(lot.remaining_quantity or 0) < quantity:
            raise HTTPException(status_code=400, detail=f"指定库存批次库存不足：需要 {quantity}，剩余 {int(lot.remaining_quantity or 0)}")
    old_quantity = int(component.quantity or 0)
    component.quantity = old_quantity - quantity
    item.quantity = component.quantity
    component.last_outbound_at = datetime.utcnow()
    try:
        record_stock_delta(
            db,
            component,
            -quantity,
            movement_type="team_manual_consume",
            reason=payload.remark or f"团队库 {library_id} 扣减库存",
            actor_user_id=auth.user_id,
            lot_id=payload.lot_id,
        )
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    item.frozen_snapshot_json = json.dumps(component_snapshot(db, component), ensure_ascii=False, default=str)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.quantity.consume",
        "component",
        f"扣减器件 {component.name} x {quantity}",
        entity_id=item.id,
        before={"quantity": old_quantity},
        after={"quantity": component.quantity, "lot_id": payload.lot_id},
    )
    db.commit()
    db.refresh(item)
    return team_component_out(db, item, auth)


@router.post("/libraries/{library_id}/components/{item_id}/ai/ask", response_model=ComponentAiAskOut)
def ask_team_component_ai(
    library_id: str,
    item_id: str,
    payload: ComponentAiAskRequest,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    item, component, _ = require_team_source_component(db, library_id, item_id, auth, mutate=False)
    context = {
        "team_library_id": library_id,
        "team_component_id": item.id,
        "team_location": item.location,
        "team_tags": item.tags,
        "team_remark": item.remark,
        "inventory_lots": [
            inventory_lot_out(row)
            for row in db.query(InventoryLot)
            .filter(InventoryLot.component_id == component.id, InventoryLot.status == "active")
            .order_by(InventoryLot.remaining_quantity.desc(), InventoryLot.received_at.desc())
            .limit(8)
            .all()
        ],
    }
    try:
        result = component_question(
            component_to_dict(component),
            payload.question.strip(),
            context,
            "auto" if payload.use_web_search else "off",
        )
    except MimoNotConfiguredError as error:
        raise HTTPException(status_code=503, detail="AI 服务尚未配置") from error
    except MimoRequestError as error:
        raise HTTPException(status_code=502, detail=f"AI 查询失败：{error}") from error
    return {
        "answer": result.get("answer") or "当前资料不足，无法给出可靠回答。",
        "confidence": result.get("confidence") or "medium",
        "evidence": result.get("evidence") if isinstance(result.get("evidence"), list) else [],
        "risks": result.get("risks") if isinstance(result.get("risks"), list) else [],
        "needs_datasheet_check": bool(result.get("needs_datasheet_check", True)),
        "sources": result.get("sources") or [],
    }


@router.patch("/libraries/{library_id}/components/{item_id}/quantity")
def update_component_quantity(
    library_id: str,
    item_id: str,
    payload: TeamComponentQuantityUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    _, member = require_library_editor(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="器件不存在")
    if item.sync_status != "live" or not item.cw_component_id:
        raise HTTPException(status_code=409, detail="冻结器件需要重新绑定后才能修改数量")
    component = db.get(Component, item.cw_component_id)
    if not component or component.owner_user_id != item.source_user_id:
        raise HTTPException(status_code=409, detail="来源器件不可用，请重新绑定")
    if item.source_user_id != auth.user_id and member.role != "captain":
        raise HTTPException(status_code=403, detail="仅来源成员或队长可以修改数量")

    old_quantity = int(component.quantity or 0)
    new_quantity = int(payload.quantity)
    component.quantity = new_quantity
    item.quantity = new_quantity
    item.updated_by_user_id = auth.user_id
    if new_quantity > old_quantity:
        component.first_stocked_at = component.first_stocked_at or datetime.utcnow()
        component.last_stocked_at = datetime.utcnow()
    elif new_quantity < old_quantity:
        component.last_outbound_at = datetime.utcnow()
    item.frozen_snapshot_json = json.dumps(
        component_snapshot(db, component),
        ensure_ascii=False,
        default=str,
    )
    delta = new_quantity - old_quantity
    record_stock_delta(
        db,
        component,
        delta,
        movement_type="team_adjustment",
        reason=payload.remark or f"团队库 {library_id} 调整库存",
        actor_user_id=auth.user_id,
    )
    db.add(
        ActivityLog(
            owner_user_id=component.owner_user_id,
            action="component.quantity.team_update",
            entity_type="component",
            entity_id=component.id,
            component_id=component.id,
            quantity_delta=delta,
            summary=f"团队版修改 {component.name} 库存数量：{old_quantity} -> {new_quantity}",
            detail=json_value({
                "library_id": library_id,
                "team_component_id": item.id,
                "actor_user_id": auth.user_id,
                "remark": payload.remark,
            }),
        )
    )
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.quantity.update",
        "component",
        f"修改器件 {component.name} 数量：{old_quantity} -> {new_quantity}",
        entity_id=item.id,
        before={"quantity": old_quantity},
        after={"quantity": new_quantity, "remark": payload.remark},
    )
    db.commit()
    db.refresh(item)
    return team_component_out(db, item, auth)


@router.get("/libraries/{library_id}/components/{item_id}/usage-records")
def team_component_usage_records(
    library_id: str,
    item_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
    limit: int = Query(100, ge=1, le=300),
):
    require_library_member(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="器件不存在")
    if not item.cw_component_id:
        return []
    rows = (
        db.query(ProjectBomItem, Project)
        .join(Project, Project.id == ProjectBomItem.project_id)
        .filter(
            ProjectBomItem.component_id == item.cw_component_id,
            Project.scope_type == "team",
            Project.team_library_id == library_id,
        )
        .order_by(Project.updated_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": -bom_item.id,
            "action": "project.bom",
            "action_label": "团队项目 BOM",
            "project_id": project.id,
            "project_code": project.project_code,
            "project_name": project.name,
            "quantity_delta": bom_item.required_quantity,
            "designators": [
                value.strip()
                for value in re.split(r"[,，、\s]+", str(bom_item.remark or ""))
                if value.strip() and not value.startswith("BOM") and ":" not in value
            ],
            "summary": f"团队项目 BOM 需求 {bom_item.required_quantity}",
            "created_at": project.updated_at or project.created_at,
        }
        for bom_item, project in rows
    ]


@router.get("/libraries/{library_id}/components/{item_id}/markers")
def list_component_markers(
    library_id: str,
    item_id: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="器件不存在")
    rows = (
        db.query(CompetitionComponentMarker)
        .filter(CompetitionComponentMarker.component_id == item_id)
        .order_by(CompetitionComponentMarker.flagged.desc(), CompetitionComponentMarker.created_at.desc())
        .all()
    )
    return [marker_out(db, row) for row in rows]


@router.post("/libraries/{library_id}/components/{item_id}/markers")
def create_component_marker(
    library_id: str,
    item_id: str,
    payload: TeamMarkerCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="器件不存在")
    marker = CompetitionComponentMarker(
        id=new_uuid(),
        library_id=library_id,
        component_id=item_id,
        category=payload.category.strip(),
        color=payload.color.upper(),
        flagged=payload.flagged,
        note=(payload.note or "").strip() or None,
        created_by_user_id=auth.user_id,
        updated_by_user_id=auth.user_id,
    )
    db.add(marker)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.marker.create",
        "component_marker",
        f"给器件 {item.name} 添加标记：{marker.category}",
        entity_id=marker.id,
        after=record_dict(marker),
    )
    db.commit()
    db.refresh(marker)
    return marker_out(db, marker)


@router.put("/libraries/{library_id}/components/{item_id}/markers/{marker_id}")
def update_component_marker(
    library_id: str,
    item_id: str,
    marker_id: str,
    payload: TeamMarkerUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    marker = db.get(CompetitionComponentMarker, marker_id)
    if not marker or marker.library_id != library_id or marker.component_id != item_id:
        raise HTTPException(status_code=404, detail="标记不存在")
    before = record_dict(marker)
    for key, value in payload.model_dump(exclude_unset=True).items():
        if isinstance(value, str):
            value = value.strip()
        if key == "color" and value:
            value = value.upper()
        setattr(marker, key, value or None if key in {"category", "note"} else value)
    if not marker.category:
        raise HTTPException(status_code=400, detail="标记分类不能为空")
    marker.updated_by_user_id = auth.user_id
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.marker.update",
        "component_marker",
        f"更新器件标记：{marker.category}",
        entity_id=marker.id,
        before=before,
        after=record_dict(marker),
    )
    db.commit()
    db.refresh(marker)
    return marker_out(db, marker)


@router.delete("/libraries/{library_id}/components/{item_id}/markers/{marker_id}")
def delete_component_marker(
    library_id: str,
    item_id: str,
    marker_id: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    marker = db.get(CompetitionComponentMarker, marker_id)
    if not marker or marker.library_id != library_id or marker.component_id != item_id:
        raise HTTPException(status_code=404, detail="标记不存在")
    before = record_dict(marker)
    db.delete(marker)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.marker.delete",
        "component_marker",
        f"删除器件标记：{marker.category}",
        entity_id=marker.id,
        before=before,
    )
    db.commit()
    return {"ok": True}


@router.delete("/libraries/{library_id}/components/{item_id}")
def delete_component(
    library_id: str,
    item_id: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="物料不存在")
    before = record_dict(item)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.delete",
        "component",
        f"删除物料 {item.name}",
        entity_id=item.id,
        before=before,
    )
    db.query(CompetitionComponentMarker).filter(
        CompetitionComponentMarker.component_id == item.id
    ).delete(synchronize_session=False)
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.post("/libraries/{library_id}/components/{item_id}/link")
def link_component(
    library_id: str,
    item_id: str,
    payload: ContestComponentLink,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionLibraryComponent, item_id)
    cw = find_cw_component(
        db,
        owner_user_id=auth.user_id,
        cw_component_id=payload.cw_component_id,
    )
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="物料不存在")
    if not cw:
        raise HTTPException(status_code=404, detail="个人版器件不存在")
    existing = find_existing_team_component(
        db,
        library_id,
        cw_component_id=cw.id,
        exclude_id=item.id,
    )
    if existing:
        before = {"source": record_dict(item), "target": record_dict(existing)}
        existing.location = existing.location or item.location
        existing.remark = append_text(existing.remark, item.remark)
        existing.updated_by_user_id = auth.user_id
        db.query(CompetitionComponentMarker).filter(
            CompetitionComponentMarker.component_id == item.id
        ).update(
            {
                CompetitionComponentMarker.component_id: existing.id,
                CompetitionComponentMarker.library_id: library_id,
            },
            synchronize_session=False,
        )
        db.delete(item)
        log_action(
            db,
            library_id,
            auth,
            request,
            "component.link_merge",
            "component",
            f"重新绑定到 {cw.warehouse_code or cw.id} 并移除重复团队条目",
            entity_id=existing.id,
            before=before,
            after=record_dict(existing),
        )
        db.commit()
        db.refresh(existing)
        return {"item": team_component_out(db, existing, auth), "merged": True}

    before = record_dict(item)
    item.cw_component_id = cw.id
    item.source_user_id = auth.user_id
    item.sync_status = "live"
    item.warehouse_code_snapshot = cw.warehouse_code
    item.frozen_snapshot_json = json.dumps(component_snapshot(db, cw), ensure_ascii=False, default=str)
    item.name = cw.name
    item.model = cw.model
    item.lcsc_number = cw.lcsc_number
    item.quantity = int(cw.quantity or 0)
    item.updated_by_user_id = auth.user_id
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.link",
        "component",
        f"重新绑定个人版器件 {cw.warehouse_code or cw.id} {cw.name}",
        entity_id=item.id,
        before=before,
        after=record_dict(item),
    )
    db.commit()
    db.refresh(item)
    return {"item": team_component_out(db, item, auth), "merged": False}


@router.post("/libraries/{library_id}/components/{item_id}/rebind")
def rebind_component(
    library_id: str,
    item_id: str,
    payload: ContestComponentRebind,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_captain(db, library_id, auth)
    return link_component(
        library_id=library_id,
        item_id=item_id,
        payload=ContestComponentLink(cw_component_id=payload.cw_component_id),
        request=request,
        auth=auth,
        db=db,
    )


@router.get("/cw-components")
def search_cw_components(
    keyword: str | None = None,
    personal_only: bool = False,
    limit: int = Query(50, ge=1, le=200),
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    query = db.query(Component).filter(Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
    if keyword:
        filters = []
        for variant in keyword_unit_variants(keyword):
            like = f"%{variant}%"
            filters.extend(
                [
                    Component.name.ilike(like),
                    Component.model.ilike(like),
                    Component.lcsc_number.ilike(like),
                    Component.warehouse_code.ilike(like),
                    Component.parameters.ilike(like),
                    Component.normalized_spec.ilike(like),
                ]
            )
        if filters:
            query = query.filter(or_(*filters))
    items = query.order_by(Component.id.asc()).limit(limit).all()
    reserved = reserved_quantities(db, [item.id for item in items])
    return [cw_component_out(item, reserved.get(item.id, 0), keyword) for item in items]


@router.get("/resolve-code")
def resolve_cw_code(
    code: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    raw = str(code or "").strip()
    tail = raw.rstrip("/").split("/")[-1]
    candidates = [raw, tail]
    item = None
    for value in candidates:
        if value.isdigit():
            item = (
                db.query(Component)
                .filter(Component.id == int(value), Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
                .first()
            )
        if not item:
            item = (
                db.query(Component)
                .filter(Component.warehouse_code == value, Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
                .first()
            )
        if not item:
            item = (
                db.query(Component)
                .filter(Component.lcsc_number == value, Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
                .first()
            )
        if item:
            break
    if not item:
        raise HTTPException(status_code=404, detail="没有找到这个 CW 物料")
    reserved = reserved_quantities(db, [item.id]).get(item.id, 0)
    return cw_component_out(item, reserved)


IMPORT_HEADERS = {
    "name": ["名称", "物料名称", "name"],
    "model": ["型号", "model"],
    "lcsc_number": ["立创 ID", "立创编号", "LCSC", "lcsc_number"],
    "quantity": ["数量", "quantity"],
    "location": ["位置", "location"],
    "category": ["分类", "category"],
    "tags": ["标签", "tags"],
    "remark": ["备注", "remark"],
}


def normalized_import_row(row: dict) -> dict:
    lowered = {str(key or "").strip().lower(): value for key, value in row.items()}
    result = {}
    for target, aliases in IMPORT_HEADERS.items():
        for alias in aliases:
            key = alias.lower()
            if key in lowered:
                result[target] = lowered[key]
                break
    try:
        result["quantity"] = max(0, int(float(result.get("quantity") or 0)))
    except (TypeError, ValueError):
        result["quantity"] = 0
    for key in ["name", "model", "lcsc_number", "location", "category", "tags", "remark"]:
        result[key] = str(result.get(key) or "").strip() or None
    return result


def parse_import_file(filename: str, content: bytes) -> list[dict]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text_value = content.decode("utf-8-sig")
        return [
            normalized_import_row(row)
            for row in csv.DictReader(io.StringIO(text_value))
        ]
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        normalized_import_row(dict(zip(headers, values)))
        for values in rows[1:]
        if any(value not in (None, "") for value in values)
    ]


@router.get("/import-template.csv")
def download_import_template():
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["名称", "型号", "立创 ID", "数量", "位置", "分类", "标签", "备注"])
    writer.writerow(["示例运放", "OPA2333", "", 2, "工具箱", "模拟前端", "低功耗", ""])
    content = "\ufeff" + buffer.getvalue()
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="contest-components-template.csv"'},
    )


@router.post("/libraries/{library_id}/components/import")
async def import_components(
    library_id: str,
    request: Request,
    file: UploadFile = File(...),
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    filename = file.filename or ""
    if not filename.lower().endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="表格不能超过 5MB")
    try:
        rows = parse_import_file(filename, content)
    except Exception as error:
        raise HTTPException(status_code=400, detail="表格无法读取") from error
    created = merged = skipped = 0
    for row in rows[:1000]:
        if not row.get("name") and not row.get("lcsc_number"):
            skipped += 1
            continue
        _, was_merged = add_or_merge_component(db, library_id, auth, request, row)
        merged += int(was_merged)
        created += int(not was_merged)
    log_action(
        db,
        library_id,
        auth,
        request,
        "component.import",
        "import",
        f"导入表格：新增 {created}，合并 {merged}，跳过 {skipped}",
        after={"created": created, "merged": merged, "skipped": skipped, "file": filename},
    )
    db.commit()
    return {"created": created, "merged": merged, "skipped": skipped}


@router.get("/libraries/{library_id}/pcbs")
def list_pcbs(
    library_id: str,
    keyword: str | None = None,
    status: str | None = None,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    query = db.query(CompetitionPcb).filter(CompetitionPcb.library_id == library_id)
    if keyword:
        like = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                CompetitionPcb.name.ilike(like),
                CompetitionPcb.main_chip.ilike(like),
                CompetitionPcb.function_desc.ilike(like),
                CompetitionPcb.suitable_task.ilike(like),
                CompetitionPcb.location.ilike(like),
            )
        )
    if status:
        query = query.filter(CompetitionPcb.status == status)
    items = query.order_by(CompetitionPcb.name.asc(), CompetitionPcb.created_at.asc()).all()
    return {"items": [pcb_out(item) for item in items], "total": len(items)}


@router.post("/libraries/{library_id}/pcbs")
def create_pcb(
    library_id: str,
    payload: ContestPcbCreate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    if payload.status not in PCB_STATUSES:
        raise HTTPException(status_code=400, detail="PCB 状态不正确")
    item = CompetitionPcb(
        id=new_uuid(),
        library_id=library_id,
        created_by_user_id=auth.user_id,
        updated_by_user_id=auth.user_id,
        **payload.model_dump(),
    )
    db.add(item)
    log_action(
        db,
        library_id,
        auth,
        request,
        "pcb.create",
        "pcb",
        f"新增 PCB {item.name}",
        entity_id=item.id,
        after=record_dict(item),
    )
    db.commit()
    db.refresh(item)
    return pcb_out(item)


@router.put("/libraries/{library_id}/pcbs/{pcb_id}")
def update_pcb(
    library_id: str,
    pcb_id: str,
    payload: ContestPcbUpdate,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionPcb, pcb_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="PCB 不存在")
    before = record_dict(item)
    values = payload.model_dump(exclude_unset=True)
    if values.get("status") not in {None, *PCB_STATUSES}:
        raise HTTPException(status_code=400, detail="PCB 状态不正确")
    for key, value in values.items():
        if isinstance(value, str):
            value = value.strip() or None
        setattr(item, key, value)
    item.updated_by_user_id = auth.user_id
    log_action(
        db,
        library_id,
        auth,
        request,
        "pcb.update",
        "pcb",
        f"更新 PCB {item.name}",
        entity_id=item.id,
        before=before,
        after=record_dict(item),
    )
    db.commit()
    db.refresh(item)
    return pcb_out(item)


def remove_media_file(relative_path: str | None) -> None:
    if not relative_path:
        return
    target = (TEAM_MEDIA_ROOT / relative_path).resolve()
    root = TEAM_MEDIA_ROOT.resolve()
    if root in target.parents and target.exists():
        target.unlink()


def clear_team_media() -> None:
    if TEAM_MEDIA_ROOT.exists():
        shutil.rmtree(TEAM_MEDIA_ROOT)


@router.delete("/libraries/{library_id}/pcbs/{pcb_id}")
def delete_pcb(
    library_id: str,
    pcb_id: str,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionPcb, pcb_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="PCB 不存在")
    before = record_dict(item)
    remove_media_file(item.front_image_path)
    remove_media_file(item.back_image_path)
    log_action(
        db,
        library_id,
        auth,
        request,
        "pcb.delete",
        "pcb",
        f"删除 PCB {item.name}",
        entity_id=item.id,
        before=before,
    )
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.post("/libraries/{library_id}/pcbs/{pcb_id}/images/{side}")
async def upload_pcb_image(
    library_id: str,
    pcb_id: str,
    side: str,
    request: Request,
    file: UploadFile = File(...),
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    if side not in {"front", "back"}:
        raise HTTPException(status_code=400, detail="图片位置不正确")
    item = db.get(CompetitionPcb, pcb_id)
    if not item or item.library_id != library_id:
        raise HTTPException(status_code=404, detail="PCB 不存在")
    content_type = str(file.content_type or "").lower()
    extension = {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
    }.get(content_type)
    if not extension:
        raise HTTPException(status_code=400, detail="仅支持 JPG、PNG 或 WebP")
    content = await file.read()
    if not content or len(content) > MAX_PCB_IMAGE_BYTES:
        raise HTTPException(status_code=413, detail="单张图片不能超过 2MB")
    folder = TEAM_MEDIA_ROOT / library_id / pcb_id
    folder.mkdir(parents=True, exist_ok=True)
    relative_path = f"{library_id}/{pcb_id}/{side}{extension}"
    target = TEAM_MEDIA_ROOT / relative_path
    old_path = item.front_image_path if side == "front" else item.back_image_path
    remove_media_file(old_path)
    target.write_bytes(content)
    before = record_dict(item)
    if side == "front":
        item.front_image_path = relative_path
    else:
        item.back_image_path = relative_path
    item.updated_by_user_id = auth.user_id
    log_action(
        db,
        library_id,
        auth,
        request,
        "pcb.image",
        "pcb",
        f"更新 PCB {item.name} {'正面' if side == 'front' else '背面'}图片",
        entity_id=item.id,
        before=before,
        after=record_dict(item),
    )
    db.commit()
    db.refresh(item)
    return pcb_out(item)


@router.get("/libraries/{library_id}/pcbs/{pcb_id}/images/{side}")
def get_pcb_image(
    library_id: str,
    pcb_id: str,
    side: str,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    item = db.get(CompetitionPcb, pcb_id)
    if not item or item.library_id != library_id or side not in {"front", "back"}:
        raise HTTPException(status_code=404, detail="图片不存在")
    relative_path = item.front_image_path if side == "front" else item.back_image_path
    if not relative_path:
        raise HTTPException(status_code=404, detail="图片不存在")
    target = (TEAM_MEDIA_ROOT / relative_path).resolve()
    if TEAM_MEDIA_ROOT.resolve() not in target.parents or not target.exists():
        raise HTTPException(status_code=404, detail="图片不存在")
    return FileResponse(target, headers={"Cache-Control": "private, max-age=300"})


@router.get("/libraries/{library_id}/logs")
def list_logs(
    library_id: str,
    limit: int = Query(100, ge=1, le=300),
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    rows = (
        db.query(CompetitionActivityLog)
        .filter(CompetitionActivityLog.library_id == library_id)
        .order_by(CompetitionActivityLog.created_at.desc(), CompetitionActivityLog.id.desc())
        .limit(limit)
        .all()
    )
    return [record_dict(row) for row in rows]


@router.post("/libraries/{library_id}/usage-events")
def record_team_usage_event(
    library_id: str,
    payload: UsageEventRequest,
    request: Request,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_member(db, library_id, auth)
    event = re.sub(r"[^a-zA-Z0-9_.:-]+", ".", payload.event).strip(".")[:70]
    if not event.startswith("ui."):
        event = f"ui.{event}"[:80]
    log_action(
        db,
        library_id,
        auth,
        request,
        event,
        payload.target_type or "ui",
        f"界面操作：{event}",
        entity_id=str(payload.target_id) if payload.target_id is not None else None,
        after={
            "page": payload.page,
            "entry": payload.entry,
            "target_id": payload.target_id,
            "viewport_width": payload.viewport_width,
            "viewport_height": payload.viewport_height,
            "detail": payload.detail or {},
        },
    )
    db.commit()
    return {"ok": True}


def ai_component_payload(db: Session, item: CompetitionLibraryComponent) -> dict:
    output = team_component_out(db, item)
    cw = output.get("cw_component") or {}
    return {
        "id": item.id,
        "cw_component_id": item.cw_component_id,
        "warehouse_code": output.get("warehouse_code"),
        "name": output.get("name"),
        "model": output.get("model"),
        "lcsc_number": output.get("lcsc_number"),
        "quantity": output.get("quantity"),
        "location": item.location,
        "category": (output.get("category") or {}).get("name") or item.category,
        "tags": item.tags,
        "remark": item.remark,
        "cw_ai_summary": cw.get("ai_summary"),
        "cw_ai_usage": cw.get("ai_usage"),
        "cw_ai_risk_notes": cw.get("ai_risk_notes"),
    }


def contest_search_terms(prompt: str) -> list[str]:
    normalized = str(prompt or "").strip().lower()
    ascii_terms = re.findall(r"[a-z0-9_.+\-/]{2,}", normalized)
    chinese_runs = re.findall(r"[\u4e00-\u9fff]{2,}", normalized)
    chinese_terms = []
    for run in chinese_runs:
        chinese_terms.append(run)
        chinese_terms.extend(run[index : index + 2] for index in range(len(run) - 1))
    return list(dict.fromkeys([*ascii_terms, *chinese_terms]))


def locally_rank_components(
    db: Session,
    rows: list[CompetitionLibraryComponent],
    prompt: str,
    limit: int = 80,
) -> list[CompetitionLibraryComponent]:
    terms = contest_search_terms(prompt)
    if not terms:
        return rows[:limit]
    scored = []
    for row in rows:
        output = team_component_out(db, row)
        category = output.get("category") or {}
        primary = " ".join(
            str(value or "").lower()
            for value in [
                output.get("warehouse_code"),
                output.get("name"),
                output.get("model"),
                output.get("normalized_spec"),
                output.get("lcsc_number"),
                category.get("name") if isinstance(category, dict) else category,
                row.tags,
            ]
        )
        secondary = " ".join(
            str(value or "").lower()
            for value in [row.location, row.remark]
        )
        score = sum(3 for term in terms if term in primary)
        score += sum(1 for term in terms if term in secondary)
        if score:
            scored.append((score, row.updated_at or row.created_at, row))
    if not scored:
        return rows[:limit]
    scored.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [item[2] for item in scored[:limit]]


def cw_ai_candidate(component: Component) -> dict:
    return {
        "cw_component_id": component.id,
        "warehouse_code": component.warehouse_code,
        "name": component.name,
        "model": component.model,
        "lcsc_number": component.lcsc_number,
        "parameters": component.parameters,
        "package": component.package,
        "ai_summary": component.ai_summary,
    }


@router.post("/libraries/{library_id}/ai")
def contest_ai(
    library_id: str,
    payload: ContestAiRequest,
    auth: AuthContext = Depends(require_access),
    db: Session = Depends(get_db),
):
    require_library_editor(db, library_id, auth)
    if payload.query_type not in {
        "find_components",
        "match_cw",
        "alternatives",
        "recommend_pcbs",
    }:
        raise HTTPException(status_code=400, detail="AI 查询类型不正确")
    all_components = (
        db.query(CompetitionLibraryComponent)
        .filter(CompetitionLibraryComponent.library_id == library_id)
        .order_by(CompetitionLibraryComponent.updated_at.desc())
        .limit(160)
        .all()
    )
    components = locally_rank_components(db, all_components, payload.prompt, limit=80)
    pcbs = (
        db.query(CompetitionPcb)
        .filter(CompetitionPcb.library_id == library_id)
        .order_by(CompetitionPcb.updated_at.desc())
        .limit(100)
        .all()
    )
    component_payload = [ai_component_payload(db, item) for item in components]
    pcb_payload = [pcb_out(item) for item in pcbs]
    cw_candidates = []
    if payload.query_type == "match_cw":
        exact_lcsc = re.search(r"\bC\d{4,}\b", payload.prompt.strip(), flags=re.IGNORECASE)
        if exact_lcsc:
            exact = (
                db.query(Component)
                .filter(
                    func.lower(Component.lcsc_number) == exact_lcsc.group(0).lower(),
                    Component.owner_user_id == auth.user_id,
                    Component.revoked_at.is_(None),
                )
                .first()
            )
            if exact:
                unique_code = exact.warehouse_code or f"CW-{exact.id:08d}"
                result = {
                    "query_type": payload.query_type,
                    "prompt": payload.prompt.strip(),
                    "answer": f"立创 ID 完全一致，确定匹配 {unique_code} {exact.name}。",
                    "component_matches": [
                        {
                            "id": None,
                            "cw_component_id": exact.id,
                            "reason": "立创 ID 完全一致",
                            "warning": "",
                        }
                    ],
                    "pcb_matches": [],
                    "next_steps": ["请人工确认后执行关联"],
                    "requires_confirmation": True,
                    "deterministic": True,
                }
                return {"cached": False, "result": result}
        candidate_query = db.query(Component).filter(Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
        terms = contest_search_terms(payload.prompt)
        if terms:
            filters = []
            for term in terms[:8]:
                like = f"%{term}%"
                filters.extend(
                    [
                        Component.name.ilike(like),
                        Component.model.ilike(like),
                        Component.lcsc_number.ilike(like),
                        Component.parameters.ilike(like),
                    ]
                )
            candidate_query = candidate_query.filter(or_(*filters))
        cw_candidates = [
            cw_ai_candidate(item)
            for item in candidate_query.order_by(Component.updated_at.desc()).limit(40).all()
        ]
    fingerprint = hashlib.sha256(
        json.dumps(
            {
                "query_type": payload.query_type,
                "prompt": payload.prompt.strip(),
                "components": component_payload,
                "pcbs": pcb_payload,
                "cw_candidates": cw_candidates,
            },
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        ).encode("utf-8")
    ).hexdigest()
    cached = (
        db.query(CompetitionAiResult)
        .filter(
            CompetitionAiResult.library_id == library_id,
            CompetitionAiResult.query_type == payload.query_type,
            CompetitionAiResult.input_hash == fingerprint,
        )
        .first()
    )
    if cached and not payload.force:
        return {"cached": True, "result": json.loads(cached.result_json)}
    try:
        result = contest_library_assist(
            payload.query_type,
            payload.prompt.strip(),
            component_payload,
            pcb_payload,
            cw_candidates,
        )
    except MimoNotConfiguredError as error:
        raise HTTPException(status_code=503, detail="AI 服务尚未配置") from error
    except MimoRequestError as error:
        raise HTTPException(status_code=502, detail=f"AI 查询失败：{error}") from error
    if cached:
        cached.prompt_text = payload.prompt.strip()
        cached.result_json = json.dumps(result, ensure_ascii=False)
        cached.created_by_user_id = auth.user_id
    else:
        cached = CompetitionAiResult(
            id=new_uuid(),
            library_id=library_id,
            query_type=payload.query_type,
            input_hash=fingerprint,
            prompt_text=payload.prompt.strip(),
            result_json=json.dumps(result, ensure_ascii=False),
            created_by_user_id=auth.user_id,
        )
        db.add(cached)
    db.commit()
    return {"cached": False, "result": result}
