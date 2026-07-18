import csv
from dataclasses import dataclass
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session
import xlrd

from ..models import Component, SupplierPart
from .component_normalizer import clean_lcsc_keyword
from .mimo_ai import component_to_dict, lcsc_search_url
from .inventory import reserved_quantities


HEADER_ALIASES = {
    "no": ["no.", "no", "序号", "编号"],
    "quantity": ["quantity", "qty", "数量", "用量"],
    "comment": ["comment", "备注", "注释", "描述"],
    "designator": ["designator", "refdes", "位号", "器件位号"],
    "footprint": ["footprint", "package", "封装", "封装格式"],
    "value": ["value", "参数", "阻值", "容值"],
    "manufacturer_part": ["manufacturer part", "mpn", "型号", "制造商型号", "商品型号"],
    "manufacturer": ["manufacturer", "brand", "品牌", "制造商"],
    "supplier_part": ["supplier part", "lcsc", "lcsc part", "立创 ID", "立创编号", "商品编号", "物料编号"],
    "supplier": ["supplier", "供应商"],
    "category": ["category", "分类", "商品类型"],
    "primary_category": ["primary category", "一级分类", "主分类"],
    "mounting_style": ["mounting style", "安装方式"],
    "layer": ["layer", "层"],
}


@dataclass
class BomRow:
    source_row: int
    data: dict[str, Any]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return default


def _normalize(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9.\u4e00-\u9fff]+", "", value.lower())


def _normalize_package(value: str | None) -> str:
    text = _normalize(value)
    text = re.sub(r"^[crl]", "", text)
    return text


def _supplier_matches(expected: str, actual: str | None) -> bool:
    if not expected:
        return True
    return _normalize(expected) == _normalize(actual)


def _lcsc_supplier(value: str | None) -> bool:
    normalized = _normalize(value)
    return not normalized or normalized in {"lcsc", "立创", "立创商城", "jlcpcb"}


def _scope_supplier_query(
    query,
    scope_type: str | None,
    owner_user_id: int | None,
    team_library_id: str | None,
):
    if not scope_type:
        return query
    query = query.filter(SupplierPart.scope_type == scope_type)
    if scope_type == "team":
        return query.filter(SupplierPart.team_library_id == team_library_id)
    return query.filter(SupplierPart.owner_user_id == owner_user_id)


def _pin_counts(*values: str | None) -> set[int]:
    text = " ".join(str(value or "") for value in values).lower()
    counts: set[int] = set()
    for match in re.finditer(r"(?<![a-z0-9])(\d{1,3})\s*(?:pin|pins|p|针|位)(?![a-z0-9])", text):
        count = int(match.group(1))
        if 1 <= count <= 200:
            counts.add(count)
    for match in re.finditer(r"(?:x|×)\s*(\d{1,3})\s*p(?![a-z0-9])", text):
        count = int(match.group(1))
        if 1 <= count <= 200:
            counts.add(count)
    return counts


def _connectorish(*values: str | None) -> bool:
    text = " ".join(str(value or "") for value in values).lower()
    return any(
        token in text
        for token in [
            "type-c",
            "typec",
            "usb",
            "connector",
            "socket",
            "header",
            "hdr",
            "排针",
            "排母",
            "座",
            "接口",
            "连接器",
            "端子",
        ]
    )


def _passive_kind(row: dict[str, Any] | None = None, component: Component | None = None) -> str | None:
    parts = []
    if row:
        parts.extend(str(row.get(key) or "") for key in ["primary_category", "category", "comment", "value", "manufacturer_part"])
        designator = str(row.get("designator") or "").strip().upper()
        if re.match(r"^R\d", designator):
            parts.append("电阻")
        elif re.match(r"^C\d", designator):
            parts.append("电容")
        elif re.match(r"^L\d", designator):
            parts.append("电感")
    if component:
        parts.extend([component.category.name if component.category else "", component.name or "", component.parameters or "", component.tags or ""])
    text = " ".join(parts).lower()
    if any(token in text for token in ["电阻", "resistor", "ohm", "Ω", "ω", "kohm", "mohm"]):
        return "resistance"
    if any(token in text for token in ["电容", "capacitor", "uf", "µf", "nf", "pf"]):
        return "capacitance"
    if any(token in text for token in ["电感", "inductor", "uh", "µh", "mh"]):
        return "inductance"
    return None


def _parse_passive_value(text: str | None, kind: str | None) -> float | None:
    if not text or not kind:
        return None
    value = (
        str(text)
        .lower()
        .replace("µ", "u")
        .replace("μ", "u")
        .replace("ω", "ohm")
        .replace("Ω", "ohm")
    )
    if kind == "resistance":
        explicit = re.search(r"(?<![a-z0-9.])(\d+(?:\.\d+)?)\s*(megohm|mohm|kohm|ohm)(?![a-z0-9])", value)
        if explicit:
            number = float(explicit.group(1))
            unit = explicit.group(2)
            if unit in {"mohm", "megohm"}:
                return number * 1_000_000
            if unit == "kohm":
                return number * 1_000
            return number
        code = re.search(r"(?<![a-z0-9.])(\d+(?:\.\d+)?)([rkm])(\d*)(?![a-z0-9])", value)
        if code and "ohm" not in value:
            whole, marker, frac = code.groups()
            number = float(f"{whole}.{frac}") if frac else float(whole)
            return number * {"r": 1, "k": 1_000, "m": 1_000_000}[marker]
        return None
    if kind == "capacitance":
        match = re.search(r"(?<![a-z0-9.])(\d+(?:\.\d+)?)\s*(uf|nf|pf|f)(?![a-z0-9])", value)
        if not match:
            return None
        number = float(match.group(1))
        return number * {"f": 1, "uf": 1e-6, "nf": 1e-9, "pf": 1e-12}[match.group(2)]
    if kind == "inductance":
        match = re.search(r"(?<![a-z0-9.])(\d+(?:\.\d+)?)\s*(uh|mh|nh|h)(?![a-z0-9])", value)
        if not match:
            return None
        number = float(match.group(1))
        return number * {"h": 1, "mh": 1e-3, "uh": 1e-6, "nh": 1e-9}[match.group(2)]
    return None


def _component_passive_value(component: Component, kind: str | None) -> float | None:
    texts = [component.normalized_spec, component.parameters, component.name, component.tags]
    for text in texts:
        parsed = _parse_passive_value(text, kind)
        if parsed is not None:
            return parsed
    return None


def _format_passive_value(value: float, kind: str) -> str:
    if kind == "resistance":
        if value >= 1_000_000:
            return f"{value / 1_000_000:g}MΩ"
        if value >= 1_000:
            return f"{value / 1_000:g}kΩ"
        return f"{value:g}Ω"
    if kind == "capacitance":
        if value >= 1e-6:
            return f"{value / 1e-6:g}uF"
        if value >= 1e-9:
            return f"{value / 1e-9:g}nF"
        return f"{value / 1e-12:g}pF"
    if kind == "inductance":
        if value >= 1e-3:
            return f"{value / 1e-3:g}mH"
        if value >= 1e-6:
            return f"{value / 1e-6:g}uH"
        return f"{value / 1e-9:g}nH"
    return f"{value:g}"


def _values_match(left: float | None, right: float | None, tolerance: float = 0.02) -> bool | None:
    if left is None or right is None:
        return None
    base = max(abs(left), abs(right), 1e-30)
    return abs(left - right) / base <= tolerance


def _header_map(headers: list[str]) -> dict[str, int]:
    normalized = {_normalize(header): idx for idx, header in enumerate(headers)}
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = _normalize(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def _all_rows_from_xlsx(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    return [(idx, list(row)) for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1)]


def _all_rows_from_xls(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [
        (idx + 1, [sheet.cell_value(idx, col_idx) for col_idx in range(sheet.ncols)])
        for idx in range(sheet.nrows)
    ]


def _all_rows_from_csv(content: bytes) -> list[tuple[int, list[Any]]]:
    text = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 编码不支持，请使用 UTF-8 或 GB18030")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [(idx, row) for idx, row in enumerate(csv.reader(text.splitlines(), dialect), start=1)]


def _all_rows(content: bytes, filename: str | None) -> list[tuple[int, list[Any]]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        return _all_rows_from_csv(content)
    if lower_name.endswith(".xls") and not lower_name.endswith(".xlsx"):
        return _all_rows_from_xls(content)
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return _all_rows_from_xls(content)
    return _all_rows_from_xlsx(content)


def _find_header(
    rows: list[tuple[int, list[Any]]],
    field_mapping: dict[str, str | int] | None = None,
) -> tuple[int | None, dict[str, int]]:
    if field_mapping:
        requested_header_row = field_mapping.get("__header_row")
        try:
            requested_header_row = int(requested_header_row) if requested_header_row else None
        except (TypeError, ValueError):
            requested_header_row = None
        for row_idx, values in rows[:30]:
            if requested_header_row and row_idx != requested_header_row:
                continue
            headers = [_cell_text(value) for value in values]
            exact_headers = {header: idx for idx, header in enumerate(headers)}
            mapping: dict[str, int] = {}
            for field, selected in field_mapping.items():
                if field.startswith("__") or selected in (None, ""):
                    continue
                if isinstance(selected, int) or str(selected).isdigit():
                    index = int(selected)
                    if 0 <= index < len(headers):
                        mapping[field] = index
                elif str(selected) in exact_headers:
                    mapping[field] = exact_headers[str(selected)]
            if mapping:
                return row_idx, mapping
    best_row = None
    best_mapping: dict[str, int] = {}
    best_score = 0
    for row_idx, values in rows[:30]:
        mapping = _header_map([_cell_text(value) for value in values])
        score = sum(1 for field in ["quantity", "designator", "footprint", "manufacturer_part", "supplier_part"] if field in mapping)
        if score > best_score:
            best_row = row_idx
            best_mapping = mapping
            best_score = score
        if score >= 3 and "quantity" in mapping:
            return row_idx, mapping
    return (best_row, best_mapping) if best_score >= 2 else (None, {})


def _pick(values: list[Any], mapping: dict[str, int], field: str) -> str:
    col = mapping.get(field)
    if col is None or col >= len(values):
        return ""
    return _cell_text(values[col])


def inspect_bom_fields(content: bytes, filename: str | None = None) -> dict[str, Any]:
    raw_rows = _all_rows(content, filename)
    header_row, mapping = _find_header(raw_rows)
    if header_row is None:
        first_nonempty = next(
            ((row_idx, values) for row_idx, values in raw_rows[:30] if any(_cell_text(value) for value in values)),
            None,
        )
        if first_nonempty:
            header_row = first_nonempty[0]
            mapping = {}
    headers: list[str] = []
    preview: list[list[str]] = []
    if header_row is not None:
        header_values = next((values for row_idx, values in raw_rows if row_idx == header_row), [])
        headers = [_cell_text(value) for value in header_values]
        preview = [
            [_cell_text(value) for value in values]
            for row_idx, values in raw_rows
            if header_row < row_idx <= header_row + 5
        ]
    return {
        "header_row": header_row,
        "headers": headers,
        "detected_mapping": {
            field: headers[index]
            for field, index in mapping.items()
            if 0 <= index < len(headers)
        },
        "preview": preview,
    }


def parse_bom_excel(
    content: bytes,
    filename: str | None = None,
    field_mapping: dict[str, str | int] | None = None,
) -> list[BomRow]:
    raw_rows = _all_rows(content, filename)
    header_row, mapping = _find_header(raw_rows, field_mapping)
    if header_row is None:
        return []

    rows: list[BomRow] = []
    for row_idx, values in raw_rows:
        if row_idx <= header_row:
            continue
        if not any(_cell_text(value) for value in values):
            continue
        quantity = _to_int(_pick(values, mapping, "quantity"), 0)
        manufacturer_part = _pick(values, mapping, "manufacturer_part")
        supplier_part = _pick(values, mapping, "supplier_part")
        comment = _pick(values, mapping, "comment")
        value = _pick(values, mapping, "value")
        footprint = _pick(values, mapping, "footprint")
        designator = _pick(values, mapping, "designator")
        if quantity <= 0 or not any([manufacturer_part, supplier_part, comment, value, footprint, designator]):
            continue
        rows.append(
            BomRow(
                source_row=row_idx,
                data={
                    "source_row": row_idx,
                    "no": _pick(values, mapping, "no"),
                    "required_quantity": quantity,
                    "comment": comment,
                    "designator": designator,
                    "footprint": footprint,
                    "value": value,
                    "manufacturer_part": manufacturer_part,
                    "manufacturer": _pick(values, mapping, "manufacturer"),
                    "supplier_part": supplier_part,
                    "supplier": _pick(values, mapping, "supplier"),
                    "category": _pick(values, mapping, "category"),
                    "primary_category": _pick(values, mapping, "primary_category"),
                    "mounting_style": _pick(values, mapping, "mounting_style"),
                    "layer": _pick(values, mapping, "layer"),
                },
            )
        )
    return rows


def _candidate_query(
    db: Session,
    row: dict[str, Any],
    limit: int = 120,
    component_ids: list[int] | None = None,
    supplier_scope_type: str | None = None,
    supplier_owner_user_id: int | None = None,
    supplier_team_library_id: str | None = None,
) -> list[Component]:
    supplier_part = str(row.get("supplier_part") or "").strip()
    supplier = str(row.get("supplier") or "").strip()
    manufacturer_part = str(row.get("manufacturer_part") or "").strip()
    manufacturer = str(row.get("manufacturer") or "").strip()
    footprint = str(row.get("footprint") or "").strip()
    ids: set[int] = set()
    if supplier_part:
        if _lcsc_supplier(supplier):
            ids.update(
                component_id
                for (component_id,) in db.query(Component.id)
                .filter(func.lower(Component.lcsc_number) == supplier_part.lower())
                .all()
            )
        supplier_rows = (
            _scope_supplier_query(
                db.query(SupplierPart),
                supplier_scope_type,
                supplier_owner_user_id,
                supplier_team_library_id,
            )
            .filter(
                SupplierPart.status == "active",
                func.lower(SupplierPart.supplier_part_number) == supplier_part.lower(),
            )
            .all()
        )
        ids.update(item.component_id for item in supplier_rows if _supplier_matches(supplier, item.supplier))
    if manufacturer_part:
        model_query = db.query(Component).filter(Component.model.isnot(None))
        if component_ids is not None:
            model_query = model_query.filter(Component.id.in_(component_ids or [0]))
        mpn_key = _normalize(manufacturer_part)
        manufacturer_key = _normalize(manufacturer)
        ids.update(
            item.id
            for item in model_query.all()
            if _normalize(item.model) == mpn_key
            and (not manufacturer_key or _normalize(item.manufacturer) == manufacturer_key)
        )
    query = db.query(Component)
    if component_ids is not None:
        query = query.filter(Component.id.in_(component_ids or [0]))
    if ids:
        query = query.filter(Component.id.in_(ids))
    elif footprint:
        package_key = _normalize_package(footprint)
        candidates = query.order_by(Component.quantity.desc(), Component.updated_at.desc()).limit(500).all()
        return [item for item in candidates if _normalize_package(item.package) == package_key][:limit]
    else:
        return []
    return query.order_by(Component.quantity.desc(), Component.updated_at.desc()).limit(limit).all()


def _score_match(row: dict[str, Any], component: Component) -> tuple[int, str, list[str]]:
    supplier_part = _normalize(row.get("supplier_part"))
    manufacturer_part = _normalize(row.get("manufacturer_part"))
    value = _normalize(row.get("value") or row.get("comment"))
    footprint = _normalize_package(row.get("footprint"))
    category = _normalize(row.get("primary_category") or row.get("category"))
    passive_kind = _passive_kind(row, component)
    bom_value = _parse_passive_value(" ".join(str(row.get(key) or "") for key in ["value", "comment", "manufacturer_part"]), passive_kind)

    lcsc = _normalize(component.lcsc_number)
    model = _normalize(component.model)
    name = _normalize(component.name)
    params = _normalize(component.parameters)
    package = _normalize_package(component.package)
    tags = _normalize(component.tags)
    component_value = _component_passive_value(component, passive_kind)
    row_passive_kind = _passive_kind(row)
    component_passive_kind = _passive_kind(component=component)
    row_pin_counts = _pin_counts(row.get("manufacturer_part"), row.get("comment"), row.get("value"), row.get("footprint"))
    component_pin_counts = _pin_counts(component.model, component.name, component.parameters, component.package, component.tags)
    connectorish = _connectorish(
        row.get("manufacturer_part"),
        row.get("comment"),
        row.get("value"),
        row.get("footprint"),
        component.model,
        component.name,
        component.parameters,
        component.package,
        component.tags,
    )

    reasons: list[str] = []
    flags: list[str] = []
    score = 0
    if supplier_part and lcsc and supplier_part == lcsc:
        flags.append("编号一致")
        if passive_kind and bom_value is not None and component_value is not None:
            flags.append("标称值一致" if _values_match(bom_value, component_value) else "标称值不一致")
        if footprint and package:
            flags.append("封装一致" if (footprint == package or footprint in package or package in footprint) else "封装不一致")
        return 100, "立创 ID 精确匹配", flags
    if manufacturer_part and model and manufacturer_part == model:
        score += 88
        reasons.append("型号精确匹配")

    haystack = " ".join([name, model, params, tags])
    if value and value in haystack:
        score += 22
        reasons.append("参数/标称值匹配")
    if passive_kind and bom_value is not None and component_value is not None:
        if _values_match(bom_value, component_value):
            score += 28
            flags.append("标称值一致")
        else:
            flags.append("标称值不一致")
            return min(score, 42), f"标称值不一致：BOM {_format_passive_value(bom_value, passive_kind)}，库存 {_format_passive_value(component_value, passive_kind)}", flags
    if footprint and package:
        if footprint == package or footprint in package or package in footprint:
            score += 18
            reasons.append("封装匹配")
            flags.append("封装一致")
        else:
            flags.append("封装不一致")
            if passive_kind:
                return min(score, 42), "封装不一致，基础器件不自动替代", flags
    if (
        row_passive_kind
        and row_passive_kind == component_passive_kind
        and bom_value is not None
        and component_value is not None
        and _values_match(bom_value, component_value)
        and footprint
        and package
        and footprint == package
    ):
        score = max(score, 96)
        reasons.append("类别、规范值和标准封装精确一致")
        flags.append("唯一复合键候选")
    if connectorish and row_pin_counts and component_pin_counts:
        if row_pin_counts & component_pin_counts:
            score += 16
            reasons.append("脚数匹配")
            flags.append("脚数一致")
        else:
            flags.append("脚数不一致")
            return min(score, 35), f"脚数不一致：BOM {sorted(row_pin_counts)[0]}P，库存 {sorted(component_pin_counts)[0]}P", flags
    if category and category in tags + name + params:
        score += 10
        reasons.append("分类接近")
    if component.quantity > 0:
        score += 4
    return min(score, 99), "，".join(reasons) or "仅精确字段候选", flags


def _bom_role(row: dict[str, Any]) -> str:
    text = " ".join(str(row.get(key) or "") for key in ["comment", "value", "manufacturer_part", "category", "primary_category"]).lower()
    if any(word in text for word in ["电容", "capacitor", "uf", "nf", "pf"]):
        return "用于去耦、滤波、储能或时序相关电路，需按位置确认耐压和介质。"
    if any(word in text for word in ["电阻", "resistor", "ohm", "ω", "k"]):
        return "用于限流、分压、上拉下拉或反馈网络，需核对阻值、精度和功耗。"
    if any(word in text for word in ["电感", "inductor", "uh"]):
        return "用于电源储能、滤波或 EMI 抑制，需核对饱和电流和 DCR。"
    if any(word in text for word in ["tvs", "esd", "保护", "浪涌"]):
        return "用于接口或电源保护，需核对工作电压、钳位电压和封装功率。"
    if any(word in text for word in ["usb", "type-c", "connector", "排针", "端子", "插座"]):
        return "用于板级或外部连接，需核对脚位、间距、耐流和机械方向。"
    if any(word in text for word in ["ldo", "dc-dc", "稳压", "电源"]):
        return "用于电源转换或稳压，需核对输入输出范围、电流和热设计。"
    if any(word in text for word in ["mcu", "芯片", "ic", "esp32"]):
        return "用于核心控制或功能实现，需核对供电、封装、外设和启动条件。"
    return "BOM 行用途需结合原理图位号确认，建议先核对型号、封装和参数。"


def _missing_suggestion(row: dict[str, Any]) -> dict[str, Any]:
    keyword = clean_lcsc_keyword(row.get("supplier_part") or row.get("manufacturer_part") or row.get("value") or row.get("comment"))
    description = row.get("manufacturer_part") or row.get("comment") or row.get("value") or "未匹配物料"
    return {
        "description": description,
        "reason": "库存中没有找到足够可靠的匹配项，请按型号/参数/封装确认后采购或手动匹配。",
        "lcsc_search_keyword": keyword,
        "lcsc_search_url": lcsc_search_url(keyword),
    }


def _combination_suggestions(
    db: Session,
    row: dict[str, Any],
    limit: int = 3,
    component_ids: list[int] | None = None,
) -> list[dict[str, str]]:
    kind = _passive_kind(row)
    target = _parse_passive_value(" ".join(str(row.get(key) or "") for key in ["value", "comment", "manufacturer_part"]), kind)
    if not kind or target is None:
        return []
    query = db.query(Component).filter(Component.quantity > 0, Component.revoked_at.is_(None))
    if component_ids is not None:
        query = query.filter(Component.id.in_(component_ids or [0]))
    components = query.order_by(Component.quantity.desc(), Component.id.asc()).limit(300).all()
    candidates = []
    for component in components:
        if _passive_kind(component=component) != kind:
            continue
        value = _component_passive_value(component, kind)
        if value is None:
            continue
        candidates.append((component, value))
    suggestions: list[dict[str, str]] = []
    seen = set()
    for idx, (left_component, left_value) in enumerate(candidates):
        for right_component, right_value in candidates[idx:]:
            modes = []
            if kind in {"resistance", "inductance"}:
                modes.append(("串联", left_value + right_value))
            if kind in {"resistance", "capacitance"} and left_value + right_value > 0:
                parallel = (left_value * right_value) / (left_value + right_value) if kind == "resistance" else left_value + right_value
                modes.append(("并联", parallel))
            for mode, combined in modes:
                if not _values_match(target, combined, tolerance=0.03):
                    continue
                key = tuple(sorted([left_component.id, right_component.id]) + [mode])
                if key in seen:
                    continue
                seen.add(key)
                left_label = left_component.model or left_component.name
                right_label = right_component.model or right_component.name
                suggestions.append(
                    {
                        "description": f"{left_label} + {right_label} {mode} ≈ {_format_passive_value(target, kind)}",
                        "reason": "组合替代仅作临时建议，需核对精度、功耗/纹波、温漂、空间和可靠性。",
                    }
                )
                if len(suggestions) >= limit:
                    return suggestions
    return suggestions


def match_bom_rows(
    db: Session,
    rows: list[BomRow],
    top_n: int = 5,
    component_ids: list[int] | None = None,
    supplier_scope_type: str | None = None,
    supplier_owner_user_id: int | None = None,
    supplier_team_library_id: str | None = None,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    reserved_by_component = reserved_quantities(db, component_ids)
    for row in rows:
        matches = []
        seen = set()
        supplier_part = str(row.data.get("supplier_part") or "").strip()
        supplier = str(row.data.get("supplier") or "").strip()
        exact_supplier_ids = {
            item.component_id
            for item in _scope_supplier_query(
                db.query(SupplierPart),
                supplier_scope_type,
                supplier_owner_user_id,
                supplier_team_library_id,
            )
            .filter(
                SupplierPart.status == "active",
                func.lower(SupplierPart.supplier_part_number) == supplier_part.lower(),
            )
            .all()
            if _supplier_matches(supplier, item.supplier)
        } if supplier_part else set()
        for component in _candidate_query(
            db,
            row.data,
            component_ids=component_ids,
            supplier_scope_type=supplier_scope_type,
            supplier_owner_user_id=supplier_owner_user_id,
            supplier_team_library_id=supplier_team_library_id,
        ):
            if component.id in seen:
                continue
            seen.add(component.id)
            score, reason, flags = _score_match(row.data, component)
            if component.id in exact_supplier_ids:
                score = 100
                reason = "供应商料号精确匹配"
                flags = ["编号一致", *[flag for flag in flags if flag != "编号一致"]]
            if score < 18:
                continue
            available = max(0, int(component.quantity or 0) - int(reserved_by_component.get(component.id, 0)))
            required = row.data["required_quantity"]
            matches.append(
                {
                    "component": component_to_dict(component),
                    "score": score,
                    "match_type": "exact" if score >= 95 else "candidate",
                    "reason": reason,
                    "flags": flags,
                    "available_quantity": available,
                    "shortage_quantity": max(0, required - available),
                    "enough": available >= required,
                }
            )
        matches.sort(key=lambda item: item["score"], reverse=True)
        top_matches = matches[:top_n]
        has_supplier_part = bool(str(row.data.get("supplier_part") or "").strip())
        exact_lcsc_match = bool(top_matches and top_matches[0]["score"] >= 100 and "编号一致" in top_matches[0].get("flags", []))
        selected = None
        if exact_lcsc_match:
            selected = top_matches[0]["component"]["id"]
            status = "exact_lcsc"
        elif has_supplier_part:
            status = "supplier_missing"
        else:
            exact_candidates = [item for item in top_matches if item["score"] >= 95]
            if len(exact_candidates) == 1:
                selected = exact_candidates[0]["component"]["id"]
                status = "exact"
            elif top_matches:
                status = "review"
            else:
                status = "missing"
        missing_suggestion = _missing_suggestion(row.data) if status in {"missing", "review", "supplier_missing"} else None
        alternatives = (
            _combination_suggestions(db, row.data, component_ids=component_ids)
            if status in {"missing", "review", "supplier_missing"}
            else []
        )
        if status == "supplier_missing" and missing_suggestion:
            missing_suggestion["reason"] = "BOM 指定立创 ID 未在库存中找到；如不替换，采购时优先购买 BOM 中的该立创 ID。"
            similar = []
            for match in top_matches[:3]:
                component = match["component"]
                label = component.get("model") or component.get("name") or f"库存 #{component.get('id')}"
                similar.append(
                    {
                        "description": f"库内相似：{label}，{match['score']}%（{match['reason']}）",
                        "reason": "编号不同，仅作为避免重复采购的替换提醒，需人工确认参数、封装和风险。",
                    }
                )
            alternatives = similar + alternatives
        if missing_suggestion and alternatives:
            missing_suggestion["alternatives"] = alternatives
        result.append(
            {
                **row.data,
                "status": status,
                "selected_component_id": selected,
                "match_confidence": top_matches[0]["score"] if top_matches else 0,
                "matches": top_matches[: max(3, top_n)],
                "role": _bom_role(row.data),
                "missing_suggestion": missing_suggestion,
                "lcsc_search_url": lcsc_search_url(row.data.get("supplier_part") or row.data.get("manufacturer_part") or row.data.get("value")),
                "ai_reason": (
                    "编号一致自动匹配。"
                    if status == "exact_lcsc"
                    else "BOM 指定供应商料号未入库，候选库存只用于人工核对，系统不会自动替换。"
                    if status == "supplier_missing"
                    else "规则匹配生成，低置信行可由 AI 补充判断。"
                ),
            }
        )
    return result
