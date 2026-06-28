from dataclasses import dataclass
from io import BytesIO
import json
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session
import xlrd

from ..models import Component, ImportRecord
from .component_normalizer import normalize_component_values
from .category_governance import category_from_order_text


HEADER_ALIASES = {
    "name": ["物料名称", "商品名称", "名称", "器件名称", "产品名称"],
    "model": ["型号", "规格型号", "商品型号", "产品型号"],
    "quantity": ["数量", "购买数量", "库存数量", "实发数量", "订购数量"],
    "package": ["封装", "封装规格", "封装/规格", "封装格式"],
    "lcsc_number": ["物料编号", "商品编号", "立创 ID", "立创编号", "LCSC", "LCSC编号", "产品编号"],
    "parameters": ["参数", "规格参数", "描述", "商品描述"],
    "source": ["来源", "供应商", "商城"],
    "order_number": ["订单编号", "订单号", "单号"],
    "brand": ["品牌", "制造商", "厂牌"],
    "product_type": ["商品类型", "物料类型", "分类", "品类"],
    "unit": ["商品单位", "单位"],
    "shipping_no": ["快递单号", "物流单号", "运单号"],
    "order_time": ["下单时间", "购买时间"],
    "shipment_date": ["发货日期", "发货时间"],
    "customer_part_number": ["客户料号"],
}

COMPONENT_FIELDS = {
    "warehouse_code",
    "name",
    "model",
    "category_id",
    "parameters",
    "package",
    "quantity",
    "source",
    "lcsc_number",
    "tags",
    "source_title",
    "part_family",
    "count_mode",
    "normalized_spec",
    "status",
    "location",
    "remark",
    "datasheet_url",
}


def _without_pending_purchase_tags(tags: str | None) -> str | None:
    tokens = [part.strip() for part in str(tags or "").replace("，", ",").replace("；", ",").replace(";", ",").split(",")]
    blocked = {"待采购", "bom待采购", "bom待采购库", "待入库", "pending_purchase"}
    result: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        key = token.replace(" ", "").lower()
        if not token or key in blocked or key in seen:
            continue
        seen.add(key)
        result.append(token)
    return ",".join(result) or None


@dataclass
class ParsedRow:
    data: dict[str, Any]
    source_row: int


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return default


def _header_map(headers: list[str]) -> dict[str, int]:
    normalized = {header.strip().lower(): idx for idx, header in enumerate(headers)}
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = alias.strip().lower()
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def _all_rows_from_xlsx(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    return [(idx, list(row)) for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1)]


def _all_rows_from_xls(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    rows: list[tuple[int, list[Any]]] = []
    for row_idx in range(sheet.nrows):
        rows.append((row_idx + 1, [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]))
    return rows


def _all_rows(content: bytes, filename: str | None = None) -> list[tuple[int, list[Any]]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".xls") and not lower_name.endswith(".xlsx"):
        return _all_rows_from_xls(content)
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return _all_rows_from_xls(content)
    return _all_rows_from_xlsx(content)


def _find_header(rows: list[tuple[int, list[Any]]]) -> tuple[int | None, list[str], dict[str, int]]:
    best: tuple[int | None, list[str], dict[str, int]] = (None, [], {})
    best_score = 0
    for row_idx, row in rows[:30]:
        headers = [_cell_text(value) for value in row]
        mapping = _header_map(headers)
        score = sum(1 for field in ["name", "model", "quantity", "lcsc_number"] if field in mapping)
        if score > best_score:
            best = (row_idx, headers, mapping)
            best_score = score
        if score >= 3 and "quantity" in mapping:
            return best
    return best if best_score >= 2 else (None, [], {})


def _pick(values: list[Any], mapping: dict[str, int], field: str) -> str:
    col = mapping.get(field)
    if col is None or col >= len(values):
        return ""
    return _cell_text(values[col])


def guess_category(order_category: str, name: str | None, parameters: str | None, db: Session) -> int | None:
    category = category_from_order_text(db, order_category)
    if str(order_category or "").strip():
        return category.id if category else None
    category = category_from_order_text(db, name) or category_from_order_text(db, parameters)
    return category.id if category else None


def parse_excel(content: bytes, db: Session, filename: str | None = None) -> list[ParsedRow]:
    raw_rows = _all_rows(content, filename)
    header_row_idx, _, mapping = _find_header(raw_rows)
    if header_row_idx is None:
        return []

    rows: list[ParsedRow] = []
    for row_idx, values in raw_rows:
        if row_idx <= header_row_idx:
            continue
        values = list(values)
        if not any(_cell_text(value) for value in values):
            continue

        name = _pick(values, mapping, "name")
        model = _pick(values, mapping, "model")
        package = _pick(values, mapping, "package")
        product_type = _pick(values, mapping, "product_type")
        parameters = _pick(values, mapping, "parameters") or product_type
        lcsc_number = _pick(values, mapping, "lcsc_number")
        quantity = _to_int(values[mapping["quantity"]], 0) if "quantity" in mapping else 0
        if not name and not model and not lcsc_number:
            continue
        if not model and not lcsc_number and quantity == 0:
            continue

        order_number = _pick(values, mapping, "order_number")
        brand = _pick(values, mapping, "brand")
        shipping_no = _pick(values, mapping, "shipping_no")
        import_key = f"{order_number}:{lcsc_number}" if order_number and lcsc_number else None
        data = normalize_component_values({
            "name": name or model or lcsc_number,
            "model": model or None,
            "category_id": guess_category(product_type, name or model, parameters, db),
            "parameters": parameters or None,
            "package": package or None,
            "quantity": quantity,
            "source": _pick(values, mapping, "source") or "立创",
            "lcsc_number": lcsc_number or None,
            "tags": product_type or None,
            "status": "in_stock" if quantity > 0 else "pending",
            "location": None,
            "remark": None,
            "datasheet_url": None,
            "source_file": filename,
            "order_number": order_number or None,
            "brand": brand or None,
            "product_type": product_type or None,
            "unit": _pick(values, mapping, "unit") or None,
            "shipping_no": shipping_no or None,
            "order_time": _pick(values, mapping, "order_time") or None,
            "shipment_date": _pick(values, mapping, "shipment_date") or None,
            "import_key": import_key,
        })
        rows.append(ParsedRow(data=data, source_row=row_idx))

    return rows


def component_values(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key in COMPONENT_FIELDS}


def find_import_record(db: Session, row: dict[str, Any]) -> ImportRecord | None:
    order_number = row.get("order_number")
    lcsc_number = row.get("lcsc_number")
    if not order_number or not lcsc_number:
        return None
    return (
        db.query(ImportRecord)
        .filter(ImportRecord.order_number == order_number, ImportRecord.lcsc_number == lcsc_number)
        .first()
    )


def create_import_record(db: Session, row: dict[str, Any], component_id: int | None, batch_id: int | None = None) -> ImportRecord | None:
    order_number = row.get("order_number")
    lcsc_number = row.get("lcsc_number")
    if not order_number or not lcsc_number or find_import_record(db, row):
        return None
    record = ImportRecord(
        batch_id=batch_id,
        order_number=order_number,
        lcsc_number=lcsc_number,
        component_id=component_id,
        quantity=int(row.get("quantity") or 0),
        source_file=row.get("source_file"),
        source_row=row.get("source_row"),
        raw_data=json.dumps(row, ensure_ascii=False, default=str),
    )
    db.add(record)
    db.flush()
    return record


def find_duplicate(db: Session, row: dict[str, Any]) -> Component | None:
    lcsc_number = row.get("lcsc_number")
    if lcsc_number:
        found = db.query(Component).filter(Component.lcsc_number == lcsc_number).first()
        if found:
            return found

    name = row.get("name")
    model = row.get("model")
    package = row.get("package")
    if not name and not model:
        return None
    return (
        db.query(Component)
        .filter(
            and_(
                Component.name == name,
                or_(Component.model == model, Component.model.is_(None) if model is None else False),
                or_(Component.package == package, Component.package.is_(None) if package is None else False),
            )
        )
        .first()
    )


def merge_component(component: Component, row: dict[str, Any]):
    incoming_quantity = int(row.get("quantity") or 0)
    was_pending_purchase = component.status == "pending_purchase"
    previous_source = component.source
    component.quantity += incoming_quantity
    for field in [
        "model",
        "category_id",
        "parameters",
        "package",
        "source",
        "lcsc_number",
        "tags",
        "source_title",
        "part_family",
        "count_mode",
        "normalized_spec",
        "status",
        "location",
        "remark",
        "datasheet_url",
    ]:
        if getattr(component, field) in (None, "") and row.get(field) not in (None, ""):
            setattr(component, field, row[field])
    if was_pending_purchase and component.quantity > 0:
        component.status = "in_stock"
        if previous_source == "BOM 待采购库":
            component.source = row.get("source") or "立创"
        component.location = None if component.location == "待采购" else component.location
        component.tags = _without_pending_purchase_tags(component.tags)
        arrival_note = f"立创订单到货自动抵消待采购 x {incoming_quantity}"
        component.remark = f"{component.remark or ''}\n{arrival_note}".strip()
