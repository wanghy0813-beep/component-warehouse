from dataclasses import dataclass
from hashlib import sha1
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session
import xlrd

from ..models import Component, ImportRecord
from .category_governance import canonical_order_category_name
from .hardware_categories import classify_hardware_category
from .mimo_ai import MimoRequestError, analyze_external_order_table


@dataclass
class ParsedExternalRow:
    data: dict[str, Any]
    source_row: int


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_int(value: Any, default: int = 1) -> int:
    try:
        if value is None or value == "":
            return default
        return max(0, int(float(str(value).strip())))
    except (TypeError, ValueError):
        return default


def _all_rows_from_xlsx(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    return [(idx, list(row)) for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1)]


def _all_rows_from_xls(content: bytes) -> list[tuple[int, list[Any]]]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [(idx + 1, [sheet.cell_value(idx, col) for col in range(sheet.ncols)]) for idx in range(sheet.nrows)]


def _all_rows(content: bytes, filename: str | None = None) -> list[tuple[int, list[Any]]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".xls") and not lower_name.endswith(".xlsx"):
        return _all_rows_from_xls(content)
    if content.startswith(b"\xd0\xcf\x11\xe0"):
        return _all_rows_from_xls(content)
    return _all_rows_from_xlsx(content)


def _headers_from_row(values: list[Any], width: int) -> list[str]:
    return [(_cell_text(values[idx]) if idx < len(values) and _cell_text(values[idx]) else f"列{idx + 1}") for idx in range(width)]


def _payload_rows(non_empty: list[tuple[int, list[Any]]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not non_empty:
        return [], []
    width = max(len(row) for _, row in non_empty)
    possible_header_values = [_cell_text(value) for value in non_empty[0][1]]
    if sum(1 for value in possible_header_values if value) >= max(2, min(4, width)):
        headers = _headers_from_row(non_empty[0][1], width)
        row_values = non_empty[1:]
    else:
        headers = [f"列{idx + 1}" for idx in range(width)]
        row_values = non_empty
    return _payload_rows_with_headers(headers, row_values)


def _payload_rows_with_headers(headers: list[str], values_rows: list[tuple[int, list[Any]]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not values_rows:
        return headers, []
    width = max(len(headers), max(len(row) for _, row in values_rows))
    normalized_headers = [(headers[idx] if idx < len(headers) and headers[idx] else f"列{idx + 1}") for idx in range(width)]
    rows: list[dict[str, Any]] = []
    for source_row, values in values_rows:
        cells = []
        for idx in range(width):
            text = _cell_text(values[idx] if idx < len(values) else "")
            if text:
                cells.append({"column": idx + 1, "header": normalized_headers[idx], "value": text[:240]})
        rows.append({"source_row": source_row, "cells": cells})
    return normalized_headers, rows


def _find_header_and_data_rows(non_empty: list[tuple[int, list[Any]]]) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    if not non_empty:
        return [], []
    header_idx = 0
    best_score = -1
    for offset, (_, row) in enumerate(non_empty[:8]):
        cells = [_cell_text(cell) for cell in row]
        non_blank = sum(1 for cell in cells if cell)
        text = " ".join(cells)
        score = non_blank
        for marker in ["订单", "商品", "型号", "规格", "数量", "金额", "链接", "店铺", "时间", "状态"]:
            if marker in text:
                score += 3
        if score > best_score:
            best_score = score
            header_idx = offset
    header_row = non_empty[header_idx][1]
    data_rows = non_empty[header_idx + 1 :]
    if not data_rows:
        data_rows = non_empty[header_idx:]
        headers = [f"列{idx + 1}" for idx in range(max(len(row) for _, row in data_rows))]
        return headers, data_rows
    width = max(len(header_row), max(len(row) for _, row in data_rows))
    return _headers_from_row(header_row, width), data_rows


def _analyze_external_rows(headers: list[str], rows: list[dict[str, Any]], categories: list[str]) -> list[dict[str, Any]]:
    if not rows:
        return []
    try:
        analyzed = analyze_external_order_table(headers, rows, categories)
        result = analyzed.get("rows") if isinstance(analyzed, dict) else None
        if isinstance(result, list):
            return result
        return []
    except MimoRequestError:
        if len(rows) <= 1:
            raise
    recovered: list[dict[str, Any]] = []
    errors: list[str] = []
    for row in rows:
        try:
            analyzed = analyze_external_order_table(headers, [row], categories)
            result = analyzed.get("rows") if isinstance(analyzed, dict) else None
            if isinstance(result, list):
                recovered.extend(result)
        except MimoRequestError as exc:
            errors.append(f"第 {row.get('source_row')} 行: {exc}")
    if recovered:
        return recovered
    raise MimoRequestError("外部订单 AI 小批解析和单行重试均失败：" + "；".join(errors[:3]))


def _dedupe_key(*parts: Any) -> str:
    raw = "|".join(str(part or "").strip() for part in parts)
    return sha1(raw.encode("utf-8")).hexdigest()[:16]


def _join_tags(tags: Any) -> str | None:
    if isinstance(tags, str):
        values = [tags]
    elif isinstance(tags, list):
        values = [str(item) for item in tags if item]
    else:
        values = []
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        for token in value.replace("，", ",").replace("；", ",").replace(";", ",").split(","):
            tag = token.strip()
            key = tag.replace(" ", "").lower()
            if not tag or key in seen:
                continue
            seen.add(key)
            result.append(tag[:40])
            if len(result) >= 8:
                return ",".join(result)
    return ",".join(result) or None


def parse_external_order(content: bytes, filename: str | None, categories: list[str]) -> list[ParsedExternalRow]:
    raw_rows = _all_rows(content, filename)
    non_empty = [(idx, row) for idx, row in raw_rows if any(_cell_text(cell) for cell in row)]
    if not non_empty:
        return []

    result: list[ParsedExternalRow] = []
    valid_categories = set(categories)
    seen_source_rows: set[int] = set()
    headers, data_rows = _find_header_and_data_rows(non_empty)
    category_headers = {"分类", "品类", "商品类型", "物料类型"}
    category_column = next((index for index, header in enumerate(headers) if _cell_text(header) in category_headers), None)
    order_categories = {
        row_index: _cell_text(values[category_column])
        for row_index, values in data_rows
        if category_column is not None and category_column < len(values)
    }
    chunks = [data_rows[idx: idx + 8] for idx in range(0, len(data_rows), 8)] or [data_rows]
    for chunk in chunks:
        _, ai_rows = _payload_rows_with_headers(headers, chunk)
        rows = _analyze_external_rows(headers, ai_rows, categories)
        for item in rows:
            item["order_category"] = order_categories.get(_to_int(item.get("source_row"), 0), "")
            parsed = _parsed_ai_external_row(item, filename, valid_categories)
            if not parsed or parsed.source_row in seen_source_rows:
                continue
            seen_source_rows.add(parsed.source_row)
            result.append(parsed)
    return result


def _parsed_ai_external_row(item: dict[str, Any], filename: str | None, valid_categories: set[str]) -> ParsedExternalRow | None:
    if not isinstance(item, dict):
        return None
    source_row = _to_int(item.get("source_row"), 0)
    if source_row <= 0 or item.get("skip"):
        return None
    name = str(item.get("normalized_name") or item.get("model") or "").strip()
    if not name:
        return None
    order_number = str(item.get("order_number") or "").strip() or None
    model = str(item.get("model") or "").strip() or None
    package = str(item.get("package") or "").strip() or None
    normalized_spec = str(item.get("normalized_spec") or "").strip() or None
    quantity = _to_int(item.get("actual_quantity"), 1)
    order_category = canonical_order_category_name(item.get("order_category"), valid_categories)
    deterministic_category, deterministic_reason = classify_hardware_category(
        item.get("normalized_name"), item.get("model"), item.get("parameters"),
        item.get("package"), item.get("product_title"), item.get("sku_text"),
    )
    ai_category = str(item.get("category") or "").strip()
    confidence = str(item.get("confidence") or "").strip().lower()
    requires_confirmation = bool(item.get("requires_confirmation"))
    category = order_category or (deterministic_category if deterministic_category in valid_categories else "") or (
        ai_category
        if ai_category in valid_categories and confidence == "high" and not requires_confirmation
        else ""
    )
    product_title = str(item.get("product_title") or "").strip()
    sku_text = str(item.get("sku_text") or "").strip()
    import_key = f"EXT-{_dedupe_key(filename, source_row, order_number, product_title, sku_text, name, model, normalized_spec)}"
    source_title = " / ".join(part for part in [product_title, sku_text] if part) or name
    remark_parts = [
        f"外部订单 AI 导入行: {source_row}",
        f"订单分类: {order_category}" if order_category else None,
        f"AI 识别依据: {item.get('reason')}" if item.get("reason") else None,
        f"17区规则: {deterministic_reason}" if deterministic_reason else None,
        f"订单购买件数: {item.get('order_quantity')}" if item.get("order_quantity") not in (None, "") else None,
        f"每件包含数量: {item.get('component_quantity_per_order')}" if item.get("component_quantity_per_order") not in (None, "") else None,
        f"原始标题: {source_title}",
        f"商品链接: {item.get('product_link')}" if item.get("product_link") else None,
    ]
    data = {
        "name": name[:200],
        "model": model,
        "category_name": category if category in valid_categories else None,
        "category_id": None,
        "parameters": str(item.get("parameters") or "").strip() or None,
        "package": package,
        "quantity": quantity,
        "source": "外部订单 AI 导入",
        "lcsc_number": None,
        "tags": _join_tags(item.get("tags")),
        "source_title": source_title,
        "part_family": "component",
        "count_mode": "exact",
        "normalized_spec": normalized_spec,
        "status": "in_stock" if quantity > 0 else "pending",
        "location": None,
        "remark": "\n".join(part for part in remark_parts if part),
        "datasheet_url": None,
        "source_file": filename,
        "source_row": source_row,
        "order_number": order_number,
        "order_time": str(item.get("order_time") or "").strip() or None,
        "store_name": str(item.get("store_name") or "").strip() or None,
        "product_link": str(item.get("product_link") or "").strip() or None,
        "import_key": import_key,
        "external_import_key": import_key,
        "ai_confidence": str(item.get("confidence") or "medium"),
        "ai_reason": str(item.get("reason") or "").strip() or None,
        "order_quantity": item.get("order_quantity"),
        "component_quantity_per_order": item.get("component_quantity_per_order"),
    }
    return ParsedExternalRow(data=data, source_row=source_row)


def find_external_duplicate(db: Session, row: dict[str, Any]) -> Component | None:
    model = row.get("model")
    package = row.get("package")
    normalized_spec = row.get("normalized_spec")
    name = row.get("name")
    if model:
        query = db.query(Component).filter(Component.model == model)
        if package:
            query = query.filter(or_(Component.package == package, Component.package.is_(None)))
        if normalized_spec:
            query = query.filter(or_(Component.normalized_spec == normalized_spec, Component.normalized_spec.is_(None)))
        found = query.order_by(Component.quantity.desc(), Component.id.asc()).first()
        if found:
            return found
    if name and package:
        return db.query(Component).filter(and_(Component.name == name, Component.package == package)).first()
    return None


def find_external_import_record(db: Session, row: dict[str, Any]) -> ImportRecord | None:
    order_number = row.get("order_number") or "external-order"
    import_key = row.get("external_import_key") or row.get("import_key")
    if not import_key:
        return None
    return db.query(ImportRecord).filter(ImportRecord.order_number == order_number, ImportRecord.lcsc_number == import_key).first()


def external_import_record_payload(row: dict[str, Any]) -> dict[str, Any]:
    order_number = row.get("order_number") or "external-order"
    import_key = row.get("external_import_key") or row.get("import_key") or _dedupe_key(row.get("name"), row.get("source_row"))
    return {**row, "order_number": order_number, "lcsc_number": import_key}
