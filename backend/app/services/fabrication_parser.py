from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import math
import multiprocessing
import os
from queue import Empty
import re
import resource
import stat
import tempfile
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from .mimo_ai import MimoNotConfiguredError, MimoRequestError, suggest_fabrication_mapping


MAX_ARCHIVE_BYTES = 200 * 1024**2
MAX_EXPANDED_BYTES = 512 * 1024**2
MAX_ENTRY_BYTES = 128 * 1024**2
MAX_ENTRIES = 500
MAX_DEPTH = 8
MAX_RATIO = 100
MAX_AI_TABLE_CHARS = 250_000
MAX_SVG_BYTES = 8 * 1024**2

GERBER_EXTENSIONS = {
    ".gbr", ".ger", ".pho", ".art", ".gtl", ".gbl", ".gts", ".gbs",
    ".gto", ".gbo", ".gtp", ".gbp", ".gpt", ".gpb", ".gko", ".gm1", ".g1", ".g2", ".g3",
}
TABLE_EXTENSIONS = {".csv", ".txt", ".pos", ".xlsx", ".xls"}
NESTED_ARCHIVE_EXTENSIONS = {".zip", ".7z", ".rar", ".tar", ".gz", ".bz2", ".xz"}

HEADER_ALIASES = {
    "designator": {
        "designator", "designators", "reference", "references", "ref", "refdes",
        "reference designator", "位号", "器件位号",
    },
    "value": {"value", "comment", "description", "参数", "值", "comment/value"},
    "model": {"model", "mpn", "manufacturer part number", "part number", "型号"},
    "footprint": {"footprint", "package", "封装", "pattern"},
    "x": {"mid x", "center x", "center-x(mm)", "center x(mm)", "posx", "x", "x(mm)", "x coordinate"},
    "y": {"mid y", "center y", "center-y(mm)", "center y(mm)", "posy", "y", "y(mm)", "y coordinate"},
    "rotation": {"rotation", "rotation(deg)", "rot", "angle", "角度"},
    "side": {"layer", "side", "board side", "面", "板面"},
    "dnp": {"dnp", "do not populate", "populate", "fitted", "mount", "不贴", "不装"},
}


class FabricationParseError(ValueError):
    pass


@dataclass
class TableData:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]
    raw_text: str


def normalized_designator(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()[:80]


def split_designators(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    values = []
    seen = set()
    for part in re.split(r"[,，、;；\s]+", text):
        item = part.strip()
        key = normalized_designator(item)
        if item and key and key not in seen:
            seen.add(key)
            values.append(item[:80])
    return values


def _safe_member(info: zipfile.ZipInfo) -> PurePosixPath:
    raw = str(info.filename or "").replace("\\", "/")
    path = PurePosixPath(raw)
    if not raw or raw.startswith("/") or path.is_absolute() or ".." in path.parts:
        raise FabricationParseError(f"ZIP 包含不安全路径：{raw or '-'}")
    if len(path.parts) > MAX_DEPTH:
        raise FabricationParseError(f"ZIP 目录层级超过 {MAX_DEPTH} 层：{raw}")
    mode = (info.external_attr >> 16) & 0xFFFF
    if mode and stat.S_ISLNK(mode):
        raise FabricationParseError(f"ZIP 不允许符号链接：{raw}")
    if info.flag_bits & 0x1:
        raise FabricationParseError(f"ZIP 不允许加密条目：{raw}")
    if Path(path.name).suffix.lower() in NESTED_ARCHIVE_EXTENSIONS:
        raise FabricationParseError(f"ZIP 不允许嵌套压缩包：{raw}")
    if info.file_size > MAX_ENTRY_BYTES:
        raise FabricationParseError(f"ZIP 单个文件超过 {MAX_ENTRY_BYTES // 1024**2}MB：{raw}")
    if info.compress_size and info.file_size / max(1, info.compress_size) > MAX_RATIO:
        raise FabricationParseError(f"ZIP 条目压缩比异常：{raw}")
    return path


def read_safe_zip(path: Path) -> dict[str, bytes]:
    if path.stat().st_size > MAX_ARCHIVE_BYTES:
        raise FabricationParseError("制造包超过 200MB 上限")
    result: dict[str, bytes] = {}
    total = 0
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as exc:
        raise FabricationParseError("制造包不是有效 ZIP") from exc
    with archive:
        infos = [item for item in archive.infolist() if not item.is_dir()]
        if not infos:
            raise FabricationParseError("制造包为空")
        if len(infos) > MAX_ENTRIES:
            raise FabricationParseError(f"制造包文件数超过 {MAX_ENTRIES}")
        for info in infos:
            safe_path = _safe_member(info)
            total += int(info.file_size or 0)
            if total > MAX_EXPANDED_BYTES:
                raise FabricationParseError("制造包展开后超过 512MB 上限")
            data = archive.read(info)
            if len(data) != info.file_size:
                raise FabricationParseError(f"ZIP 条目长度异常：{safe_path}")
            result[safe_path.as_posix()] = data
    return result


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _table_from_delimited(name: str, data: bytes) -> TableData:
    text = _decode_text(data)
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return TableData(name, [], [], text)
    suffix = Path(name).suffix.lower()
    if suffix == ".pos":
        header_index = next(
            (
                index
                for index, line in enumerate(lines[:40])
                if "ref" in line.lower()
                and any(token in line.lower() for token in ("posx", "center x", "mid x", "x(mm)"))
            ),
            0,
        )
        header_line = lines[header_index].lstrip("# ").strip()
        headers = [part.strip() for part in re.split(r"\s+", header_line)]
        rows = []
        for line in lines[header_index + 1 :]:
            if line.lstrip().startswith("#"):
                continue
            values = re.split(r"\s+", line.strip(), maxsplit=max(0, len(headers) - 1))
            if len(values) >= len(headers):
                rows.append(dict(zip(headers, values)))
        return TableData(name, headers, rows, text)
    sample = "\n".join(lines[:20])
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        if "\t" in lines[0]:
            delimiter = "\t"
        elif ";" in lines[0]:
            delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [str(item or "").strip() for item in (reader.fieldnames or [])]
    rows = [
        {str(key or "").strip(): value for key, value in row.items() if key is not None}
        for row in reader
        if any(str(value or "").strip() for value in row.values())
    ]
    return TableData(name, headers, rows, text)


def _table_from_xlsx(name: str, data: bytes) -> TableData:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise FabricationParseError(f"无法读取表格 {name}：{exc}") from exc
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return TableData(name, [], [], "")
    header_index = next((index for index, row in enumerate(values[:20]) if sum(value is not None for value in row) >= 2), 0)
    headers = [str(value or "").strip() for value in values[header_index]]
    rows = []
    for row in values[header_index + 1 :]:
        item = {headers[index]: value for index, value in enumerate(row[: len(headers)]) if headers[index]}
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    preview = json.dumps({"headers": headers, "rows": rows}, ensure_ascii=False, default=str)
    return TableData(name, headers, rows, preview)


def read_table(name: str, data: bytes) -> TableData:
    suffix = Path(name).suffix.lower()
    if suffix == ".xlsx":
        return _table_from_xlsx(name, data)
    if suffix == ".xls":
        raise FabricationParseError(f"旧版 XLS 暂不用于制造包自动映射，请另存为 XLSX 或 CSV：{name}")
    return _table_from_delimited(name, data)


def _header_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def detect_columns(headers: list[str]) -> dict[str, str]:
    normalized = {_header_key(header): header for header in headers}
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[field] = normalized[alias]
                break
    return result


def _name_role(name: str) -> str | None:
    value = Path(name).name.lower()
    suffix = Path(value).suffix.lower()
    if suffix not in TABLE_EXTENSIONS:
        return None
    if any(token in value for token in ("pick", "place", "cpl", "centroid", "position", "坐标")) or suffix == ".pos":
        return "cpl"
    if any(token in value for token in ("bom", "bill_of_material", "bill-of-material", "物料")):
        return "bom"
    return None


def detect_layer_role(name: str, data: bytes) -> tuple[str, str]:
    lower = Path(name).name.lower()
    text = data[:4096].decode("ascii", errors="ignore").lower()
    side = "both"
    if any(token in lower for token in (".gtl", "f_cu", "top_copper")) or "copper,l1,top" in text:
        return "copper", "top"
    if any(token in lower for token in (".gbl", "b_cu", "bottom_copper")) or "copper,l2,bot" in text:
        return "copper", "bottom"
    if Path(lower).suffix in {".g1", ".g2", ".g3"}:
        return "copper", "both"
    if any(token in lower for token in (".gts", "f_mask", "topmask", "top_mask")) or "soldermask,top" in text:
        return "mask", "top"
    if any(token in lower for token in (".gbs", "b_mask", "bottommask", "bottom_mask")) or "soldermask,bot" in text:
        return "mask", "bottom"
    if any(token in lower for token in (".gto", "f_silk", "topsilk", "top_silk")) or "legend,top" in text:
        return "silk", "top"
    if any(token in lower for token in (".gbo", "b_silk", "bottomsilk", "bottom_silk")) or "legend,bot" in text:
        return "silk", "bottom"
    if Path(lower).suffix in {".gtp", ".gpt"}:
        return "other", "top"
    if Path(lower).suffix in {".gbp", ".gpb"}:
        return "other", "bottom"
    if any(token in lower for token in (".gko", ".gm1", "edge_cuts", "outline", "boardoutline")) or "profile" in text:
        return "outline", "both"
    if "top" in lower or "front" in lower:
        side = "top"
    elif "bottom" in lower or "back" in lower or "bot" in lower:
        side = "bottom"
    return "other", side


def _parse_number(value: Any, default_units: str = "mm") -> float | None:
    text = str(value or "").strip().lower().replace(",", "")
    if not text:
        return None
    units = default_units
    if text.endswith("mm"):
        units = "mm"
        text = text[:-2].strip()
    elif text.endswith("mil"):
        units = "mil"
        text = text[:-3].strip()
    elif text.endswith("inch"):
        units = "inch"
        text = text[:-4].strip()
    elif text.endswith("in"):
        units = "inch"
        text = text[:-2].strip()
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    if units == "inch":
        return number * 25.4
    if units == "mil":
        return number * 0.0254
    return number


def _normalize_side(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"bottom", "bot", "b", "back", "bottomlayer", "底层", "底面"} or "bottom" in text:
        return "bottom"
    return "top"


def _is_dnp(value: Any, header: str | None = None) -> bool:
    text = str(value or "").strip().lower()
    header_key = _header_key(header)
    if header_key in {"populate", "fitted", "mount"}:
        return text in {"0", "false", "no", "n", "not fitted", "dnp", "dnf", "不贴", "不装"}
    return text in {"1", "true", "yes", "y", "dnp", "dnf", "not fitted", "不贴", "不装"}


def _table_payload(table: TableData) -> dict[str, Any]:
    text = table.raw_text
    truncated = len(text) > MAX_AI_TABLE_CHARS
    return {
        "name": table.name,
        "headers": table.headers,
        "content": text[:MAX_AI_TABLE_CHARS],
        "truncated": truncated,
    }


def _ai_mapping(tables: list[TableData]) -> dict[str, Any] | None:
    try:
        return suggest_fabrication_mapping([_table_payload(table) for table in tables])
    except (MimoNotConfiguredError, MimoRequestError, ValueError, KeyError):
        return None


def _sanitize_svg_markup(
    name: str,
    markup: str,
    geometry: dict[str, float] | None = None,
) -> tuple[str | None, dict[str, Any] | None, str | None]:
    """Return inert, inline-only SVG markup suitable for client-side rendering."""
    try:
        root = ET.fromstring(markup)
    except ET.ParseError:
        return None, None, f"图层 {name} 生成的 SVG 无效"
    allowed_tags = {
        "svg", "g", "path", "circle", "rect", "line", "polyline", "polygon",
        "ellipse", "defs", "clipPath", "mask", "use",
    }
    allowed_attrs = {
        "viewBox", "width", "height", "x", "y", "x1", "x2", "y1", "y2", "cx", "cy", "r", "rx", "ry",
        "d", "points", "fill", "fill-opacity", "stroke", "stroke-width", "stroke-opacity", "opacity", "transform",
        "clip-path", "mask", "href", "id", "fill-rule", "clip-rule", "stroke-linecap", "stroke-linejoin",
    }

    def safe_attribute(attr: str, value: str) -> bool:
        lowered = value.strip().lower()
        if "javascript:" in lowered or "data:" in lowered or "http:" in lowered or "https:" in lowered:
            return False
        if attr == "href":
            return bool(re.fullmatch(r"#[A-Za-z_][A-Za-z0-9_.:-]*", value.strip()))
        if attr in {"clip-path", "mask"}:
            return bool(re.fullmatch(r"url\(#[A-Za-z_][A-Za-z0-9_.:-]*\)", value.strip()))
        if "url(" in lowered:
            return bool(re.fullmatch(r"url\(#[A-Za-z_][A-Za-z0-9_.:-]*\)", value.strip()))
        return True

    def clean(element: ET.Element) -> ET.Element | None:
        tag = element.tag.rsplit("}", 1)[-1]
        if tag not in allowed_tags:
            return None
        cleaned = ET.Element(tag)
        for key, value in element.attrib.items():
            attr = key.rsplit("}", 1)[-1]
            text = str(value)
            if attr in allowed_attrs and safe_attribute(attr, text):
                cleaned.set(attr, text)
        for child in list(element):
            child_clean = clean(child)
            if child_clean is not None:
                cleaned.append(child_clean)
        return cleaned

    cleaned_root = clean(root)
    if cleaned_root is None:
        return None, None, f"图层 {name} 未生成可用 SVG"
    bounds: dict[str, Any] = {
        "view_box": cleaned_root.attrib.get("viewBox"),
        "width": cleaned_root.attrib.get("width"),
        "height": cleaned_root.attrib.get("height"),
    }
    bounds.update(geometry or {})
    return ET.tostring(cleaned_root, encoding="unicode"), bounds, None


def _render_and_sanitize_svg(name: str, data: bytes) -> tuple[str | None, dict[str, Any] | None, str | None]:
    try:
        from pygerber.gerberx3.api.v2 import GerberFile
    except Exception:
        return None, None, "PyGerber 未安装，已保留图层元数据但未生成矢量图"
    try:
        with tempfile.TemporaryDirectory(prefix="cw-gerber-") as temp:
            source = Path(temp) / Path(name).name
            output = Path(temp) / "layer.svg"
            source.write_bytes(data)
            previous_logging_disable = logging.root.manager.disable
            logging.disable(logging.WARNING)
            try:
                parsed = GerberFile.from_file(source).parse()
                info = parsed.get_info()
                geometry = {
                    "min_x": float(info.min_x_mm),
                    "min_y": float(info.min_y_mm),
                    "max_x": float(info.max_x_mm),
                    "max_y": float(info.max_y_mm),
                    "width_mm": float(info.width_mm),
                    "height_mm": float(info.height_mm),
                }
                if geometry["width_mm"] <= 0 or geometry["height_mm"] <= 0:
                    return None, geometry, f"图层 {name} 为空，已省略矢量预览"
                parsed.render_svg(output)
            finally:
                logging.disable(previous_logging_disable)
            if output.stat().st_size > MAX_SVG_BYTES:
                return None, None, f"图层 {name} 生成的 SVG 超过 8MB，已省略矢量预览"
            markup = output.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return None, None, f"图层 {name} 渲染失败：{str(exc)[:180]}"
    return _sanitize_svg_markup(name, markup, geometry)


def parse_fabrication_package(
    path: Path,
    mapping: dict[str, Any] | None = None,
    *,
    allow_ai: bool = True,
    supplements: list[tuple[str, bytes]] | None = None,
) -> dict[str, Any]:
    files = read_safe_zip(path)
    for name, data in supplements or []:
        safe_name = Path(str(name or "supplement")).name
        if Path(safe_name).suffix.lower() not in TABLE_EXTENSIONS - {".xls"}:
            raise FabricationParseError(f"补传文件类型不支持：{safe_name}")
        if len(data) > 20 * 1024**2:
            raise FabricationParseError(f"补传表格超过 20MB：{safe_name}")
        files[f"supplements/{safe_name}"] = data
    warnings: list[str] = []
    tables: list[TableData] = []
    gerbers: list[tuple[str, bytes]] = []
    for name, data in files.items():
        suffix = Path(name).suffix.lower()
        if suffix in GERBER_EXTENSIONS:
            gerbers.append((name, data))
        elif suffix in TABLE_EXTENSIONS:
            try:
                table = read_table(name, data)
                columns = detect_columns(table.headers)
                if suffix != ".txt" or _name_role(name) or name.startswith("supplements/") or "designator" in columns:
                    tables.append(table)
            except FabricationParseError as exc:
                warnings.append(str(exc))

    explicit = mapping or {}
    table_by_name = {table.name: table for table in tables}
    bom_table = table_by_name.get(str(explicit.get("bom_file") or ""))
    cpl_table = table_by_name.get(str(explicit.get("cpl_file") or ""))
    if not bom_table:
        bom_table = next((table for table in tables if _name_role(table.name) == "bom"), None)
    if not cpl_table:
        cpl_table = next((table for table in tables if _name_role(table.name) == "cpl"), None)
    if not cpl_table:
        cpl_table = next((table for table in tables if {"designator", "x", "y"}.issubset(detect_columns(table.headers))), None)
    if not bom_table:
        bom_table = next((table for table in tables if "designator" in detect_columns(table.headers) and table is not cpl_table), None)

    ai_mapping = None
    if (not cpl_table or not bom_table) and tables and allow_ai:
        ai_mapping = _ai_mapping(tables)
        if ai_mapping:
            bom_table = bom_table or table_by_name.get(str(ai_mapping.get("bom_file") or ""))
            cpl_table = cpl_table or table_by_name.get(str(ai_mapping.get("cpl_file") or ""))
            warnings.append("本地适配器未完全识别，已使用 AI 生成表格映射建议；提交前必须人工确认")

    column_overrides = explicit.get("columns") if isinstance(explicit.get("columns"), dict) else {}
    ai_columns = ai_mapping.get("columns") if isinstance(ai_mapping, dict) and isinstance(ai_mapping.get("columns"), dict) else {}
    cpl_columns = detect_columns(cpl_table.headers) if cpl_table else {}
    bom_columns = detect_columns(bom_table.headers) if bom_table else {}
    for field, value in {**ai_columns, **column_overrides}.items():
        if value:
            if cpl_table and value in cpl_table.headers:
                cpl_columns[field] = value
            if bom_table and value in bom_table.headers:
                bom_columns[field] = value

    detected_units = "mm"
    cpl_unit_hint = str(cpl_table.raw_text if cpl_table else "").lower()
    if re.search(r"unit\s*=\s*(inch|inches)|\bunits?\s*[:=]\s*(inch|inches)", cpl_unit_hint):
        detected_units = "inch"
    elif re.search(r"unit\s*=\s*mil|\bunits?\s*[:=]\s*mil", cpl_unit_hint):
        detected_units = "mil"
    default_units = str(explicit.get("units") or (ai_mapping or {}).get("units") or detected_units).lower()
    if default_units not in {"mm", "inch", "mil"}:
        default_units = "mm"

    bom_rows: dict[str, dict[str, Any]] = {}
    if bom_table and "designator" in bom_columns:
        for row in bom_table.rows:
            for designator in split_designators(row.get(bom_columns["designator"])):
                key = normalized_designator(designator)
                bom_rows[key] = {
                    "designator": designator,
                    "value": row.get(bom_columns.get("value", "")) if bom_columns.get("value") else None,
                    "model": row.get(bom_columns.get("model", "")) if bom_columns.get("model") else None,
                    "footprint": row.get(bom_columns.get("footprint", "")) if bom_columns.get("footprint") else None,
                    "dnp": _is_dnp(
                        row.get(bom_columns.get("dnp", "")), bom_columns.get("dnp")
                    ) if bom_columns.get("dnp") else False,
                }
    elif bom_table:
        warnings.append(f"BOM {bom_table.name} 缺少可识别位号列")

    placements: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    if cpl_table and {"designator", "x", "y"}.issubset(cpl_columns):
        for row in cpl_table.rows:
            designators = split_designators(row.get(cpl_columns["designator"]))
            for designator in designators:
                key = normalized_designator(designator)
                side = _normalize_side(row.get(cpl_columns.get("side", ""))) if cpl_columns.get("side") else "top"
                stable = (side, key)
                if stable in seen:
                    warnings.append(f"CPL 出现重复位号：{side}/{designator}")
                    continue
                seen.add(stable)
                bom = bom_rows.get(key, {})
                x = _parse_number(row.get(cpl_columns["x"]), default_units)
                y = _parse_number(row.get(cpl_columns["y"]), default_units)
                rotation = _parse_number(row.get(cpl_columns.get("rotation", "")), "mm") if cpl_columns.get("rotation") else 0
                placements.append(
                    {
                        "designator": designator,
                        "designator_key": key,
                        "board_side": side,
                        "x_mm": x,
                        "y_mm": y,
                        "rotation_deg": float(rotation or 0) % 360,
                        "value": bom.get("value") or (row.get(cpl_columns.get("value", "")) if cpl_columns.get("value") else None),
                        "model": bom.get("model") or (row.get(cpl_columns.get("model", "")) if cpl_columns.get("model") else None),
                        "footprint": bom.get("footprint") or (row.get(cpl_columns.get("footprint", "")) if cpl_columns.get("footprint") else None),
                        "dnp": bool(bom.get("dnp")) or (
                            _is_dnp(row.get(cpl_columns.get("dnp", "")), cpl_columns.get("dnp"))
                            if cpl_columns.get("dnp") else False
                        ),
                        "positioned": x is not None and y is not None,
                        "match_status": "matched" if key in bom_rows else "cpl_only",
                        "source": "cpl",
                        "confidence": "ai" if ai_mapping else "deterministic",
                    }
                )
    elif cpl_table:
        warnings.append(f"CPL {cpl_table.name} 缺少位号或 X/Y 列，需要人工映射")

    placed_keys = {item["designator_key"] for item in placements}
    for key, bom in bom_rows.items():
        if key in placed_keys:
            continue
        placements.append(
            {
                **bom,
                "designator_key": key,
                "board_side": "top",
                "x_mm": None,
                "y_mm": None,
                "rotation_deg": 0,
                "positioned": False,
                "match_status": "bom_only",
                "source": "bom",
                "confidence": "ai" if ai_mapping else "deterministic",
            }
        )

    layers = []
    for name, data in gerbers[:24]:
        role, side = detect_layer_role(name, data)
        svg, svg_bounds, warning = _render_and_sanitize_svg(name, data)
        if warning:
            warnings.append(warning)
        layers.append(
            {
                "source_name": name,
                "role": role,
                "side": side,
                "svg_markup": svg,
                "bounds": svg_bounds,
                "byte_size": len(data),
                "sha256": hashlib.sha256(data).hexdigest(),
            }
        )
    if len(gerbers) > 24:
        warnings.append("Gerber 图层超过 24 个，仅渲染前 24 个")

    outline_bounds = [
        item.get("bounds") for item in layers
        if item.get("role") == "outline" and item.get("bounds") and "min_x" in item["bounds"]
    ]
    gerber_bounds = [
        item.get("bounds") for item in layers
        if item.get("bounds") and "min_x" in item["bounds"]
    ]
    geometry_bounds = outline_bounds or gerber_bounds
    positioned = [item for item in placements if item.get("positioned")]
    if geometry_bounds:
        bounds = {
            "min_x": min(float(item["min_x"]) for item in geometry_bounds),
            "min_y": min(float(item["min_y"]) for item in geometry_bounds),
            "max_x": max(float(item["max_x"]) for item in geometry_bounds),
            "max_y": max(float(item["max_y"]) for item in geometry_bounds),
        }
    elif positioned:
        xs = [float(item["x_mm"]) for item in positioned]
        ys = [float(item["y_mm"]) for item in positioned]
        margin = 5.0
        bounds = {
            "min_x": min(xs) - margin,
            "min_y": min(ys) - margin,
            "max_x": max(xs) + margin,
            "max_y": max(ys) + margin,
        }
    else:
        bounds = {"min_x": 0, "min_y": 0, "max_x": 100, "max_y": 80}

    mapping_required = bool(ai_mapping) or not cpl_table or not {"designator", "x", "y"}.issubset(cpl_columns)
    if not bom_table:
        warnings.append("未找到可识别 BOM；请补传 BOM 后确认库存匹配")
    if not cpl_table:
        warnings.append("未找到可识别 CPL/Pick-and-Place；普通 Gerber 无法可靠恢复位号坐标")
    if not gerbers:
        warnings.append("未找到可识别 Gerber 图层")
    profile = "manual"
    names = " ".join(files).lower()
    if "jlc" in names or "gerber_job" in names or (cpl_table and "pickandplace" in cpl_table.name.lower()):
        profile = "jlc-easyeda"
    elif any(Path(name).suffix.lower() == ".pos" for name in files) or "kicad" in names:
        profile = "kicad"
    elif (
        cpl_table
        and {"center x", "center y"}.issubset({_header_key(header) for header in cpl_table.headers})
        and bom_table
        and "manufacturer part number" in {_header_key(header) for header in bom_table.headers}
    ) or (cpl_table and any(header.lower() == "ref x" for header in cpl_table.headers)):
        profile = "altium"
    elif any(Path(name).suffix.lower() in {".apr", ".apr_lib", ".drr", ".ldp", ".extrep", ".rul"} for name in files):
        profile = "altium"

    return {
        "profile": profile,
        "files": [{"name": name, "size": len(data), "role": _name_role(name) or (detect_layer_role(name, data)[0] if Path(name).suffix.lower() in GERBER_EXTENSIONS else "other")} for name, data in files.items()],
        "mapping": {
            "bom_file": bom_table.name if bom_table else None,
            "cpl_file": cpl_table.name if cpl_table else None,
            "columns": {**bom_columns, **cpl_columns},
            "units": default_units,
            "ai_suggestion": ai_mapping,
        },
        "mapping_required": mapping_required,
        "placements": placements,
        "layers": layers,
        "bounds": bounds,
        "warnings": list(dict.fromkeys(warnings)),
        "summary": {
            "file_count": len(files),
            "layer_count": len(layers),
            "placement_count": len(placements),
            "positioned_count": sum(1 for item in placements if item.get("positioned")),
            "unpositioned_count": sum(1 for item in placements if not item.get("positioned")),
            "dnp_count": sum(1 for item in placements if item.get("dnp")),
            "bom_only_count": sum(1 for item in placements if item.get("match_status") == "bom_only"),
            "cpl_only_count": sum(1 for item in placements if item.get("match_status") == "cpl_only"),
        },
        "ai_assisted": bool(ai_mapping),
    }


def _isolated_parse_target(
    queue,
    source: str,
    mapping: dict[str, Any] | None,
    allow_ai: bool,
    supplements: list[tuple[str, bytes]] | None,
) -> None:
    try:
        memory_limit = int(os.getenv("FABRICATION_PARSE_MEMORY_MB", "1536")) * 1024**2
        cpu_limit = max(10, int(os.getenv("FABRICATION_PARSE_CPU_SECONDS", "90")))
        temp_limit = int(os.getenv("FABRICATION_PARSE_TEMP_MB", "768")) * 1024**2
        resource.setrlimit(resource.RLIMIT_AS, (memory_limit, memory_limit))
        resource.setrlimit(resource.RLIMIT_CPU, (cpu_limit, cpu_limit + 5))
        resource.setrlimit(resource.RLIMIT_FSIZE, (temp_limit, temp_limit))
        queue.put({
            "ok": True,
            "result": parse_fabrication_package(
                Path(source), mapping, allow_ai=allow_ai, supplements=supplements
            ),
        })
    except BaseException as exc:
        queue.put({"ok": False, "error": f"{type(exc).__name__}: {str(exc)[:1000]}"})


def parse_fabrication_package_isolated(
    path: Path,
    mapping: dict[str, Any] | None = None,
    *,
    allow_ai: bool = True,
    supplements: list[tuple[str, bytes]] | None = None,
) -> dict[str, Any]:
    """Parse one package in a resource-limited child process with a hard timeout."""
    context = multiprocessing.get_context("spawn")
    queue = context.Queue(maxsize=1)
    process = context.Process(
        target=_isolated_parse_target,
        args=(queue, str(path), mapping, allow_ai, supplements),
        daemon=True,
    )
    process.start()
    timeout = max(15, int(os.getenv("FABRICATION_PARSE_TIMEOUT_SECONDS", "120")))
    deadline = time.monotonic() + timeout
    payload = None
    while time.monotonic() < deadline:
        try:
            payload = queue.get(timeout=min(0.5, max(0.01, deadline - time.monotonic())))
            break
        except Empty:
            if not process.is_alive():
                try:
                    payload = queue.get(timeout=1)
                except Empty as exc:
                    raise FabricationParseError(
                        f"制造包解析进程异常退出（exit={process.exitcode}）"
                    ) from exc
                break
    if payload is None:
        process.kill()
        process.join(5)
        raise FabricationParseError(f"制造包解析超过 {timeout} 秒，任务已终止")
    process.join(5)
    if process.is_alive():
        process.kill()
        process.join(5)
    if not payload.get("ok"):
        raise FabricationParseError(str(payload.get("error") or "制造包解析失败"))
    return payload["result"]
