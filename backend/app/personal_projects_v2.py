from __future__ import annotations

import csv
import hashlib
import io
import os
import re
from collections import defaultdict
from datetime import date, datetime, time as datetime_time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Annotated, Literal
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from .auth import AuthContext, require_access
from .database import get_db
from .models import (
    Component,
    PersonalProjectBoardV2,
    PersonalProjectBomItemV2,
    PersonalProjectCostEventV2,
    PersonalProjectExpenseV2,
    PersonalProjectFileV2,
    PersonalProjectRiskV2,
    PersonalProjectSolderPointV2,
    PersonalProjectStatusEventV2,
    PersonalProjectV2,
    PersonalProjectVersionV2,
)
from .services.stock_ledger import record_stock_delta


router = APIRouter(prefix="/api/project-workspace", tags=["personal-project-workspace-v2"])
Protected = Annotated[AuthContext, Depends(require_access)]
SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
CODE_PATTERN = re.compile(r"^[A-Z0-9]+(?:-[A-Z0-9]+)*$")
VERSION_PATTERN = re.compile(r"^V[A-Z0-9]+(?:[._-][A-Z0-9]+)*$")

PROJECT_STATUSES = {
    "planning": "方案规划",
    "component_selection": "零件选型",
    "schematic": "原理图准备",
    "pcb_design": "PCB 设计",
    "fabricating": "打板中",
    "assembly_testing": "装配调试",
    "validated": "验证完成",
    "delivered": "已交付",
    "paused": "暂停",
    "cancelled": "取消",
}
ACTIVE_STATUSES = {
    "planning", "component_selection", "schematic", "pcb_design", "fabricating", "assembly_testing"
}
LIFECYCLE_STATUSES = (
    "planning",
    "component_selection",
    "schematic",
    "pcb_design",
    "fabricating",
    "assembly_testing",
    "validated",
    "delivered",
)
INITIAL_TIMELINE_SOURCES = {
    "create", "timeline_backfill", "timeline_estimate", "timeline_actual", "chatgpt_approval"
}
VERSION_STATUSES = {
    "designing": "设计中",
    "fabricating": "打板中",
    "assembly_testing": "装配测试",
    "passed": "验证通过",
    "failed": "验证失败",
    "retired": "已停用",
}
BOARD_STATUSES = {"assembly": "装配中", "testing": "测试中", "passed": "通过", "failed": "异常", "retired": "停用"}
EXPENSE_CATEGORIES = {
    "pcb_fabrication": "PCB 打样",
    "assembly_smt": "贴片 / 装配",
    "shipping_tax": "运费税费",
    "enclosure_mechanical": "外壳与机械",
    "tooling": "工装",
    "other": "其他",
}
RISK_SEVERITIES = {"low": "低", "medium": "中", "high": "高"}
PROJECT_FILE_ROOT = Path(os.getenv("PROJECT_V2_FILE_ROOT", "./data/project-v2-files"))
MAX_PROJECT_FILE_BYTES = 20 * 1024 * 1024
MAX_BOM_FILE_BYTES = 8 * 1024 * 1024


class ProjectCreate(BaseModel):
    project_code: str = Field(min_length=1, max_length=80)
    name: str = Field(min_length=1, max_length=200)
    description: str | None = Field(default=None, max_length=5000)
    status: str = "planning"
    start_date: date | None = None
    lifecycle_dates: dict[str, date] | None = None


class ProjectUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=200)
    description: str | None = Field(default=None, max_length=5000)
    start_date: date | None = None
    end_date: date | None = None


class StatusChange(BaseModel):
    status: str
    note: str | None = Field(default=None, max_length=1000)
    source: str = Field(default="web", max_length=32)
    clear_end_date: bool = False


class TimelineBackfill(BaseModel):
    start_date: date


class TimelineActualUpdate(BaseModel):
    lifecycle_dates: dict[str, date]


class VersionCreate(BaseModel):
    version_code: str | None = Field(default=None, max_length=40)
    change_summary: str | None = Field(default=None, max_length=3000)
    copy_from_version_id: str | None = None


class VersionUpdate(BaseModel):
    status: str | None = None
    change_summary: str | None = Field(default=None, max_length=3000)


class BomCreate(BaseModel):
    component_id: int
    quantity_per_board: int = Field(default=1, ge=1, le=10000)
    designators: str | None = Field(default=None, max_length=10000)
    note: str | None = Field(default=None, max_length=2000)


class BoardCreate(BaseModel):
    name: str | None = Field(default=None, max_length=120)
    note: str | None = Field(default=None, max_length=2000)


class BoardUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=120)
    status: str | None = None
    note: str | None = Field(default=None, max_length=2000)


class SolderAction(BaseModel):
    action: Literal["solder", "unsolder", "loss", "undo_loss"]
    expected_version: int = Field(ge=1)


class ExpenseCreate(BaseModel):
    version_id: str | None = None
    category: str
    amount: Decimal = Field(gt=0, max_digits=14, decimal_places=2)
    occurred_on: date
    vendor: str | None = Field(default=None, max_length=200)
    note: str | None = Field(default=None, max_length=2000)


class RiskCreate(BaseModel):
    severity: str = "medium"
    title: str = Field(min_length=1, max_length=240)
    detail: str | None = Field(default=None, max_length=5000)


class RiskUpdate(BaseModel):
    severity: str | None = None
    status: str | None = None
    title: str | None = Field(default=None, min_length=1, max_length=240)
    detail: str | None = Field(default=None, max_length=5000)


def new_id() -> str:
    return str(uuid4())


def today() -> date:
    return datetime.now(SHANGHAI_TZ).date()


def local_now() -> datetime:
    """Return a naive Shanghai timestamp, matching the project's SQLite DateTime columns."""
    return datetime.now(SHANGHAI_TZ).replace(tzinfo=None)


def iso_week(value: date | None) -> str | None:
    if not value:
        return None
    year, week, _ = value.isocalendar()
    return f"{year}W{week:02d}"


def validate_project_start_date(value: date) -> date:
    if value > today():
        raise HTTPException(status_code=422, detail="开始日期不能晚于今天")
    return value


def add_initial_timeline_events(
    db: Session,
    project: PersonalProjectV2,
    owner_user_id: int,
    *,
    source: str = "create",
    cutoff: datetime | None = None,
) -> list[PersonalProjectStatusEventV2]:
    """Create the initial reached lifecycle as dated, immutable audit events.

    A project may be created after work has already started. When exact dates
    are not supplied, reached primary stages are explicitly marked as estimates.
    Paused/cancelled are side states and remain one event.
    """
    effective_cutoff = cutoff or local_now()
    if project.status not in LIFECYCLE_STATUSES:
        row = PersonalProjectStatusEventV2(
            id=new_id(), project_id=project.id, from_status=None, to_status=project.status,
            source=source, note="创建项目", created_by_user_id=owner_user_id,
            created_at=effective_cutoff,
        )
        db.add(row)
        return [row]

    reached = LIFECYCLE_STATUSES[:LIFECYCLE_STATUSES.index(project.status) + 1]
    start_clock = datetime_time(hour=9) if project.start_date < effective_cutoff.date() else datetime_time.min
    start_at = datetime.combine(project.start_date, start_clock)
    if start_at > effective_cutoff:
        start_at = effective_cutoff
    span = effective_cutoff - start_at
    rows: list[PersonalProjectStatusEventV2] = []
    previous: str | None = None
    for index, status in enumerate(reached):
        fraction = index / (len(reached) - 1) if len(reached) > 1 else 0
        occurred_at = start_at + span * fraction
        is_last = index == len(reached) - 1
        if len(reached) == 1:
            event_source = source
            note = "项目立项"
        elif is_last:
            event_source = "timeline_estimate"
            note = f"创建时估算进入 {PROJECT_STATUSES[status]}"
        else:
            event_source = "timeline_estimate"
            note = "创建时按开始日期估算阶段" if index else "项目立项（估算日期）"
        row = PersonalProjectStatusEventV2(
            id=new_id(), project_id=project.id, from_status=previous, to_status=status,
            source=event_source, note=note, created_by_user_id=owner_user_id,
            created_at=occurred_at,
        )
        db.add(row)
        rows.append(row)
        previous = status
    return rows


def normalized_actual_lifecycle_dates(
    project_status: str,
    start_date: date,
    lifecycle_dates: dict[str, date] | None,
) -> list[tuple[str, date]]:
    if project_status not in LIFECYCLE_STATUSES:
        raise HTTPException(status_code=422, detail="暂停或取消状态不能设置研发节点日期")
    reached = LIFECYCLE_STATUSES[:LIFECYCLE_STATUSES.index(project_status) + 1]
    values = lifecycle_dates or {}
    unknown = [key for key in values if key not in reached]
    if unknown:
        raise HTTPException(status_code=422, detail=f"节点日期包含未到达阶段：{unknown[0]}")
    missing = [status for status in reached if status not in values]
    if missing:
        raise HTTPException(status_code=422, detail=f"请填写 {PROJECT_STATUSES[missing[0]]} 的实际日期")
    ordered = [(status, values[status]) for status in reached]
    if ordered[0][1] != start_date:
        raise HTTPException(status_code=422, detail="方案规划日期必须与项目开始日期一致")
    if any(value > today() for _, value in ordered):
        raise HTTPException(status_code=422, detail="实际节点日期不能晚于今天")
    if any(ordered[index][1] > ordered[index + 1][1] for index in range(len(ordered) - 1)):
        raise HTTPException(status_code=422, detail="实际节点日期必须按项目阶段依次递增")
    return ordered


def add_actual_timeline_events(
    db: Session,
    project: PersonalProjectV2,
    owner_user_id: int,
    lifecycle_dates: dict[str, date],
) -> list[PersonalProjectStatusEventV2]:
    ordered = normalized_actual_lifecycle_dates(project.status, project.start_date, lifecycle_dates)
    rows: list[PersonalProjectStatusEventV2] = []
    previous: str | None = None
    for index, (status, occurred_on) in enumerate(ordered):
        occurred_at = datetime.combine(occurred_on, datetime_time(hour=9)) + timedelta(minutes=index)
        note = "项目立项（实际日期）" if index == 0 else f"进入 {PROJECT_STATUSES[status]}（实际日期）"
        row = PersonalProjectStatusEventV2(
            id=new_id(), project_id=project.id, from_status=previous, to_status=status,
            source="timeline_actual", note=note, created_by_user_id=owner_user_id,
            created_at=occurred_at,
        )
        db.add(row)
        rows.append(row)
        previous = status
    return rows


def status_history_out(rows: list[PersonalProjectStatusEventV2]) -> list[dict]:
    return [{
        "id": row.id, "from_status": row.from_status, "to_status": row.to_status,
        "from_label": PROJECT_STATUSES.get(row.from_status, row.from_status),
        "to_label": PROJECT_STATUSES.get(row.to_status, row.to_status),
        "source": row.source, "note": row.note, "created_at": row.created_at,
        "occurred_precision": "date" if row.source == "timeline_actual" else "datetime",
    } for row in rows]


def lifecycle_out(project: PersonalProjectV2, rows: list[PersonalProjectStatusEventV2]) -> dict:
    first_reached: dict[str, PersonalProjectStatusEventV2] = {}
    for row in sorted(rows, key=lambda item: item.created_at):
        if row.to_status in LIFECYCLE_STATUSES and (
            row.to_status not in first_reached or row.source == "timeline_actual"
        ):
            first_reached[row.to_status] = row
    current_index = LIFECYCLE_STATUSES.index(project.status) if project.status in LIFECYCLE_STATUSES else -1
    nodes = []
    ordered_reached = sorted(first_reached.values(), key=lambda item: item.created_at)
    for index, status in enumerate(LIFECYCLE_STATUSES):
        event = first_reached.get(status)
        next_event = next((item for item in ordered_reached if event and item.created_at > event.created_at), None)
        state = (
            "current" if index == current_index
            else "completed" if event
            else "skipped" if current_index >= 0 and index < current_index
            else "upcoming"
        )
        nodes.append({
            "status": status,
            "label": PROJECT_STATUSES[status],
            "position": index + 1,
            "state": state,
            "occurred_at": event.created_at if event else None,
            "occurred_on": event.created_at.date() if event else None,
            "occurred_precision": "date" if event and event.source == "timeline_actual" else "datetime",
            "iso_week": iso_week(event.created_at.date()) if event else None,
            "ended_at": next_event.created_at if next_event else None,
            "ended_on": next_event.created_at.date() if next_event else None,
            "end_iso_week": iso_week(next_event.created_at.date()) if next_event else None,
            "duration_days": max(1, (next_event.created_at.date() - event.created_at.date()).days + 1) if event and next_event else None,
            "ongoing": bool(event and status == project.status and not next_event),
            "source": event.source if event else None,
            "note": event.note if event else None,
        })
    return {
        "nodes": nodes,
        "current_status": project.status,
        "current_label": PROJECT_STATUSES.get(project.status, project.status),
        "side_state": project.status if project.status in {"paused", "cancelled"} else None,
    }


def normalize_code(value: str) -> str:
    code = str(value or "").strip().upper()
    if not CODE_PATTERN.fullmatch(code):
        raise HTTPException(status_code=422, detail="项目编号只能包含字母、数字和连字符，且不能以连字符开头或结尾")
    return code


def normalize_version(value: str) -> str:
    code = str(value or "").strip().upper()
    if not code.startswith("V"):
        code = f"V{code}"
    if not VERSION_PATTERN.fullmatch(code):
        raise HTTPException(status_code=422, detail="PCB 版本号格式不正确")
    return code


def require_project(db: Session, project_id: str, auth: AuthContext) -> PersonalProjectV2:
    project = db.query(PersonalProjectV2).filter(
        PersonalProjectV2.id == project_id,
        PersonalProjectV2.owner_user_id == auth.user_id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return project


def require_version(db: Session, project: PersonalProjectV2, version_id: str) -> PersonalProjectVersionV2:
    version = db.query(PersonalProjectVersionV2).filter(
        PersonalProjectVersionV2.id == version_id,
        PersonalProjectVersionV2.project_id == project.id,
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="PCB 版本不存在")
    return version


def require_component(db: Session, component_id: int, auth: AuthContext) -> Component:
    component = db.query(Component).filter(
        Component.id == component_id,
        Component.owner_user_id == auth.user_id,
        Component.revoked_at.is_(None),
    ).first()
    if not component:
        raise HTTPException(status_code=404, detail="元器件不存在")
    return component


def clean_designators(value: str | None, quantity: int, prefix: str = "P") -> list[str]:
    raw = re.split(r"[,，;；\s]+", str(value or "").strip()) if value else []
    items: list[str] = []
    for item in raw:
        key = item.strip().upper()
        if key and key not in items:
            items.append(key)
    while len(items) < quantity:
        candidate = f"{prefix}{len(items) + 1}"
        if candidate not in items:
            items.append(candidate)
    return items[:quantity]


def period_out(project: PersonalProjectV2) -> dict:
    effective_end = project.end_date or today()
    days = max(1, (effective_end - project.start_date).days + 1)
    return {
        "start_date": project.start_date,
        "end_date": project.end_date,
        "start_week": iso_week(project.start_date),
        "end_week": iso_week(project.end_date),
        "current_week": iso_week(effective_end),
        "actual_days": days,
        "actual_weeks": round(days / 7, 1),
    }


def cost_summary(db: Session, project: PersonalProjectV2, version_id: str | None = None) -> dict:
    material_query = db.query(PersonalProjectCostEventV2).filter(PersonalProjectCostEventV2.project_id == project.id)
    expense_query = db.query(PersonalProjectExpenseV2).filter(
        PersonalProjectExpenseV2.project_id == project.id,
        PersonalProjectExpenseV2.archived_at.is_(None),
    )
    if version_id:
        material_query = material_query.filter(PersonalProjectCostEventV2.version_id == version_id)
        expense_query = expense_query.filter(PersonalProjectExpenseV2.version_id == version_id)
    material_events = material_query.order_by(PersonalProjectCostEventV2.created_at.asc()).all()
    expenses = expense_query.order_by(PersonalProjectExpenseV2.occurred_on.asc()).all()
    material = sum((Decimal(str(row.amount)) for row in material_events if row.amount is not None), Decimal("0"))
    direct = sum((Decimal(str(row.amount)) for row in expenses), Decimal("0"))
    unpriced = max(0, sum(int(row.quantity_delta or 0) for row in material_events if row.unpriced))

    bom_estimate = Decimal("0")
    unpriced_bom = 0
    active_version_id = version_id or project.current_version_id
    if active_version_id:
        rows = db.query(PersonalProjectBomItemV2, Component).join(
            Component, Component.id == PersonalProjectBomItemV2.component_id
        ).filter(
            PersonalProjectBomItemV2.version_id == active_version_id,
            PersonalProjectBomItemV2.archived_at.is_(None),
        ).all()
        for item, component in rows:
            if component.average_unit_price is None:
                unpriced_bom += 1
            else:
                bom_estimate += Decimal(str(component.average_unit_price)) * int(item.quantity_per_board or 0)

    category_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for expense in expenses:
        category_totals[expense.category] += Decimal(str(expense.amount))
    return {
        "bom_estimate": bom_estimate,
        "actual_material_cost": material,
        "direct_expense": direct,
        "comprehensive_cost": material + direct,
        "unpriced_material_events": unpriced,
        "unpriced_bom_items": unpriced_bom,
        "unpriced_count": unpriced + unpriced_bom,
        "expense_breakdown": [
            {"category": key, "label": EXPENSE_CATEGORIES.get(key, key), "amount": value}
            for key, value in sorted(category_totals.items(), key=lambda item: item[1], reverse=True)
        ],
    }


def version_out(db: Session, version: PersonalProjectVersionV2) -> dict:
    bom_count = db.query(func.count(PersonalProjectBomItemV2.id)).filter(
        PersonalProjectBomItemV2.version_id == version.id,
        PersonalProjectBomItemV2.archived_at.is_(None),
    ).scalar() or 0
    board_count = db.query(func.count(PersonalProjectBoardV2.id)).filter(PersonalProjectBoardV2.version_id == version.id).scalar() or 0
    total_points = db.query(func.count(PersonalProjectSolderPointV2.id)).filter(PersonalProjectSolderPointV2.version_id == version.id).scalar() or 0
    completed_points = db.query(func.count(PersonalProjectSolderPointV2.id)).filter(
        PersonalProjectSolderPointV2.version_id == version.id,
        PersonalProjectSolderPointV2.state == "soldered",
    ).scalar() or 0
    return {
        "id": version.id,
        "project_id": version.project_id,
        "sequence_number": version.sequence_number,
        "version_code": version.version_code,
        "status": version.status,
        "status_label": VERSION_STATUSES.get(version.status, version.status),
        "change_summary": version.change_summary,
        "bom_item_count": bom_count,
        "board_count": board_count,
        "solder_total": total_points,
        "soldered_count": completed_points,
        "solder_progress": round(completed_points * 100 / total_points) if total_points else 0,
        "created_at": version.created_at,
        "updated_at": version.updated_at,
    }


def project_out(db: Session, project: PersonalProjectV2) -> dict:
    version = db.get(PersonalProjectVersionV2, project.current_version_id) if project.current_version_id else None
    costs = cost_summary(db, project)
    open_risks = db.query(func.count(PersonalProjectRiskV2.id)).filter(
        PersonalProjectRiskV2.project_id == project.id,
        PersonalProjectRiskV2.status == "open",
    ).scalar() or 0
    return {
        "id": project.id,
        "project_code": project.project_code,
        "name": project.name,
        "description": project.description,
        "status": project.status,
        "status_label": PROJECT_STATUSES.get(project.status, project.status),
        "period": period_out(project),
        "current_version": version_out(db, version) if version else None,
        "cost": costs,
        "open_risk_count": open_risks,
        "archived": project.archived_at is not None,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
    }


def create_points_for_board(db: Session, board: PersonalProjectBoardV2) -> int:
    created = 0
    rows = db.query(PersonalProjectBomItemV2).filter(
        PersonalProjectBomItemV2.version_id == board.version_id,
        PersonalProjectBomItemV2.archived_at.is_(None),
    ).all()
    for row in rows:
        prefix = f"I{row.id[:4].upper()}-"
        for designator in clean_designators(row.designators, int(row.quantity_per_board or 1), prefix):
            db.add(PersonalProjectSolderPointV2(
                id=new_id(), project_id=board.project_id, version_id=board.version_id,
                board_id=board.id, bom_item_id=row.id, designator=designator,
            ))
            created += 1
    return created


def bom_out(db: Session, item: PersonalProjectBomItemV2) -> dict:
    component = db.get(Component, item.component_id)
    price = Decimal(str(component.average_unit_price)) if component and component.average_unit_price is not None else None
    total = price * int(item.quantity_per_board or 0) if price is not None else None
    return {
        "id": item.id,
        "project_id": item.project_id,
        "version_id": item.version_id,
        "component_id": item.component_id,
        "warehouse_code": component.warehouse_code if component else None,
        "name": component.name if component else "已删除器件",
        "model": component.model if component else None,
        "package": component.package if component else None,
        "inventory_quantity": int(component.quantity or 0) if component else 0,
        "average_unit_price": price,
        "quantity_per_board": item.quantity_per_board,
        "designators": clean_designators(item.designators, int(item.quantity_per_board or 1), f"I{item.id[:4].upper()}-"),
        "note": item.note,
        "estimated_cost": total,
        "unpriced": price is None,
    }


def board_out(db: Session, board: PersonalProjectBoardV2) -> dict:
    points = db.query(PersonalProjectSolderPointV2, PersonalProjectBomItemV2, Component).join(
        PersonalProjectBomItemV2, PersonalProjectBomItemV2.id == PersonalProjectSolderPointV2.bom_item_id
    ).join(Component, Component.id == PersonalProjectBomItemV2.component_id).filter(
        PersonalProjectSolderPointV2.board_id == board.id,
        PersonalProjectBomItemV2.archived_at.is_(None),
    ).order_by(PersonalProjectSolderPointV2.designator.asc()).all()
    soldered = sum(1 for point, _, _ in points if point.state == "soldered")
    return {
        "id": board.id,
        "project_id": board.project_id,
        "version_id": board.version_id,
        "board_number": board.board_number,
        "name": board.name,
        "status": board.status,
        "status_label": BOARD_STATUSES.get(board.status, board.status),
        "note": board.note,
        "point_count": len(points),
        "soldered_count": soldered,
        "progress": round(soldered * 100 / len(points)) if points else 0,
        "points": [
            {
                "id": point.id, "bom_item_id": item.id, "component_id": component.id,
                "component_name": component.name, "warehouse_code": component.warehouse_code,
                "designator": point.designator, "state": point.state,
                "state_version": point.state_version, "unit_cost_snapshot": point.unit_cost_snapshot,
            }
            for point, item, component in points
        ],
        "created_at": board.created_at,
    }


@router.get("/bootstrap")
def bootstrap(
    auth: Protected,
    db: Session = Depends(get_db),
    search: str | None = None,
    status: str | None = None,
    include_archived: bool = False,
):
    query = db.query(PersonalProjectV2).filter(PersonalProjectV2.owner_user_id == auth.user_id)
    all_projects = query.order_by(PersonalProjectV2.updated_at.desc()).all()
    visible = [row for row in all_projects if include_archived or row.archived_at is None]
    if search:
        needle = search.strip().lower()
        visible = [row for row in visible if needle in row.project_code.lower() or needle in row.name.lower()]
    if status:
        visible = [row for row in visible if row.status == status]
    rows = [project_out(db, row) for row in visible]
    all_rows = [project_out(db, row) for row in all_projects if row.archived_at is None]
    status_distribution = [
        {"status": key, "label": label, "count": sum(1 for row in all_rows if row["status"] == key)}
        for key, label in PROJECT_STATUSES.items()
        if any(row["status"] == key for row in all_rows)
    ]
    weekly: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    project_ids = [row.id for row in all_projects]
    if project_ids:
        for event in db.query(PersonalProjectCostEventV2).filter(PersonalProjectCostEventV2.project_id.in_(project_ids)).all():
            if event.amount is not None:
                weekly[iso_week(event.created_at.date()) or ""] += Decimal(str(event.amount))
        for expense in db.query(PersonalProjectExpenseV2).filter(
            PersonalProjectExpenseV2.project_id.in_(project_ids), PersonalProjectExpenseV2.archived_at.is_(None)
        ).all():
            weekly[iso_week(expense.occurred_on) or ""] += Decimal(str(expense.amount))
    expense_breakdown: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in all_rows:
        for item in row["cost"]["expense_breakdown"]:
            expense_breakdown[item["category"]] += Decimal(str(item["amount"]))
    return {
        "schema_version": "project-workspace-v2",
        "metrics": {
            "active_count": sum(1 for row in all_rows if row["status"] in ACTIVE_STATUSES),
            "paused_count": sum(1 for row in all_rows if row["status"] == "paused"),
            "completed_count": sum(1 for row in all_rows if row["status"] in {"validated", "delivered"}),
            "comprehensive_cost": sum((Decimal(str(row["cost"]["comprehensive_cost"])) for row in all_rows), Decimal("0")),
            "unpriced_count": sum(int(row["cost"]["unpriced_count"] or 0) for row in all_rows),
        },
        "status_distribution": status_distribution,
        "weekly_cost": [{"week": key, "amount": value} for key, value in sorted(weekly.items())[-10:]],
        "expense_breakdown": [
            {"category": key, "label": EXPENSE_CATEGORIES.get(key, key), "amount": value}
            for key, value in sorted(expense_breakdown.items(), key=lambda item: item[1], reverse=True)
        ],
        "projects": rows,
        "options": {
            "project_statuses": [{"value": key, "label": label} for key, label in PROJECT_STATUSES.items()],
            "version_statuses": [{"value": key, "label": label} for key, label in VERSION_STATUSES.items()],
            "board_statuses": [{"value": key, "label": label} for key, label in BOARD_STATUSES.items()],
            "expense_categories": [{"value": key, "label": label} for key, label in EXPENSE_CATEGORIES.items()],
            "risk_severities": [{"value": key, "label": label} for key, label in RISK_SEVERITIES.items()],
        },
    }


@router.post("/projects")
def create_project(payload: ProjectCreate, auth: Protected, db: Session = Depends(get_db)):
    code = normalize_code(payload.project_code)
    if payload.status not in PROJECT_STATUSES:
        raise HTTPException(status_code=422, detail="不支持的项目状态")
    if db.query(PersonalProjectV2.id).filter(PersonalProjectV2.project_code == code).first():
        raise HTTPException(status_code=409, detail="项目编号已存在")
    start_date = validate_project_start_date(payload.start_date or today())
    project = PersonalProjectV2(
        id=new_id(), owner_user_id=auth.user_id, project_code=code,
        name=payload.name.strip(), description=(payload.description or "").strip() or None,
        status=payload.status, start_date=start_date,
    )
    db.add(project)
    db.flush()
    version = PersonalProjectVersionV2(
        id=new_id(), project_id=project.id, sequence_number=1, version_code="V1", status="designing"
    )
    db.add(version)
    db.flush()
    project.current_version_id = version.id
    if payload.lifecycle_dates:
        add_actual_timeline_events(db, project, auth.user_id, payload.lifecycle_dates)
    else:
        add_initial_timeline_events(db, project, auth.user_id, source="create")
    db.commit()
    db.refresh(project)
    return project_out(db, project)


@router.get("/projects/{project_id}")
def get_project(project_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    versions = db.query(PersonalProjectVersionV2).filter(PersonalProjectVersionV2.project_id == project.id).order_by(
        PersonalProjectVersionV2.sequence_number.desc()
    ).all()
    history = db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).order_by(PersonalProjectStatusEventV2.created_at.desc()).limit(100).all()
    expenses = db.query(PersonalProjectExpenseV2).filter(
        PersonalProjectExpenseV2.project_id == project.id, PersonalProjectExpenseV2.archived_at.is_(None)
    ).order_by(PersonalProjectExpenseV2.occurred_on.desc()).all()
    risks = db.query(PersonalProjectRiskV2).filter(PersonalProjectRiskV2.project_id == project.id).order_by(
        PersonalProjectRiskV2.status.asc(), PersonalProjectRiskV2.created_at.desc()
    ).all()
    files = db.query(PersonalProjectFileV2).filter(PersonalProjectFileV2.project_id == project.id).order_by(
        PersonalProjectFileV2.created_at.desc()
    ).all()
    data = project_out(db, project)
    data.update({
        "versions": [version_out(db, row) for row in versions],
        "lifecycle": lifecycle_out(project, history),
        "status_history": status_history_out(history),
        "expenses": [{
            "id": row.id, "version_id": row.version_id, "category": row.category,
            "category_label": EXPENSE_CATEGORIES.get(row.category, row.category), "amount": row.amount,
            "occurred_on": row.occurred_on, "vendor": row.vendor, "note": row.note, "created_at": row.created_at,
        } for row in expenses],
        "risks": [{
            "id": row.id, "severity": row.severity, "severity_label": RISK_SEVERITIES.get(row.severity, row.severity),
            "status": row.status, "title": row.title, "detail": row.detail, "created_at": row.created_at,
        } for row in risks],
        "files": [{
            "id": row.id, "version_id": row.version_id, "name": row.original_name, "mime_type": row.mime_type,
            "size_bytes": row.size_bytes, "download_url": f"/project-workspace/projects/{project.id}/files/{row.id}",
            "created_at": row.created_at,
        } for row in files],
    })
    return data


@router.patch("/projects/{project_id}")
def update_project(project_id: str, payload: ProjectUpdate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    values = payload.model_dump(exclude_unset=True)
    if "name" in values:
        project.name = str(values["name"]).strip()
    if "description" in values:
        project.description = str(values["description"] or "").strip() or None
    if "start_date" in values and values["start_date"]:
        project.start_date = validate_project_start_date(values["start_date"])
    if "end_date" in values:
        project.end_date = values["end_date"]
    if project.end_date and project.end_date < project.start_date:
        raise HTTPException(status_code=422, detail="结束日期不能早于开始日期")
    db.commit()
    db.refresh(project)
    return project_out(db, project)


@router.post("/projects/{project_id}/timeline/backfill")
def backfill_project_timeline(
    project_id: str,
    payload: TimelineBackfill,
    auth: Protected,
    db: Session = Depends(get_db),
):
    """Rebuild only the untouched initial timeline using an earlier real start date."""
    project = require_project(db, project_id, auth)
    start_date = validate_project_start_date(payload.start_date)
    if project.status not in LIFECYCLE_STATUSES:
        raise HTTPException(status_code=409, detail="暂停或取消状态不能自动补齐研发时间线")
    history = db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).order_by(PersonalProjectStatusEventV2.created_at.asc()).all()
    if any(row.source not in INITIAL_TIMELINE_SOURCES for row in history):
        raise HTTPException(status_code=409, detail="项目已有手工状态变更，不能覆盖原始审计记录")
    cutoff = max((row.created_at for row in history), default=project.created_at or local_now())
    if start_date > cutoff.date():
        raise HTTPException(status_code=422, detail="开始日期不能晚于项目创建日期")
    db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).delete(synchronize_session=False)
    project.start_date = start_date
    add_initial_timeline_events(db, project, auth.user_id, source="create", cutoff=cutoff)
    db.commit()
    db.refresh(project)
    return get_project(project_id, auth, db)


@router.put("/projects/{project_id}/timeline/actual")
def replace_project_actual_timeline(
    project_id: str,
    payload: TimelineActualUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    """Replace only untouched initial/estimated nodes with user-supplied actual dates."""
    project = require_project(db, project_id, auth)
    ordered = normalized_actual_lifecycle_dates(project.status, project.start_date, payload.lifecycle_dates)
    history = db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).order_by(PersonalProjectStatusEventV2.created_at.asc()).all()
    if any(row.source not in INITIAL_TIMELINE_SOURCES for row in history):
        raise HTTPException(status_code=409, detail="项目已有后续手工状态变更，不能覆盖原始审计记录")
    db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).delete(synchronize_session=False)
    add_actual_timeline_events(db, project, auth.user_id, dict(ordered))
    db.commit()
    db.refresh(project)
    return get_project(project_id, auth, db)


@router.patch("/projects/{project_id}/timeline/actual")
def backfill_project_actual_timeline_nodes(
    project_id: str,
    payload: TimelineActualUpdate,
    auth: Protected,
    db: Session = Depends(get_db),
):
    """Add or correct individual reached lifecycle dates without rewriting audit history."""
    project = require_project(db, project_id, auth)
    history = db.query(PersonalProjectStatusEventV2).filter(
        PersonalProjectStatusEventV2.project_id == project.id
    ).order_by(PersonalProjectStatusEventV2.created_at.asc()).all()
    reached_indexes = [
        LIFECYCLE_STATUSES.index(row.to_status)
        for row in history
        if row.to_status in LIFECYCLE_STATUSES
    ]
    if project.status in LIFECYCLE_STATUSES:
        reached_indexes.append(LIFECYCLE_STATUSES.index(project.status))
    if not reached_indexes:
        raise HTTPException(status_code=409, detail="项目还没有可补录的研发节点")
    max_reached = max(reached_indexes)
    values = payload.lifecycle_dates or {}
    unknown = [status for status in values if status not in LIFECYCLE_STATUSES]
    if unknown:
        raise HTTPException(status_code=422, detail=f"未知研发节点：{unknown[0]}")
    future = [status for status, occurred_on in values.items() if occurred_on > today()]
    if future:
        raise HTTPException(status_code=422, detail="实际节点日期不能晚于今天")
    unreached = [status for status in values if LIFECYCLE_STATUSES.index(status) > max_reached]
    if unreached:
        raise HTTPException(status_code=422, detail=f"尚未到达 {PROJECT_STATUSES[unreached[0]]}，不能补录日期")
    if not values:
        raise HTTPException(status_code=422, detail="请至少填写一个节点日期")

    existing_dates: dict[str, date] = {}
    for row in history:
        if row.to_status not in LIFECYCLE_STATUSES:
            continue
        if row.to_status not in existing_dates or row.source == "timeline_actual":
            existing_dates[row.to_status] = row.created_at.date()
    effective = {**existing_dates, **values}
    ordered_known = [
        (status, effective[status])
        for status in LIFECYCLE_STATUSES[:max_reached + 1]
        if status in effective
    ]
    if any(ordered_known[index][1] > ordered_known[index + 1][1] for index in range(len(ordered_known) - 1)):
        raise HTTPException(status_code=422, detail="节点日期必须按研发阶段依次递增")

    if "planning" in values:
        project.start_date = validate_project_start_date(values["planning"])
    for status, occurred_on in values.items():
        db.query(PersonalProjectStatusEventV2).filter(
            PersonalProjectStatusEventV2.project_id == project.id,
            PersonalProjectStatusEventV2.to_status == status,
            PersonalProjectStatusEventV2.source == "timeline_actual",
        ).delete(synchronize_session=False)
        index = LIFECYCLE_STATUSES.index(status)
        db.add(PersonalProjectStatusEventV2(
            id=new_id(), project_id=project.id,
            from_status=LIFECYCLE_STATUSES[index - 1] if index else None,
            to_status=status, source="timeline_actual",
            note="项目立项（后补实际日期）" if index == 0 else f"进入 {PROJECT_STATUSES[status]}（后补实际日期）",
            created_by_user_id=auth.user_id,
            created_at=datetime.combine(occurred_on, datetime_time(hour=9)) + timedelta(minutes=index),
        ))
    db.commit()
    db.refresh(project)
    return get_project(project_id, auth, db)


@router.post("/projects/{project_id}/status")
def change_status(project_id: str, payload: StatusChange, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    if payload.status not in PROJECT_STATUSES:
        raise HTTPException(status_code=422, detail="不支持的项目状态")
    previous = project.status
    if payload.status == previous:
        return project_out(db, project)
    project.status = payload.status
    if payload.status == "validated" and not project.end_date:
        project.end_date = today()
    elif payload.status in ACTIVE_STATUSES and payload.clear_end_date:
        project.end_date = None
    db.add(PersonalProjectStatusEventV2(
        id=new_id(), project_id=project.id, from_status=previous, to_status=payload.status,
        source=payload.source, note=(payload.note or "").strip() or None, created_by_user_id=auth.user_id,
    ))
    db.commit()
    db.refresh(project)
    return project_out(db, project)


@router.post("/projects/{project_id}/archive")
def archive_project(project_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    project.archived_at = datetime.utcnow()
    db.commit()
    return {"archived": True, "project_id": project.id}


@router.post("/projects/{project_id}/restore")
def restore_project(project_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    project.archived_at = None
    db.commit()
    return project_out(db, project)


@router.post("/projects/{project_id}/versions")
def create_version(project_id: str, payload: VersionCreate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    sequence = (db.query(func.max(PersonalProjectVersionV2.sequence_number)).filter(
        PersonalProjectVersionV2.project_id == project.id
    ).scalar() or 0) + 1
    code = normalize_version(payload.version_code or f"V{sequence}")
    if sequence > 1 and not str(payload.change_summary or "").strip():
        raise HTTPException(status_code=422, detail="V2 及后续版本必须填写变更说明")
    if db.query(PersonalProjectVersionV2.id).filter(
        PersonalProjectVersionV2.project_id == project.id, PersonalProjectVersionV2.version_code == code
    ).first():
        raise HTTPException(status_code=409, detail="版本号已存在")
    version = PersonalProjectVersionV2(
        id=new_id(), project_id=project.id, sequence_number=sequence, version_code=code,
        status="designing", change_summary=(payload.change_summary or "").strip() or None,
    )
    db.add(version)
    db.flush()
    source_id = payload.copy_from_version_id or project.current_version_id
    if source_id:
        source = require_version(db, project, source_id)
        for row in db.query(PersonalProjectBomItemV2).filter(
            PersonalProjectBomItemV2.version_id == source.id,
            PersonalProjectBomItemV2.archived_at.is_(None),
        ).all():
            db.add(PersonalProjectBomItemV2(
                id=new_id(), project_id=project.id, version_id=version.id, component_id=row.component_id,
                quantity_per_board=row.quantity_per_board, designators=row.designators, note=row.note,
            ))
    project.current_version_id = version.id
    db.commit()
    db.refresh(version)
    return version_out(db, version)


@router.patch("/projects/{project_id}/versions/{version_id}")
def update_version(project_id: str, version_id: str, payload: VersionUpdate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    values = payload.model_dump(exclude_unset=True)
    if "status" in values:
        if values["status"] not in VERSION_STATUSES:
            raise HTTPException(status_code=422, detail="不支持的版本状态")
        version.status = values["status"]
    if "change_summary" in values:
        version.change_summary = str(values["change_summary"] or "").strip() or None
    project.current_version_id = version.id
    db.commit()
    return version_out(db, version)


@router.get("/projects/{project_id}/versions/{version_id}/workspace")
def version_workspace(project_id: str, version_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    bom = db.query(PersonalProjectBomItemV2).filter(
        PersonalProjectBomItemV2.version_id == version.id,
        PersonalProjectBomItemV2.archived_at.is_(None),
    ).order_by(
        PersonalProjectBomItemV2.created_at.asc()
    ).all()
    boards = db.query(PersonalProjectBoardV2).filter(PersonalProjectBoardV2.version_id == version.id).order_by(
        PersonalProjectBoardV2.board_number.asc()
    ).all()
    return {
        "version": version_out(db, version),
        "bom": [bom_out(db, row) for row in bom],
        "boards": [board_out(db, row) for row in boards],
        "cost": cost_summary(db, project, version.id),
    }


@router.get("/components")
def component_options(auth: Protected, db: Session = Depends(get_db), q: str | None = None, limit: int = Query(30, ge=1, le=100)):
    query = db.query(Component).filter(Component.owner_user_id == auth.user_id, Component.revoked_at.is_(None))
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(or_(Component.name.ilike(like), Component.model.ilike(like), Component.warehouse_code.ilike(like)))
    rows = query.order_by(Component.quantity.desc(), Component.name.asc()).limit(limit).all()
    return [{
        "id": row.id, "warehouse_code": row.warehouse_code, "name": row.name, "model": row.model,
        "package": row.package, "quantity": int(row.quantity or 0), "average_unit_price": row.average_unit_price,
    } for row in rows]


def add_bom_item(db: Session, project: PersonalProjectV2, version: PersonalProjectVersionV2, component: Component, payload: BomCreate) -> PersonalProjectBomItemV2:
    if db.query(PersonalProjectBomItemV2.id).filter(
        PersonalProjectBomItemV2.version_id == version.id,
        PersonalProjectBomItemV2.component_id == component.id,
    ).first():
        raise HTTPException(status_code=409, detail=f"{component.name} 已在当前版本 BOM 中")
    refs = clean_designators(payload.designators, payload.quantity_per_board, f"{(component.warehouse_code or 'P')[:8]}-")
    item = PersonalProjectBomItemV2(
        id=new_id(), project_id=project.id, version_id=version.id, component_id=component.id,
        quantity_per_board=payload.quantity_per_board, designators=",".join(refs), note=(payload.note or "").strip() or None,
    )
    db.add(item)
    db.flush()
    for board in db.query(PersonalProjectBoardV2).filter(PersonalProjectBoardV2.version_id == version.id).all():
        for ref in refs:
            db.add(PersonalProjectSolderPointV2(
                id=new_id(), project_id=project.id, version_id=version.id, board_id=board.id,
                bom_item_id=item.id, designator=ref,
            ))
    return item


@router.post("/projects/{project_id}/versions/{version_id}/bom")
def create_bom_item(project_id: str, version_id: str, payload: BomCreate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    component = require_component(db, payload.component_id, auth)
    item = add_bom_item(db, project, version, component, payload)
    db.commit()
    return bom_out(db, item)


@router.delete("/projects/{project_id}/versions/{version_id}/bom/{item_id}")
def delete_bom_item(project_id: str, version_id: str, item_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    require_version(db, project, version_id)
    item = db.query(PersonalProjectBomItemV2).filter(
        PersonalProjectBomItemV2.id == item_id, PersonalProjectBomItemV2.version_id == version_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM 条目不存在")
    acted = db.query(PersonalProjectSolderPointV2.id).filter(
        PersonalProjectSolderPointV2.bom_item_id == item.id,
        PersonalProjectSolderPointV2.state != "pending",
    ).first()
    if acted:
        raise HTTPException(status_code=409, detail="已有焊接或报损记录，不能删除该 BOM 条目")
    db.query(PersonalProjectSolderPointV2).filter(PersonalProjectSolderPointV2.bom_item_id == item.id).delete()
    db.delete(item)
    db.commit()
    return {"deleted": True}


def read_bom_rows(filename: str, content: bytes) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    elif suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        values = list(csv.reader(io.StringIO(text)))
    else:
        raise HTTPException(status_code=422, detail="仅支持 CSV、XLSX 或 XLSM BOM 文件")
    if not values:
        return []
    headers = [str(value or "").strip().lower().replace(" ", "") for value in values[0]]
    aliases = {
        "code": {"仓库编号", "器件编号", "物料编号", "warehousecode", "componentcode", "code"},
        "quantity": {"数量", "单板数量", "quantity", "qty"},
        "designators": {"位号", "参考位号", "designator", "designators", "reference"},
        "note": {"备注", "说明", "note", "remark"},
    }
    indexes = {key: next((i for i, header in enumerate(headers) if header in names), None) for key, names in aliases.items()}
    if indexes["code"] is None:
        raise HTTPException(status_code=422, detail="BOM 缺少仓库编号列")
    rows = []
    for number, values_row in enumerate(values[1:], start=2):
        row = list(values_row)
        def value(key):
            index = indexes[key]
            return row[index] if index is not None and index < len(row) else None
        code = str(value("code") or "").strip()
        if not code:
            continue
        try:
            quantity = max(1, int(float(value("quantity") or 1)))
        except (TypeError, ValueError):
            quantity = 1
        rows.append({"row": number, "code": code, "quantity": quantity, "designators": value("designators"), "note": value("note")})
    return rows


@router.post("/projects/{project_id}/versions/{version_id}/bom/import")
async def import_bom(project_id: str, version_id: str, auth: Protected, db: Session = Depends(get_db), file: UploadFile = File(...)):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    content = await file.read(MAX_BOM_FILE_BYTES + 1)
    if len(content) > MAX_BOM_FILE_BYTES:
        raise HTTPException(status_code=413, detail="BOM 文件不能超过 8 MB")
    rows = read_bom_rows(file.filename or "bom.xlsx", content)
    codes = {row["code"].upper() for row in rows}
    components = db.query(Component).filter(
        Component.owner_user_id == auth.user_id,
        Component.warehouse_code.in_(codes),
        Component.revoked_at.is_(None),
    ).all()
    by_code = {str(row.warehouse_code or "").upper(): row for row in components}
    created, skipped, unmatched = [], [], []
    for row in rows:
        component = by_code.get(row["code"].upper())
        if not component:
            unmatched.append({"row": row["row"], "warehouse_code": row["code"], "reason": "库存中未找到"})
            continue
        try:
            item = add_bom_item(db, project, version, component, BomCreate(
                component_id=component.id, quantity_per_board=row["quantity"],
                designators=str(row["designators"] or "") or None, note=str(row["note"] or "") or None,
            ))
            created.append(item)
        except HTTPException as exc:
            if exc.status_code == 409:
                skipped.append({"row": row["row"], "warehouse_code": row["code"], "reason": "已存在"})
            else:
                raise
    db.commit()
    return {"created": len(created), "skipped": skipped, "unmatched": unmatched, "items": [bom_out(db, row) for row in created]}


@router.post("/projects/{project_id}/versions/{version_id}/boards")
def create_board(project_id: str, version_id: str, payload: BoardCreate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    number = (db.query(func.max(PersonalProjectBoardV2.board_number)).filter(
        PersonalProjectBoardV2.version_id == version.id
    ).scalar() or 0) + 1
    board = PersonalProjectBoardV2(
        id=new_id(), project_id=project.id, version_id=version.id, board_number=number,
        name=(payload.name or "").strip() or f"{version.version_code} · 第 {number} 块板",
        note=(payload.note or "").strip() or None,
    )
    db.add(board)
    db.flush()
    create_points_for_board(db, board)
    db.commit()
    return board_out(db, board)


@router.patch("/projects/{project_id}/versions/{version_id}/boards/{board_id}")
def update_board(project_id: str, version_id: str, board_id: str, payload: BoardUpdate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    require_version(db, project, version_id)
    board = db.query(PersonalProjectBoardV2).filter(
        PersonalProjectBoardV2.id == board_id, PersonalProjectBoardV2.version_id == version_id
    ).first()
    if not board:
        raise HTTPException(status_code=404, detail="实物板不存在")
    values = payload.model_dump(exclude_unset=True)
    if "status" in values and values["status"] not in BOARD_STATUSES:
        raise HTTPException(status_code=422, detail="不支持的实物板状态")
    for key in ("name", "status", "note"):
        if key in values:
            setattr(board, key, values[key])
    db.commit()
    return board_out(db, board)


def apply_solder_transition(
    db: Session,
    project: PersonalProjectV2,
    version: PersonalProjectVersionV2,
    point: PersonalProjectSolderPointV2,
    action: str,
    expected_version: int,
    auth: AuthContext,
) -> PersonalProjectSolderPointV2:
    """Apply one audited inventory/cost transition without committing.

    Both the compact per-board controls and the Gerber workbench use this
    helper, so a click on the board canvas has exactly the same stock and cost
    semantics as the regular BOM view.
    """

    if point.state_version != expected_version:
        raise HTTPException(status_code=409, detail="焊点已在其他操作中更新，请刷新后重试")
    item = db.get(PersonalProjectBomItemV2, point.bom_item_id)
    if not item or item.project_id != project.id or item.version_id != version.id:
        raise HTTPException(status_code=404, detail="焊点对应的 BOM 条目不存在")
    component = require_component(db, item.component_id, auth)
    transitions = {
        "solder": ("pending", "soldered", -1, "solder_consume_v2", "solder"),
        "unsolder": ("soldered", "pending", 1, "solder_restore_v2", "unsolder"),
        "loss": ("pending", "lost", -1, "assembly_loss_v2", "loss"),
        "undo_loss": ("lost", "pending", 1, "assembly_loss_restore_v2", "undo_loss"),
    }
    if action not in transitions:
        raise HTTPException(status_code=422, detail="不支持的装配操作")
    expected, target, inventory_delta, movement_type, event_type = transitions[action]
    if point.state != expected:
        raise HTTPException(status_code=409, detail=f"当前焊点状态为 {point.state}，不能执行 {action}")
    if inventory_delta < 0 and int(component.quantity or 0) < 1:
        raise HTTPException(status_code=409, detail=f"{component.name} 库存不足，本次未扣减")

    original = None
    if inventory_delta > 0:
        original = db.query(PersonalProjectCostEventV2).filter(
            PersonalProjectCostEventV2.solder_point_id == point.id,
            PersonalProjectCostEventV2.quantity_delta == 1,
        ).order_by(PersonalProjectCostEventV2.created_at.desc()).first()
    snapshot = (original.unit_cost_snapshot if original else None) if inventory_delta > 0 else component.average_unit_price
    amount = Decimal(str(snapshot)) if snapshot is not None else None
    if inventory_delta > 0 and amount is not None:
        amount = -amount
    component.quantity = int(component.quantity or 0) + inventory_delta
    now = datetime.utcnow()
    if inventory_delta < 0:
        component.last_outbound_at = now
    else:
        component.last_stocked_at = now
        component.first_stocked_at = component.first_stocked_at or now
    record_stock_delta(
        db, component, inventory_delta, movement_type=movement_type,
        reason=f"项目 {project.project_code} · {point.designator}", project_id=None,
        actor_user_id=auth.user_id, source_reference=project.project_code,
    )
    point.state = target
    point.state_version += 1
    point.unit_cost_snapshot = snapshot if target in {"soldered", "lost"} else None
    point.soldered_at = now if target == "soldered" else None
    point.lost_at = now if target == "lost" else None
    db.add(PersonalProjectCostEventV2(
        id=new_id(), project_id=project.id, version_id=version.id, board_id=point.board_id,
        bom_item_id=item.id, solder_point_id=point.id, component_id=component.id,
        event_type=event_type, quantity_delta=1 if inventory_delta < 0 else -1,
        unit_cost_snapshot=snapshot, amount=amount, unpriced=snapshot is None,
        reversal_of_event_id=original.id if original and inventory_delta > 0 else None,
        created_by_user_id=auth.user_id,
    ))
    return point


@router.post("/projects/{project_id}/versions/{version_id}/boards/{board_id}/points/{point_id}/action")
def solder_action(
    project_id: str, version_id: str, board_id: str, point_id: str, payload: SolderAction,
    auth: Protected, db: Session = Depends(get_db),
):
    project = require_project(db, project_id, auth)
    version = require_version(db, project, version_id)
    point = db.query(PersonalProjectSolderPointV2).filter(
        PersonalProjectSolderPointV2.id == point_id,
        PersonalProjectSolderPointV2.board_id == board_id,
        PersonalProjectSolderPointV2.version_id == version_id,
    ).first()
    if not point:
        raise HTTPException(status_code=404, detail="焊点不存在")
    apply_solder_transition(db, project, version, point, payload.action, payload.expected_version, auth)
    db.commit()
    board = db.get(PersonalProjectBoardV2, board_id)
    return {"point": next(row for row in board_out(db, board)["points"] if row["id"] == point.id), "cost": cost_summary(db, project, version_id)}


@router.post("/projects/{project_id}/expenses")
def create_expense(project_id: str, payload: ExpenseCreate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    if payload.category not in EXPENSE_CATEGORIES:
        raise HTTPException(status_code=422, detail="不支持的费用分类")
    if payload.version_id:
        require_version(db, project, payload.version_id)
    expense = PersonalProjectExpenseV2(
        id=new_id(), project_id=project.id, version_id=payload.version_id, category=payload.category,
        amount=payload.amount, occurred_on=payload.occurred_on, vendor=(payload.vendor or "").strip() or None,
        note=(payload.note or "").strip() or None, created_by_user_id=auth.user_id,
    )
    db.add(expense)
    db.commit()
    return {"id": expense.id, "category": expense.category, "category_label": EXPENSE_CATEGORIES[expense.category], "amount": expense.amount}


@router.post("/projects/{project_id}/expenses/{expense_id}/archive")
def archive_expense(project_id: str, expense_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    expense = db.query(PersonalProjectExpenseV2).filter(
        PersonalProjectExpenseV2.id == expense_id, PersonalProjectExpenseV2.project_id == project.id
    ).first()
    if not expense:
        raise HTTPException(status_code=404, detail="费用不存在")
    expense.archived_at = datetime.utcnow()
    db.commit()
    return {"archived": True}


@router.post("/projects/{project_id}/cost/fill-unpriced")
def fill_unpriced(project_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    events = db.query(PersonalProjectCostEventV2).filter(
        PersonalProjectCostEventV2.project_id == project.id,
        PersonalProjectCostEventV2.unpriced.is_(True),
    ).order_by(PersonalProjectCostEventV2.quantity_delta.desc(), PersonalProjectCostEventV2.created_at.asc()).all()
    filled = 0
    for event in events:
        component = db.get(Component, event.component_id)
        if not component or component.average_unit_price is None:
            continue
        original = db.get(PersonalProjectCostEventV2, event.reversal_of_event_id) if event.reversal_of_event_id else None
        snapshot = original.unit_cost_snapshot if original and original.unit_cost_snapshot is not None else Decimal(str(component.average_unit_price))
        event.unit_cost_snapshot = snapshot
        event.amount = snapshot if event.quantity_delta > 0 else -snapshot
        event.unpriced = False
        filled += 1
    db.commit()
    return {"filled": filled, "cost": cost_summary(db, project)}


@router.post("/projects/{project_id}/risks")
def create_risk(project_id: str, payload: RiskCreate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    if payload.severity not in RISK_SEVERITIES:
        raise HTTPException(status_code=422, detail="风险等级不正确")
    risk = PersonalProjectRiskV2(
        id=new_id(), project_id=project.id, severity=payload.severity, title=payload.title.strip(),
        detail=(payload.detail or "").strip() or None, created_by_user_id=auth.user_id,
    )
    db.add(risk)
    db.commit()
    return {"id": risk.id, "severity": risk.severity, "status": risk.status, "title": risk.title, "detail": risk.detail}


@router.patch("/projects/{project_id}/risks/{risk_id}")
def update_risk(project_id: str, risk_id: str, payload: RiskUpdate, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    risk = db.query(PersonalProjectRiskV2).filter(
        PersonalProjectRiskV2.id == risk_id, PersonalProjectRiskV2.project_id == project.id
    ).first()
    if not risk:
        raise HTTPException(status_code=404, detail="风险不存在")
    values = payload.model_dump(exclude_unset=True)
    if "severity" in values and values["severity"] not in RISK_SEVERITIES:
        raise HTTPException(status_code=422, detail="风险等级不正确")
    if "status" in values and values["status"] not in {"open", "resolved"}:
        raise HTTPException(status_code=422, detail="风险状态不正确")
    for key, value in values.items():
        setattr(risk, key, value)
    db.commit()
    return {"id": risk.id, "severity": risk.severity, "status": risk.status, "title": risk.title, "detail": risk.detail}


def project_file_dir(auth: AuthContext, project: PersonalProjectV2) -> Path:
    directory = PROJECT_FILE_ROOT / str(auth.user_id) / project.id
    directory.mkdir(parents=True, exist_ok=True)
    return directory


@router.post("/projects/{project_id}/files")
async def upload_file(
    project_id: str, auth: Protected, db: Session = Depends(get_db),
    version_id: str | None = None, file: UploadFile = File(...),
):
    project = require_project(db, project_id, auth)
    if version_id:
        require_version(db, project, version_id)
    content = await file.read(MAX_PROJECT_FILE_BYTES + 1)
    if not content:
        raise HTTPException(status_code=422, detail="文件为空")
    if len(content) > MAX_PROJECT_FILE_BYTES:
        raise HTTPException(status_code=413, detail="单个文件不能超过 20 MB")
    original = Path(file.filename or "project-file").name[:300]
    suffix = Path(original).suffix.lower()[:12]
    file_id = new_id()
    target = project_file_dir(auth, project) / f"{file_id}{suffix}"
    target.write_bytes(content)
    row = PersonalProjectFileV2(
        id=file_id, project_id=project.id, version_id=version_id, original_name=original,
        storage_path=str(target.resolve()), mime_type=file.content_type or "application/octet-stream",
        size_bytes=len(content), sha256=hashlib.sha256(content).hexdigest(), created_by_user_id=auth.user_id,
    )
    db.add(row)
    db.commit()
    return {"id": row.id, "name": row.original_name, "size_bytes": row.size_bytes}


@router.get("/projects/{project_id}/files/{file_id}")
def download_file(project_id: str, file_id: str, auth: Protected, db: Session = Depends(get_db)):
    project = require_project(db, project_id, auth)
    row = db.query(PersonalProjectFileV2).filter(
        PersonalProjectFileV2.id == file_id, PersonalProjectFileV2.project_id == project.id
    ).first()
    if not row or not Path(row.storage_path).is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(row.storage_path, media_type=row.mime_type, filename=row.original_name)
