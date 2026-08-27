#!/usr/bin/env python3
"""Apply the 17-zone inventory taxonomy and import audited purchase rows."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import json
from pathlib import Path
import sqlite3
import sys

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import joinedload, sessionmaker


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "backend"))

from app.component_identity import allocate_component_identity, refresh_identity_snapshot  # noqa: E402
from app.models import (  # noqa: E402
    ActivityLog, AppMigration, Category, Component, ComponentIdentityRegistry,
    ComponentPriceEntry, ImportRecord, InventoryLot, OrderImportBatch, OrderImportLine,
)
from app.services.hardware_categories import (  # noqa: E402
    CATEGORY_BY_NAME, HARDWARE_CATEGORIES, classify_hardware_category,
)
from app.services.stock_ledger import record_stock_delta  # noqa: E402


MIGRATION_KEY = "v1.4.0-inventory-17-zone-and-20260827-purchases"
MONEY = Decimal("0.000001")


def money(value: object) -> Decimal:
    return Decimal(str(value or 0)).quantize(MONEY, rounding=ROUND_HALF_UP)


def read_workbook_mapping(path: Path) -> dict[str, dict]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["完整分类"]
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))]
    result: dict[str, dict] = {}
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw))
        code = str(row.get("器件ID/批次") or "").strip()
        if not code or code.startswith("NEW-"):
            continue
        category = str(row.get("收纳区") or "").strip()
        if category not in CATEGORY_BY_NAME:
            raise RuntimeError(f"{code} 的分类不属于 17 区：{category}")
        if code in result:
            raise RuntimeError(f"分类表存在重复器件编号：{code}")
        result[code] = {
            "category": category,
            "location": str(row.get("实体位置") or CATEGORY_BY_NAME[category].location).strip(),
            "subcategory": str(row.get("内部子分类") or "").strip(),
        }
    return result


def read_purchases(path: Path) -> list[dict]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(rows, list) or not rows:
        raise RuntimeError("采购数据为空")
    seen: set[tuple[str, str]] = set()
    for index, row in enumerate(rows, start=1):
        ref = str(row.get("lcsc_number") or row.get("external_ref") or "").strip().upper()
        key = (str(row.get("order_number") or "").strip(), ref)
        if not all(key) or key in seen:
            raise RuntimeError(f"采购数据第 {index} 行缺少唯一订单键或发生重复")
        if row.get("category") not in CATEGORY_BY_NAME:
            raise RuntimeError(f"采购数据第 {index} 行分类无效")
        if int(row.get("quantity") or 0) <= 0:
            raise RuntimeError(f"采购数据第 {index} 行数量无效")
        seen.add(key)
    return rows


def allocate_shipping(rows: list[dict]) -> None:
    by_order: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_order[row["order_number"]].append(row)
    for order_rows in by_order.values():
        explicit_paid = [row for row in order_rows if row.get("paid_total") is not None]
        if len(order_rows) == 1 and explicit_paid:
            row = order_rows[0]
            row["allocated_shipping"] = money(row["paid_total"]) - money(row["merchandise_total"])
            row["landed_total"] = money(row["paid_total"])
            continue
        shipping = max((money(row.get("order_shipping")) for row in order_rows), default=Decimal("0"))
        total = sum((money(row["merchandise_total"]) for row in order_rows), Decimal("0"))
        allocated = Decimal("0")
        for index, row in enumerate(sorted(order_rows, key=lambda item: str(item.get("external_ref") or item.get("lcsc_number")))):
            share = money(shipping - allocated) if index == len(order_rows) - 1 else money(shipping * money(row["merchandise_total"]) / total) if total else Decimal("0")
            row["allocated_shipping"] = share
            row["landed_total"] = money(row["merchandise_total"]) + share
            allocated += share


def backup_database(database: Path) -> Path:
    root = database.parent / "backups"
    root.mkdir(parents=True, exist_ok=True)
    target = root / f"pre-17-zone-import-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db"
    with sqlite3.connect(database) as source, sqlite3.connect(target) as destination:
        source.backup(destination)
    return target


def component_snapshot(component: Component) -> dict:
    return {
        field: getattr(component, field, None)
        for field in ("warehouse_code", "name", "model", "category_id", "parameters", "package", "quantity", "source", "lcsc_number", "tags", "status", "location", "remark", "datasheet_url")
    }


def fallback_category(component: Component, old_name: str) -> str:
    classified, _ = classify_hardware_category(
        component.name, component.model, component.description, component.parameters,
        component.package, component.tags, component.source_title, old_name,
    )
    if classified:
        return classified
    old_fallback = {
        "电阻": "贴片电阻", "电容": "MLCC", "电感": "电感/晶振", "时钟源": "电感/晶振",
        "二极管": "二极管/保护", "保护器件": "二极管/保护", "三极管": "BJT/MOS", "MOS管": "BJT/MOS",
        "电源": "电源IC", "芯片": "数字/接口IC", "接口": "数字/接口IC", "传感器": "传感器",
        "连接件": "USB/XT/线束", "开关": "开关/机电", "机电件": "开关/机电",
        "开发板": "模块/开发板/显示", "功能模块": "模块/开发板/显示", "通信模块": "模块/开发板/显示", "显示模块": "模块/开发板/显示",
        "设备": "结构/工具/电池", "散热件": "结构/工具/电池", "结构件": "结构/工具/电池", "其他": "结构/工具/电池",
    }
    return old_fallback.get(old_name, "结构/工具/电池")


def ensure_categories(db) -> tuple[dict[str, Category], list[Category]]:
    old = db.query(Category).order_by(Category.id).all()
    for index, category in enumerate(old, start=1):
        category.code_prefix = f"Z{index:02d}"[-3:]
        category.code_prefix_locked = True
    db.flush()
    result: dict[str, Category] = {}
    for definition in HARDWARE_CATEGORIES:
        category = next((item for item in old if item.name == definition.name), None)
        if not category:
            category = Category(name=definition.name)
            db.add(category)
            db.flush()
        category.color = definition.color
        category.code_prefix = definition.prefix
        category.code_prefix_locked = True
        result[definition.name] = category
    db.flush()
    return result, old


def migrate_categories(db, mapping: dict[str, dict]) -> dict:
    active = db.query(Component).options(joinedload(Component.category)).filter(Component.revoked_at.is_(None)).all()
    active_codes = {str(component.warehouse_code or "").strip() for component in active}
    missing = sorted(active_codes - set(mapping))
    extra = sorted(set(mapping) - active_codes)
    if missing or extra:
        raise RuntimeError(f"分类表与现役库存不一致：缺失 {missing[:5]}，多出 {extra[:5]}")
    old_name_by_id = {component.id: component.category.name if component.category else "" for component in db.query(Component).options(joinedload(Component.category)).all()}
    categories, old_categories = ensure_categories(db)
    exact = fallback = 0
    for component in db.query(Component).order_by(Component.id).all():
        item = mapping.get(str(component.warehouse_code or "").strip())
        category_name = item["category"] if item else fallback_category(component, old_name_by_id.get(component.id, ""))
        component.category_id = categories[category_name].id
        component.location = item["location"] if item else CATEGORY_BY_NAME[category_name].location
        if item and item.get("subcategory"):
            tags = [part.strip() for part in str(component.tags or "").replace("，", ",").split(",") if part.strip()]
            if item["subcategory"] not in tags:
                tags.append(item["subcategory"])
            component.tags = ",".join(tags)[:300]
        exact += int(item is not None)
        fallback += int(item is None)
    db.flush()
    for identity in db.query(ComponentIdentityRegistry).filter(ComponentIdentityRegistry.component_id.isnot(None)).all():
        component = db.get(Component, identity.component_id)
        if component:
            refresh_identity_snapshot(identity, component)
    target_ids = {category.id for category in categories.values()}
    for category in old_categories:
        if category.id not in target_ids:
            db.delete(category)
    db.flush()
    return {"exact": exact, "fallback": fallback, "categories": len(categories)}


def recompute_price(db, component: Component) -> None:
    entries = db.query(ComponentPriceEntry).filter_by(owner_user_id=component.owner_user_id, component_id=component.id, active=True, order_status="已发货").all()
    quantity = sum(max(0, int(entry.quantity or 0)) for entry in entries)
    if quantity:
        component.average_unit_price = money(sum((Decimal(entry.landed_total or 0) for entry in entries), Decimal("0")) / Decimal(quantity))


def import_purchases(db, rows: list[dict], owner_user_id: int) -> dict:
    allocate_shipping(rows)
    categories = {category.name: category for category in db.query(Category).all()}
    batch = OrderImportBatch(owner_user_id=owner_user_id, source_file="2026-08-27 采购单与截图", order_number="MULTI-20260827", status="active")
    db.add(batch)
    db.flush()
    created = merged = skipped = 0
    affected: set[int] = set()
    for row in rows:
        reference = str(row.get("lcsc_number") or row["external_ref"]).strip().upper()
        exists = db.query(ImportRecord).filter_by(order_number=row["order_number"], lcsc_number=reference).first()
        if exists:
            skipped += 1
            continue
        query = db.query(Component).filter(Component.owner_user_id == owner_user_id, Component.revoked_at.is_(None))
        component = query.filter(Component.lcsc_number == row.get("lcsc_number")).first() if row.get("lcsc_number") else None
        if component is None:
            component = query.filter(Component.model == row["model"]).first()
        previous = component_snapshot(component) if component else None
        quantity = int(row["quantity"])
        received_at = datetime.fromisoformat(str(row.get("received_at") or row["ordered_at"]))
        if component:
            component.quantity = int(component.quantity or 0) + quantity
            component.category_id = categories[row["category"]].id
            component.location = CATEGORY_BY_NAME[row["category"]].location
            for field in ("manufacturer", "description", "parameters", "package", "datasheet_url"):
                incoming = row.get("function") if field == "description" else row.get(field)
                if not getattr(component, field, None) and incoming:
                    setattr(component, field, incoming)
            operation = "merge"
            merged += 1
        else:
            component = Component(
                owner_user_id=owner_user_id, name=row["name"], model=row["model"], manufacturer=row.get("manufacturer"),
                description=row.get("function"), parameters=row.get("parameters"), package=row.get("package"), quantity=quantity,
                category_id=categories[row["category"]].id, source=row["source"], lcsc_number=row.get("lcsc_number"),
                tags="采购入库,17区自动分类", source_title=row["name"], part_family="component", count_mode="exact",
                status="in_stock", location=CATEGORY_BY_NAME[row["category"]].location, datasheet_url=row.get("datasheet_url"),
                remark=f"来源：{row['source']}；订单：{row['order_number']}；功能：{row.get('function') or '—'}",
                ai_status="stale", first_stocked_at=received_at, last_stocked_at=received_at,
            )
            db.add(component)
            db.flush()
            allocate_component_identity(db, component)
            operation = "create"
            created += 1
        component.last_stocked_at = received_at
        movements = record_stock_delta(
            db, component, quantity, movement_type=f"purchase_{operation}", reason="2026-08-27 采购数据直接入库",
            actor_user_id=owner_user_id, location=component.location, unit_cost=float(money(row["landed_total"]) / quantity),
            source_type="purchase", source_reference=row["order_number"],
        )
        for movement in movements:
            if movement.lot_id:
                movement_lot = db.get(InventoryLot, movement.lot_id)
                if movement_lot is not None:
                    movement_lot.received_at = received_at
        record = ImportRecord(
            owner_user_id=owner_user_id, batch_id=batch.id, order_number=row["order_number"], lcsc_number=reference,
            component_id=component.id, quantity=quantity, source_file=row["source"], source_row=row.get("source_row"),
            raw_data=json.dumps(row, ensure_ascii=False, default=str),
        )
        db.add(record)
        db.flush()
        db.add(OrderImportLine(
            owner_user_id=owner_user_id, batch_id=batch.id, import_record_id=record.id, component_id=component.id,
            source_row=row.get("source_row"), order_number=row["order_number"], lcsc_number=reference,
            operation=operation, quantity_delta=quantity,
            previous_component=json.dumps(previous, ensure_ascii=False, default=str) if previous else None,
            row_data=json.dumps(row, ensure_ascii=False, default=str), note="直接入库并记录库存批次及价格历史",
        ))
        entry = ComponentPriceEntry(
            owner_user_id=owner_user_id, component_id=component.id, order_number=row["order_number"], lcsc_number=reference,
            order_status="已发货", ordered_at=row.get("ordered_at"), quantity=quantity,
            merchandise_total=money(row["merchandise_total"]), allocated_shipping=money(row["allocated_shipping"]),
            landed_total=money(row["landed_total"]), source_file=row["source"], source_row=row.get("source_row"), active=True,
        )
        db.add(entry)
        db.add(ActivityLog(
            owner_user_id=owner_user_id, action="import.purchase.direct", entity_type="component", component_id=component.id,
            quantity_delta=quantity, summary=f"采购入库 {row['model']} x {quantity}",
            detail=json.dumps({"order_number": row["order_number"], "reference": reference, "batch_id": batch.id, "operation": operation}, ensure_ascii=False),
        ))
        affected.add(component.id)
    db.flush()
    for component_id in affected:
        recompute_price(db, db.get(Component, component_id))
    batch.created_count = created
    batch.merged_count = merged
    batch.skipped_count = skipped
    return {"created": created, "merged": merged, "skipped": skipped, "batch_id": batch.id}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", type=Path, default=REPO_ROOT / "data/component_warehouse.db")
    parser.add_argument("--classification-workbook", type=Path, required=True)
    parser.add_argument("--purchases-json", type=Path, default=REPO_ROOT / "scripts/data/2026-08-27-purchases.json")
    parser.add_argument("--owner-user-id", type=int, default=1)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    mapping = read_workbook_mapping(args.classification_workbook)
    purchases = read_purchases(args.purchases_json)
    engine = create_engine(f"sqlite:///{args.database}")
    Session = sessionmaker(bind=engine, autoflush=False)
    with Session() as db:
        active_count = db.query(Component).filter(Component.revoked_at.is_(None)).count()
        existing = db.get(AppMigration, MIGRATION_KEY)
        print(json.dumps({"mode": "apply" if args.apply else "dry-run", "active_components": active_count, "mapped_components": len(mapping), "purchase_rows": len(purchases), "already_applied": bool(existing)}, ensure_ascii=False))
        if not args.apply or existing:
            return 0
    backup = backup_database(args.database)
    with Session() as db:
        try:
            category_result = migrate_categories(db, mapping)
            import_result = import_purchases(db, purchases, args.owner_user_id)
            detail = {"category": category_result, "purchases": import_result, "backup": str(backup)}
            db.add(AppMigration(key=MIGRATION_KEY, detail=json.dumps(detail, ensure_ascii=False)))
            db.commit()
        except Exception:
            db.rollback()
            raise
    with sqlite3.connect(args.database) as connection:
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
    if integrity != "ok" or foreign_keys:
        raise RuntimeError(f"数据库校验失败：integrity={integrity}, foreign_keys={foreign_keys[:3]}")
    print(json.dumps({"backup": str(backup), "category": category_result, "purchases": import_result, "integrity": integrity, "foreign_key_errors": len(foreign_keys)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
